"""
Hotel P&L Builder
=================

Reads a folder of Operating Statement .xls exports and builds one formatted
Excel P&L workbook per hotel, showing Actual vs Budget vs Projected for every
year you have a statement for, with large variances highlighted.

    python hotel_pl_tool.py --input "C:\\path\\to\\statements" --output "C:\\path\\out"

One source file = one hotel, one year. The year and hotel name are read out of
the report header, so files can be named anything.

Notes on the source format (fixed layout, column J holds the line-item label):
    A  PTD actual        G  PTD budget
    K  YTD actual        Q  YTD budget        N  YTD last year
The YTD-last-year column is unreliable in these exports (it repeats the PTD
value), so history is built from separate annual files instead.
"""
from __future__ import annotations

import argparse
import io
import re
import sys
import zipfile
from collections import OrderedDict, defaultdict
from pathlib import Path

import xlrd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as gcl

# --------------------------------------------------------------------------
# configuration
# --------------------------------------------------------------------------
MAX_YEARS = 10                  # keep at most this many years, most recent first
VAR_PCT = 0.10                  # flag when |actual-budget| / budget exceeds this
VAR_MIN = 5000                  # ...and the dollar swing exceeds this

COL_LABEL = 9                   # J
COL_PTD_ACT, COL_PTD_BUD = 0, 6         # A, G
COL_YTD_ACT, COL_YTD_BUD, COL_YTD_LY = 10, 16, 13   # K, Q, N

# page title (as it appears in the statement)  ->  output tab
TAB_MAP = OrderedDict([
    ("Rooms Department",      "Rooms"),
    ("Food Department",       "Food"),
    ("Beverage Department",   "Beverage"),
    ("Spa",                   "Miscellaneous"),
    ("Miscellaneous Revenue", "Miscellaneous"),
    ("Rental Income",         "Miscellaneous"),
])

# lines pulled onto the Fixed Expenses tab (matched case-insensitively)
FIXED_LINES = ["Real Estates Taxes", "Real Estate Taxes", "Insurance"]

# Summary tab, grouped. Every line comes off the statement's own Summary page.
SUMMARY_GROUPS = [
    ("REVENUE", ["Room", "Food & Beverage", "Spa Revenue", "Miscellaneous",
                 "Rental Income", "Total Revenue"]),
    ("EXPENSES", ["Total Departmental Expenses", "Total Undistributed Expenses",
                  "Non-Operating Expenses", "Total Non-OperatingExpenses"]),
    ("PROFIT", ["Total Department Profit or Loss", "Operating Profit or Loss",
                "Net Income or Loss"]),
    ("OPERATING STATISTICS", ["Rooms Available", "Rooms Sold", "A.D.R.",
                              "Occupancy", "REV PAR"]),
]
SUMMARY_LINES = [l for _, g in SUMMARY_GROUPS for l in g]
RATIO_LINES = {"a.d.r.", "occupancy", "rev par"}   # not dollars

# --------------------------------------------------------------------------
# styling
# --------------------------------------------------------------------------
FONT = "Calibri"
# palette read straight out of "Hilton Memphis - P&L (7) - CC.xlsx"
# (theme colours with their tints resolved to flat RGB)
NAVY = "17375E"        # dk2 -25%   section bands + column headers
TEAL = "93CDDD"        # accent5 +40%  total-row label
PALE = "DCE6F2"        # accent1 +80%  total-row figures
INK, WHITE, GREEN, MUTED = "000000", "FFFFFF", "14532D", "6B6B6B"

F_TITLE = Font(name=FONT, size=14, color=INK)
F_YEAR = Font(name=FONT, size=12, color=WHITE)
F_HDR = Font(name=FONT, size=12, color=WHITE)
F_SECT = Font(name=FONT, size=11, color=WHITE)
F_LBL = Font(name=FONT, size=11, color=INK)
F_TOT = Font(name=FONT, size=11, bold=True, color=INK)      # total figures
F_TOTL = Font(name=FONT, size=12, bold=True, color=INK)     # total label
F_NOTE = Font(name=FONT, size=9, italic=True, color=MUTED)

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_TEAL = PatternFill("solid", fgColor=TEAL)
FILL_PALE = PatternFill("solid", fgColor=PALE)
# Conditional formatting uses DIFFERENTIAL formats, and Excel takes a dxf solid
# fill's colour from bgColor, not fgColor. openpyxl writes only fgColor, which
# makes Excel paint the cell solid WHITE - covering the gridlines. Setting both
# start_color and end_color emits fgColor AND bgColor, so it renders correctly.
FILL_BAD = PatternFill(start_color="F8D2D2", end_color="F8D2D2", fill_type="solid")
FILL_GOOD = PatternFill(start_color="D8EDDF", end_color="D8EDDF", fill_type="solid")

# accounting format, exactly as in the CC file - no borders anywhere in this design
MONEY = r'_("$"* #,##0_);_("$"* \(#,##0\);_("$"* "-"??_);_(@_)'
RATIO = r'_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
PCT = '0.0%'
COUNT = '#,##0'
COUNT_LINES = {"rooms available", "rooms sold"}
PCT_LINES = {"occupancy"}
W_LABEL, W_DATA = 46.3, 13
H_TITLE, H_BAND = 18.75, 15.75


# --------------------------------------------------------------------------
# parsing
# --------------------------------------------------------------------------
class Line:
    __slots__ = ("page", "section", "label", "act", "bud", "is_total")

    def __init__(self, page, section, label, act, bud, is_total):
        self.page, self.section, self.label = page, section, label
        self.act, self.bud, self.is_total = act, bud, is_total

    @property
    def key(self):
        return (self.page, self.section, self.label)


def _txt(sheet, r, c):
    try:
        v = sheet.cell_value(r, c)
    except IndexError:
        return ""
    return v.strip() if isinstance(v, str) else v


def _num(sheet, r, c):
    try:
        v = sheet.cell_value(r, c)
    except IndexError:
        return None
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def parse_statement(path=None, *, data: bytes = None, name: str = ""):
    """-> (hotel, year, [Line, ...], (month, day))  from a path or raw .xls bytes.

    The 4th value is the statement's "As of" month/day. If someone uploads twelve
    monthly exports for one year they all carry the same year, so the caller keeps
    whichever has the latest as-of date - that is the most complete YTD."""
    if data is not None:
        book = xlrd.open_workbook(file_contents=data)
        path = Path(name or "uploaded.xls")
    else:
        book = xlrd.open_workbook(str(path), on_demand=False)
    sh = book.sheet_by_index(0)

    hotel, year, asof = None, None, (0, 0)
    for r in range(min(sh.nrows, 12)):
        g = str(_txt(sh, r, 6))
        if "Property:" in g and "Operating" in g and hotel is None:
            hotel = g.split("Property:")[-1].strip()
        m = re.search(r"As of\s+(\d{1,2})/(\d{1,2})/(\d{4})", g)
        if m:
            asof = (int(m.group(1)), int(m.group(2)))
            year = int(m.group(3))
    if hotel is None:
        for r in range(min(sh.nrows, 12)):
            g = str(_txt(sh, r, 6))
            if "Property:" in g:
                hotel = g.split("Property:")[-1].strip()
                break
    if hotel is None or year is None:
        raise ValueError(f"{path.name}: could not read hotel name / year from header")
    hotel = re.sub(r"\s+", " ", hotel).strip()   # header spacing varies between exports

    lines, page_title, section = [], None, None
    seen_page_header = False
    for r in range(sh.nrows):
        if str(_txt(sh, r, 0)) == "PTD":          # column-header row = new page
            page_title, section, seen_page_header = None, None, True
            continue
        if not seen_page_header:
            continue
        label = _txt(sh, r, COL_LABEL)
        if not isinstance(label, str) or not label:
            continue
        act = _num(sh, r, COL_YTD_ACT)
        bud = _num(sh, r, COL_YTD_BUD)
        if act is None and bud is None:           # a heading, not a data row
            if page_title is None:
                page_title = label
            section = label
            continue
        if page_title is None:
            page_title = "Summary"
        low = label.lower()
        lines.append(Line(page_title, section or page_title, label,
                          act or 0.0, bud or 0.0,
                          low.startswith("total") or "profit or loss" in low
                          or low.startswith("net income")))
    return hotel, year, lines, asof


# --------------------------------------------------------------------------
# workbook writing
# --------------------------------------------------------------------------
def _year_header(ws, years, start_row, first_col=2):
    """Year banner over Actual / Budget / Projected / Var, both rows navy."""
    ws.row_dimensions[start_row].height = H_BAND
    ws.row_dimensions[start_row + 1].height = H_BAND
    hc = ws.cell(row=start_row + 1, column=1)
    hc.fill = FILL_NAVY
    for i, yr in enumerate(years):
        c0 = first_col + i * 4
        ws.merge_cells(start_row=start_row, start_column=c0,
                       end_row=start_row, end_column=c0 + 3)
        for j in range(4):
            ws.cell(row=start_row, column=c0 + j).fill = FILL_NAVY
        cell = ws.cell(row=start_row, column=c0, value=yr)
        cell.font, cell.fill = F_YEAR, FILL_NAVY
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = '0'
        for j, h in enumerate(["Actual", "Budget", "Projected", "Var vs Bud"]):
            c = ws.cell(row=start_row + 1, column=c0 + j, value=h)
            c.font, c.fill = F_HDR, FILL_NAVY
            c.alignment = Alignment(horizontal="left")
            ws.column_dimensions[gcl(c0 + j)].width = W_DATA


def _write_grid(ws, rows, years, data, start_row, first_col=2, at=None):
    """rows: list of (key, label, is_total). data: {year: {key: (act, bud)}}
    `at` (optional dict) is filled with label.lower() -> row for chart wiring."""
    r = start_row
    for key, label, is_total in rows:
        if key is None:                                   # section band
            c = ws.cell(row=r, column=1, value=label)
            c.font, c.fill = F_SECT, FILL_NAVY
            r += 1
            continue
        low = label.strip().lower()
        fmt = (PCT if low in PCT_LINES else COUNT if low in COUNT_LINES
               else RATIO if low in RATIO_LINES else MONEY)
        lc = ws.cell(row=r, column=1, value=label)
        if is_total:
            lc.font, lc.fill = F_TOTL, FILL_TEAL
            ws.row_dimensions[r].height = H_BAND
        else:
            lc.font = F_LBL
        for i, yr in enumerate(years):
            c0 = first_col + i * 4
            act, bud = data.get(yr, {}).get(key, (0.0, 0.0))
            a = ws.cell(row=r, column=c0, value=act)
            b = ws.cell(row=r, column=c0 + 1, value=bud)
            p = ws.cell(row=r, column=c0 + 2, value=None)      # you fill this in
            v = ws.cell(row=r, column=c0 + 3,
                        value=f"={gcl(c0)}{r}-{gcl(c0+1)}{r}")
            for cell in (a, b, p, v):
                cell.number_format = fmt
                cell.font = F_TOT if is_total else F_LBL
                if is_total:
                    cell.fill = FILL_PALE
        if at is not None:
            at.setdefault(low, r)
        r += 1
    return r


def _apply_variance_rules(ws, years, first_data_row, last_row, first_col=2):
    for i in range(len(years)):
        c0 = first_col + i * 4
        var, bud = gcl(c0 + 3), gcl(c0 + 1)
        rng = f"{var}{first_data_row}:{var}{last_row}"
        over = (f'=AND({bud}{first_data_row}<>0,'
                f'ABS({var}{first_data_row})>{VAR_MIN},'
                f'ABS({var}{first_data_row}/{bud}{first_data_row})>{VAR_PCT},'
                f'{var}{first_data_row}<0)')
        under = over.replace(f'{var}{first_data_row}<0)', f'{var}{first_data_row}>0)')
        ws.conditional_formatting.add(rng, FormulaRule(formula=[over[1:]], fill=FILL_BAD))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[under[1:]], fill=FILL_GOOD))



def _fix_styles(raw: bytes) -> bytes:
    """Repair two things openpyxl gets wrong in styles.xml.

    1. It writes the default fill as a bare <patternFill/> with no patternType.
       Excel requires fill 0 to be patternType="none"; without it the cell is
       treated as a pattern fill with no colour and renders as an opaque white
       block that paints over the gridlines. That is the "white boxes with no
       outline" artefact.
    2. It pads 6-digit colours to 00RRGGBB - a 00 ALPHA channel, i.e. fully
       transparent. Excel's own files use FF.
    """
    zin = zipfile.ZipFile(io.BytesIO(raw))
    parts = {n: zin.read(n) for n in zin.namelist()}
    # empty cells must not carry t="n" - Excel omits the type for numbers, and a
    # numeric cell with no <v> is what makes the blank Projected column render as
    # a white block instead of an ordinary empty cell
    for name in parts:
        if name.startswith("xl/worksheets/sheet"):
            sx = parts[name].decode("utf8")
            sx = sx.replace(' t="n"', "")
            parts[name] = sx.encode("utf8")
    st = parts["xl/styles.xml"].decode("utf8")
    # any dxf solid fill missing bgColor would render white over the gridlines
    def _dxf_bg(m):
        block = m.group(0)
        if "bgColor" in block:
            return block
        fg = re.search(r'<fgColor rgb="([0-9A-Fa-f]{8})"', block)
        if not fg:
            return block
        return block.replace("</patternFill>",
                             '<bgColor rgb="%s"/></patternFill>' % fg.group(1))
    st = re.sub(r"<dxf>.*?</dxf>", _dxf_bg, st, flags=re.S)
    st = re.sub(r"<patternFill\s*/>", '<patternFill patternType="none"/>', st)
    st = re.sub(r'rgb="00([0-9A-Fa-f]{6})"',
                lambda m: 'rgb="FF%s"' % m.group(1), st)
    parts["xl/styles.xml"] = st.encode("utf8")
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zo:
        for n in zin.namelist():
            zo.writestr(n, parts[n])
    return out.getvalue()



def _axes(chart, money=True):
    """Zero-based, labelled, dollar-formatted axes.

    Two things openpyxl gets wrong by default: it leaves `delete` unset, so Excel
    hides the axis labels entirely, and it lets the value axis auto-scale. On a
    19.2M-vs-20.1M comparison auto-scaling starts the axis near 18M and turns a
    4.6% miss into a bar half the height of its neighbour. Anchoring at zero
    makes the bars proportional to the actual numbers.
    """
    from openpyxl.chart.axis import ChartLines
    chart.y_axis.delete = False
    chart.x_axis.delete = False
    chart.y_axis.scaling.min = 0            # <- bars now read to scale
    chart.y_axis.numFmt = '"$"#,##0'      # the $ on the tick labels is enough
    chart.y_axis.majorGridlines = ChartLines()
    chart.visible_cells_only = False       # feed rows are hidden - still plot them
    chart.display_blanks = "gap"



def _tab_chart(ws, years, at, label, title, feed_row, anchor):
    """One Actual-vs-Budget bar chart for this tab's headline line.

    Values are copied out of the grid as literals - a feed of formulas would
    render as an empty chart, because openpyxl cannot write cached results.
    """
    from openpyxl.chart import BarChart, Reference

    row = at.get(label.lower())
    if row is None:
        return
    hdr = feed_row
    hc = ws.cell(row=hdr, column=1, value="CHART DATA")
    hc.font, hc.fill = F_SECT, FILL_NAVY
    for j, t in enumerate(["Year", "Actual", "Budget"]):
        c = ws.cell(row=hdr + 1, column=1 + j, value=t)
        c.font, c.fill = F_HDR, FILL_NAVY
    for i, yr in enumerate(years):
        c0 = 2 + i * 4
        r = hdr + 2 + i
        ws.cell(row=r, column=1, value=yr).number_format = '0'
        ws.cell(row=r, column=1).font = F_LBL
        for j, src in enumerate((c0, c0 + 1)):
            v = ws.cell(row=row, column=src).value
            c = ws.cell(row=r, column=2 + j,
                        value=v if isinstance(v, (int, float)) else 0)
            c.number_format, c.font = MONEY, F_LBL

    for rr in range(hdr, hdr + 2 + len(years)):
        ws.row_dimensions[rr].hidden = True

    first, last = hdr + 2, hdr + 1 + len(years)
    ch = BarChart()
    ch.type = "col"
    ch.title = title
    ch.height, ch.width = 8.5, 17
    _axes(ch)
    ch.legend.position = "b"
    for ci in (2, 3):
        ch.add_data(Reference(ws, min_col=ci, min_row=first - 1, max_row=last),
                    titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=first, max_row=last))
    ws.add_chart(ch, anchor)


def _summary_value(data, order, yr, label, budget=False):
    """Look up one Summary-page line for one year. Returns a real number."""
    for key, ln in order.items():
        if ln.page == "Summary" and ln.label.lower() == label.lower():
            a, b = data.get(yr, {}).get(key, (0.0, 0.0))
            return b if budget else a
    return 0.0


def _add_charts(ws, years, data, order, start_row):
    """Chart-feed block plus the charts.

    The feed holds LITERAL numbers, not formulas. openpyxl cannot write cached
    formula results, and an Excel chart plots the cached value - so a feed built
    from formulas renders as a completely empty chart.
    """
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.marker import Marker

    METRICS = [("Total Revenue", "Total Revenue", False),
               ("Room", "Room", False),
               ("Food & Beverage", "Food & Beverage", False),
               ("Miscellaneous", "Miscellaneous", False),
               ("Rental Income", "Rental Income", False),
               ("Operating Profit", "Operating Profit or Loss", False),
               ("Net Income", "Net Income or Loss", False),
               ("Total Revenue Budget", "Total Revenue", True)]

    hdr = start_row
    hc = ws.cell(row=hdr, column=1, value="CHART DATA")
    hc.font, hc.fill = F_SECT, FILL_NAVY
    yc = ws.cell(row=hdr + 1, column=1, value="Year")
    yc.font, yc.fill = F_HDR, FILL_NAVY
    for j, (title, _lab, _b) in enumerate(METRICS):
        c = ws.cell(row=hdr + 1, column=2 + j, value=title)
        c.font, c.fill = F_HDR, FILL_NAVY
        c.alignment = Alignment(horizontal="left", wrap_text=True)
        ws.column_dimensions[gcl(2 + j)].width = 15

    for i, yr in enumerate(years):
        r = hdr + 2 + i
        ws.cell(row=r, column=1, value=yr).number_format = '0'
        ws.cell(row=r, column=1).font = F_LBL
        for j, (title, label, is_bud) in enumerate(METRICS):
            cell = ws.cell(row=r, column=2 + j,
                           value=_summary_value(data, order, yr, label, is_bud))
            cell.number_format, cell.font = MONEY, F_LBL

    for rr in range(hdr, hdr + 2 + len(years)):
        ws.row_dimensions[rr].hidden = True

    first, last = hdr + 2, hdr + 1 + len(years)
    cats = Reference(ws, min_col=1, min_row=first, max_row=last)
    anchor_row = last + 2
    single = len(years) < 2          # one data point cannot draw a line

    def place(chart, title, col_idxs, anchor, stacked=False):
        chart.title = title
        chart.height, chart.width = 8.5, 17
        _axes(chart)
        chart.legend.position = "b"
        for ci in col_idxs:
            chart.add_data(Reference(ws, min_col=ci, min_row=first - 1, max_row=last),
                           titles_from_data=True)
        chart.set_categories(cats)
        if isinstance(chart, BarChart):
            chart.type = "col"
            if stacked:
                chart.grouping, chart.overlap = "stacked", 100
        else:
            for ser in chart.series:          # markers, so single points show
                ser.marker = Marker(symbol="circle", size=7)
                ser.smooth = False
        ws.add_chart(chart, anchor)

    place(BarChart(), "TOTAL REVENUE  -  ACTUAL vs BUDGET", [2, 9], f"A{anchor_row}")
    place(BarChart(), "REVENUE MIX BY DEPARTMENT", [3, 4, 5, 6],
          f"K{anchor_row}", stacked=True)
    place(BarChart() if single else LineChart(),
          "OPERATING PROFIT & NET INCOME", [7, 8], f"A{anchor_row + 18}")


def _sheet_header(ws, hotel, title):
    ws["A1"] = f"{hotel}  -  {title}"
    ws["A1"].font = F_TITLE
    ws.row_dimensions[1].height = H_TITLE
    ws.column_dimensions["A"].width = W_LABEL
    ws.sheet_view.showGridLines = True


def build_workbook(hotel, per_year, out_path: Path, sources=None):
    # only the years we actually have a statement for - never zero-filled,
    # capped at the most recent MAX_YEARS
    years = sorted(per_year)
    span = years[-MAX_YEARS:]

    data = defaultdict(dict)
    order = OrderedDict()          # key -> (page, section, label, is_total)
    for yr in years:
        for ln in per_year[yr]:
            data[yr][ln.key] = (ln.act, ln.bud)
            order.setdefault(ln.key, ln)

    wb = Workbook()
    wb.remove(wb.active)

    # ---- Summary ---------------------------------------------------------
    ws = wb.create_sheet("Summary")
    _sheet_header(ws, hotel, "SUMMARY")
    _year_header(ws, span, 4)
    rows, used = [], set()
    for heading, labels in SUMMARY_GROUPS:
        rows.append((None, heading, False))
        for name in labels:
            for key, ln in order.items():
                if (ln.page == "Summary" and key not in used
                        and ln.label.lower() == name.lower()):
                    rows.append((key, ln.label, ln.is_total))
                    used.add(key)
                    break
    at = {}
    end = _write_grid(ws, rows, span, data, 6, at=at)
    _apply_variance_rules(ws, span, 6, end - 1)

    # margins, computed off the rows just written
    mc = ws.cell(row=end + 1, column=1, value="MARGINS")
    mc.font, mc.fill = F_SECT, FILL_NAVY
    rev, op, ni = at.get("total revenue"), at.get("operating profit or loss"),         at.get("net income or loss")
    for k, (lab, src) in enumerate([("Operating Profit Margin", op),
                                    ("Net Income Margin", ni)]):
        r = end + 2 + k
        ws.cell(row=r, column=1, value=lab).font = F_LBL
        if not (src and rev):
            continue
        for i in range(len(span)):
            for j in (0, 1):                       # actual and budget columns
                c = gcl(2 + i * 4 + j)
                cell = ws.cell(row=r, column=2 + i * 4 + j,
                               value=f'=IF({c}{rev}=0,"",{c}{src}/{c}{rev})')
                cell.number_format, cell.font = '0.0%', F_LBL
    ws.freeze_panes = "A6"
    _add_charts(ws, span, data, order, end + 5)

    # ---- department tabs -------------------------------------------------
    tabs = OrderedDict()
    for key, ln in order.items():
        tab = TAB_MAP.get(ln.page)
        if tab:
            tabs.setdefault(tab, []).append((key, ln))
    HEADLINE = {"Rooms": ("Total Room Revenue", "ROOM REVENUE  -  ACTUAL vs BUDGET"),
                "Food": ("Total Food Revenue", "FOOD REVENUE  -  ACTUAL vs BUDGET"),
                "Beverage": ("Total Beverage Revenue",
                             "BEVERAGE REVENUE  -  ACTUAL vs BUDGET"),
                "Miscellaneous": ("Total Miscellaneous Revenue",
                                  "MISC REVENUE  -  ACTUAL vs BUDGET")}
    for tab in ["Rooms", "Food", "Beverage", "Miscellaneous"]:
        items = tabs.get(tab, [])
        ws = wb.create_sheet(tab)
        _sheet_header(ws, hotel, tab.upper())
        _year_header(ws, span, 4)
        rows, cur = [], None
        for key, ln in items:
            if ln.section != cur:
                rows.append((None, ln.section, False))
                cur = ln.section
            rows.append((key, ln.label, ln.is_total))
        at_tab = {}
        end = _write_grid(ws, rows, span, data, 6, at=at_tab)
        _apply_variance_rules(ws, span, 6, max(end - 1, 6))
        ws.freeze_panes = "A6"
        lab, ttl = HEADLINE[tab]
        _tab_chart(ws, span, at_tab, lab, ttl, end + 2,
                   f"{gcl(2 + len(span) * 4 + 1)}7")

    # ---- Fixed Expenses --------------------------------------------------
    ws = wb.create_sheet("Fixed Expenses")
    _sheet_header(ws, hotel, "FIXED EXPENSES")
    _year_header(ws, span, 4)
    rows = [(None, "FIXED EXPENSES", False)]
    fx = [f.lower() for f in FIXED_LINES]
    for key, ln in order.items():
        if ln.label.lower() in fx and ln.page == "Non-Operating Expenses":
            rows.append((key, ln.label, False))
    end = _write_grid(ws, rows, span, data, 6)
    tot = end
    tl = ws.cell(row=tot, column=1, value="TOTAL FIXED EXPENSES")
    tl.font, tl.fill = F_TOTL, FILL_TEAL
    ws.row_dimensions[tot].height = H_BAND
    for i in range(len(span)):
        c0 = 2 + i * 4
        for j in range(4):
            col = gcl(c0 + j)
            c = ws.cell(row=tot, column=c0 + j,
                        value=f"=SUM({col}7:{col}{tot-1})")
            c.number_format, c.font, c.fill = MONEY, F_TOT, FILL_PALE
    _apply_variance_rules(ws, span, 6, tot)
    ws.freeze_panes = "A6"
    fx_at = {"total fixed expenses": tot}
    for i in range(len(span)):                  # the SUM row holds formulas, so
        c0 = 2 + i * 4                          # bake literals for the chart feed
        for j in (0, 1):
            ws.cell(row=tot, column=c0 + j).value = sum(
                ws.cell(row=rr, column=c0 + j).value or 0
                for rr in range(7, tot)
                if isinstance(ws.cell(row=rr, column=c0 + j).value, (int, float)))
    _tab_chart(ws, span, fx_at, "total fixed expenses",
               "FIXED EXPENSES  -  ACTUAL vs BUDGET", tot + 3,
               f"{gcl(2 + len(span) * 4 + 1)}7")

    for s in wb.worksheets:
        s.sheet_properties.tabColor = GREEN if s.title != "Summary" else "1C1C1C"
    # the chart-feed cells are formulas with no cached result; without this the
    # charts open empty until something forces a recalculation
    wb.calculation.fullCalcOnLoad = True
    tmp = io.BytesIO()
    wb.save(tmp)
    data = _fix_styles(tmp.getvalue())
    if hasattr(out_path, "write"):
        out_path.write(data)
    else:
        Path(out_path).write_bytes(data)
    return span


def main():
    ap = argparse.ArgumentParser(description="Build hotel P&L workbooks from operating statements.")
    ap.add_argument("--input", required=True, help="folder containing .xls operating statements")
    ap.add_argument("--output", default=None, help="output folder (default: alongside input)")
    a = ap.parse_args()

    src = Path(a.input)
    out = Path(a.output) if a.output else src
    out.mkdir(parents=True, exist_ok=True)

    files = sorted(p for p in src.glob("*.xls") if not p.name.startswith("~$"))
    if not files:
        sys.exit(f"no .xls files found in {src}")

    hotels = defaultdict(dict)
    srcmap = defaultdict(dict)
    asofmap = defaultdict(dict)
    for p in files:
        try:
            hotel, year, lines, asof = parse_statement(p)
        except Exception as e:                                  # noqa: BLE001
            print(f"  SKIPPED {p.name}: {e}")
            continue
        prev = asofmap[hotel].get(year)
        if prev is not None and prev >= asof:
            print(f"  read {p.name:52} -> {hotel} {year} (older as-of, ignored)")
            continue
        hotels[hotel][year] = lines
        srcmap[hotel][year] = p.name
        asofmap[hotel][year] = asof
        print(f"  read {p.name:52} -> {hotel} {year} ({len(lines)} lines)")

    for hotel, per_year in hotels.items():
        safe = re.sub(r"[^\w\- ]", "", hotel).strip() or "Hotel"
        dest = out / f"{safe} - P&L.xlsx"
        span = build_workbook(hotel, per_year, dest, srcmap[hotel])
        print(f"\n{hotel}: {len(span)} year(s)  ->  {dest.name}")
        print(f"  years included : {span}")


if __name__ == "__main__":
    main()
