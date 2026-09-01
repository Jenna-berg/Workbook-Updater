import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import column_index_from_string, get_column_letter
import io
import csv
import re
import collections
import zipfile
import datetime
import hashlib
import json
import bcrypt
from pathlib import Path
from copy import copy, deepcopy
from xml.sax.saxutils import escape
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

# ── CSV parsing ───────────────────────────────────────────────────────────────

MONTH_ABBR = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}
DAILY_RE = re.compile(r"^\d{1,2}[/-]\d{1,2}[/-]\d{4}")


def is_formula(value) -> bool:
    return isinstance(value, str) and value.strip().startswith("=")


def is_datelike(value) -> bool:
    return isinstance(value, (datetime.datetime, datetime.date))


def parse_csv(file_bytes: bytes) -> pd.DataFrame:
    raw = pd.read_csv(io.BytesIO(file_bytes), header=None, dtype=str, encoding="utf-8-sig")
    df = raw.iloc[2:].reset_index(drop=True)
    df.columns = range(df.shape[1])
    return df


def classify_row(date_str: str):
    """Return ('daily', date) | ('monthly', (year, month)) | (None, None)"""
    if not isinstance(date_str, str):
        return None, None
    date_str = date_str.strip()
    if DAILY_RE.match(date_str):
        raw = date_str[:10].replace("/", "-")
        try:
            d = datetime.datetime.strptime(raw, "%m-%d-%Y").date()
            return "daily", d
        except ValueError:
            return None, None
    parts = date_str.split()
    if len(parts) == 2 and parts[0][:3].lower() in MONTH_ABBR:
        try:
            month = MONTH_ABBR[parts[0][:3].lower()]
            year = int(parts[1])
            return "monthly", (year, month)
        except ValueError:
            pass
    return None, None


def safe_float(val):
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


# Margaritaville's PMS exports an "Occupancy Statistics" .xlsx instead of the
# standard "Business on the Books" CSV every other hotel uses. Confirmed the
# SAME export (same "DATE PRINTED" timestamp, same values) is used for both
# SR and Forecast, matching how one CSV already feeds ROB/SR/Forecast
# together for every other hotel — so this is ONE parser covering every
# field either flow needs, not a separate one per workbook type. It
# normalizes the export into a DataFrame with the exact same column
# positions as parse_csv() (0=date, 1=Rms Sold, 4=OOO, 5=Room Revenue,
# 6=ADR, 7=Grp PU TY, 8=Grp N/PU TY, 9=Grp Rev TY, 15=Trans count,
# 16=Trans Rev) — mapping confirmed against real exports — so
# STRATEGY_CSV_COLS / build_strategy_change_plan / build_forecast_change_plan
# need no changes at all.
MARGARITAVILLE_SOURCE_FIELDS = {
    "rms sold":     1,   # -> Forecast Rooms Sold (both future & actual)
    "ooo rms":      4,   # -> SR ooo_rms
    "room revenue": 5,   # -> Forecast Revenue (actual/past dates)
    "adr":          6,   # -> Forecast ADR OTB (future dates)
    "grp pkup rms": 7,   # -> SR grp_pu_ty
    "grp rem":      8,   # -> SR grp_npu_ty ("remaining" = not yet picked up)
    "grp rm rev":   9,   # -> SR grp_rev_ty
    "trans rms":    15,  # -> SR otb_trans
    "trans rm rev": 16,  # -> SR trans_rev_ty
}


def parse_margaritaville_source(file_bytes: bytes) -> pd.DataFrame:
    """Parse Margaritaville's 'Occupancy Statistics' PMS export (feeds ROB*/
    SR/Forecast — see MARGARITAVILLE_SOURCE_FIELDS). Detects the header row
    and field columns by their text labels — never by color; the source
    file's color-coding was only for human reference while this mapping was
    being worked out, not something to parse at runtime (this app never uses
    cell color to find targets). Skips 'History Total' / 'Forecasted Total' /
    'Total' summary rows and the trailing filter/timestamp/hotel-name rows at
    the bottom of the sheet (any row whose date column doesn't parse as a
    real date).
    * ROB mapping not wired up yet — pending a small tweak to be confirmed.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.worksheets[0]

    header_row = None
    for r in range(1, min(ws.max_row, 30) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "history/forecasted" in v.strip().lower():
                header_row = r
                break
        if header_row:
            break
    if header_row is None:
        raise ValueError("Could not find the 'History/Forecasted' header row in the source file.")

    col_for_field = {}
    for c in range(1, ws.max_column + 1):
        label = str(ws.cell(header_row, c).value or "").strip().lower()
        for field_label, dest_col in MARGARITAVILLE_SOURCE_FIELDS.items():
            if label == field_label:
                col_for_field[dest_col] = c
    missing = [label for label, dest_col in MARGARITAVILLE_SOURCE_FIELDS.items() if dest_col not in col_for_field]
    if missing:
        raise ValueError(f"Could not find expected column(s) in source file: {', '.join(missing)}.")

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        label = str(ws.cell(r, 1).value or "").strip()
        if not label or "total" in label.lower():
            continue
        date_val = ws.cell(r, 2).value
        # The top data row is sometimes frozen/pinned in Margaritaville's
        # source (confirmed real case) and can retain a native Excel date
        # value there while every other row's date is plain text — a
        # str-only check silently dropped that one row every time.
        if isinstance(date_val, (datetime.datetime, datetime.date)):
            date_str = date_val.strftime("%m/%d/%Y")
        elif isinstance(date_val, str) and DAILY_RE.match(date_val.strip()):
            date_str = date_val.strip()
        else:
            continue
        row_data = {0: date_str}
        for dest_col, src_col in col_for_field.items():
            row_data[dest_col] = safe_float(ws.cell(r, src_col).value)
        rows.append(row_data)

    if not rows:
        raise ValueError("No daily rows found in source file.")

    max_col = max(max(r.keys()) for r in rows)
    df = pd.DataFrame(rows).reindex(columns=range(max_col + 1))
    return _add_margaritaville_monthly_totals(df)


# Columns build_rob_change_plan reads for a "monthly" row: Revenue, Room
# Nights, Grp PU, Grp N/PU, Grp Rev. Same column positions the standard
# Business on the Books CSV already provides monthly totals for directly —
# Margaritaville's source has no such totals, so they're synthesized here by
# summing the daily rows for each calendar month present in the data.
ROB_MONTHLY_SUM_COLS = [1, 5, 7, 8, 9]


def _add_margaritaville_monthly_totals(df: pd.DataFrame) -> pd.DataFrame:
    """Append one synthetic 'monthly' row (e.g. 'Jul 2026') per calendar
    month present in the daily rows, summing ROB_MONTHLY_SUM_COLS — so
    build_rob_change_plan (which only reads rows classify_row calls
    'monthly') works unchanged, the same way it already does for every other
    hotel's CSV, which provides these totals directly."""
    sums = {}  # (year, month) -> {col: running sum}
    for _, row in df.iterrows():
        date_str = str(row[0]).strip() if row[0] else ""
        kind, d = classify_row(date_str)
        if kind != "daily":
            continue
        key = (d.year, d.month)
        bucket = sums.setdefault(key, {c: 0.0 for c in ROB_MONTHLY_SUM_COLS})
        for c in ROB_MONTHLY_SUM_COLS:
            v = row.get(c)
            if v is not None and not pd.isna(v):
                bucket[c] += v

    if not sums:
        return df

    monthly_rows = []
    for (year, month), bucket in sums.items():
        month_name = datetime.date(year, month, 1).strftime("%b")
        row_data = {0: f"{month_name} {year}"}
        row_data.update(bucket)
        monthly_rows.append(row_data)

    monthly_df = pd.DataFrame(monthly_rows).reindex(columns=df.columns)
    return pd.concat([df, monthly_df], ignore_index=True)


def parse_bob_source(uploaded_file) -> pd.DataFrame:
    """Dispatch on file extension: .csv is the standard Business on the
    Books export every hotel uses; .xlsx is Margaritaville's differently-
    formatted PMS export (SR + Forecast wired up so far — ROB needs a small
    additional tweak once that's confirmed)."""
    file_bytes = uploaded_file.read()
    if uploaded_file.name.lower().endswith(".xlsx"):
        return parse_margaritaville_source(file_bytes)
    return parse_csv(file_bytes)


# ── Hilton portfolio source files ────────────────────────────────────────────
# Two exports feed a Hilton run, and neither can produce the ROB on its own:
#
#   SRP Activity      one sheet covering every Hilton property, stay-level,
#                     identified by 'Property - InnCode'
#   Group Wash        one file PER hotel, group-block level, per occupancy date
#
# Group has to come from the Wash report, not from SRP's own 'convention' SRP
# Type — that flag badly undercounts group at some properties (Kansas City
# September: 270 rooms by SRP against 1,017 by Wash). So the ROB totals are
# assembled as SRP transient + Wash pick-up rather than taken from SRP whole.

WASH_PERM_SEGMENT = "PERM"   # Market Segment marking the airline/crew blocks


def _find_header_row(df_raw, first_col_name, limit=40):
    """Row index of the header inside a raw (header=None) export.

    These exports print a filter block above the table and the number of
    filters varies between pulls — the same report has landed on row 11 and
    row 13 — so the header position must be found, never assumed.
    """
    col0 = df_raw[0].astype(str).str.strip()
    hits = df_raw.index[col0 == first_col_name]
    if len(hits) == 0:
        raise ValueError(f"Could not find a '{first_col_name}' header row in the export")
    return int(hits[0])


def _spread_stay(arrival, nights):
    """Yield each occupancy date of a stay. A stay is booked once but occupies
    a room on every night of its span, and month totals are by occupancy — a
    stay arriving 30 Aug for five nights puts two nights in August and three
    in September, not five in August."""
    for i in range(int(nights)):
        yield (arrival + datetime.timedelta(days=i)).date()


def _srp_bucket():
    return collections.defaultdict(
        lambda: collections.defaultdict(lambda: [0, 0.0]))


def parse_srp_filters(file_like):
    """The FILTERS block the SRP Activity export prints above its data.

    Worth reading, because two of those filters quietly change what the numbers
    mean and neither is visible once the rows are parsed:

      Booked Date     excludes anything booked on or after it. One real export
                      was generated on 12 Aug with this set to 5 Aug, so a whole
                      week of bookings was missing. That is what made Kansas
                      City's September read 773 rooms on the books against 1,017
                      already picked up in the wash report — group appeared to
                      exceed the total, when really the total was a week stale.

      Departure Date  excludes stays that checked out before it. Set to two days
                      before the run, it drops nearly every completed day: one
                      hotel's 1 Aug read 4 rooms against a real 245.

    Returns {"booked_to", "departure_from", "run_date", "inncodes", "lines"},
    with dates as date or None.
    """
    head = pd.read_excel(file_like, sheet_name=0, header=None, nrows=13)
    lines = []
    for r in range(len(head)):
        vals = [str(v).strip() for v in head.iloc[r].tolist()
                if str(v) != "nan" and str(v).strip()]
        if vals:
            lines.append(" ".join(vals))

    def _date_after(label):
        for line in lines:
            if line.lower().startswith(label.lower()):
                m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", line)
                if m:
                    mo, dy, yr = (int(g) for g in m.groups())
                    return datetime.date(yr, mo, dy)
        return None

    codes = []
    for line in lines:
        if line.lower().startswith("property - inncode"):
            codes = re.findall(r"\b[A-Z]{5}\b", line.split(":", 1)[-1])

    # The report carries its own run date in the data. Comparing the Booked Date
    # filter against that, rather than against today, means a file uploaded a
    # day late isn't accused of being stale.
    run_date = None
    try:
        hr = _find_header_row(head, "Stay ID")
        first = pd.read_excel(file_like, sheet_name=0, header=hr, nrows=1)
        rd = first["Run Date"].iloc[0] if "Run Date" in first.columns else None
        if isinstance(rd, (datetime.datetime, pd.Timestamp)):
            run_date = rd.date()
        elif isinstance(rd, datetime.date):
            run_date = rd
    except Exception:
        pass

    return {
        "booked_to":      _date_after("Booked Date"),
        "departure_from": _date_after("Departure Date"),
        "run_date":       run_date,
        "inncodes":       codes,
        "lines":          [l for l in lines if ":" in l and not l.startswith("Stay ID")],
    }


def srp_filter_warnings(filters):
    """Filter settings that will skew the figures, in plain terms.

    Caught at upload rather than left to show up as odd numbers in a workbook
    days later, which is how both of these were originally found.
    """
    if not filters:
        return []
    out = []
    ref = filters.get("run_date") or datetime.date.today()

    booked = filters.get("booked_to")
    if booked and (ref - booked).days >= 2:
        behind = (ref - booked).days
        out.append(
            f"**This export is missing {behind} days of bookings.** Its "
            f"'Booked Date' filter is set to {booked.day} {booked:%b} but the "
            f"report ran on {ref.day} {ref:%b}, so anything booked in between is "
            f"absent and every on-the-books total will read low. Re-run it with "
            f"Booked Date set to the day you run it.")

    return out


def parse_srp_activity(file_like):
    """SRP Activity export → {inncode: {"name", "months", "days"}}

    'months' is {(year, month): {seg: [nights, revenue]}} and 'days' the same
    keyed by occupancy date. seg is 'TOT', 'GRP', 'PRM' or 'TRN' (transient =
    everything that is neither group nor permanent). Only TRN is ultimately
    used, since group and permanent come from the Wash report, but the rest is
    kept for reconciliation display.

    The ROB wants months and the Forecast wants days. Both views are built in
    the same pass over the same stays, so the two workbooks cannot end up
    disagreeing about what a booking contributed.

    'name' is the export's own 'Property - Name', kept so a hotel can still be
    identified when its InnCode isn't one we know.
    """
    raw = pd.read_excel(file_like, sheet_name=0, header=None)
    hr = _find_header_row(raw, "Stay ID")
    df = pd.read_excel(file_like, sheet_name=0, header=hr)
    df = df[df["Stay ID"].notna()]

    out = {}
    for rec in df.to_dict("records"):
        # Every export seen so far arrives pre-filtered to live bookings, but
        # the column is there and nothing promises it always will be.
        if str(rec.get("Cancelled Flag", "")).strip().lower() in ("yes", "y", "true"):
            continue
        nights = int(safe_float(rec.get("Room Nights")) or 0)
        if nights <= 0:
            continue
        arrival = rec.get("Arrival Date")
        if not isinstance(arrival, (datetime.datetime, datetime.date)):
            continue
        per = (safe_float(rec.get("Room Revenue *")) or 0.0) / nights
        is_perm = str(rec.get("MCAT", "")).strip().lower() == "permanent"
        is_grp = str(rec.get("SRP Type", "")).strip().lower() == "convention" and not is_perm
        kind = "PRM" if is_perm else ("GRP" if is_grp else "TRN")
        inn = str(rec.get("Property - InnCode", "")).strip().upper()
        prop = out.setdefault(
            inn, {"name": "", "months": _srp_bucket(), "days": _srp_bucket()})
        if not prop["name"]:
            prop["name"] = str(rec.get("Property - Name", "") or "").strip()
        for d in _spread_stay(arrival, nights):
            for view, key in ((prop["months"], (d.year, d.month)), (prop["days"], d)):
                b = view[key]
                b["TOT"][0] += 1;   b["TOT"][1] += per
                b[kind][0] += 1;    b[kind][1] += per
    return out


def _wash_bucket():
    return collections.defaultdict(
        lambda: collections.defaultdict(
            lambda: {"pu_rooms": 0.0, "pu_rev": 0.0, "av_rooms": 0.0, "av_rev": 0.0}))


def parse_group_wash(file_like):
    """Individual Group Wash export → {"months": ..., "days": ...}

    'months' is {(year, month): {seg: {...}}} and 'days' the same keyed by
    occupancy date, built in one pass for the same reason parse_srp_activity
    builds both: the ROB reads months, the Forecast reads days, and they must
    describe the same blocks.

    seg is 'GRP' (Market Segment != PERM) or 'PRM' (== PERM); each holds
    pu_rooms / pu_rev / av_rooms / av_rev.

    'Pick Up' is what has actually been reserved out of the block and belongs
    in the ROB's column E. 'Available Block' is the unpicked remainder and
    belongs in column G ('not p/u'). Market Segment is the discriminator, not
    Forecast Group — one real property has a block named 'Group_PERM_SMRF'
    whose segment is SMRF, i.e. ordinary group business despite the name.
    """
    raw = pd.read_excel(file_like, sheet_name=0, header=None)
    hr = _find_header_row(raw, "Group Code")
    df = pd.read_excel(file_like, sheet_name=0, header=hr)
    df.columns = [re.sub(r"\s+", " ", str(c)).strip() for c in df.columns]

    months, days = _wash_bucket(), _wash_bucket()
    for rec in df.to_dict("records"):
        occ = rec.get("Occupancy Date")
        if not isinstance(occ, (datetime.datetime, datetime.date)):
            continue
        pu = safe_float(rec.get("Pick Up")) or 0.0
        av = safe_float(rec.get("Available Block")) or 0.0
        rate = safe_float(rec.get("Rate")) or 0.0
        seg = "PRM" if str(rec.get("Market Segment", "")).strip().upper() == WASH_PERM_SEGMENT else "GRP"
        day = occ.date() if isinstance(occ, datetime.datetime) else occ
        for view, key in ((months, (occ.year, occ.month)), (days, day)):
            b = view[key][seg]
            b["pu_rooms"] += pu
            b["pu_rev"] += pu * rate
            b["av_rooms"] += av
            b["av_rev"] += av * rate
    return {"months": months, "days": days}


# ── IHG portfolio source file ────────────────────────────────────────────────
# "History and Forecast Business Block" — a PDF covering the current month
# end to end, one row per date plus two Subtotal rows and a Total. The rows
# above the first Subtotal are days that have already happened; the rows below
# it are still on the books. That single split drives everything: the ROB
# takes the Total row, the Forecast takes the daily rows on both sides of it,
# and the Strategy Report takes the first Subtotal.
#
# Columns, left to right:
#   Date, Total Occ., Arr. Rms.,
#   Individual: Deduct Rms., Room Revenue, Non-D Rms., Non-Deduct Room Revenue
#   Blocks:     Deduct Rms., Room Revenue, Non-D Rms., Non-Deduct Room Revenue
#   Total Hotel: Occ %, Room Revenue, Average Rate

def _ihg_num(s):
    try:
        return float(str(s).replace(",", "").rstrip("%"))
    except (TypeError, ValueError):
        return 0.0


# ── Reading these PDFs by column title ───────────────────────────────────────
# Values are read by matching them to the heading above them, never by counting
# along the row. Counting worked only for as long as the reports kept exactly
# the columns they have today: insert one, drop one, or reorder them and every
# figure after that point silently shifts to the wrong field, which is the kind
# of error that produces a plausible-looking workbook rather than a crash.
#
# The anchor is the right edge. Both reports right-align their numbers under
# right-aligned headings, so a value's right edge sits within a point or two of
# its column's — far more reliable than centres, which drift with digit count.
#
# History and Forecast stacks its headings three deep and repeats names across
# groups ('Room Revenue' appears under Individual, Blocks and Total Hotel), so
# the spanning group label is needed to tell them apart. Business on the Books
# has one flat row of unique names.

def _ihg_text_lines(page, tolerance=2.5):
    """Words grouped into visual lines, each sorted left to right.

    Grouped by how close the tops are rather than by an exact rounded value.
    Words on one printed row are not always at an identical top — a real Total
    row in these reports drifts by a fraction of a point partway across, and
    bucketing on round(top) split it in two, silently dropping every figure
    after the break.
    """
    lines = []
    for word in sorted(page.extract_words(), key=lambda w: (w["top"], w["x0"])):
        if lines and abs(word["top"] - lines[-1][0]) <= tolerance:
            lines[-1][1].append(word)
            # track the row's running centre so a long row can't drift away
            lines[-1][0] = (lines[-1][0] * (len(lines[-1][1]) - 1)
                            + word["top"]) / len(lines[-1][1])
        else:
            lines.append([word["top"], [word]])
    return [(top, sorted(words, key=lambda w: w["x0"])) for top, words in lines]


def _ihg_is_data_line(words):
    """A dated row carrying figures — not the report date printed in the corner."""
    if not words or not re.match(r"^\d{2}-\d{2}-\d{2}$", words[0]["text"]):
        return False
    return sum(1 for w in words[1:]
               if re.fullmatch(r"[\d,]+\.?\d*%?", w["text"])) >= 6


_IHG_PAGE_FURNITURE = re.compile(
    r"\d{2}:\d{2}|^\d{2}-\d{2}-\d{2}$|Holiday|History and Forecast|"
    r"Business on the Books", re.IGNORECASE)


def _ihg_columns(page):
    """({right_edge: (group, label)}, lines, index_of_first_data_line)."""
    lines = _ihg_text_lines(page)
    first_data = next((i for i, (_t, ws) in enumerate(lines)
                       if _ihg_is_data_line(ws)), None)
    if first_data is None:
        return {}, lines, None

    header = [(t, ws) for t, ws in lines[:first_data]
              if not _IHG_PAGE_FURNITURE.search(" ".join(w["text"] for w in ws).strip())]
    if not header:
        return {}, lines, first_data

    # Columns come from the heading line the figures align to — the lowest one.
    cols = []
    for word in header[-1][1]:
        if cols and word["x0"] - cols[-1][-1]["x1"] < 6:
            cols[-1].append(word)
        else:
            cols.append([word])
    columns = {max(w["x1"] for w in grp): {
        "left": min(w["x0"] for w in grp),
        "right": max(w["x1"] for w in grp),
        "label": " ".join(w["text"] for w in grp),
        "above": [],
        "group": "",
    } for grp in cols}

    for _t, words in reversed(header[:-1]):
        pairs = [(w, min(columns.values(), key=lambda c: abs(c["right"] - w["x1"])))
                 for w in words]
        aligned = [(w, c) for w, c in pairs if abs(c["right"] - w["x1"]) <= 3]
        # Decide per row, not per word: 'Total Hotel' ends within 3pt of its own
        # column, enough to mistake the spanning group row for a stacked one if
        # a single match were taken as proof.
        if len(aligned) >= max(2, len(words) * 0.6):
            for word, col in aligned:
                col["above"].insert(0, word["text"])
            continue
        spans, run = [], []
        for word in words:
            if run and word["x0"] - run[-1]["x1"] < 8:
                run.append(word)
            else:
                if run:
                    spans.append(run)
                run = [word]
        if run:
            spans.append(run)
        centres = [((min(w["x0"] for w in s) + max(w["x1"] for w in s)) / 2,
                    " ".join(w["text"] for w in s)) for s in spans]
        if centres:
            for col in columns.values():
                mid = (col["left"] + col["right"]) / 2
                col["group"] = min(centres, key=lambda p: abs(p[0] - mid))[1]

    named = {}
    for right, col in columns.items():
        label = re.sub(r"\s+", " ", " ".join(col["above"] + [col["label"]])).strip()
        named[right] = (col["group"].strip().lower(), label.lower())
    return named, lines, first_data


def _ihg_locate(columns, wanted, source_name):
    """{field: right_edge} from {field: (group_contains, exact_label)}.

    Raises rather than guessing: a missing heading means the report has changed
    shape, and quietly carrying on would write whatever happened to be nearby.
    """
    found, missing = {}, []
    for field, (group_hint, label) in wanted.items():
        hits = [r for r, (g, lab) in columns.items()
                if lab == label and (not group_hint or group_hint in g)]
        if len(hits) == 1:
            found[field] = hits[0]
        else:
            missing.append(f"{field} ({group_hint + ' | ' if group_hint else ''}{label})"
                           + (f" — {len(hits)} matches" if hits else ""))
    if missing:
        raise ValueError(
            f"Couldn't find these columns in the {source_name} report by heading: "
            + "; ".join(missing)
            + ". Headings read: "
            + ", ".join(sorted(f"{g}|{lab}" for g, lab in columns.values()))
        )
    return found


def _ihg_row_values(words, columns):
    """{right_edge: text} for one line, each value matched to its column."""
    out = {}
    for word in words:
        nearest = min(columns, key=lambda r: abs(r - word["x1"]))
        if abs(nearest - word["x1"]) <= 6:
            out[nearest] = word["text"]
    return out


def parse_ihg_history_forecast(file_like):
    """History and Forecast Business Block PDF → parsed rows.

    Returns {report_date, days: [...], subtotals: [...], total: {...}} where
    each day carries date, total_occ, blk_rms, blk_rev, total_rev, avg_rate
    and an `is_past` flag set from the position of the first Subtotal — the
    report prints completed days above it and on-the-books days below.
    """
    import pdfplumber

    # (group contains, exact heading) for each figure taken from this report.
    wanted = {
        "total_occ": ("",            "total occ."),
        "ind_rms":   ("individual",  "deduct rms."),
        "ind_rev":   ("individual",  "deduct room revenue"),
        "blk_rms":   ("blocks",      "deduct rms."),
        "blk_rev":   ("blocks",      "deduct room revenue"),
        "occ_pct":   ("total hotel", "occ %"),
        "total_rev": ("total hotel", "room revenue"),
        "avg_rate":  ("total hotel", "average rate"),
    }

    report_date, days, subtotals, total = None, [], [], None
    seen_subtotal = False
    with pdfplumber.open(file_like) as pdf:
        for page in pdf.pages:
            if report_date is None:
                head = (page.extract_text() or "").split("\n")
                m = re.search(r"(\d{2})-(\d{2})-(\d{2})\s*$",
                              head[0].strip() if head else "")
                if m:
                    mm, dd, yy = (int(x) for x in m.groups())
                    report_date = datetime.date(2000 + yy, mm, dd)

            columns, lines, first_data = _ihg_columns(page)
            if not columns or first_data is None:
                continue
            where = _ihg_locate(columns, wanted, "History and Forecast")

            for _top, words in lines[first_data:]:
                head_text = words[0]["text"].strip()
                is_day = bool(re.match(r"^\d{2}-\d{2}-\d{2}$", head_text))
                if not is_day and head_text not in ("Subtotal", "Total"):
                    continue
                cells = _ihg_row_values(words, columns)
                rec = {f: _ihg_num(cells.get(col)) for f, col in where.items()}
                if head_text == "Subtotal":
                    subtotals.append(rec)
                    seen_subtotal = True
                elif head_text == "Total":
                    total = rec
                else:
                    mm, dd, yy = (int(x) for x in head_text.split("-"))
                    rec["date"] = datetime.date(2000 + yy, mm, dd)
                    # Everything above the first Subtotal has already happened.
                    rec["is_past"] = not seen_subtotal
                    days.append(rec)

    if total is None:
        raise ValueError(
            "No 'Total' row found — is this the History and Forecast Business "
            "Block report?")
    return {"report_date": report_date, "days": days,
            "subtotals": subtotals, "total": total}


_BOB_MONTHS = {m.upper(): i + 1 for i, m in enumerate(
    ["JAN", "FEB", "MAR", "APR", "MAY", "JUN",
     "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"])}


def parse_ihg_business_on_books(file_like):
    """Business on the Books PDF → {report_date, months: {(year, month): {...}}}

    One page per month, running twelve-odd months out from the report date,
    each ending in a Total row. Per month it yields rooms, group_rooms,
    group_rev and total_rev.

    Group rooms is Blk-R + D-Blk. That pair together equals the single
    'Blocks Deduct Rms.' column of the History and Forecast report over the
    same span, which is what lets the two reports be used side by side.

    Note the first month is partial — this report starts at the report date,
    not the 1st — so the current month must come from History and Forecast
    instead, which covers it end to end.
    """
    import pdfplumber

    # One flat heading row here, every name unique, so no group is needed.
    wanted = {
        "ind_rms":   ("", "ind-r"),
        "blk_rms":   ("", "blk-r"),
        "npu_rms":   ("", "d-blk"),
        "rooms":     ("", "t-rms"),
        "ooo":       ("", "ooo"),
        "turn":      ("", "turn"),
        "ind_rev":   ("", "ind-revenue"),
        "blk_rev":   ("", "blk-revenue"),
        "blk_avg":   ("", "blk-avg. rate"),
        "total_rev": ("", "total revenue"),
        "avg_rate":  ("", "total avg. rate"),
    }

    months, report_date = {}, None
    with pdfplumber.open(file_like) as pdf:
        for page in pdf.pages:
            text_lines = (page.extract_text() or "").split("\n")
            if report_date is None and text_lines:
                m = re.search(r"(\d{2})-(\d{2})-(\d{2})\s*$", text_lines[0].strip())
                if m:
                    mm, dd, yy = (int(x) for x in m.groups())
                    report_date = datetime.date(2000 + yy, mm, dd)

            columns, lines, first_data = _ihg_columns(page)
            if not columns or first_data is None:
                continue
            # The month this page covers is named in the heading's first cell,
            # e.g. 'AUG 26' sitting above the date column.
            month_cell = min(columns, key=lambda r: r)
            mh = re.match(r"^([a-z]{3})\s+(\d{2})$", columns[month_cell][1].strip())
            if not mh:
                continue
            mon = _BOB_MONTHS.get(mh.group(1).upper())
            if not mon:
                continue
            year = 2000 + int(mh.group(2))
            where = _ihg_locate(columns, wanted, "Business on the Books")

            entry = None
            days = []
            for _top, words in lines[first_data:]:
                head_text = words[0]["text"].strip()
                is_day = bool(re.match(r"^\d{2}-\d{2}-\d{2}$", head_text))
                if not is_day and head_text != "Total":
                    continue
                cells = _ihg_row_values(words, columns)
                v = {f: _ihg_num(cells.get(col)) for f, col in where.items()}
                if is_day:
                    mm, dd, yy = (int(x) for x in head_text.split("-"))
                    days.append(dict(v, date=datetime.date(2000 + yy, mm, dd)))
                else:
                    entry = {
                        "rooms":       v["rooms"],
                        # The ROB wants the whole block, which is the picked-up
                        # part plus the remainder still unpicked.
                        "group_rooms": v["blk_rms"] + v["npu_rms"],
                        "group_rev":   v["blk_rev"],
                        "total_rev":   v["total_rev"],
                        # D-Blk alone is that unpicked remainder, which is what
                        # the ROB's pink 'Group rooms NPU' column asks for.
                        "npu_rooms":   v["npu_rms"],
                        "blk_avg":     v["blk_avg"],
                    }
            if entry is not None:
                # Every day in this report is still on the books — it starts at
                # the report date, so nothing in it has happened yet.
                entry["days"] = days
                months[(year, mon)] = entry
    if not months:
        raise ValueError(
            "No month totals found — is this the Business on the Books report?")
    return {"report_date": report_date, "months": months}


def build_ihg_rob_plan(parsed, ws, as_of=None, bob=None):
    """ROB changes for one IHG hotel from the two IHG PDFs.

    History and Forecast covers the current month end to end and supplies it;
    Business on the Books starts at the report date, so its own current month
    is partial and is deliberately skipped in favour of that. Every later
    month comes from Business on the Books.

    Months already closed are never touched, and neither is any cell holding a
    formula.
    """
    as_of = as_of or parsed.get("report_date") or datetime.date.today()
    blocks = rob_month_blocks(ws)
    changes = [{"row": 4, "col": 5, "label": "As-of date", "month": None,
                "new_value": as_of, "skip_reason": None}]

    def put(row, col, name, value, month, source):
        if row is None:
            return
        changes.append({
            "row": row, "col": col, "label": f"{name} ({source})", "month": month,
            "new_value": value,
            "skip_reason": "formula" if is_formula(ws.cell(row, col).value) else None,
        })

    def emit(labels, month, rooms, rev, grp_rooms, grp_rev, source):
        put(labels.get("revenue"),        5, "Revenue",        round(rev, 2),         month, source)
        put(labels.get("room nights"),    5, "Room Nights",    int(round(rooms)),     month, source)
        put(labels.get("group rms sold"), 5, "Group Rms sold", int(round(grp_rooms)), month, source)
        put(labels.get("group rm rev"),   5, "Group Rm Rev",   round(grp_rev, 2),     month, source)

    def emit_npu(labels, month, m):
        """Column G — the pink 'Group rooms NPU' pair.

        Rooms are D-Blk, the part of the block still unpicked; revenue is that
        at the block's own average rate, which is how the column has been kept
        by hand (Nov: 55 x 112.16 = 6,168.80 to the cent).

        This comes from Business on the Books even for the current month:
        History and Forecast has no unpicked figure, and only the days still
        ahead can have one anyway, which is exactly the span Business on the
        Books starts from.
        """
        npu = m.get("npu_rooms", 0.0)
        put(labels.get("group rms sold"), 7, "Group NPU rms",
            int(round(npu)), month, "BoB")
        put(labels.get("group rm rev"), 7, "Group NPU rev",
            round(npu * m.get("blk_avg", 0.0), 2), month, "BoB")

    cur = blocks.get(as_of.month - 1)
    if cur:
        t = parsed["total"]
        emit(cur, as_of.month, t["total_occ"], t["total_rev"],
             t["blk_rms"], t["blk_rev"], "H&F")
        if bob:
            m = bob["months"].get((as_of.year, as_of.month))
            if m:
                emit_npu(cur, as_of.month, m)

    if bob:
        for (year, month), m in sorted(bob["months"].items()):
            if year != as_of.year or month <= as_of.month:
                continue          # current month comes from H&F; past is closed
            labels = blocks.get(month - 1)
            if not labels:
                continue
            emit(labels, month, m["rooms"], m["total_rev"],
                 m["group_rooms"], m["group_rev"], "BoB")
            emit_npu(labels, month, m)
    return changes


def build_ihg_forecast_plan(parsed, ws, wb=None):
    """Forecast changes for one IHG hotel from the same PDF.

    A day that has already happened is an actual and goes to the ACTUAL block
    as rooms + revenue; a day still ahead is on the books and goes to the OTB
    block as rooms + rate. Which side a day falls on comes from the report's
    own Subtotal split rather than from today's date, so a report pulled for
    an earlier as-of still lands correctly.

    `wb` is needed because an untouched week tab carries its dates as formulas
    pointing at WK1 rather than as literal dates; without the workbook there is
    nothing to resolve them against and no column would match.
    """
    rows = locate_forecast_rows(ws)
    if not rows:
        return []
    otb_row       = rows["otb_rooms_row"]
    adr_row       = rows["adr_otb_row"]
    act_rooms_row = rows["actual_rooms_row"]
    act_rev_row   = rows["actual_revenue_row"]

    col_of = {}
    for date_val, col in build_forecast_date_col_map(
            ws, wb=wb, date_row=rows["date_row"]).items():
        col_of[date_val.date() if isinstance(date_val, datetime.datetime) else date_val] = col

    changes = []

    def put(row, col, label, value):
        if row is None or col is None:
            return
        changes.append({
            "row": row, "col": col, "label": label, "month": None,
            "new_value": value,
            "skip_reason": "formula" if is_formula(ws.cell(row, col).value) else None,
        })

    for d in parsed["days"]:
        col = col_of.get(d["date"])
        if col is None:
            continue
        stamp = d["date"].strftime("%b %d").replace(" 0", " ")
        if d["is_past"]:
            put(act_rooms_row, col, f"{stamp} actual rooms",   int(round(d["total_occ"])))
            put(act_rev_row,   col, f"{stamp} actual revenue", round(d["total_rev"], 2))
        else:
            put(otb_row, col, f"{stamp} OTB rooms", int(round(d["total_occ"])))
            put(adr_row, col, f"{stamp} OTB ADR",   round(d["avg_rate"], 2))

    snap = find_forecast_snapshot_row(ws)
    if snap:
        for d in parsed["days"]:
            col = col_of.get(d["date"])
            if col is None:
                continue
            stamp = d["date"].strftime("%b %d").replace(" 0", " ")
            put(snap, col, f"{stamp} wk/wk rooms", int(round(d["total_occ"])))
    return changes


def detect_ihg_sr_columns(ws):
    """{field: column} for an IHG Strategy Report, read off the header.

    Headings are split across rows 3 and 4 and only the pair is unique — 'TRAN'
    alone is both the LY transient rooms column and the start of the transient
    revenue heading — so the two rows are joined before matching.
    """
    wanted = {
        ("trans", "sold"):     "trans_rms",
        ("group", "sold"):     "group_rms",
        ("prop.", "regrets"):  "regrets",
        ("grp rms", "n/pu"):   "group_npu",
        ("ooo", "rooms"):      "ooo",
        ("tran", "rev ty"):    "trans_rev",
        ("grp", "rev ty"):     "group_rev",
    }
    out = {}
    for c in range(1, 70):
        h3 = str(ws.cell(3, c).value or "").strip().lower()
        h4 = str(ws.cell(4, c).value or "").strip().lower()
        key = wanted.get((h3, h4))
        if key and key not in out:
            out[key] = c
    return out


def build_ihg_sr_plan(hf, bob, ws, ws_values=None):
    """Strategy Report changes for one IHG hotel from the two PDFs.

    Each dated row takes its figures from whichever report covers that day:
    History and Forecast for days already behind the Business on the Books
    start, Business on the Books from that date on. The boundary is read from
    the report itself rather than assumed, since it moves with the pull date.

    History and Forecast carries no OOO or turnaway columns, so for the earlier
    days those two are left as they are rather than zeroed.

    Monthly total rows never match a date and so are skipped, which is what
    keeps them from being written over.

    `ws_values` is the same sheet from a data_only load. The date column is not
    always literal dates — one hotel's later week tabs carry '=WKONE!C5' and
    '=C5+1' instead, which read back as formula text and match no day at all.
    Dates are taken from the cached values when that view is supplied.
    """
    cols = detect_ihg_sr_columns(ws)
    if "trans_rms" not in cols:
        return []
    dates_from = ws_values if ws_values is not None else ws
    date_col = detect_date_column(dates_from) or 3

    hf_by_date = {d["date"]: d for d in (hf or {}).get("days", [])}
    bob_by_date = {}
    for m in (bob or {}).get("months", {}).values():
        for d in m.get("days", []):
            bob_by_date[d["date"]] = d
    if not bob_by_date and not hf_by_date:
        return []
    bob_start = min(bob_by_date) if bob_by_date else None

    changes = []

    def put(row, col, label, value, when):
        if col is None or value is None:
            return
        changes.append({
            "row": row, "col": col, "label": f"{when:%b %d} {label}".replace(" 0", " "),
            "month": when.month, "new_value": value,
            "skip_reason": "formula" if is_formula(ws.cell(row, col).value) else None,
        })

    for r in range(5, max(ws.max_row, dates_from.max_row) + 1):
        v = dates_from.cell(r, date_col).value
        if isinstance(v, datetime.datetime):
            day = v.date()
        elif isinstance(v, datetime.date):
            day = v
        else:
            continue                     # blank rows and monthly totals

        if bob_start is not None and day >= bob_start:
            d = bob_by_date.get(day)
            if not d:
                continue
            put(r, cols.get("trans_rms"), "trans sold",  int(round(d["ind_rms"])), day)
            put(r, cols.get("group_rms"), "group sold",
                int(round(d["blk_rms"] + d["npu_rms"])), day)
            put(r, cols.get("group_npu"), "grp n/pu",    int(round(d["npu_rms"])), day)
            put(r, cols.get("ooo"),       "OOO",         int(round(d["ooo"])), day)
            put(r, cols.get("regrets"),   "regrets",     int(round(d["turn"])), day)
            put(r, cols.get("trans_rev"), "trans rev",   round(d["ind_rev"], 2), day)
            put(r, cols.get("group_rev"), "group rev",   round(d["blk_rev"], 2), day)
        else:
            d = hf_by_date.get(day)
            if not d:
                continue
            put(r, cols.get("trans_rms"), "trans sold", int(round(d["ind_rms"])), day)
            put(r, cols.get("group_rms"), "group sold", int(round(d["blk_rms"])), day)
            put(r, cols.get("trans_rev"), "trans rev",  round(d["ind_rev"], 2), day)
            put(r, cols.get("group_rev"), "group rev",  round(d["blk_rev"], 2), day)
            # A day that has already happened can have nothing left unpicked.
            put(r, cols.get("group_npu"), "grp n/pu", 0, day)
            # History and Forecast carries no OOO or turnaway columns, so these
            # keep whatever earlier weeks put there — unless it's text, which
            # would feed a #VALUE! into the row's demand and occupancy formulas.
            for key in ("ooo", "regrets"):
                col = cols.get(key)
                if col and isinstance(ws.cell(r, col).value, str) \
                        and not is_formula(ws.cell(r, col).value):
                    put(r, col, f"{key} (cleared non-numeric)", 0, day)
    return changes


def find_forecast_snapshot_row(ws, search_from=45, search_to=90):
    """Row in the week-over-week pickup grid that belongs to THIS week tab.

    The grid stacks one row per weekly snapshot with a 'Pick UP' row between
    each pair, computed as this snapshot minus the one above. Every row except
    one is a formula reaching back into an earlier week's tab; the row for the
    tab you're on is the one whose column A is '=A2' — a self-reference to its
    own as-of date. That is the only row that should ever be written, and it
    moves down by two with each successive week (WK2 row 57, WK3 59, WK4 61).

    Returns None if no such row is found, so a template that doesn't carry
    this grid is skipped rather than written to at a guessed row.
    """
    for r in range(search_from, search_to):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().replace("+", "").replace(" ", "").upper() == "=A2":
            return r
    return None


def build_ihg_next_month_forecast_plan(bob, ws, target_month, wb=None):
    """Forecast changes for a month other than the current one, from the
    Business on the Books daily rows.

    Everything in that report is still on the books — it starts at the report
    date — so every day goes to the OTB rows and none to the actual rows.
    """
    key = (target_month.year, target_month.month)
    month = (bob or {}).get("months", {}).get(key)
    if not month or not month.get("days"):
        return []

    rows = locate_forecast_rows(ws)
    if not rows:
        return []

    col_of = {}
    for date_val, col in build_forecast_date_col_map(
            ws, wb=wb, date_row=rows["date_row"]).items():
        col_of[date_val.date() if isinstance(date_val, datetime.datetime) else date_val] = col

    changes = []
    for d in month["days"]:
        col = col_of.get(d["date"])
        if col is None:
            continue
        stamp = d["date"].strftime("%b %d").replace(" 0", " ")
        for row, label, value in [
            (rows["otb_rooms_row"], f"{stamp} OTB rooms", int(round(d["rooms"]))),
            (rows["adr_otb_row"],   f"{stamp} OTB ADR",   round(d["avg_rate"], 2)),
        ]:
            changes.append({
                "row": row, "col": col, "label": label, "month": target_month.month,
                "new_value": value,
                "skip_reason": "formula" if is_formula(ws.cell(row, col).value) else None,
            })
    return changes


# ── ROB Update ───────────────────────────────────────────────────────────────

ROB_SHEETS = ["wk one", "wk two", "wk three", "wk four", "wk five", "wk six"]


def find_secondary_col(ws, block_start):
    candidates = []
    for cell in ws[block_start]:
        if cell.column <= 5:
            continue
        if isinstance(cell.value, str) and "variance" in cell.value.strip().lower():
            candidates.append(cell.column)
    return min(candidates) if candidates else None


def _srp_seg(srp_period, seg):
    """(rooms, revenue) for one segment of one period, or zeros.

    seg is 'TOT', 'GRP', 'PRM' or 'TRN'.

    Only TOT is used to write with. It is the ROB's total and the Forecast's
    on-the-books rooms, taken exactly as the export states it: checked against
    three hand-corrected workbooks, every Revenue and Room Nights cell in them
    equals the plain SRP figure, and Andover reproduces cell for cell.

    Group and permanent are NOT read from here even though SRP carries them —
    its 'convention' flag undercounts group badly. Those come from the Wash
    report; see build_hilton_rob_plan.
    """
    rooms, rev = (srp_period or {}).get(seg, [0, 0.0])
    return rooms, rev


def extract_hilton_mtd_actuals_from_forecast(raw_bytes, as_of):
    """Return completed current-month actual rooms/revenue through as_of - 2 days.

    Hilton's daily workflow intentionally overlaps by one day:
      * Forecast actuals are populated through yesterday.
      * ROB current-month total uses actuals only through the day before yesterday.
      * SRP contributes yesterday through month-end.

    The one-day overlap lets the SRP export supply the freshest value for
    yesterday, while the Forecast provides the already-actualized beginning
    of the month.

    Returns {"rooms", "revenue", "sheet", "through"} or None.
    """
    if not as_of:
        return None

    cutoff = as_of - datetime.timedelta(days=2)
    if cutoff.month != as_of.month or cutoff.year != as_of.year:
        # On the first/second day of a month there may be no current-month
        # actualized portion yet.
        return {
            "rooms": 0,
            "revenue": 0.0,
            "sheet": None,
            "through": cutoff,
        }

    wb_formulas = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=False)
    wb_values = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)

    candidates = []

    for sname in FORECAST_SHEETS:
        if sname not in wb_formulas.sheetnames or sname not in wb_values.sheetnames:
            continue

        wsf = wb_formulas[sname]
        wsv = wb_values[sname]
        rows = locate_forecast_rows(wsf)
        if not rows:
            continue

        sheet_as_of = (
            parse_any_date(wsv.cell(rows["as_of_row"], 1).value)
            or parse_any_date(wsf.cell(rows["as_of_row"], 1).value)
        )
        if sheet_as_of is None or sheet_as_of > as_of:
            continue

        col_map = build_forecast_date_col_map(
            wsv, wb_values, date_row=rows["date_row"]
        )
        if not col_map:
            continue

        # Only consider tabs that contain at least one completed actual value.
        has_actual = any(
            d.year == as_of.year
            and d.month == as_of.month
            and d <= cutoff
            and (
                safe_float(wsv.cell(rows["actual_rooms_row"], col).value)
                or safe_float(wsv.cell(rows["actual_revenue_row"], col).value)
            )
            for d, col in col_map.items()
        )
        if has_actual:
            candidates.append((sheet_as_of, sname, rows, col_map))

    if not candidates:
        return None

    # The most recent filled Forecast snapshot is the trusted source.
    _, sname, rows, col_map = max(candidates, key=lambda x: x[0])
    ws = wb_values[sname]

    rooms = 0.0
    revenue = 0.0
    for d, col in col_map.items():
        if (
            d.year == as_of.year
            and d.month == as_of.month
            and d <= cutoff
        ):
            rooms += safe_float(ws.cell(rows["actual_rooms_row"], col).value) or 0.0
            revenue += safe_float(ws.cell(rows["actual_revenue_row"], col).value) or 0.0

    return {
        "rooms": rooms,
        "revenue": revenue,
        "sheet": sname,
        "through": cutoff,
    }


def hilton_current_month_total(srp_days, forecast_actuals, as_of):
    """Combine Forecast actuals + SRP exactly as the Hilton daily ROB workflow.

    Current month =
      Forecast actuals: month start through as_of - 2
      SRP OTB/live stay values: as_of - 1 through month end
    """
    if not as_of:
        return None
    if forecast_actuals is None:
        return None

    month_end = (
        (as_of.replace(day=28) + datetime.timedelta(days=4)).replace(day=1)
        - datetime.timedelta(days=1)
    )
    srp_start = as_of - datetime.timedelta(days=1)

    srp_rooms = 0.0
    srp_revenue = 0.0
    d = srp_start
    while d <= month_end:
        rooms, rev = _srp_seg(srp_days.get(d), "TOT")
        srp_rooms += rooms
        srp_revenue += rev
        d += datetime.timedelta(days=1)

    return {
        "rooms": forecast_actuals["rooms"] + srp_rooms,
        "revenue": forecast_actuals["revenue"] + srp_revenue,
        "actual_rooms": forecast_actuals["rooms"],
        "actual_revenue": forecast_actuals["revenue"],
        "srp_rooms": srp_rooms,
        "srp_revenue": srp_revenue,
        "actual_through": forecast_actuals["through"],
        "srp_from": srp_start,
        "forecast_sheet": forecast_actuals["sheet"],
    }


def build_hilton_rob_plan(srp_months, wash_months, ws, as_of=None, current_month_total=None):
    """ROB changes for one Hilton hotel from the two Hilton exports.

    srp_months  — parse_srp_activity()[inncode]["months"]
    wash_months — parse_group_wash()["months"] for that same hotel

    Each figure comes from the report that measures it best:

      total (col E)    SRP's own TOT — every room on the books
      group/perm (E)   the Wash report's Pick Up
      not p/u (col G)  the Wash report's Available Block

    The total used to be assembled as SRP transient + Wash pick-up, which
    double-counted: pick-up *reclassifies* rooms that SRP has already counted
    inside TOT — SRP files a block room under whatever code the individual
    booked with — so adding the two put every month out by a few thousand.
    Group still has to come from the Wash report, because SRP's 'convention'
    flag genuinely undercounts it (Kansas City September: 270 rooms by SRP
    against 789 by the wash report).

    Rows are located by label. A cell holding a hand-written reconciliation
    like '=294767+55017' is left alone, but a cross-sheet mirror like
    "='wk one'!E110" is overwritten: that is the template pointing at last
    week, not a figure anyone entered, and leaving it stranded a stale Perm
    row on a real workbook.
    """
    as_of = as_of or datetime.date.today()
    blocks = rob_month_blocks(ws)
    changes, warns = [], []

    def put(row, col, label, value, month):
        if row is None or value is None:
            return
        existing = ws.cell(row, col).value
        # Same rule _rob_week_taken_reason uses to decide whether a week is
        # really filled: a formula reaching into another sheet is a placeholder,
        # anything else someone typed on purpose.
        skip = "formula" if (is_formula(existing) and "!" not in str(existing)) else None
        changes.append({"row": row, "col": col, "label": label, "month": month,
                        "new_value": value, "skip_reason": skip})

    changes.append({"row": 4, "col": 5, "label": "As-of date", "month": None,
                    "new_value": as_of, "skip_reason": None})

    for mi, labels in sorted(blocks.items()):
        month = mi + 1
        if month < as_of.month:
            continue                      # never rewrite a closed month
        key = (as_of.year, month)
        srp = srp_months.get(key) or {}
        wash = wash_months.get(key) or {}
        tot_rooms, tot_rev = _srp_seg(srp, "TOT")

        # The current month is different from future months. The daily SRP
        # export does not contain a reliable full-month history once completed
        # stays have fallen outside its Departure Date filter. For the current
        # month, use the established Hilton workflow:
        #   actuals through day-before-yesterday + SRP yesterday through EOM.
        if month == as_of.month and current_month_total is not None:
            tot_rooms = current_month_total["rooms"]
            tot_rev = current_month_total["revenue"]

        g = wash.get("GRP") or {}
        p = wash.get("PRM") or {}

        # Zero is a real ROB value, not "missing data". A month with no rooms
        # on the books must still write 0 to Rooms and Revenue so the workbook
        # does not misleadingly retain blanks/stale values.
        #
        # _srp_seg() returns (0, 0.0) for a genuinely empty month, so do not
        # skip merely because tot_rooms is zero.

        # The two reports have to describe the same hotel. Pick-up is a subset
        # of what is on the books, so it cannot exceed it — when it does, one
        # of the two exports is for a different date range or a different
        # property, and the month is wrong whichever figure you believe.
        # Confirmed real case: Kansas City September read 773 rooms on the
        # books against 1,017 picked up.
        picked = g.get("pu_rooms", 0.0) + p.get("pu_rooms", 0.0)
        if picked > tot_rooms:
            warns.append(
                f"{datetime.date(as_of.year, month, 1):%B}: the Wash report picks up "
                f"{picked:,.0f} rooms but SRP only has {tot_rooms:,.0f} on the books. "
                f"Group can't exceed the total — check the two exports cover the "
                f"same dates and the same property.")

        L = labels.get
        put(L("revenue"),        5, "Revenue",        round(tot_rev, 2), month)
        put(L("room nights"),    5, "Room Nights",    int(round(tot_rooms)), month)
        put(L("group rms sold"), 5, "Group Rms sold", int(round(g.get("pu_rooms", 0.0))), month)
        put(L("group rm rev"),   5, "Group Rm Rev",   round(g.get("pu_rev", 0.0), 2), month)
        put(L("group rms sold"), 7, "Group not p/u rms", int(round(g.get("av_rooms", 0.0))), month)
        put(L("group rm rev"),   7, "Group not p/u rev", round(g.get("av_rev", 0.0), 2), month)
        if L("perm rms sold"):
            put(L("perm rms sold"), 5, "Perm Rms Sold", int(round(p.get("pu_rooms", 0.0))), month)
            put(L("perm rm rev"),   5, "Perm Rm Rev",   round(p.get("pu_rev", 0.0), 2), month)
            put(L("perm rms sold"), 7, "Perm not p/u rms", int(round(p.get("av_rooms", 0.0))), month)
            put(L("perm rm rev"),   7, "Perm not p/u rev", round(p.get("av_rev", 0.0), 2), month)
    return changes, warns


def build_hilton_forecast_plan(srp_days, ws, as_of=None):
    """Forecast changes for one Hilton hotel from the SRP export.

    The Hilton Forecast is the same shape as every other Forecast the app
    writes — dates running across row 4, with OTB rooms, OTB ADR, actual rooms
    and actual revenue each on their own titled row — so it is located with the
    shared helpers rather than a second Hilton-specific set.

    The rule, confirmed with the user: this fills today to month end and
    nothing else. Days that have actualised are entered by hand.

    That matches what the export can actually support. SRP Activity lists
    reservations that are still live, so a completed day's rooms decay as it
    recedes: on a 17 Aug export one hotel's 1 Aug read 4 rooms against a real
    245, 8 Aug read 11, 13 Aug read 83, and only 14–16 Aug came back right.
    Writing those into the actuals row is worse than leaving it empty — the day
    looks filled and reads twenty times low, which is how this was found.

    'Estimated Pick Up' and the forecast ADR are likewise never written: they
    are the revenue manager's judgement, not anything the export knows.
    """
    as_of = as_of or datetime.date.today()
    rows = locate_forecast_rows(ws)
    if not rows:
        return [], ["could not read its row titles (As-of date / Rooms Sold / "
                    "ADR OTB / Revenue)."]
    col_map = build_forecast_date_col_map(ws, ws.parent, date_row=rows["date_row"])
    if not col_map:
        return [], ["could not read its date row."]

    changes, warns = [], []

    def put(label, row, col, value):
        changes.append({
            "label": label, "row": row, "col": col, "new_value": value,
            "skip_reason": "formula" if is_formula(ws.cell(row, col).value) else None,
        })

    put("As-of date", rows["as_of_row"], 1, as_of)

    dated, past = 0, 0
    for d, col in sorted(col_map.items()):
        if d < as_of:
            past += 1
            continue
        rooms, rev = _srp_seg(srp_days.get(d), "TOT")
        if not rooms:
            continue
        dated += 1
        put(f"Rooms Sold (OTB) {d}", rows["otb_rooms_row"], col, int(round(rooms)))
        put(f"ADR OTB {d}", rows["adr_otb_row"], col, round(rev / rooms, 2))

    if not dated:
        return [], [f"none of its dates ({min(col_map):%b %Y}) carry any rooms still "
                    f"on the books. Is this the right month's workbook?"]
    if past:
        warns.append(f"filled {dated} day(s) from {as_of:%b %d} to month end. The "
                     f"{past} day(s) before that have actualised and are left for "
                     f"you to enter by hand, as agreed.")

    # Pick-up tracking chart: this week's on-the-books rooms written under the
    # previous weeks', which is what makes the week-on-week build visible. Same
    # cutoff — a completed day's figure here would be as wrong as it is above.
    track = find_next_pickup_data_row(ws)
    if track:
        put("Pickup tracking: date", track, 1, as_of)
        for d, col in sorted(col_map.items()):
            if d < as_of:
                continue
            rooms, _ = _srp_seg(srp_days.get(d), "TOT")
            put(f"Pickup tracking: Rooms Sold {d}", track, col, int(round(rooms)))
    else:
        warns.append("no free row left in its pick-up tracking chart.")

    return changes, warns


def build_rob_change_plan(df, ws, grp_npu_rev_override: dict = None):
    """grp_npu_rev_override: optional {(year, month): dollar_value} — when
    present for a given month, writes that literal value into the 'Group Not
    P/U rev' secondary-column cell instead of the standard count*ADR formula.
    Used for Margaritaville, whose source data doesn't include a reliable
    Not-P/U room count to build that formula from — instead the value is
    computed elsewhere as the difference between two comparable PMS exports
    (one including not-yet-picked-up group revenue, one excluding it)."""
    today = datetime.date.today()
    current_month = today.month
    current_year = today.year
    changes = []
    # Rows per month block, read off this sheet rather than assumed — hotels
    # with a Permanent-rooms section use 11-row blocks, not 8, and every row
    # below is relative to it.
    block_step = rob_block_step(ws)
    # Last writable row = the final row of the Dec block. For the usual 8-row
    # layout this is 99, so `>= 100` behaves exactly as the old hardcoded
    # guard did; an 11-row sheet correctly allows through row 135.
    max_data_row = 4 + block_step * 12

    # E4 = as-of date
    changes.append({
        "row": 4, "col": 5, "label": "As-of date", "month": None,
        "new_value": today, "skip_reason": None,
    })

    for _, row in df.iterrows():
        date_str = str(row[0]).strip() if row[0] else ""
        kind, info = classify_row(date_str)
        if kind != "monthly":
            continue
        year, month = info
        prev_month = current_month - 1 if current_month > 1 else 12
        prev_year  = current_year if current_month > 1 else current_year - 1
        if year == prev_year and month == prev_month:
            pass  # allow previous month (final numbers come in on the 1st)
        elif year != current_year or month < current_month:
            continue

        month_index = month - 1
        block_start = 4 + block_step * month_index

        rev_raw = row[5]
        rms_raw = row[1]
        rev     = safe_float(rev_raw)
        rms     = safe_float(rms_raw)

        # Preserve an explicit 0 from the source as a literal zero in Excel.
        # Only truly blank/unparseable source cells remain None.
        if rev is None and str(rev_raw).strip() in {"0", "0.0", "0.00", "$0", "$0.00"}:
            rev = 0.0
        if rms is None and str(rms_raw).strip() in {"0", "0.0", "0.00"}:
            rms = 0.0
        grp_pu  = safe_float(row[7])
        grp_npu = safe_float(row[8])
        grp_rvn = safe_float(row[9])

        grp_sold = (grp_pu or 0) + (grp_npu or 0) if grp_pu is not None and grp_npu is not None else None
        sec_col = find_secondary_col(ws, block_start)

        entries = [
            (block_start + 1, 5, "Revenue",        rev,       False),
            (block_start + 2, 5, "Room Nights",     rms,       False),
            (block_start + 4, 5, "Group Rms Sold",  grp_sold,  False),
            (block_start + 5, 5, "Group Rm Rev",    grp_rvn,   False),
        ]
        if sec_col:
            npu_row    = block_start + 4
            entries.append((npu_row, sec_col, "Group Not P/U rooms", grp_npu, False))

            override_val = grp_npu_rev_override.get((year, month)) if grp_npu_rev_override else None
            if override_val is not None:
                entries.append((npu_row + 1, sec_col, "Group Not P/U rev (computed)", override_val, False))
            else:
                from openpyxl.utils import get_column_letter
                sec_letter = get_column_letter(sec_col)
                adr_row    = block_start + 6
                npu_formula = f"={sec_letter}{npu_row}*E{adr_row}"
                entries.append((npu_row + 1, sec_col, "Group Not P/U rev (formula)", npu_formula, True))

        for r, c, label, val, is_formula_write in entries:
            skip = None
            if r >= max_data_row:
                skip = f"row≥{max_data_row}"
            elif not is_formula_write and is_formula(ws.cell(r, c).value):
                skip = "formula"
            changes.append({"row": r, "col": c, "label": label, "month": month,
                             "new_value": val, "skip_reason": skip})

    return changes


def compute_grp_npu_rev_override(df, npu_compare_df):
    """Margaritaville ROB only: npu_compare_df is the same source format but
    from a second PMS export that includes not-yet-picked-up group revenue
    (df itself is the export that excludes it — confirmed the smaller of the
    two is used everywhere else). The difference per month is the dollar
    value for the "Group Not P/U rev" bright-green box on the ROB. Returns
    None if either input is missing (i.e. every hotel except Margaritaville)."""
    if npu_compare_df is None or df is None:
        return None

    def _monthly_col5(source_df):
        out = {}
        for _, row in source_df.iterrows():
            kind, info = classify_row(str(row[0]).strip() if row[0] else "")
            if kind == "monthly":
                out[info] = safe_float(row[5])
        return out

    main_sums    = _monthly_col5(df)
    compare_sums = _monthly_col5(npu_compare_df)
    return {
        key: compare_sums[key] - main_sums[key]
        for key in main_sums
        if key in compare_sums and main_sums[key] is not None and compare_sums[key] is not None
    }


def apply_rob_changes(wb, sheet_name, changes):
    ws = wb[sheet_name]
    for ch in changes:
        if ch["skip_reason"]:
            continue
        ws.cell(ch["row"], ch["col"]).value = ch["new_value"]


DONE_TAB_RGB = "FF00B050"  # green — set by color_tab_done() when a week is genuinely complete
DONE_TAB_HEX = "00B050"    # same green, without the alpha channel


def _is_done_color(rgb_value) -> bool:
    """True if rgb_value is our green 'done' marker. Matches on the trailing
    6 hex digits (case-insensitive) so it recognizes a tab colored directly in
    Excel's own Tab Color picker (which often stores 6-digit RGB with no alpha
    prefix) as well as ones this app set (8-digit ARGB) — a real week that a
    hotel had already filled in and marked green by hand was being treated as
    'not done' by a strict 8-char match, so the app kept re-picking it.
    """
    return isinstance(rgb_value, str) and rgb_value[-6:].upper() == DONE_TAB_HEX


_MONTH_LABELS = {m.lower(): i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])}
_MONTH_LABELS.update({m.lower(): i for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June", "July",
     "August", "September", "October", "November", "December"])})
ROB_DEFAULT_BLOCK_STEP = 8


def rob_block_step(ws, default=ROB_DEFAULT_BLOCK_STEP):
    """Rows per month block on a ROB week tab, read off the sheet.

    Most ROBs use 8 rows per month (Revenue/Room Nights/ADR/Group x3/Pickup),
    but hotels with a Permanent-rooms section (airline crew contracts) use 11
    — they carry three extra rows (Perm Rms Sold / Perm Rm Rev / Perm ADR).
    Assuming 8 everywhere makes every row calculation land inside the wrong
    month on those sheets: on an 11-row ROB the "August" block computes to
    row 60, which is actually inside June.

    Derived from the Jan/Feb month labels in column A and cross-checked
    against Mar, so a sheet that doesn't look like a ROB falls back to the
    historical constant rather than guessing.
    """
    found = {}
    for r in range(1, 200):
        v = ws.cell(r, 1).value
        if not isinstance(v, str):
            continue
        idx = _MONTH_LABELS.get(v.strip().lower())
        if idx is not None and idx not in found:
            found[idx] = r
    if 0 in found and 1 in found:
        step = found[1] - found[0]
        if 2 in found and found[2] - found[1] != step:
            return default            # spacing isn't uniform — don't trust it
        if 4 <= step <= 20:
            return step
    return default


def _rob_week_taken_reason(ws, block_start):
    """Why this week tab is unavailable, or None if it's free to write to.

    Two things this has to get right, both confirmed against real files:

    A zero is not data. These templates render empty currency cells as "$ -"
    (a literal 0), so testing `isinstance(v, (int, float))` alone counts an
    untouched month as filled and silently skips the tab.

    A formula usually IS data. Weeks are routinely reconciled by hand into an
    expression — real examples are '=294767+55017+2028' and
    '=320153-11611+6135' sitting in the Revenue cell of an already-completed
    week. Treating only literal numbers as data marks those weeks available
    and overwrites the reconciliation. The exception is a bare cross-sheet
    reference like ="wk one"!E82, which is the template's own placeholder
    mirroring week one rather than this week's figures.
    """
    tc = ws.sheet_properties.tabColor
    if tc is not None and _is_done_color(getattr(tc, "rgb", None)):
        return "tab marked done (green)"
    rev = ws.cell(block_start + 1, 5).value
    rms = ws.cell(block_start + 2, 5).value

    def filled(v):
        if isinstance(v, bool) or v is None:
            return False
        if isinstance(v, (int, float)):
            return v != 0
        if is_formula(str(v)):
            return "!" not in str(v)   # cross-sheet mirror = still a placeholder
        return False

    if filled(rev) or filled(rms):
        return (f"already has this month's data "
                f"(E{block_start+1}={rev!r}, E{block_start+2}={rms!r})")
    return None


def rob_month_blocks(ws):
    """{month_index: {row_label_lower: row_number}} for a ROB week tab.

    Locates every row by its column-A label instead of by a fixed offset from
    the block start. Offsets are not safe to assume even within one workbook:
    a real file has 'wk one' on 11-row blocks with Perm at offsets 7/8 while
    'wk two'..'wk six' are 12-row with Perm at 8/9, so an offset that is right
    on one tab writes into the wrong row on the next.

    Labels are lower-cased and whitespace-collapsed, e.g. 'revenue',
    'room nights', 'group rms sold', 'group rm rev', 'perm rms sold',
    'perm rm rev', 'adr', 'pickup wow'.
    """
    starts = {}
    for r in range(1, 260):
        v = ws.cell(r, 1).value
        if not isinstance(v, str):
            continue
        s = v.strip()
        # month header cells hold just the month name — guard against a long
        # sentence that merely begins with one
        if len(s) <= 12 and s[:3].lower() in _MONTH_LABELS:
            starts.setdefault(_MONTH_LABELS[s[:3].lower()], r)
    ordered = sorted(starts.items(), key=lambda kv: kv[1])
    if not ordered:
        return {}
    span = (ordered[1][1] - ordered[0][1]) if len(ordered) > 1 else ROB_DEFAULT_BLOCK_STEP
    out = {}
    for n, (mi, r0) in enumerate(ordered):
        end = ordered[n + 1][1] if n + 1 < len(ordered) else r0 + span
        labels = {}
        for r in range(r0 + 1, end):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and v.strip():
                labels.setdefault(re.sub(r"\s+", " ", v.strip().lower()), r)
        out[mi] = labels
    return out


def rob_week_status(wb, sheet_names):
    """[(sheet_name, reason_or_None), ...] — why each week tab was passed over.
    Purely diagnostic; drives the 'why did it skip a week' caption in the UI.
    """
    month = datetime.date.today().month
    out = []
    for name in sheet_names:
        ws = wb[name]
        block_start = 4 + rob_block_step(ws) * (month - 1)
        out.append((name, _rob_week_taken_reason(ws, block_start)))
    return out


def first_uncolored_sheet(wb, sheet_names):
    """Return the first ROB week tab that is neither marked done (our green)
    nor already holding real data in this month's block.

    Color alone isn't enough in either direction: a master template can carry
    its own unrelated baked-in tab color on a sheet that's never been touched
    (so "any color = done" produces false positives), while a real, already-
    filled week can fail to carry our exact green (so "only our green = done"
    produces false negatives — confirmed on a real hotel: week one already had
    this month's Revenue/Room Nights filled in but got silently re-picked and
    overwritten because its tab color didn't match). Checking actual data in
    the cells this month's update is about to write closes that gap.
    """
    month = datetime.date.today().month
    for name in sheet_names:
        ws = wb[name]
        block_start = 4 + rob_block_step(ws) * (month - 1)
        if _rob_week_taken_reason(ws, block_start) is None:
            return name
    return sheet_names[-1]  # fallback: last sheet


def first_unhighlighted_forecast_sheet(wb, sheet_names):
    """Return the first Forecast week tab that is neither marked done (our
    green) nor already holding real OTB Rooms Sold data.

    Tab color alone can't be trusted in either direction here: hotels color-
    code Forecast tabs by hand with whatever color they like, but the SAME
    color can also be a stray artifact baked into a never-used master
    template — confirmed on real files: Hampton's untouched master has ALL 9
    week tabs pre-colored, and Provincetown Brass's untouched master has
    FCST-WK1 pre-colored magenta despite being completely blank. Checking the
    OTB Rooms Sold row for actual filled-in numbers is the reliable signal;
    our own exact green is kept as a secondary check for weeks we marked done
    ourselves but that ended up with no literal numbers written (e.g. every
    cell in that row happened to be formula-protected).
    """
    for name in sheet_names:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        tc = ws.sheet_properties.tabColor
        if tc is not None and _is_done_color(getattr(tc, "rgb", None)):
            continue
        rows = locate_forecast_rows(ws)
        if not rows:
            return name  # can't verify — treat as available rather than guess
        otb_row = rows["otb_rooms_row"]
        # Check the actual date-mapped columns for this sheet rather than a
        # hardcoded 2-9 range — confirmed real case (Tybee): a filled week
        # got silently treated as available and overwritten, most likely
        # because its real data sat outside that fixed range. Falls back to
        # the old range if the date row itself can't be read. safe_float
        # (not a strict isinstance check) also catches numbers stored as
        # text, which a strict int/float check would miss entirely.
        col_map = build_forecast_date_col_map(ws, ws.parent, date_row=rows["date_row"])
        data_cols = list(col_map.values()) if col_map else list(range(2, 10))
        # The actuals row counts as data too. A week whose OTB row was never
        # filled but whose actuals were entered by hand is a used week —
        # confirmed on a real Hilton file, where checking OTB alone offered up
        # an already-written week as the next one to fill.
        has_data = any(safe_float(ws.cell(r, c).value) is not None
                       for r in (otb_row, rows["actual_rooms_row"])
                       for c in data_cols)
        if has_data:
            continue
        return name
    return sheet_names[-1]  # fallback: last sheet


def first_undone_strategy_sheet(wb, sheet_names):
    """Return the first Strategy Report week tab that is neither marked done
    (our green) nor already holding real OTB TY Trans data.

    Same reasoning as first_unhighlighted_forecast_sheet: tab color alone
    can't be trusted — confirmed on real files that the identical purple
    (FF9900FF) marks a genuinely completed Surfside week AND sits untouched
    on Wolfboro's never-used master template. Check actual filled data as the
    reliable signal; our own exact green is a secondary check for weeks we
    marked done ourselves but that ended up with no literal numbers written.
    """
    for name in sheet_names:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        tc = ws.sheet_properties.tabColor
        if tc is not None and _is_done_color(getattr(tc, "rgb", None)):
            continue
        col_map = detect_strategy_columns(ws)
        otb_col = col_map.get("otb_trans")
        if not otb_col:
            return name  # can't verify — treat as available rather than guess
        has_data = any(isinstance(ws.cell(r, otb_col).value, (int, float)) for r in range(5, 15))
        if has_data:
            continue
        return name
    return sheet_names[-1]  # fallback: last sheet


def color_tab_done(wb, sheet_name):
    """Mark a sheet tab green to indicate it has been completed."""
    from openpyxl.styles.colors import Color
    wb[sheet_name].sheet_properties.tabColor = Color(rgb=DONE_TAB_RGB)


def clear_tab_colors(wb, sheet_names):
    """Reset tab color to none for every listed sheet. New-month setup copies
    the master/prior file, which can carry over a stale 'done' tab color —
    without this, a week that hasn't been touched yet can look completed and
    get skipped by the 'first uncolored sheet' auto-detect."""
    for name in sheet_names:
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = None


def strip_tables(wb):
    """Remove Excel table definitions to prevent openpyxl save corruption."""
    for ws in wb.worksheets:
        ws.tables.clear()


def apply_portfolio_plans(svc, jobs, undo_key):
    """Write, mark done, upload — shared by the Hilton and IHG runs.

    Each job is {key, file_id, file_name, wb_bytes, sheet, changes}. The
    pre-write value of every cell is snapshotted first so the run can be undone
    in one step, and the clean pre-write bytes are kept alongside so the undo
    rebuilds from the original rather than from what was just written.

    A macro-enabled workbook is reopened with keep_vba so uploading it back
    doesn't strip the macros.
    """
    saved, errors, snapshot = [], [], {}
    for job in jobs:
        try:
            keep_vba = str(job.get("file_name", "")).lower().endswith(".xlsm")
            wb = openpyxl.load_workbook(io.BytesIO(job["wb_bytes"]),
                                        data_only=False, keep_vba=keep_vba)
            ws = wb[job["sheet"]]
            writes = [c for c in job["changes"] if not c.get("skip_reason")]
            prev_tab = ws.sheet_properties.tabColor
            snapshot[job["key"]] = {
                "file_id":   job["file_id"],
                "file_name": job["file_name"],
                "wb_bytes":  job["wb_bytes"],
                "sheet":     job["sheet"],
                "keep_vba":  keep_vba,
                # The tab colour is part of what a run changes — undoing the
                # cells but leaving the tab green would leave the week looking
                # complete and get it skipped on the next run.
                "tab_rgb":   getattr(prev_tab, "rgb", None) if prev_tab is not None else None,
                "cells":     {(job["sheet"], c["row"], c["col"]):
                              ws.cell(c["row"], c["col"]).value for c in writes},
            }
            for c in writes:
                ws.cell(c["row"], c["col"]).value = c["new_value"]
                if isinstance(c["new_value"], (datetime.date, datetime.datetime)):
                    ws.cell(c["row"], c["col"]).number_format = "m/d/yyyy"
            color_tab_done(wb, job["sheet"])
            strip_tables(wb)
            out = io.BytesIO()
            wb.save(out)
            drive_upload(svc, job["file_id"], out.getvalue(), job["file_name"])
            saved.append(f'{job["file_name"]} → {job["sheet"]} ({len(writes)} cells)')
        except Exception as e:
            errors.append(f'{job["key"]}: {e}')
    if snapshot:
        st.session_state[undo_key] = snapshot
    return saved, errors


def undo_portfolio_plans(svc, undo_key):
    """Put every snapshotted cell back and re-upload."""
    snapshot = st.session_state.get(undo_key) or {}
    if not snapshot:
        return [], ["Nothing to undo — no snapshot from this session."]
    saved, errors = [], []
    for key, info in snapshot.items():
        try:
            wb = openpyxl.load_workbook(io.BytesIO(info["wb_bytes"]),
                                        data_only=False,
                                        keep_vba=info.get("keep_vba", False))
            ws = wb[info["sheet"]]
            for (_sheet, row, col), original in info["cells"].items():
                ws.cell(row, col).value = original
            rgb = info.get("tab_rgb")
            if rgb:
                from openpyxl.styles.colors import Color
                ws.sheet_properties.tabColor = Color(rgb=rgb)
            else:
                ws.sheet_properties.tabColor = None
            strip_tables(wb)
            out = io.BytesIO()
            wb.save(out)
            drive_upload(svc, info["file_id"], out.getvalue(), info["file_name"])
            saved.append(info["file_name"])
        except Exception as e:
            errors.append(f"{key}: {e}")
    if not errors:
        st.session_state.pop(undo_key, None)
    return saved, errors


# ── Strategy Report ───────────────────────────────────────────────────────────

STRATEGY_SHEETS = ["WKONE", "WKTWO", "WKTHREE", "WKFOUR", "WKFIVE"]
FORECAST_SHEETS = ["FCST-WK1", "FCST-WK2", "FCST-WK3", "FCST-WK4",
                   "FCST-WK5", "FCST-WK6", "FCST-WK7", "FCST-WK8", "FCST-WK9"]

# CSV column index for each field (0-based) — source data never changes
STRATEGY_CSV_COLS = {
    "otb_trans":    (15, "OTB TY Trans (Indiv Count)"),
    "grp_pu_ty":    ( 7, "GRP PU TY"),
    "grp_npu_ty":   ( 8, "GRP N/PU TY"),
    "ooo_rms":      ( 4, "OOO RMS"),
    "trans_rev_ty": (16, "Trans Rev TY"),
    "grp_rev_ty":   ( 9, "Grp Rev TY"),
}

# Each field: list of (row3_keyword, row4_keyword) pairs to try in order.
# Match = both keywords found (case-insensitive) in their respective rows of that column.
# A None keyword means "don't check that row."
# "!WORD" suffix(es) on a keyword mean the combined headers must NOT contain WORD.
STRATEGY_FIELD_PATTERNS = {
    # ── TY columns (written from CSV) ──────────────────────────────────────────
    "otb_trans":       [("OTB TY", "TRANS"),            ("TRANS!LY", "SOLD!LY")],
    "grp_pu_ty":       [("GRP PU", "TY!LY"),            ("GROUP!LY", "SOLD!LY")],
    "grp_npu_ty":      [("GRP N/PU", "TY!LY"),          ("GRP RMS", "N/PU"),       ("N/PU!LY", None)],
    "ooo_rms":         [("OOO", None)],
    "trans_rev_ty":    [("TY TRANS", "REV"),             ("TRAN!LY", "REV TY")],
    "grp_rev_ty":      [("GRP TY", "REV"),               ("GRP!LY!N/PU", "REV TY")],
    "otb_lst_wk":      [("OTB", "LST WEK"),               ("OTB", "LST WK"),         ("OTB", "LAST WK"), ("OTB LST", None)],
    "casino_ballroom": [("CASINO", "BALLROOM!LY"),       ("CASINO BALLROOM", None)],
    # ── LY columns (written from last year's SR) ───────────────────────────────
    "otb_ly_trans":    [("LY", "TRAN"),                  ("OTB LY", "TRANS"),       ("TRANS!TY", "SOLD!TY"), ("LY", "TRANS!TY")],
    "grp_pu_ly":       [("LY", "GRP"),                   ("GRP PU", "LY"),          ("GROUP!TY", "LY"),        ("GRP PU LY", None)],
    "grp_npu_ly":      [("GRP N/PU", "LY"),              ("N/PU LY", None),         ("GRP RMS", "LY")],
    "trans_rev_ly":    [("LY TRANS", "REV"),             ("TRAN!TY", "REV LY"),     ("LY", "TRANS REV")],
    "grp_rev_ly":      [("GRP LY", "REV"),               ("GRP!TY!N/PU", "REV LY")],
    "grp_npu_rev_ly":  [("GRP N/PU", "REV LY"),         ("N/PU LY", "REV"),        ("N/PU", "REV LY")],
    "casino_ballroom_ly": [("CASINO", "BALLROOM!TY"),    ("CASINO", "LY")],
}

# Maps LY destination field → TY source field in last year's SR
LY_FROM_TY = {
    "otb_ly_trans":    "otb_trans",
    "grp_pu_ly":       "grp_pu_ty",
    "grp_npu_ly":      "grp_npu_ty",
    "trans_rev_ly":    "trans_rev_ty",
    "grp_rev_ly":      "grp_rev_ty",
    "grp_npu_rev_ly":  "grp_npu_ty",  # source is GRP N/PU TY
    "casino_ballroom_ly": "casino_ballroom",
}

def _kw_matches(cell_val, keyword, r3_val, r4_val):
    """Check if keyword matches cell_val.
    Supports !WORD suffixes — the combined headers must NOT contain those words.
    e.g. 'TRAN!LY!ADR' matches if cell contains 'TRAN' and neither header contains 'LY' or 'ADR'.
    """
    parts = keyword.split("!")
    kw = parts[0].strip()
    excludes = [p.strip().upper() for p in parts[1:]]
    if kw and kw.upper() not in str(cell_val or "").upper():
        return False
    combined = (str(r3_val or "") + " " + str(r4_val or "")).upper()
    for excl in excludes:
        if excl in combined:
            return False
    return True


def detect_strategy_columns(ws):
    """Scan rows 3+4 of THIS sheet and return {field_key: col_index} for each
    field. Every sheet is re-scanned independently — never assume two sheets
    (even in the same workbook/hotel) share column positions. Week-1 vs
    week-2+ tabs can differ (e.g. an extra pickup-tracking column shifts
    everything after it), so a value pinned from one sheet can silently be
    wrong on another.

    If row 4 is blank under a TY/LY column, infer it from context (e.g., blank
    under "OTB TY" → assume "TRANS", blank under "GRP PU TY" → assume "TY").
    """
    max_col = ws.max_column
    # Build lookup: col → (r3_text, r4_text)
    headers = {}
    for c in range(1, max_col + 1):
        r3_val = str(ws.cell(3, c).value or "").strip()
        r4_val = str(ws.cell(4, c).value or "").strip()

        # If row 4 is blank or "None", infer from row 3 context
        if (not r4_val or r4_val == "None") and r3_val:
            if "OTB TY" in r3_val:
                r4_val = "TRANS"  # OTB TY column always has TRANS
            elif "TY" in r3_val and "LY" not in r3_val:
                r4_val = "TY"  # Other TY columns infer "TY"
            elif "LY" in r3_val:
                r4_val = "LY"  # LY columns infer "LY"

        headers[c] = (r3_val, r4_val)

    col_map = {}
    for field, patterns in STRATEGY_FIELD_PATTERNS.items():
        found = None
        for r3_kw, r4_kw in patterns:
            for c, (r3v, r4v) in headers.items():
                # Match each keyword against the combined row3+row4 text,
                # not just its "own" row — confirmed real case: a header
                # like "OTB TY TRANS" doesn't always split the same way
                # across the two rows on every sheet (e.g. "OTB" / "TY
                # TRANS" instead of "OTB TY" / "TRANS"), and a strict
                # own-row-only check missed it entirely.
                combined = f"{r3v} {r4v}"
                r3_ok = r3_kw is None or _kw_matches(combined, r3_kw, r3v, r4v)
                r4_ok = r4_kw is None or _kw_matches(combined, r4_kw, r3v, r4v)
                if r3_ok and r4_ok and (r3_kw or r4_kw):
                    found = c
                    break
            if found:
                break
        if found:
            col_map[field] = found
        else:
            col_map[field] = None  # will surface as a warning, not a crash

    # Casino Ballroom is a special two-column section on Ashworth/Hampton.
    casino_cols = []
    for c, (r3v, r4v) in headers.items():
        combined = f"{r3v} {r4v}".upper()
        if "CASINO" in combined and "BALLROOM" in combined:
            casino_cols.append(c)

    if casino_cols:
        casino_cols = sorted(set(casino_cols))
        col_map["casino_ballroom"] = casino_cols[0]
        if len(casino_cols) >= 2:
            col_map["casino_ballroom_ly"] = casino_cols[1]
        else:
            next_col = casino_cols[0] + 1
            col_map["casino_ballroom_ly"] = next_col if next_col <= max_col else None

    return col_map


_CROSS_SHEET_REF_RE = re.compile(r"^=(?:'([^']+)'|([A-Za-z0-9_]+))!\$?([A-Z]+)\$?(\d+)$")
_LOCAL_OFFSET_RE    = re.compile(r"^=\$?([A-Z]+)\$?(\d+)([+-]\d+)$")


def _resolve_formula_date(wb, sheet_name, formula, _depth=0):
    """Resolve a date-continuation formula to an actual date without a real
    spreadsheet engine — handles the two real patterns this app's SR
    templates use: a cross-sheet single-cell reference (e.g. '=WKONE!C5',
    used by every non-WKONE week tab to mirror WKONE's own calendar
    column — the same "other weeks reference wk one" pattern ROB uses) and
    a local same-column +N day increment (e.g. '=C5+1'). Recurses through
    a chain (a local offset can itself point at a cross-sheet ref and vice
    versa) with a depth cap as a cycle guard. Confirmed real case:
    Provincetown Crowne Pointe's WKTHREE date column is 100% cross-sheet
    references with no literal cell of its own anywhere — a resolver that
    only trusted local '=prevRow+1' chains, requiring a literal anchor
    first, could never resolve a single date from it and always came back
    empty, blocking all CSV data from being written.
    """
    if _depth > 20 or not isinstance(formula, str) or not formula.startswith("="):
        return None
    formula = formula.strip()

    m = _CROSS_SHEET_REF_RE.match(formula)
    if m:
        target_sheet = m.group(1) or m.group(2)
        if target_sheet not in wb.sheetnames:
            return None
        col_num = column_index_from_string(m.group(3))
        row_num = int(m.group(4))
        val = wb[target_sheet].cell(row_num, col_num).value
        if isinstance(val, datetime.datetime):
            return val.date()
        if isinstance(val, datetime.date):
            return val
        return _resolve_formula_date(wb, target_sheet, val, _depth + 1)

    m = _LOCAL_OFFSET_RE.match(formula)
    if m:
        col_num = column_index_from_string(m.group(1))
        row_num = int(m.group(2))
        offset  = int(m.group(3))
        val = wb[sheet_name].cell(row_num, col_num).value
        if isinstance(val, datetime.datetime):
            base = val.date()
        elif isinstance(val, datetime.date):
            base = val
        else:
            base = _resolve_formula_date(wb, sheet_name, val, _depth + 1)
            if base is None:
                return None
        return base + datetime.timedelta(days=offset)

    return None


def detect_date_column(ws, wb=None):
    """Find the column whose data rows (5+) contain the earliest daily dates —
    i.e. the column that maps to each row's actual calendar date.
    Scans cols 1-10 only (dates are always on the left side).

    A sheet commonly has a Last Year date column right next to the This Year
    one (e.g. col 1 = LY dates, col 3 = TY dates), both starting on the 1st
    of the same month in different years — under a pure "most consecutive,
    prefer earliest day-of-month" score they tie exactly, and confirmed on a
    real file (Anchor In) the tie silently kept the LY column, so the whole
    date-to-row map was built a year off and almost nothing matched the CSV.
    Breaking ties by which column's dates start closest to today reliably
    picks the current/forward-looking column instead.

    That "closest to today" tie-break only works when candidates are
    genuinely a year or more apart — it breaks down when two candidates are
    just a day or two apart (which happens when a sheet has more than one
    plausible calendar column, e.g. one mislabeled/off by a day), because
    whichever one starts *later* will almost always look "closer to today"
    and win even when it's the wrong one. Confirmed real case: Provincetown
    Crowne Pointe's WKTHREE mirrors two of WKONE's columns via cross-sheet
    formula — one starting the 2nd (genuinely wrong data in WKONE itself),
    one starting the 1st (correct) — and pure proximity picked the 2nd on
    almost every day of the month. This app's own convention
    (restructure_sr_dates) always starts a TY column on day 1 of the
    target month, so preferring a candidate that starts on the 1st is a
    more reliable tie-break than raw proximity, checked before falling
    back to proximity for the remaining genuine (same-day-of-month) ties.

    Formula cells are resolved via _resolve_formula_date (handles both
    local +N increments and cross-sheet references like '=WKONE!C5') when
    wb is provided, instead of just being ignored or blindly assumed to
    continue by exactly 1 day.

    Scans 60 rows (not just the first 10) — a 10-row window is too easily
    fooled by a sheet whose real calendar column doesn't fully establish
    itself that close to the top, or where some other left-side column
    coincidentally looks date-like in just that narrow band. A wider
    window gives every genuine candidate room to prove real consecutive
    length before scoring, at negligible extra cost.
    """
    today = datetime.date.today()
    sheet_name = ws.title
    best_col, best_key = 3, None  # fallback to col 3
    for c in range(1, 11):
        anchor_date = None   # first date of the longest verified-consecutive run
        last_date   = None   # most recently confirmed date in the current run
        count = 0
        best_run = 0
        run_anchor = None
        for r in range(5, min(ws.max_row + 1, 65)):
            v = ws.cell(r, c).value
            if isinstance(v, datetime.datetime):
                d = v.date()
            elif isinstance(v, datetime.date):
                d = v
            elif isinstance(v, str) and v.startswith("=") and wb is not None:
                d = _resolve_formula_date(wb, sheet_name, v)
                if d is None:
                    last_date = None
                    continue
            else:
                last_date = None  # non-date cell — any run in progress is broken
                continue

            if last_date is not None and (d - last_date).days == 1:
                count += 1
            else:
                run_anchor = d  # gap or scattered date — start a fresh run here
                count = 1
            last_date = d
            if count > best_run:
                best_run = count
                anchor_date = run_anchor
        if best_run < 3 or anchor_date is None:
            continue
        proximity  = abs((anchor_date - today).days)
        starts_1st = 1 if anchor_date.day == 1 else 0
        key = (best_run, starts_1st, -proximity)
        if best_key is None or key > best_key:
            best_key = key
            best_col = c
    return best_col


def detect_comp_set_columns(ws, col_map):
    """Find the comp set chart's TY (far-left) and LY (far-right) columns.
    The chart sits between Restrictions and TY TRANS REV.
    LY col  = last non-empty text header to the LEFT of TY TRANS REV.
    TY col  = first non-empty text header to the RIGHT of Restrictions.
    Scans rows 2-6 for headers (tolerates slight layout shifts).
    Returns (ty_col, ly_col) or (None, None) if not found.
    """
    trans_rev_col = col_map.get("trans_rev_ty")
    if not trans_rev_col:
        return None, None

    restrict_col = find_restrictions_col(ws, upto_col=trans_rev_col - 1)

    left_bound  = (restrict_col + 1) if restrict_col else max(1, trans_rev_col - 30)
    right_bound = trans_rev_col - 1

    def _has_text_header(c):
        for scan_r in range(2, 7):
            v = str(ws.cell(scan_r, c).value or "").strip()
            if v and not v.replace(".", "").isdigit():
                return True
        return False

    # LY col: scan left from just before TY TRANS REV
    ly_col = None
    for c in range(right_bound, left_bound - 1, -1):
        if _has_text_header(c):
            ly_col = c
            break

    # TY col: scan right from just after Restrictions
    ty_col = None
    for c in range(left_bound, right_bound + 1):
        if _has_text_header(c):
            ty_col = c
            break

    if ty_col == ly_col:
        return None, None  # same column — chart not found

    return ty_col, ly_col


def get_ly_sr_data(service, hotel_id, hotel_name, current_month, sheet_name):
    """Fetch LY data from last year's SR (same month, same week tab).
    Returns {date: {ly_field: value}} where dates are mapped to this year.
    Also returns (comp_ty_col_in_ly_ws, comp_ly_col_in_current_ws_placeholder)
    via a separate 'comp_set' key: {this_year_date: value}.
    """
    ly_month = current_month.replace(year=current_month.year - 1)
    result, err = resolve_drive_workbook(service, hotel_id, hotel_name,
                                         "Strategy Report", month_date=ly_month)
    if err or not result:
        return {}

    file_id, _ = result
    try:
        file_bytes = drive_download(service, file_id)
        ly_wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    except Exception:
        return {}

    if sheet_name not in ly_wb.sheetnames:
        return {}

    ly_ws = ly_wb[sheet_name]
    ly_col_map  = detect_strategy_columns(ly_ws)
    ly_date_col = detect_date_column(ly_ws, wb=ly_wb)

    # Build date→row for last year's sheet
    ly_date_row = {}
    for r in range(5, ly_ws.max_row + 1):
        v = ly_ws.cell(r, ly_date_col).value
        if isinstance(v, datetime.datetime):
            ly_date_row[v.date()] = r
        elif isinstance(v, datetime.date):
            ly_date_row[v] = r

    # Comp set: far-left TY col in LY sheet
    comp_ty_col, _ = detect_comp_set_columns(ly_ws, ly_col_map)

    out = {}  # {this_year_date: {field: value}}
    for ly_date, r in ly_date_row.items():
        this_year_date = ly_date.replace(year=ly_date.year + 1)
        row_data = {}

        # Pull each TY source field from last year
        for ly_dest_field, ty_src_field in LY_FROM_TY.items():
            src_col = ly_col_map.get(ty_src_field)
            if src_col:
                v = ly_ws.cell(r, src_col).value
                if v is not None and not is_formula(v):
                    row_data[ly_dest_field] = safe_float(v)

        # Comp set TY value (far-left hotel col in LY sheet) — keep text as-is (e.g. "Sold out", "LOS2")
        if comp_ty_col:
            v = ly_ws.cell(r, comp_ty_col).value
            if v is not None and not is_formula(v):
                row_data["comp_set_ly"] = v

        if row_data:
            out[this_year_date] = row_data

    return out


def build_date_row_map(wb, prefer_sheet=None, fallback_to_wkone=True):
    """Build {date: row_number} using auto-detected date column, trying
    prefer_sheet (the tab actually being updated) first and falling back to
    WKONE. When subsequent rows contain formulas (=C5+1 style), extrapolates
    from the first real date so the full year is mapped correctly.

    Confirmed real case: Provincetown Harbor Hotel's WKONE date column was
    stuck a year behind (2024/2025) while WKFOUR — the sheet actually being
    updated — had its own correct, independently populated 2025/2026
    column. Always reading WKONE broke every week's update, even weeks
    whose own calendar was perfectly fine, because their own date column
    was never even looked at.

    Set fallback_to_wkone=False when the caller is about to WRITE
    current-cycle data and prefer_sheet's own column must be the sole
    source of truth — no cross-sheet fallback allowed, since writing based
    on a different sheet's calendar than the one being edited risks
    silently landing data on the wrong row if the two ever drift apart.
    """
    candidates = [prefer_sheet] if prefer_sheet and prefer_sheet in wb.sheetnames else []
    if fallback_to_wkone and "WKONE" in wb.sheetnames:
        candidates.append("WKONE")

    for sheet_name in candidates:
        ws = wb[sheet_name]
        date_col = detect_date_column(ws, wb=wb)
        mapping = {}
        anchor_date = None
        anchor_row  = None
        # Scan at least 400 rows regardless of this sheet's own max_row —
        # confirmed real case: a week tab whose date column is entirely
        # cross-sheet formulas referencing WKONE can have its own "used
        # range" end far short of WKONE's actual full-year extent (e.g. if
        # the formulas were only ever dragged down ~33 rows), truncating
        # every month after that even though WKONE and the CSV both cover
        # the full year. Reading past a sheet's real content just returns
        # blank cells (openpyxl doesn't error), so this is always safe.
        scan_end = max(ws.max_row, 400)
        for row_num in range(5, scan_end + 1):
            val = ws.cell(row_num, date_col).value
            if isinstance(val, datetime.datetime):
                d = val.date()
            elif isinstance(val, datetime.date):
                d = val
            elif isinstance(val, str) and val.startswith("="):
                # Resolves both local +N increments and cross-sheet
                # references (e.g. '=WKONE!C5') — non-WKONE week tabs
                # mirror WKONE's calendar this way by design, so this must
                # actually follow the reference, not just assume +1 day.
                d = _resolve_formula_date(wb, sheet_name, val)
                if d is None and anchor_row is not None:
                    # Unreadable cell, but the row-to-date relationship is a
                    # fixed, already-proven pattern (row N is always N minus
                    # anchor_row days after the anchor) — confirmed real
                    # case: Crowne Pointe's date formula chain was only ever
                    # filled down partway, leaving later rows genuinely
                    # blank even though the CSV and WKONE both have a full
                    # year of real data. Extrapolate instead of stopping.
                    d = anchor_date + datetime.timedelta(days=row_num - anchor_row)
                elif d is None:
                    continue
            elif anchor_row is not None:
                # Truly blank cell past the established anchor — same
                # extrapolation as above.
                d = anchor_date + datetime.timedelta(days=row_num - anchor_row)
            else:
                continue
            if anchor_date is None:
                anchor_date = d
                anchor_row  = row_num
            mapping[d] = row_num
        if mapping:
            # Auto-correct a stale year instead of requiring the sheet's
            # own cells to be manually fixed every cycle — confirmed real
            # case: Harbor Hotel's WKONE date column was genuinely stuck at
            # 2025 (never advanced when the year rolled to 2026). If the
            # detected column's anchor year doesn't match the current year,
            # shift every mapped date forward by that many years. A
            # correctly-dated column's anchor is already close to today
            # (that's what detect_date_column's proximity scoring picks
            # for), so this is a no-op in the normal case.
            year_shift = datetime.date.today().year - anchor_date.year
            if year_shift != 0:
                shifted = {}
                for d, r in mapping.items():
                    try:
                        shifted[d.replace(year=d.year + year_shift)] = r
                    except ValueError:
                        # Feb 29 with no Feb 29 in the shifted year
                        shifted[d.replace(month=2, day=28, year=d.year + year_shift)] = r
                mapping = shifted
            return mapping
    return {}


def find_otb_date_cell(ws):
    """Return (row, col) of the as-of date cell — one row above the OTB/TRANS header.
    Tries both Plymouth-style (OTB TY / TRANS) and Long Beach-style (TRANS / SOLD).
    """
    col_map = detect_strategy_columns(ws)
    otb_col = col_map.get("otb_trans")
    if otb_col:
        # Find which of rows 3 or 4 has the header, then go one above
        for r in range(2, 6):
            v = str(ws.cell(r, otb_col).value or "").strip().upper()
            if "OTB" in v or "TRANS" in v or "SOLD" in v:
                return r - 1, otb_col
    return 2, 4  # fallback


def _extract_otb_trans_by_date(wb, sheet_name, from_date):
    """Read OTB TY Trans keyed by date from an in-memory workbook.
    Uses build_date_row_map for date→row resolution so formula-based date
    cells (=C5+1 style) are handled via anchor extrapolation.
    Only returns rows where date >= from_date.
    """
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    col_map = detect_strategy_columns(ws)
    otb_col = col_map.get("otb_trans")
    if not otb_col:
        return {}
    date_row_map = build_date_row_map(wb, prefer_sheet=sheet_name)
    out = {}
    for d, r in date_row_map.items():
        if d < from_date:
            continue
        val = ws.cell(r, otb_col).value
        if val is not None and not is_formula(val):
            out[d] = safe_float(val)
    return out


def _extract_ly_data_from_wb(ly_wb, sheet_name, ty_wb=None):
    """Read prior-year Strategy values and align them to TY by weekday.

    Portfolio-wide STLY rule: a TY date receives the value from the nearest
    date in the prior calendar year that falls on the SAME weekday.
    Example: Tue 09/01/2026 <- Tue 09/02/2025.
    """
    if sheet_name not in ly_wb.sheetnames:
        return {}

    ly_ws = ly_wb[sheet_name]
    ly_col_map = detect_strategy_columns(ly_ws)
    ly_date_map = build_date_row_map(
        ly_wb, prefer_sheet=sheet_name, fallback_to_wkone=False
    )
    if not ly_date_map:
        return {}

    ty_date_map = {}
    if ty_wb and sheet_name in ty_wb.sheetnames:
        ty_date_map = build_date_row_map(
            ty_wb, prefer_sheet=sheet_name, fallback_to_wkone=False
        )

    if not ty_date_map:
        # Defensive fallback for callers without TY workbook context.
        for src_date in ly_date_map:
            try:
                nominal = datetime.date(src_date.year + 1, src_date.month, src_date.day)
            except ValueError:
                nominal = datetime.date(src_date.year + 1, 2, 28)
            for offset in range(-3, 4):
                candidate = nominal + datetime.timedelta(days=offset)
                if candidate.weekday() == src_date.weekday():
                    ty_date_map[candidate] = None
                    break

    comp_ty_col, _ = detect_comp_set_columns(ly_ws, ly_col_map)
    source_dates = sorted(ly_date_map.keys())
    out = {}

    for ty_date in sorted(ty_date_map.keys()):
        prior_year = ty_date.year - 1
        try:
            nominal = datetime.date(prior_year, ty_date.month, ty_date.day)
        except ValueError:
            nominal = datetime.date(prior_year, 2, 28)

        candidates = [
            d for d in source_dates
            if d.year == prior_year
            and d.weekday() == ty_date.weekday()
            and abs((d - nominal).days) <= 7
        ]
        if candidates:
            src_date = min(candidates, key=lambda d: (abs((d - nominal).days), d))
        else:
            # If the source workbook has a bad printed year, preserve weekday
            # alignment and choose the nearest month/day as a fallback.
            candidates = [d for d in source_dates if d.weekday() == ty_date.weekday()]
            if not candidates:
                continue
            src_date = min(
                candidates,
                key=lambda d: (
                    abs(d.month - nominal.month) * 31 + abs(d.day - nominal.day),
                    abs(d.year - prior_year),
                ),
            )

        src_row = ly_date_map[src_date]
        row_data = {}
        for ly_dest, ty_src in LY_FROM_TY.items():
            src_col = ly_col_map.get(ty_src)
            if not src_col:
                continue
            val = ly_ws.cell(src_row, src_col).value
            if val is not None and not is_formula(val):
                if ly_dest == "casino_ballroom_ly":
                    row_data[ly_dest] = val
                else:
                    row_data[ly_dest] = safe_float(val)

        if comp_ty_col:
            val = ly_ws.cell(src_row, comp_ty_col).value
            if val is not None and not is_formula(val):
                row_data["comp_set_ly"] = val

        if row_data:
            out[ty_date] = row_data

    return out


def build_strategy_change_plan(df, wb, sheet_name, prev_month_wb=None, ly_wb=None,
                               scope_start=None, scope_end=None):
    """Build strategy changes.
    prev_month_wb: in-memory previous month's SR workbook (for OTB Lst Wek on WKONE)
    ly_wb:         in-memory last year's SR workbook (for all LY columns, every week)
    """
    today = datetime.date.today()
    if scope_start is None:
        scope_start = today.replace(day=1)
    if scope_end is None:
        scope_end = datetime.date(today.year + 1, 12, 31)

    date_row_map = build_date_row_map(wb, prefer_sheet=sheet_name)
    ws = wb[sheet_name]

    # Detect actual column positions from headers — no guessing
    col_map = detect_strategy_columns(ws)
    ly_only_fields = {"otb_lst_wk", "otb_ly_trans", "grp_pu_ly", "grp_npu_ly",
                      "trans_rev_ly", "grp_rev_ly", "grp_npu_rev_ly", "casino_ballroom_ly"}
    # Casino ballroom is optional — only some hotels (e.g., Hampton Beach) have it
    optional_fields = {"casino_ballroom", "casino_ballroom_ly"}
    missing = [f for f, c in col_map.items() if c is None and f not in ly_only_fields and f not in optional_fields]
    if missing:
        st.warning(f"Strategy: could not locate columns for: {', '.join(missing)}")

    # Comp set columns in current sheet
    comp_ty_col_cur, comp_ly_col_cur = detect_comp_set_columns(ws, col_map)

    # OTB Lst Wek — WKONE only, from previous month's SR (already in memory)
    prev_otb_map = {}
    src_sheet = None
    if sheet_name == "WKONE" and col_map.get("otb_lst_wk") and prev_month_wb:
        # Use last FILLED tab (opposite of first_undone) — checked by actual
        # OTB data, not tab color (see first_undone_strategy_sheet: the same
        # color can mark a genuinely completed week OR sit untouched on a
        # never-used master template, so color alone can't tell "filled").
        last_filled = None
        for s in STRATEGY_SHEETS:
            if s not in prev_month_wb.sheetnames:
                continue
            pcol_map = detect_strategy_columns(prev_month_wb[s])
            potb_col = pcol_map.get("otb_trans")
            if potb_col and any(isinstance(prev_month_wb[s].cell(r, potb_col).value, (int, float)) for r in range(5, 15)):
                last_filled = s
        src_sheet = last_filled or (STRATEGY_SHEETS[-1] if STRATEGY_SHEETS[-1] in prev_month_wb.sheetnames else None)
        if src_sheet:
            prev_otb_map = _extract_otb_trans_by_date(prev_month_wb, src_sheet, scope_start)

    # LY data — every week, aligned by weekday rather than calendar date.
    ly_data = {}
    if ly_wb:
        ly_data = _extract_ly_data_from_wb(ly_wb, sheet_name, ty_wb=wb)

    # Ashworth Casino Ballroom TY/current events carry forward week-over-week.
    # Prior-year Casino events are sourced separately from ly_wb above.
    casino_prev_map = {}
    is_ashworth_sheet = "ASHWORTH" in _strategy_norm(ws["A1"].value)
    if is_ashworth_sheet and sheet_name in STRATEGY_SHEETS:
        sheet_idx = STRATEGY_SHEETS.index(sheet_name)
        if sheet_idx > 0:
            prev_sheet_name = STRATEGY_SHEETS[sheet_idx - 1]
            if prev_sheet_name in wb.sheetnames:
                prev_ws_same_month = wb[prev_sheet_name]
                prev_cols = detect_strategy_columns(prev_ws_same_month)
                prev_casino_col = prev_cols.get("casino_ballroom")
                if prev_casino_col:
                    prev_date_rows = build_date_row_map(
                        wb, prefer_sheet=prev_sheet_name, fallback_to_wkone=False
                    )
                    for prev_date, prev_row in prev_date_rows.items():
                        val = prev_ws_same_month.cell(prev_row, prev_casino_col).value
                        if val is not None and not is_formula(val):
                            casino_prev_map[prev_date] = val

    changes = []

    # Today's date above the OTB TY TRANS header
    date_row, date_col = find_otb_date_cell(ws)
    if date_row >= 1:
        changes.append({
            "date": today, "row": date_row, "col": date_col,
            "label": "As-of date", "new_value": today,
            "skip_reason": "formula" if is_formula(ws.cell(date_row, date_col).value) else None,
        })

    # As-of date above OTB Lst Wek header (row above row 3, i.e. row 2)
    lst_wk_col = col_map.get("otb_lst_wk")
    if lst_wk_col and prev_month_wb and src_sheet:
        prev_ws = prev_month_wb[src_sheet]
        prev_date_row, _ = find_otb_date_cell(prev_ws)
        src_date = prev_ws.cell(prev_date_row, date_col).value if prev_date_row >= 1 else None
        if src_date is None:
            # fallback: look in row above the lst_wk header in source
            for hr in range(2, 6):
                if str(prev_ws.cell(hr, lst_wk_col).value or "").strip():
                    src_date = prev_ws.cell(hr - 1, lst_wk_col).value
                    break
        if src_date:
            hdr_row = next((r for r in range(2, 6) if str(ws.cell(r, lst_wk_col).value or "").strip()), 3)
            label_row = hdr_row - 1
            if label_row >= 1:
                changes.append({
                    "date": None, "row": label_row, "col": lst_wk_col,
                    "label": "OTB Lst Wek as-of date",
                    "new_value": src_date,
                    "skip_reason": "formula" if is_formula(ws.cell(label_row, lst_wk_col).value) else None,
                })

    # As-of date above OTB LY TRANS header
    ly_trans_col = col_map.get("otb_ly_trans")
    if ly_trans_col and ly_wb and sheet_name in ly_wb.sheetnames:
        ly_ws_src = ly_wb[sheet_name]
        ly_date_row, _ = find_otb_date_cell(ly_ws_src)
        ly_src_date = ly_ws_src.cell(ly_date_row, date_col).value if ly_date_row >= 1 else None
        if ly_src_date:
            if isinstance(ly_src_date, datetime.datetime):
                ly_src_date = ly_src_date.date()
            hdr_row = next((r for r in range(2, 6) if str(ws.cell(r, ly_trans_col).value or "").strip()), 3)
            label_row = hdr_row - 1
            if label_row >= 1:
                changes.append({
                    "date": None, "row": label_row, "col": ly_trans_col,
                    "label": "OTB LY Trans as-of date",
                    "new_value": ly_src_date,
                    "skip_reason": None,
                })

    # ── CSV-sourced TY columns (only when BOB uploaded) ──────────────────────
    # Uses ONLY this sheet's own date column (no WKONE fallback) — this is
    # current-cycle data being written, and it must always be placed
    # against the calendar actually printed on the sheet being edited.
    own_date_row_map = build_date_row_map(wb, prefer_sheet=sheet_name, fallback_to_wkone=False)
    for _, row in (df.iterrows() if df is not None else []):
        date_str = str(row[0]).strip() if row[0] else ""
        kind, info = classify_row(date_str)
        if kind != "daily":
            continue
        d = info
        if d < scope_start or d > scope_end:
            continue
        if d not in own_date_row_map:
            continue
        excel_row = own_date_row_map[d]
        for field, (csv_col, label) in STRATEGY_CSV_COLS.items():
            excel_col = col_map.get(field)
            if excel_col is None:
                continue
            val = safe_float(row[csv_col])
            # BOB data is always authoritative — write even if cell has a template formula
            changes.append({
                "date": d, "row": excel_row, "col": excel_col,
                "label": label, "new_value": val, "skip_reason": None,
            })

    # ── Drive-sourced columns (run every time, no CSV needed) ─────────────────
    all_dates = set(date_row_map.keys()) & (set(prev_otb_map) | set(ly_data) | set(casino_prev_map))
    for d in sorted(all_dates):
        if d < scope_start or d > scope_end:
            continue
        excel_row = date_row_map[d]

        # OTB Lst Wek — WKONE only
        lst_wk_col = col_map.get("otb_lst_wk")
        if lst_wk_col and d in prev_otb_map:
            skip = "formula" if is_formula(ws.cell(excel_row, lst_wk_col).value) else None
            changes.append({
                "date": d, "row": excel_row, "col": lst_wk_col,
                "label": "OTB Lst Wek", "new_value": prev_otb_map[d], "skip_reason": skip,
            })

        # Ashworth Casino Ballroom TY/current event from the prior week tab.
        casino_ty_col = col_map.get("casino_ballroom")
        if casino_ty_col and d in casino_prev_map:
            changes.append({
                "date": d, "row": excel_row, "col": casino_ty_col,
                "label": "Casino Ballroom TY (week-over-week)",
                "new_value": casino_prev_map[d], "skip_reason": None,
            })

        # LY columns
        if d in ly_data:
            row_ly = ly_data[d]
            for ly_field, ly_label in [
                ("otb_ly_trans",   "OTB LY Trans"),
                ("grp_pu_ly",      "GRP PU LY"),
                ("grp_npu_ly",     "GRP N/PU LY"),
                ("trans_rev_ly",   "LY Trans Rev"),
                ("grp_rev_ly",     "GRP LY Rev"),
                ("grp_npu_rev_ly", "GRP N/PU LY Rev"),
                ("casino_ballroom_ly", "Casino Ballroom LY"),
            ]:
                dest_col = col_map.get(ly_field)
                if dest_col and ly_field in row_ly:
                    changes.append({
                        "date": d, "row": excel_row, "col": dest_col,
                        "label": ly_label, "new_value": row_ly[ly_field], "skip_reason": None,
                    })

            # Comp set far-right (LY) ← far-left (TY) from last year
            if comp_ly_col_cur and "comp_set_ly" in row_ly:
                changes.append({
                    "date": d, "row": excel_row, "col": comp_ly_col_cur,
                    "label": "Comp Set LY", "new_value": row_ly["comp_set_ly"], "skip_reason": None,
                })

    # ── Blank LY cells with no confirmed LY data — don't leave stale leftovers
    # from whatever the template last held (e.g. a day with no LY rates). Only
    # runs when last year's SR actually loaded AND we extracted LY data (not
    # just a regular weekly upload). Skip if ly_data is empty (extraction failed).
    if ly_wb and ly_data:
        ly_field_labels = [
            ("otb_ly_trans",   "OTB LY Trans"),
            ("grp_pu_ly",      "GRP PU LY"),
            ("grp_npu_ly",     "GRP N/PU LY"),
            ("trans_rev_ly",   "LY Trans Rev"),
            ("grp_rev_ly",     "GRP LY Rev"),
            ("grp_npu_rev_ly", "GRP N/PU LY Rev"),
            ("casino_ballroom_ly", "Casino Ballroom LY"),
        ]
        for d, excel_row in date_row_map.items():
            if d < scope_start or d > scope_end:
                continue
            row_ly = ly_data.get(d, {})
            for ly_field, ly_label in ly_field_labels:
                dest_col = col_map.get(ly_field)
                if not dest_col or ly_field in row_ly:
                    continue
                cur_val = ws.cell(excel_row, dest_col).value
                if cur_val is None or is_formula(cur_val):
                    continue
                changes.append({
                    "date": d, "row": excel_row, "col": dest_col,
                    "label": f"{ly_label} (no LY data — cleared)",
                    "new_value": None, "skip_reason": None,
                })
            if comp_ly_col_cur and "comp_set_ly" not in row_ly:
                cur_val = ws.cell(excel_row, comp_ly_col_cur).value
                if cur_val is not None and not is_formula(cur_val):
                    changes.append({
                        "date": d, "row": excel_row, "col": comp_ly_col_cur,
                        "label": "Comp Set LY (no LY data — cleared)",
                        "new_value": None, "skip_reason": None,
                    })

    return changes


def apply_strategy_changes(wb, sheet_name, changes):
    ws = wb[sheet_name]
    for ch in changes:
        if ch["skip_reason"]:
            continue
        ws.cell(ch["row"], ch["col"]).value = ch["new_value"]


def parse_rate_csv(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, encoding="utf-8-sig")
    df.columns = [c.strip() for c in df.columns]
    return df


def find_header_col(ws, keyword, header_rows=(2, 3, 4)):
    """Find the leftmost column whose concatenated header text (rows 2-4) contains keyword."""
    keyword = keyword.strip().lower()
    col_texts = {}
    for row_num in header_rows:
        for cell in ws[row_num]:
            col = cell.column
            if col not in col_texts:
                col_texts[col] = ""
            if isinstance(cell.value, str):
                col_texts[col] += cell.value.strip()
    matches = [col for col, text in col_texts.items() if keyword in text.lower()]
    return min(matches) if matches else None


def find_restrictions_col(ws, upto_col=None):
    """Find the 'Restrictions' header column, scanning rows 3-4.
    The header can be split across those two rows AND hyphenated for
    word-wrap (e.g. 'Restric'+'tions', or 'Restri-'+'ctions' — confirmed on
    a real hotel's sheet, ALLEGRIA/'Long Beach'). Stripping spaces and
    hyphens before matching handles both forms; a plain substring search on
    the raw concatenation misses the hyphenated one and returns nothing.

    Also confirmed a master template with the header simply misspelled —
    'Restictions' (missing the second 'r') on Hampton/Ashworth By The Sea, in
    a single cell, no split at all. Rather than chase every possible typo,
    match on 'REST' + 'TION' both present, which covers correct spelling,
    the hyphenated/split form, and this typo alike.
    """
    upto_col = upto_col or ws.max_column
    for c in range(1, upto_col + 1):
        combined_hdr = (str(ws.cell(3, c).value or "") + str(ws.cell(4, c).value or "")).upper().replace(" ", "").replace("-", "")
        if "REST" in combined_hdr and "TION" in combined_hdr:
            return c
    return None


def _strategy_norm(v):
    return re.sub(r"[^A-Z0-9]", "", str(v or "").upper())


def _strategy_hotel_aliases(hotel_name):
    """Return safe aliases for the selected Strategy hotel."""
    base = _strategy_norm(hotel_name)
    aliases = {base}

    for _portfolio, members in PORTFOLIO_HOTELS.items():
        for label, keywords in members.items():
            label_norm = _strategy_norm(label)
            keyword_norms = {_strategy_norm(k) for k in keywords}

            # The UI label may be shorter than the workbook header, e.g.
            # "Middletown" vs "Inn at Middletown". Match the portfolio entry
            # by containment as well as exact label equality.
            belongs = (
                label == hotel_name
                or (base and (base in label_norm or label_norm in base))
                or any(base and (base in k or k in base) for k in keyword_norms)
            )
            if belongs:
                aliases.add(label_norm)
                aliases.update(keyword_norms)

    return {a for a in aliases if len(a) >= 4}


def find_strategy_hotel_rate_restriction_cols(ws, hotel_name):
    """Find selected hotel's SNT Rate and Restrictions columns.

    Real SNT Strategy layout uses rows 1-4 as headers. Some labels are split
    vertically, e.g. X3='Restric' and X4='tions'. The selected hotel's own
    rate is the column headed by that hotel's name.
    """
    aliases = _strategy_hotel_aliases(hotel_name)
    header_max_row = min(4, ws.max_row)

    def col_header_text(col):
        parts = []
        for r in range(1, header_max_row + 1):
            v = ws.cell(r, col).value
            if v not in (None, ""):
                parts.append(str(v).strip())
        return " ".join(parts).strip()

    def matches_hotel(value):
        n = _strategy_norm(value)
        return bool(n) and any(
            a == n or a in n or n in a
            for a in aliases
        )

    # Ashworth/Hampton has a Casino Ballroom block before its far-right
    # rate-shopping section. Its own SNT rate belongs in the exact ASH column.
    sheet_title = _strategy_norm(ws["A1"].value)
    requested = _strategy_norm(hotel_name)
    is_ashworth = (
        "ASHWORTH" in sheet_title
        or "ASHWORTH" in requested
        or "HAMPTON" in requested
    )
    if is_ashworth:
        rate_col = None
        for c in range(1, ws.max_column + 1):
            hdr = " ".join(
                str(ws.cell(r, c).value).strip()
                for r in range(1, header_max_row + 1)
                if ws.cell(r, c).value not in (None, "")
            )
            if _strategy_norm(hdr) == "ASH":
                rate_col = c
                break
        restric_col = _find_restrictions_col(ws)
        if rate_col:
            return rate_col, restric_col, (
                f"Matched Ashworth Strategy rate to ASH column "
                f"(rate col {rate_col}, restrictions col {restric_col})"
            )

    hotel_candidates = []
    for c in range(1, ws.max_column + 1):
        header = col_header_text(c)
        if matches_hotel(header):
            hotel_candidates.append((c, header))

    # Merged hotel headings, if present.
    for rng in ws.merged_cells.ranges:
        if rng.min_row > header_max_row:
            continue
        value = ws.cell(rng.min_row, rng.min_col).value
        if matches_hotel(value):
            for c in range(rng.min_col, rng.max_col + 1):
                hotel_candidates.append((c, str(value or "").strip()))

    if not hotel_candidates:
        return None, None, (
            f"Could not match selected hotel '{hotel_name}' to a Strategy Report header."
        )

    rate_col, matched_header = min(hotel_candidates, key=lambda x: x[0])

    restriction_candidates = []
    for c in range(1, ws.max_column + 1):
        compact = re.sub(r"[^A-Z]", "", col_header_text(c).upper())
        if "RESTRICTIONS" in compact or compact.startswith("RESTRIC"):
            restriction_candidates.append(c)

    restric_col = None
    left = [c for c in restriction_candidates if c < rate_col]
    if left:
        restric_col = max(left)
    elif restriction_candidates:
        restric_col = min(
            restriction_candidates,
            key=lambda c: abs(c - rate_col)
        )

    diag = (
        f"Matched '{hotel_name}' to Strategy header '{matched_header}' "
        f"(rate col {rate_col}"
        + (f", restrictions col {restric_col})"
           if restric_col else ", restrictions not found)")
    )
    return rate_col, restric_col, diag


def parse_lighthouse_rates_xlsx(file_bytes):
    """Parse a Lighthouse rate-shopping workbook."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Rates" not in wb.sheetnames:
        raise ValueError("Lighthouse workbook does not contain a 'Rates' tab.")

    ws = wb["Rates"]
    header_row = None
    date_col = None
    for r in range(1, min(ws.max_row, 20) + 1):
        row_vals = [str(ws.cell(r, c).value or "").strip().lower()
                    for c in range(1, ws.max_column + 1)]
        if "date" in row_vals and "day" in row_vals:
            header_row = r
            date_col = row_vals.index("date") + 1
            break

    if header_row is None or date_col is None:
        raise ValueError(
            "Could not locate the hotel/date header row on the Lighthouse 'Rates' tab."
        )

    hotel_cols = {}
    for c in range(date_col + 2, ws.max_column + 1):
        name = str(ws.cell(header_row, c).value or "").strip()
        if not name:
            continue
        if _strategy_norm(name) in {"RATECHANGES", "GUESTS", "UPDATED", "MARKETDEMAND"}:
            continue
        hotel_cols[c] = name

    if not hotel_cols:
        raise ValueError("No hotel rate columns were found on the Lighthouse 'Rates' tab.")

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        d = parse_any_date(ws.cell(r, date_col).value)
        if not d:
            continue
        rates = {}
        for c, hotel in hotel_cols.items():
            val = ws.cell(r, c).value
            if val is None or val == "":
                continue
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                rates[hotel] = float(val)
            else:
                rates[hotel] = str(val).strip()
        rows.append({"date": d, "rates": rates})

    return {"hotels": list(hotel_cols.values()), "rows": rows}


def _strategy_name_tokens(value):
    raw = re.findall(r"[A-Z0-9]+", str(value or "").upper())
    stop = {"THE", "HOTEL", "INN", "AND", "BY", "AT", "OF", "SUITES", "RESORT", "SPA"}
    return {t for t in raw if len(t) >= 3 and t not in stop}


def _strategy_names_match(a, b):
    na = _strategy_norm(a)
    nb = _strategy_norm(b)
    if not na or not nb:
        return False
    if na == nb or na in nb or nb in na:
        return True
    ta = _strategy_name_tokens(a)
    tb = _strategy_name_tokens(b)
    if not ta or not tb:
        return False
    overlap = ta & tb
    if len(overlap) >= 2:
        return True
    if len(ta) == 1 and len(tb) == 1 and ta == tb:
        return True
    return False


def _strategy_header_candidates(ws):
    out = []
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= 4 and rng.max_row >= 1:
            label = str(ws.cell(rng.min_row, rng.min_col).value or "").strip()
            if label:
                out.append({"label": label, "start": rng.min_col, "end": rng.max_col})
    for r in range(1, 5):
        for c in range(1, ws.max_column + 1):
            label = str(ws.cell(r, c).value or "").strip()
            if label:
                out.append({"label": label, "start": c, "end": c})
    return out


def find_strategy_compset_rate_col(ws, lighthouse_hotel):
    """Find an existing Strategy competitor rate column by hotel name.

    Only rows 1-4 are considered header rows. Extra Lighthouse competitors
    that are not already present in the Strategy Report are ignored.
    """
    header_max_row = min(4, ws.max_row)

    def combined(col):
        return " ".join(
            str(ws.cell(r, col).value or "").strip()
            for r in range(1, header_max_row + 1)
            if ws.cell(r, col).value not in (None, "")
        ).strip()

    candidates = []
    for c in range(1, ws.max_column + 1):
        header = combined(c)
        if not header:
            continue
        compact = re.sub(r"[^A-Z]", "", header.upper())
        if compact.startswith("RESTRIC"):
            continue
        if _strategy_names_match(lighthouse_hotel, header):
            candidates.append((c, header))

    if not candidates:
        return None, None

    return min(candidates, key=lambda x: x[0])


def build_lighthouse_compset_change_plan(lighthouse_data, wb, sheet_name, hotel_name):
    if not lighthouse_data:
        return [], []

    ws = wb[sheet_name]
    date_row_map = build_date_row_map(wb, prefer_sheet=sheet_name)
    if not date_row_map:
        return [], ["Lighthouse: no Strategy Report date rows could be mapped."]

    selected_aliases = _strategy_hotel_aliases(hotel_name)
    changes = []
    warnings = []
    matched = {}
    ignored = []

    for lh_hotel in lighthouse_data.get("hotels", []):
        lh_norm = _strategy_norm(lh_hotel)
        if any(alias == lh_norm or alias in lh_norm or lh_norm in alias for alias in selected_aliases):
            continue

        rate_col, strategy_label = find_strategy_compset_rate_col(ws, lh_hotel)
        if not rate_col:
            ignored.append(lh_hotel)
            continue
        if rate_col in matched.values():
            ignored.append(lh_hotel)
            continue
        matched[lh_hotel] = rate_col

    for entry in lighthouse_data.get("rows", []):
        d = entry.get("date")
        if d not in date_row_map:
            continue
        excel_row = date_row_map[d]

        for lh_hotel, value in entry.get("rates", {}).items():
            rate_col = matched.get(lh_hotel)
            if not rate_col:
                continue
            skip = "formula" if is_formula(ws.cell(excel_row, rate_col).value) else None
            changes.append({
                "date": d,
                "row": excel_row,
                "col": rate_col,
                "label": f"Lighthouse Rate — {lh_hotel}",
                "new_value": value,
                "skip_reason": skip,
            })

    if matched:
        warnings.append("Lighthouse compset matched: " + ", ".join(matched.keys()))
    else:
        warnings.append(
            "Lighthouse compset: none of the competitor hotel names matched existing Strategy Report headers."
        )
    if ignored:
        warnings.append(
            "Lighthouse compset ignored (not already in Strategy Report or ambiguous): " + ", ".join(ignored)
        )
    return changes, warnings


def _snt_rate_and_mlos_from_row(row):
    """Read SNT rate + minimum-stay fields without relying on one exact header.

    Different SNT exports have used headers such as Double, Double Rate, Rate,
    BAR, Min Length of Stay, Minimum Length of Stay, and MLOS. Restrictions
    could still work while the rate silently read None if only the rate header
    changed, which is what happened on Brass Key.
    """
    def norm_header(v):
        return re.sub(r"[^a-z0-9]", "", str(v or "").lower())

    rate = None
    mlos = None

    # Exact/common names first.
    for key in ("Double", "Double Rate", "Rate", "BAR", "BAR Rate"):
        if key in row.index:
            v = safe_float(row.get(key))
            if v is not None:
                rate = v
                break

    for key in ("Min Length of Stay", "Minimum Length of Stay", "MLOS", "Min LOS"):
        if key in row.index:
            v = safe_float(row.get(key))
            if v is not None:
                mlos = v
                break

    # Fuzzy fallback for vendor-header variations.
    if rate is None or mlos is None:
        for col in row.index:
            nh = norm_header(col)
            val = row.get(col)

            if rate is None:
                is_rate = (
                    nh in {"double", "doublerate", "rate", "bar", "barrate"}
                    or ("double" in nh and "rate" in nh)
                )
                # Avoid things like "Rate Change" / "Average Rate".
                if is_rate and "change" not in nh and "average" not in nh:
                    v = safe_float(val)
                    if v is not None:
                        rate = v

            if mlos is None:
                is_mlos = (
                    "minlengthofstay" in nh
                    or "minimumlengthofstay" in nh
                    or nh == "mlos"
                    or "minlos" in nh
                )
                if is_mlos:
                    v = safe_float(val)
                    if v is not None:
                        mlos = v

    return rate, mlos


def build_rates_change_plan(rate_df, wb, sheet_name, hotel_name=None):
    today = datetime.date.today()
    # include previous month — final numbers arrive on the 1st of the following month
    prev_month_start = (today.replace(day=1) - datetime.timedelta(days=1)).replace(day=1)
    scope_start = prev_month_start

    date_row_map = build_date_row_map(wb, prefer_sheet=sheet_name)
    # Cover the sheet's actual date range, not a hardcoded Dec 31 — the SR's
    # rolling 12 months usually crosses into the next calendar year (e.g. a
    # July setup runs into next June), and a hardcoded year-end cutoff was
    # silently dropping every month past December, regardless of what real
    # data the CSV had for those dates.
    scope_end = max(date_row_map.keys()) if date_row_map else datetime.date(today.year, 12, 31)
    ws = wb[sheet_name]

    changes = []
    warnings = []
    hotel_col = restric_col = None
    if hotel_name:
        hotel_col, restric_col, diag = find_strategy_hotel_rate_restriction_cols(ws, hotel_name)
        warnings.append(diag)
    else:
        warnings.append("No hotel name supplied for Rates & Restrictions mapping.")
    if not hotel_col:
        warnings.append(f"Could not locate the Rate column for selected hotel '{hotel_name or 'Unknown'}'; rates will not be changed.")
    if not restric_col:
        warnings.append(f"Could not locate the Restrictions column for selected hotel '{hotel_name or 'Unknown'}'; restrictions will not be changed.")

    for _, row in rate_df.iterrows():
        # SNT exports can represent Date as M/D/YYYY, ISO text, an Excel
        # date, or a pandas/datetime value depending on how the CSV was saved.
        # Use the app-wide parser instead of accepting only three string formats.
        raw_date = row.get("Date", "")
        d = parse_any_date(raw_date)
        if d is None:
            # pandas may stringify Timestamp values with a time suffix.
            try:
                parsed = pd.to_datetime(raw_date, errors="coerce")
                if not pd.isna(parsed):
                    d = parsed.date()
            except Exception:
                d = None
        if d is None:
            continue
        if d < scope_start or d > scope_end:
            continue
        if d not in date_row_map:
            continue

        excel_row  = date_row_map[d]
        double_val, mlos_val = _snt_rate_and_mlos_from_row(row)

        if hotel_col and double_val is not None:
            existing = ws.cell(excel_row, hotel_col).value
            skip = (
                "formula"
                if is_formula(existing) and "!" not in str(existing)
                else None
            )
            changes.append({
                "date": d, "row": excel_row, "col": hotel_col,
                "label": "Hotel Rate (SNT)", "new_value": double_val,
                "skip_reason": skip
            })
        if restric_col and mlos_val is not None:
            skip = "formula" if is_formula(ws.cell(excel_row, restric_col).value) else None
            changes.append({
                "date": d, "row": excel_row, "col": restric_col,
                "label": "Restrictions (MLOS)", "new_value": mlos_val,
                "skip_reason": skip
            })

    return changes, warnings


# ── Forecast helpers ─────────────────────────────────────────────────────────

DATE_FORMATS = [
    "%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%m/%d/%y", "%m-%d-%y",
    "%B %d, %Y", "%b %d, %Y", "%B %d %Y", "%b %d %Y",
    "%d-%b-%Y", "%d/%b/%Y",
]

def parse_any_date(val):
    """Parse a date from a datetime object, int serial, or string in many formats."""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.date() if isinstance(val, datetime.datetime) else val
    if isinstance(val, (int, float)):
        # Excel serial date (days since 1900-01-01, with Lotus bug offset)
        return (datetime.date(1899, 12, 30) + datetime.timedelta(days=int(val)))
    if isinstance(val, str):
        val = val.strip()
        for fmt in DATE_FORMATS:
            try:
                return datetime.datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def locate_forecast_rows(ws):
    """Find the As-of date / OTB Rooms Sold / ADR OTB / actual Rooms Sold /
    actual Revenue rows by reading column-A titles, instead of assuming fixed
    row numbers. Confirmed on a real workbook that this drifts within a single
    file: WK1 has an extra 'Occupancy' row that WK4/WK8/WK9 don't, shifting
    every row below it by one — a hardcoded row 14/16/19 lands on a blank row
    or the wrong label on most week tabs.

    'Rooms Sold' and 'Revenue' each appear multiple times in this template, so
    rows are resolved in reading order relative to the unique 'ADR OTB' label:
    the first 'Rooms Sold' above it is the OTB (future) entry row, the first
    'Rooms Sold' below it is the actual entry row, and the first 'Revenue'
    below that is the actual revenue row.

    Returns None if the expected labels aren't all found (caller should warn
    and skip rather than guess a row number).
    """
    labels = []
    for r in range(1, 31):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip():
            labels.append((r, v.strip().lower()))

    def find_after(text, after_row=0):
        for r, v in labels:
            if r > after_row and text in v:
                return r
        return None

    dow_row            = find_after("day of week")
    otb_rooms_row      = find_after("rooms sold")
    adr_otb_row        = find_after("adr otb")
    actual_rooms_row   = find_after("rooms sold", adr_otb_row) if adr_otb_row else None
    actual_revenue_row = find_after("revenue", actual_rooms_row) if actual_rooms_row else None

    if not all([dow_row, otb_rooms_row, adr_otb_row, actual_rooms_row, actual_revenue_row]):
        return None

    return {
        "as_of_row":          dow_row - 1,
        "date_row":           dow_row + 1,
        "otb_rooms_row":      otb_rooms_row,
        "adr_otb_row":        adr_otb_row,
        "actual_rooms_row":   actual_rooms_row,
        "actual_revenue_row": actual_revenue_row,
    }


def build_forecast_date_col_map(ws, wb=None, date_row=4):
    """Return {date: col_index} from date_row. Falls back to WK1 for formula-only sheets."""
    month_start = parse_any_date(ws.cell(date_row, 2).value)

    # If this sheet's col B is a formula, find the start date from any WK sheet with a literal
    if month_start is None and wb is not None:
        for sname in wb.sheetnames:
            if "glance" in sname.lower():
                continue
            candidate = parse_any_date(wb[sname].cell(date_row, 2).value)
            if candidate is not None:
                month_start = candidate
                break

    if month_start is None:
        return {}

    col_map = {}
    col = 2
    while col <= ws.max_column:
        cell = ws.cell(date_row, col)
        if isinstance(cell.value, str) and "total" in cell.value.lower():
            break
        if cell.value is None and col > 2:
            break
        d = month_start + datetime.timedelta(days=col - 2)
        # One workbook covers one month. The templates carry 31 date columns
        # and most chain them straight through (=prev+1), so in a 30-day month
        # the last one computes to the 1st of the *next* month — confirmed on a
        # real June file, where the map ran to 1 Jul. Some hotels blank that
        # cell by hand and some don't, so the month end is what decides, not
        # how many columns the sheet happens to have. Left in, a foreign day
        # lands inside this month's Totals column.
        if (d.year, d.month) != (month_start.year, month_start.month):
            break
        col_map[d] = col
        col += 1
    return col_map


def row_is_filled(ws, r):
    """Return True if cols B-AF contain actual numbers OR cross-sheet formula refs."""
    for col in range(2, 33):
        v = ws.cell(r, col).value
        if isinstance(v, (int, float)):
            return True
        if isinstance(v, str) and v.startswith("='"):
            return True
    return False


def find_next_pickup_data_row(ws):
    """Find the next available row in the pick-up tracking chart.

    Strategy:
    1. Locate the 'Day of Week' section header (last one before 'Total Pick UP').
    2. Within the section, find the last row that has actual numbers in B-AF
       (cross-sheet formula refs look empty to openpyxl, so only direct numbers count).
    3. Scan forward from there for the first row with no numbers and no skip keyword
       — works for WK1/WK2 (back-to-back entries) and WK3+ (Pick UP rows between).
    """
    skip_keywords = {"pick", "day", "total", "forecast", "budget", "last year"}

    # Find Total Pick UP boundary
    total_row = None
    for r in range(40, 150):
        if "total pick" in str(ws.cell(r, 1).value or "").lower():
            total_row = r
            break
    search_end = total_row if total_row else 150

    # Find last 'Day of Week' header before the boundary
    section_start = None
    for r in range(40, search_end):
        if "day of week" in str(ws.cell(r, 1).value or "").lower():
            section_start = r
    if section_start is None:
        return None

    # Find the last row in the section that has actual numbers in B-AF
    last_filled = None
    for r in range(section_start + 2, search_end):
        if row_is_filled(ws, r):
            last_filled = r

    if last_filled is None:
        # Nothing filled yet — return first non-header row in section
        last_filled = section_start + 1

    # Scan forward from last_filled+1 for the first row with no numbers
    for r in range(last_filled + 1, search_end):
        a_val = str(ws.cell(r, 1).value or "").strip().lower()
        if any(k in a_val for k in skip_keywords):
            continue
        if not row_is_filled(ws, r):
            return r

    return None


def extract_rob_month_end_data(rob_wb, target_month):
    """Pull Budget and Last Year Room Nights + Revenue for target_month from any ROB sheet."""
    for sheet_name in rob_wb.sheetnames:
        ws = rob_wb[sheet_name]
        # Scan for the month block header row (col A = month abbreviation, e.g. "Jul")
        month_abbr = target_month.strftime("%b").lower()
        for r in range(1, ws.max_row + 1):
            cell_val = str(ws.cell(r, 1).value or "").strip().lower()
            if cell_val != month_abbr:
                continue
            # Found the block header row — find Budget and Last Year columns
            budget_col = ly_col = None
            for c in range(1, 25):
                hdr = str(ws.cell(r, c).value or "").strip()
                if not budget_col and "budget" in hdr.lower():
                    budget_col = c
                elif not ly_col and ("month end" in hdr.lower() or ("last" in hdr.lower() and "year" in hdr.lower())):
                    ly_col = c
                if budget_col and ly_col:
                    break
            # Also check row 4 (global header) if not found on block row
            if not budget_col or not ly_col:
                for c in range(1, 25):
                    hdr = str(ws.cell(4, c).value or "").strip()
                    if not budget_col and "budget" in hdr.lower():
                        budget_col = c
                    elif not ly_col and "month end" in hdr.lower():
                        ly_col = c
                    if budget_col and ly_col:
                        break
            if not budget_col or not ly_col:
                continue
            # Scan the next ~8 rows for Revenue and Room Nights labels in col A
            rev_row = rms_row = None
            for dr in range(1, 9):
                label = str(ws.cell(r + dr, 1).value or "").strip().lower()
                if "revenue" in label and rev_row is None:
                    rev_row = r + dr
                elif ("room" in label or "night" in label) and rms_row is None:
                    rms_row = r + dr
                if rev_row and rms_row:
                    break
            if not rev_row or not rms_row:
                continue
            return {
                "budget_rev": safe_float(ws.cell(rev_row, budget_col).value),
                "budget_rms": safe_float(ws.cell(rms_row, budget_col).value),
                "ly_rev":     safe_float(ws.cell(rev_row, ly_col).value),
                "ly_rms":     safe_float(ws.cell(rms_row, ly_col).value),
            }
    return None


def find_month_ending_forecast_cells(fcst_ws):
    """Return cell coords for Budget/LY Room Nts and Revenue in the Month Ending Forecast table."""
    for r in range(1, fcst_ws.max_row + 1):
        for c in range(1, fcst_ws.max_column + 1):
            if "month ending forecast" in str(fcst_ws.cell(r, c).value or "").lower():
                # Header found — scan next ~6 rows for Budget and Last Year rows
                # and the header row for column positions
                hdr_row = col_rms = col_rev = None
                budget_row = ly_row = None
                for dr in range(1, 8):
                    row_vals = {col: str(fcst_ws.cell(r + dr, col).value or "").strip()
                                for col in range(1, 8)}
                    row_text = " ".join(row_vals.values()).lower()
                    if "room" in row_text and hdr_row is None:
                        hdr_row = r + dr
                        for col, v in row_vals.items():
                            if "room" in v.lower():
                                col_rms = col
                            elif "revenue" in v.lower() or "rev" in v.lower():
                                col_rev = col
                    if "budget" in row_text and budget_row is None:
                        budget_row = r + dr
                    if "last year" in row_text and ly_row is None:
                        ly_row = r + dr
                if col_rms and col_rev and budget_row and ly_row:
                    return col_rms, col_rev, budget_row, ly_row
    return None, None, None, None


def build_forecast_change_plan(df, ws, rob_wb=None, is_wk1=False):
    """Build list of cell writes for the Forecast sheet."""
    today = datetime.date.today()
    yesterday = today - datetime.timedelta(days=1)
    month_start = today.replace(day=1)

    rows = locate_forecast_rows(ws)
    if not rows:
        return [], ["Could not read row titles (As-of date / Rooms Sold / ADR OTB / Revenue) from forecast sheet."]

    col_map = build_forecast_date_col_map(ws, ws.parent, date_row=rows["date_row"])
    if not col_map:
        return [], ["Could not read date row from forecast sheet."]

    changes = []
    warnings = []

    # As-of date (row directly above "Day of Week")
    changes.append({
        "label": "As-of date", "row": rows["as_of_row"], "col": 1,
        "new_value": today, "skip_reason": "formula" if is_formula(ws.cell(rows["as_of_row"], 1).value) else None,
    })

    # Build lookup from CSV: date -> row dict
    # Uses the same flexible date matching as ROB/SR (classify_row) — the BOB
    # CSV's date format (1-2 digit month/day, "/" or "-") wasn't matching the
    # old hyphen-only, zero-padded regex here, so Forecast silently got no data.
    daily_rows = {}
    for _, row in df.iterrows():
        date_str = str(row.iloc[0]).strip()
        kind, d = classify_row(date_str)
        if kind != "daily":
            continue
        daily_rows[d] = row

    for d, col in col_map.items():
        if d not in daily_rows:
            continue
        csv_row = daily_rows[d]
        rms = safe_float(csv_row.iloc[1])
        adr = safe_float(csv_row.iloc[6])
        rev = safe_float(csv_row.iloc[5])

        is_future = d >= today
        is_past   = d <= yesterday and d >= month_start

        # Rooms Sold (future / OTB)
        if is_future:
            skip = "formula" if is_formula(ws.cell(rows["otb_rooms_row"], col).value) else None
            changes.append({"label": f"Rooms Sold (future) {d}", "row": rows["otb_rooms_row"], "col": col,
                            "new_value": rms, "skip_reason": skip})
            # ADR OTB (future)
            skip = "formula" if is_formula(ws.cell(rows["adr_otb_row"], col).value) else None
            changes.append({"label": f"ADR OTB {d}", "row": rows["adr_otb_row"], "col": col,
                            "new_value": adr, "skip_reason": skip})

        # Rooms Sold (actuals)
        if is_past:
            skip = "formula" if is_formula(ws.cell(rows["actual_rooms_row"], col).value) else None
            changes.append({"label": f"Rooms Sold (actual) {d}", "row": rows["actual_rooms_row"], "col": col,
                            "new_value": rms, "skip_reason": skip})
            # Revenue (actuals)
            skip = "formula" if is_formula(ws.cell(rows["actual_revenue_row"], col).value) else None
            changes.append({"label": f"Revenue (actual) {d}", "row": rows["actual_revenue_row"], "col": col,
                            "new_value": rev, "skip_reason": skip})

    # Pick-up tracking row: write full month rooms sold to next available row
    target_row = find_next_pickup_data_row(ws)

    if target_row:
        changes.append({"label": "Pickup tracking: date", "row": target_row, "col": 1,
                        "new_value": today, "skip_reason": None})
        for d, col in col_map.items():
            if d not in daily_rows:
                continue
            rms = safe_float(daily_rows[d].iloc[1])
            skip = "formula" if is_formula(ws.cell(target_row, col).value) else None
            changes.append({"label": f"Pickup tracking: Rooms Sold {d}",
                            "row": target_row, "col": col,
                            "new_value": rms, "skip_reason": skip})
    else:
        warnings.append("No available row found in pick-up tracking chart.")

    # Month Ending Forecast table — only on WK1, only when ROB workbook provided
    if is_wk1 and rob_wb is not None:
        target_month = today.replace(day=1)
        rob_data = extract_rob_month_end_data(rob_wb, target_month)
        if rob_data:
            col_rms, col_rev, budget_row, ly_row = find_month_ending_forecast_cells(ws)
            if col_rms and col_rev and budget_row and ly_row:
                entries = [
                    (budget_row, col_rms, "Month End Forecast: Budget Room Nts", rob_data["budget_rms"]),
                    (budget_row, col_rev, "Month End Forecast: Budget Revenue",  rob_data["budget_rev"]),
                    (ly_row,     col_rms, "Month End Forecast: LY Room Nts",     rob_data["ly_rms"]),
                    (ly_row,     col_rev, "Month End Forecast: LY Revenue",       rob_data["ly_rev"]),
                ]
                for r, c, label, val in entries:
                    skip = "formula" if is_formula(ws.cell(r, c).value) else None
                    changes.append({"label": label, "row": r, "col": c,
                                    "new_value": val, "skip_reason": skip})
            else:
                warnings.append("Could not locate Month Ending Forecast table in forecast sheet.")
        else:
            warnings.append("Could not find month data in ROB workbook for Month Ending Forecast.")

    return changes, warnings


def build_next_month_forecast_plan(df, ws):
    """For weeks 3 & 4: write Rooms Sold and ADR OTB (rows located by title,
    not a fixed offset) for ALL dates in the next month (everything in the CSV
    beyond the current month). Also writes the As-of date and pick-up tracking row.
    """
    today = datetime.date.today()
    current_month_end = (today.replace(day=1) + datetime.timedelta(days=32)).replace(day=1) - datetime.timedelta(days=1)

    rows = locate_forecast_rows(ws)
    if not rows:
        return [], ["Could not read row titles (As-of date / Rooms Sold / ADR OTB / Revenue) from next-month forecast sheet."]

    col_map = build_forecast_date_col_map(ws, ws.parent, date_row=rows["date_row"])
    if not col_map:
        return [], ["Could not read date row from next-month forecast sheet."]

    changes = []
    warnings = []

    # As-of date (row directly above "Day of Week")
    changes.append({
        "label": "As-of date", "row": rows["as_of_row"], "col": 1,
        "new_value": today,
        "skip_reason": "formula" if is_formula(ws.cell(rows["as_of_row"], 1).value) else None,
    })

    daily_rows = {}
    for _, row in df.iterrows():
        date_str = str(row.iloc[0]).strip()
        kind, d = classify_row(date_str)
        if kind != "daily":
            continue
        daily_rows[d] = row

    for d, col in col_map.items():
        # Only next month dates (after current month end)
        if d <= current_month_end:
            continue
        if d not in daily_rows:
            continue
        csv_row = daily_rows[d]
        rms = safe_float(csv_row.iloc[1])
        adr = safe_float(csv_row.iloc[6])

        skip_rms = "formula" if is_formula(ws.cell(rows["otb_rooms_row"], col).value) else None
        skip_adr = "formula" if is_formula(ws.cell(rows["adr_otb_row"], col).value) else None
        changes.append({"label": f"Rooms Sold (future) {d}", "row": rows["otb_rooms_row"], "col": col, "new_value": rms, "skip_reason": skip_rms})
        changes.append({"label": f"ADR OTB {d}",             "row": rows["adr_otb_row"],   "col": col, "new_value": adr, "skip_reason": skip_adr})

    # Pick-up tracking row (same logic — full month of next month dates)
    target_row = find_next_pickup_data_row(ws)
    if target_row:
        changes.append({"label": "Pickup tracking: date", "row": target_row, "col": 1,
                        "new_value": today, "skip_reason": None})
        for d, col in col_map.items():
            if d not in daily_rows:
                continue
            rms = safe_float(daily_rows[d].iloc[1])
            skip = "formula" if is_formula(ws.cell(target_row, col).value) else None
            changes.append({"label": f"Pickup tracking: Rooms Sold {d}",
                            "row": target_row, "col": col, "new_value": rms, "skip_reason": skip})
    else:
        warnings.append("No available row found in next-month pick-up tracking chart.")

    return changes, warnings


def apply_forecast_changes(wb, sheet_name, changes):
    ws = wb[sheet_name]
    for ch in changes:
        if ch["skip_reason"]:
            continue
        ws.cell(ch["row"], ch["col"]).value = ch["new_value"]


# ── Google Drive ──────────────────────────────────────────────────────────────

MULTI_ID_PREFIX = "MULTI:"

# Hotels whose Drive folders had to be shared per-month/year directly (no
# common parent folder to share instead) get grouped by a known keyword
# instead of trying to generically parse the date prefix out of the folder
# name — that parsing kept breaking on real naming inconsistencies (mixed
# case, 2- vs 4-digit years, colon/period/no separator, even "Anchor In" vs
# "Hyannis Anchor In" for the same hotel). Matching on a fixed keyword that's
# guaranteed present in every one of that hotel's folder names is far more
# reliable. Add an entry here for each hotel using this sharing pattern.
KNOWN_MULTI_FOLDER_HOTELS = {
    "Hyannis Anchor In":     ["ANCHOR"],
    "Provincetown Surfside": ["SURFSIDE"],
    "Hotel 1620":            ["1620", "PLYMOUTH"],
    "Wolfeboro":             ["WOLF"],
    "Provincetown Harbor Hotel": ["HARBOR"],
    "Provincetown Brass Key":    ["BRASS"],
    "Provincetown Crowne Pointe": ["CROWN"],
}


def _strip_dedup_suffix(name):
    """Strip a trailing ' (1)', ' (2)', etc. — Drive's auto-added suffix when
    a folder name collides with an existing one (confirmed real case:
    Surfside's hotel folder is literally named 'SURFSIDE (1)' in Drive).
    Purely cosmetic for the dropdown; the real folder_id is unaffected.
    """
    return re.sub(r'\s*\(\d+\)\s*$', '', name).strip()


def _is_rev_reports_name(name):
    """True if a folder name denotes a 'Revenue Reports' folder.

    Either word gets abbreviated independently in real folder names, so all
    four combinations have to be accepted:

        REVENUE REPORTS   the usual form
        REV RPTS          Hotel 1620's convention, 'G. JUL2018 REV RPTS HOTEL 1620'
        REV REPORTS       Salem's, 'SALEM REV REPORTS'
        REVENUE RPTS      for symmetry

    Matching only the first two hid every hotel filing under the others: the
    folder is shared and correctly structured, but nothing recognises it as a
    revenue-reports folder, so the hotel never appears and the failure looks
    exactly like a permissions problem.
    """
    return bool(re.search(r'REV(?:ENUE)?\s+(?:REPORTS|RPTS)', name.upper()))


def _extract_hotel_name_from_rev_folder(name):
    """Strip date-ish prefixes and the phrase 'Revenue Reports' from a folder
    name, leaving just the hotel name. Confirmed real naming is inconsistent
    even for a single hotel (Hyannis Anchor In, sharing per-month/year folders
    since there's no common parent to share instead): 'A: APR2025 Revenue
    Reports Anchor In', 'A: JUN2025 Revenue Reports Hyannis Anchor In',
    'Jul2024 Revenue Reports Hyannis Anchor In', 'A: MAY25 Revenue Reports
    Hyannis Anchor In' — mixing a leading single-letter marker (with a colon,
    period, or dash — not just a period), 2- or 4-digit years, and even
    inconsistent inclusion of "Hyannis" in the hotel name itself.

    Only strips a leading single letter when followed by a clear separator
    (":", ".", "-") — a bare leading letter with no separator is the start of
    a month name (e.g. "Jul2024") and must be left alone, not eaten.
    """
    s = name.strip()
    s = re.sub(r'^[A-Za-z]\s*[\.\:\-]\s*', '', s)
    s = re.sub(r'^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s*\d{2,4}\s+', '', s, flags=re.IGNORECASE)
    s = re.sub(r'^\d{4}\s+', '', s)
    s = re.sub(r'revenue\s*reports|rev\s+rpts', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s+', ' ', s).strip()
    return s.strip(' -:.,')


@st.cache_data(ttl=300)
def get_hotels_from_drive():
    """Return one Drive target per known hotel.

    Drive discovery is intentionally done in two cheap stages:
      1. Fetch all visible folders in paginated batches, including parent IDs.
      2. Examine only folders whose names are Revenue Reports folders.

    Revenue-report folders are grouped directly to the canonical hotel labels
    already declared in PORTFOLIO_HOTELS. This keeps every year's folder ID
    available for prior-month / prior-year lookups without scanning thousands
    of unrelated folders or making one API request per folder.
    """
    try:
        svc = get_drive_service()
        q = "mimeType = 'application/vnd.google-apps.folder' and trashed = false"

        folders = []
        page_token = None

        while True:
            result = svc.files().list(
                q=q,
                fields="nextPageToken, files(id, name, parents)",
                pageSize=1000,
                pageToken=page_token,
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
            ).execute()

            folders.extend(result.get("files", []))
            page_token = result.get("nextPageToken")
            if not page_token:
                break

        # Parent names let us identify a generic folder named only
        # "REVENUE REPORTS" from the hotel folder it lives under.
        folder_by_id = {f["id"]: f for f in folders}

        # Canonical hotel label -> every Revenue Reports folder ID visible
        # for that hotel, across all years/months.
        grouped = {}

        for folder in folders:
            name = str(folder.get("name", "") or "")
            name_upper = name.upper()

            if "ANCILLARY" in name_upper:
                continue
            if not _is_rev_reports_name(name):
                continue

            extracted = _extract_hotel_name_from_rev_folder(name)

            parent_names = []
            for parent_id in folder.get("parents", []):
                parent = folder_by_id.get(parent_id)
                if parent:
                    parent_names.append(str(parent.get("name", "") or ""))

            # Search the folder name, extracted hotel name, and visible parent
            # name against the app's existing portfolio hotel aliases.
            haystack = " ".join(
                [name, extracted] + parent_names
            ).upper()

            canonical = None
            for _portfolio, members in PORTFOLIO_HOTELS.items():
                for hotel_label, keywords in members.items():
                    if any(keyword.upper() in haystack for keyword in keywords):
                        canonical = hotel_label
                        break
                if canonical:
                    break

            if not canonical:
                continue

            grouped.setdefault(canonical, []).append(folder["id"])

        hotels = []
        for hotel_label, ids in grouped.items():
            # De-duplicate IDs while preserving Drive discovery order.
            ids = list(dict.fromkeys(ids))
            folder_id = (
                ids[0]
                if len(ids) == 1
                else MULTI_ID_PREFIX + ",".join(ids)
            )
            hotels.append((hotel_label, folder_id))

        hotels = [
            (name, fid)
            for name, fid in hotels
            if not _is_test_folder(name)
        ]
        return sorted(hotels, key=lambda x: x[0])

    except Exception:
        return []


# Sandbox copies live alongside the real folders — 'LONG BEACH - TEST' next to
# 'LONG BEACH'. They are filtered out at discovery rather than in any one
# dropdown, so nothing downstream can reach them: not the workbook update, not
# the OOO report, not the ancillary tool.
#
# Matched as a whole word so a hotel whose name merely contains the letters
# (Testa, Contest Point) would not disappear.
_TEST_FOLDER_RE = re.compile(r"\bTEST\b", re.IGNORECASE)


def _is_test_folder(name: str) -> bool:
    return bool(_TEST_FOLDER_RE.search(str(name or "")))


WORKBOOK_TYPES = ["ROB", "Strategy Report", "Forecast"]

# ── Portfolios ───────────────────────────────────────────────────────────────
# Hotels are discovered from Drive folder names, which carry no notion of
# grouping, so membership is declared here. Each portfolio's data arrives in a
# different export, which is why they get separate tabs rather than one list.
#
# Matching is on keywords, not exact names: Drive folder naming is
# inconsistent for the same hotel and two properties are filed under a
# different name entirely (Long Beach as ALLEGRIA, Westerly as Pleasant View
# Inn). Keep keywords distinctive enough not to collide.
PORTFOLIOS = ["Stay In Touch", "Hilton", "IHG"]

PORTFOLIO_HOTELS = {
    "Stay In Touch": {
        "Wolfeboro":      ["WOLF"],
        "Harbor Hotel":   ["HARBOR"],
        "Foxberry":       ["FOXBER", "FOXBUR"],
        "Tybee":          ["TYBEE"],
        "Provincetown Inn": ["PROVINCETOWN INN", "PTOWN INN"],
        "Middletown":     ["MIDDLETOWN"],
        "Brass Key":      ["BRASS"],
        "Long Beach":     ["LONG BEACH", "ALLEGRIA"],
        "Westerly":       ["WESTERLY", "PLEASANT VIEW"],
        "Crowne Pointe":  ["CROWN", "CROWNE"],
        "Ashworth":       ["ASHWORTH", "HAMPTON BEACH"],
        "Anchor Inn":     ["ANCHOR"],
        "Surfside":       ["SURFSIDE"],
        "1620":           ["1620", "PLYMOUTH"],
    },
    "Hilton": {
        "Northbrook":     ["NORTHBROOK"],
        "Andover":        ["ANDOVER"],
        "Ann Arbor":      ["ANN ARBOR"],
        "Silver Spring":  ["SILVER SPRING"],
        "Mesa":           ["MESA"],
        "Littleton":      ["LITTLETON"],
        "Nashua":         ["NASHUA"],
        "Kansas City":    ["KANSAS"],
        "Memphis":        ["MEMPHIS"],
    },
    "IHG": {
        "Salem":          ["SALEM"],
        "Manchester":     ["MANCHESTER"],
        "Weirton":        ["WEIRTON"],
    },
}

# Property codes as they appear in the SRP Activity export's 'Property -
# InnCode' column, which is how one shared export gets split per hotel.
# Confirmed from real exports; the three blanks have not appeared in one yet.
# Every code below is the one the property's own SRP report prints in its
# "Property: <code> - <name>" header, so the mapping is the vendor's, not a
# guess. All nine are listed: a hotel missing from here used to be skipped
# with a warning, which is how a three-hotel run quietly became a one-hotel
# run. _match_inncode can now also fall back to the export's
# 'Property - Name' column, so a code changing underneath us degrades to a
# name match rather than to silence.
HILTON_INNCODES = {
    "Littleton":     "LTNNH",   # Hampton Inn Littleton
    "Nashua":        "ASHSS",   # DoubleTree by Hilton Nashua
    "Ann Arbor":     "ARBAA",   # DoubleTree by Hilton Ann Arbor North
    "Mesa":          "MESWH",   # DoubleTree by Hilton Phoenix Mesa
    "Kansas City":   "MCIAP",   # Hilton Kansas City Airport
    "Memphis":       "MEMPH",   # Hilton Memphis
    "Northbrook":    "CHINB",   # Chicago Northbrook Hilton
    "Andover":       "BOSOR",   # DoubleTree by Hilton Boston-Andover
    "Silver Spring": "DCAGS",   # DoubleTree Washington DC Silver Spring
}

# Hilton properties do not run a Strategy Report.
PORTFOLIO_WORKBOOKS = {
    "Stay In Touch": WORKBOOK_TYPES,
    "Hilton":        ["ROB", "Forecast"],
    "IHG":           WORKBOOK_TYPES,
}


def portfolio_of(hotel_display_name):
    """Which portfolio a Drive-discovered hotel belongs to, or None."""
    up = (hotel_display_name or "").upper()
    for pf, members in PORTFOLIO_HOTELS.items():
        for kws in members.values():
            if any(k in up for k in kws):
                return pf
    return None


def hotels_in_portfolio(portfolio, discovered):
    """[(display_name, folder_id)] for every hotel declared in `portfolio`,
    in declared order so the list is stable between runs.

    A hotel with no matching Drive folder is still listed, with an empty folder
    id. Leaving it out meant a hotel could vanish from the dropdown with no
    indication of why, and checking access up front to decide was slow enough
    to be worse than the problem. Selecting one now fails when it is run, with
    a message naming what is wrong.
    """
    members = PORTFOLIO_HOTELS.get(portfolio, {})
    out = []
    for label, kws in members.items():
        matched = [(name, fid) for name, fid in discovered
                   if any(k in name.upper() for k in kws)]
        if matched:
            for entry in matched:
                if entry not in out:
                    out.append(entry)
        else:
            out.append((label, ""))
    return out


@st.cache_data(ttl=300)
def _all_visible_folders():
    """Every folder the signed-in account can see, as (name, id).

    Deliberately unfiltered — the point is to separate "the folder isn't shared
    with me" from "it's shared but I don't recognise it as a hotel", which look
    identical from the hotel list alone.
    """
    try:
        svc = get_drive_service()
        out, token = [], None
        while True:
            res = svc.files().list(
                q="mimeType = 'application/vnd.google-apps.folder' and trashed = false",
                fields="nextPageToken, files(id, name)", pageSize=1000,
                pageToken=token, supportsAllDrives=True,
                includeItemsFromAllDrives=True,
            ).execute()
            out.extend((f["name"], f["id"]) for f in res.get("files", []))
            token = res.get("nextPageToken")
            if not token:
                break
        return out
    except Exception:
        return []


def render_folder_diagnostic(portfolio, missing):
    """Say, per missing hotel, whether anything matching it is visible at all."""
    folders = _all_visible_folders()
    if not folders:
        st.error("Couldn't list any folders — the Drive connection itself is "
                 "failing, so this isn't about one hotel's sharing.")
        return
    st.caption(f"Signed in as `{service_account_email() or 'unknown'}`, which can "
               f"see **{len(folders)}** folders in total.")
    members = PORTFOLIO_HOTELS.get(portfolio, {})
    for label in missing:
        kws = members.get(label, [])
        hits = [(n, i) for n, i in folders
                if any(k in n.upper() for k in kws) and not _is_test_folder(n)]
        if not hits:
            st.markdown(f"**{label}** — nothing visible with this name. "
                        f"The folder isn't shared with this account.")
            continue
        rev = [n for n, _ in hits if _is_rev_reports_name(n)]
        st.markdown(
            f"**{label}** — {len(hits)} folder(s) visible, so it *is* shared. "
            + ("None of them reads as a revenue-reports folder, which is what "
               "makes it a hotel." if not rev else
               "One of them does read as a revenue-reports folder, so this "
               "should now resolve — press ↺.")
        )
        st.code("\n".join(n for n, _ in hits[:12]) or "(none)", language=None)


def portfolio_hotels_missing(portfolio, discovered):
    """Declared hotels with no matching Drive folder."""
    members = PORTFOLIO_HOTELS.get(portfolio, {})
    names = [n.upper() for n, _ in discovered]
    return [label for label, kws in members.items()
            if not any(any(k in n for k in kws) for n in names)]

# Maps workbook type → partial filename keyword to search for in Drive
WORKBOOK_KEYWORDS = {
    "ROB":             "ROB",
    "Strategy Report": "STRATEGY",
    "Forecast":        "FORECAST",
}

CREDS_PATH = "credentials.json.json"
SCOPES     = ["https://www.googleapis.com/auth/drive"]


# ── Drive transport ──────────────────────────────────────────────────────────
# A plain build() gives the whole app one httplib2 connection, and the service
# object is cached for the process, so every session and every Streamlit thread
# shared a single socket. httplib2 is not thread-safe: two people working at
# once interleave reads on the same TLS stream and one of them gets
#
#   ssl.SSLError  ...  in recv_into / self._sslobj.read(len, buffer)
#
# with no useful message, because the record it read was the middle of somebody
# else's response. The same socket also goes stale between calls — Google closes
# idle connections, and httplib2 hands the closed one back out.
#
# googleapiclient already knows how to retry that exact error, but only
# num_retries times and the default is zero, so the first blip aborted the run.
#
# Both halves are fixed here rather than at the 28 call sites: every request is
# built with its own connection and its own retry budget.
DRIVE_RETRIES = 5


@st.cache_resource
def _drive_credentials():
    # Production (Streamlit Cloud): credentials stored in st.secrets["google_credentials"]
    if "google_credentials" in st.secrets:
        return service_account.Credentials.from_service_account_info(
            dict(st.secrets["google_credentials"]), scopes=SCOPES
        )
    # Local dev: read from file
    return service_account.Credentials.from_service_account_file(CREDS_PATH, scopes=SCOPES)


def _authorized_http():
    """A connection of its own, for one request."""
    import httplib2
    import google_auth_httplib2
    return google_auth_httplib2.AuthorizedHttp(
        _drive_credentials(), http=httplib2.Http(timeout=120))


@st.cache_resource
def get_drive_service():
    from googleapiclient.http import HttpRequest

    class _OwnConnectionRequest(HttpRequest):
        """Retries the transient transport errors the default ignores."""

        def execute(self, http=None, num_retries=DRIVE_RETRIES):
            return super().execute(http=http, num_retries=num_retries)

        def next_chunk(self, http=None, num_retries=DRIVE_RETRIES):
            return super().next_chunk(http=http, num_retries=num_retries)

    def request_builder(_http, *args, **kwargs):
        return _OwnConnectionRequest(_authorized_http(), *args, **kwargs)

    return build("drive", "v3", credentials=_drive_credentials(),
                 cache_discovery=False, requestBuilder=request_builder)


@st.cache_data(ttl=3600)
def service_account_email():
    """The address folders have to be shared with, read from the live
    credentials rather than written down anywhere.

    Worth taking from the credentials themselves: the account in use has
    changed before, and telling someone to share a folder with a stale address
    produces a hotel that stays invisible with no indication why.
    """
    try:
        if "google_credentials" in st.secrets:
            return dict(st.secrets["google_credentials"]).get("client_email")
        import json
        with open(CREDS_PATH, encoding="utf-8") as fh:
            return json.load(fh).get("client_email")
    except Exception:
        return None


# has_edit_access() / get_hotels_with_edit_access() removed. They pre-checked
# every hotel's folder permissions to decide what to show, which meant a Drive
# call per hotel before the page could draw — slow enough that it was worse
# than the problem it solved. Nothing called them. The dropdown now lists every
# declared hotel and reports a sharing problem when one is actually run.


def drive_find_folder_by_keyword(service, keyword, parent_id=None):
    """Return the first folder whose name contains keyword (case-insensitive)."""
    q = "mimeType = 'application/vnd.google-apps.folder' and trashed = false"
    if parent_id:
        q += " and '%s' in parents" % parent_id
    result = service.files().list(
        q=q, fields="files(id, name)", pageSize=100,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    for f in result.get("files", []):
        if keyword.lower() in f["name"].lower():
            return f["id"], f["name"]
    return None, None

def _explicit_folder_year(name):
    """Return an explicit 20xx year found in a folder name, or None.

    Works for both spaced names like "2026 REVENUE REPORTS" and compact
    month/year names like "JUL2027 REVENUE REPORTS".
    """
    match = re.search(r"(?<!\d)(20\d{2})(?!\d)", str(name or ""))
    return int(match.group(1)) if match else None
def _pick_rev_reports_candidate(candidates, year_kw, month_kw):
    """Rank REVENUE REPORTS folder candidates for a target month/year.

    IMPORTANT:
    Never allow a folder explicitly belonging to a different year to satisfy
    the request. Generic folders with no year in their name remain valid
    because some hotels use:
        REVENUE REPORTS > Year > Month

    Order of preference:
      1. requested-year folder with no month abbreviation
      2. exact requested-month folder
      3. other requested-year folder
      4. generic folder with no explicit year
    """
    if not candidates:
        return None

    target_year = int(year_kw)

    # Exclude folders that explicitly belong to another calendar year.
    #
    # Example during a 2026 lookup:
    #   2026 REVENUE REPORTS  -> allowed
    #   AUG2026 REVENUE ...   -> allowed
    #   REVENUE REPORTS       -> allowed
    #   2027 REVENUE REPORTS  -> EXCLUDED
    eligible = [
        f for f in candidates
        if _explicit_folder_year(f["name"]) in (None, target_year)
    ]

    if not eligible:
        return None

    month_kw_2digit = month_kw[:3] + month_kw[-2:] if month_kw else None

    # Step 1: Prefer a true year-level folder.
    year_only = []
    for f in eligible:
        name_upper = f["name"].upper()
        if (
            year_kw in name_upper
            and not any(m.upper() in name_upper for m in MONTH_ABBR)
        ):
            year_only.append(f)

    if year_only:
        return year_only[0]

    # Step 2: Exact month folder for the requested month/year.
    for f in eligible:
        name_upper = f["name"].upper()
        if month_kw and (
            month_kw in name_upper
            or (month_kw_2digit and month_kw_2digit in name_upper)
        ):
            return f

    # Step 3: Any remaining folder explicitly matching the requested year.
    for f in eligible:
        if year_kw in f["name"].upper():
            return f

    # Step 4: Generic REVENUE REPORTS folder with no explicit year.
    generic = [
        f for f in eligible
        if _explicit_folder_year(f["name"]) is None
    ]

    if generic:
        return generic[0]

    return None

def _find_rev_reports_folder_for_year(service, hotel_id, year_kw, month_kw=None):
    """Find the REVENUE REPORTS folder to use for a given year (and,
    preferably, the specific target month — see _pick_rev_reports_candidate
    for the ranking; the old first-year-substring-match behavior silently
    picked whichever same-year sibling Drive happened to list first for
    hotels with one folder per month).
    Some hotels have ONE 'REVENUE REPORTS' folder for years directly (year/month
    subfolders inside it); others have a SEPARATE '<year> REVENUE REPORTS <hotel>'
    folder per year (or per month), sitting side by side.
    """
    # Hotel has separate REVENUE REPORTS folders shared per year/month (each
    # individually shared, since there's no common parent to share instead).
    if hotel_id.startswith(MULTI_ID_PREFIX):
        candidate_ids = hotel_id[len(MULTI_ID_PREFIX):].split(",")
        candidates = []
        for cid in candidate_ids:
            try:
                info = service.files().get(fileId=cid, fields="name", supportsAllDrives=True).execute()
                candidates.append({"id": cid, "name": info["name"]})
            except Exception:
                continue
        best = _pick_rev_reports_candidate(candidates, year_kw, month_kw)
        if best:
            return best["id"], best["name"]
        return None, None

    q = ("mimeType = 'application/vnd.google-apps.folder' and trashed = false "
         "and '%s' in parents") % hotel_id
    children = service.files().list(
        q=q, fields="files(id, name)", pageSize=100,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute().get("files", [])
    # Match "revenue reports", "rev reports", or just "revenue" in folder name
    candidates = [f for f in children
                  if any(kw in f["name"].lower() for kw in ["revenue reports", "rev reports", "revenue"])]
    best = _pick_rev_reports_candidate(candidates, year_kw, month_kw)
    if best:
        return best["id"], best["name"]

    # Some hotels are shared directly at the REVENUE REPORTS folder level
    # (Shared Drive permissions don't propagate to a parent folder) — hotel_id
    # IS that folder already in that case, not its parent, so there's no
    # child to find. Check hotel_id's own name before giving up.
    try:
        self_info = service.files().get(
            fileId=hotel_id, fields="name", supportsAllDrives=True
        ).execute()
        if "revenue reports" in self_info.get("name", "").lower():
            return hotel_id, self_info["name"]
    except Exception:
        pass

    return None, None


def _find_month_folder_under_rev(service, rev_id, year_kw, month_kw, target_month, hotel_name):
    """Locate the month folder for a new-month setup, handling both layouts:
    month folders directly inside the REVENUE REPORTS folder (common when
    there's one REVENUE REPORTS folder per year), or nested under a year
    subfolder. NEVER creates a folder — the month folder is expected to
    already exist (created ahead of time as part of the standard process).
    Returns (None, None) if it can't be found, so the caller can error out
    with a clear message instead of the app silently creating a new one in
    the wrong place.

    Finding the "year subfolder" is trickier than a plain keyword search:
    a sibling MONTH folder (e.g. "JUL2026 REVENUE REPORTS X") also contains
    the year digits as a substring, so a naive year_kw match can pick a
    sibling month folder instead of a real year folder. Excluding any
    candidate whose name contains a month abbreviation avoids that ambiguity.
    """
    # Flat per-month layout (confirmed real case: Provincetown Harbor Hotel,
    # 'H: AUG2026 REVENUE REPORTS ...'): the rev folder's own name already
    # names the target month — it IS the month folder; there is no month
    # subfolder inside it to find.
    try:
        rev_info = service.files().get(fileId=rev_id, fields="name", supportsAllDrives=True).execute()
        rev_name = rev_info.get("name", "")
        month_kw_2digit = month_kw[:3] + month_kw[-2:]
        if month_kw in rev_name.upper() or month_kw_2digit in rev_name.upper():
            return rev_id, rev_name
    except Exception:
        pass

    month_id, month_name = drive_find_folder_by_keyword(service, month_kw, parent_id=rev_id)
    if month_id:
        return month_id, month_name

    q = ("mimeType = 'application/vnd.google-apps.folder' and trashed = false "
         "and '%s' in parents") % rev_id
    siblings = service.files().list(
        q=q, fields="files(id, name)", pageSize=100,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute().get("files", [])
    year_id = None
    for f in siblings:
        name_upper = f["name"].upper()
        if year_kw in f["name"] and not any(m.upper() in name_upper for m in MONTH_ABBR):
            year_id = f["id"]
            break
    if year_id:
        return drive_find_month_folder(service, year_id, month_kw)

    # Fallback: search recursively through nested folders (handles deep nesting like SALEM)
    # Some hotels have multiple intermediate folders before reaching the month folder
    def search_recursive(parent_id, depth=0):
        if depth > 3:  # Prevent infinite recursion
            return None, None
        try:
            q = ("mimeType = 'application/vnd.google-apps.folder' and trashed = false "
                 "and '%s' in parents") % parent_id
            children = service.files().list(
                q=q, fields="files(id, name)", pageSize=100,
                supportsAllDrives=True, includeItemsFromAllDrives=True,
            ).execute().get("files", [])

            best_match = None
            for child in children:
                child_name_upper = child["name"].upper()
                # Prefer folders with both month_kw and "REVENUE" (more specific)
                # to avoid stopping at intermediate folders
                if month_kw in child_name_upper and "REVENUE" in child_name_upper:
                    return child["id"], child["name"]
                # Also check for "REPORTS" to catch "AUG2026 REVENEU REPORTS SALEM"
                if month_kw in child_name_upper and "REPORTS" in child_name_upper:
                    best_match = (child["id"], child["name"])
                # Recurse into subfolders to find deeper matches
                result_id, result_name = search_recursive(child["id"], depth + 1)
                if result_id:
                    return result_id, result_name
            if best_match:
                return best_match
        except Exception:
            pass
        return None, None

    return search_recursive(rev_id)


def drive_find_file(service, keyword, parent_id):
    """Return (file_id, file_name) for first xlsx whose name contains keyword,
    excluding files whose name also contains 'copy' (to skip backup copies).

    Also matches native Google Sheets (mimeType 'application/vnd.google-apps.
    spreadsheet') — confirmed real case: Hotel 1620's Forecast workbook was
    created directly as a Google Sheet rather than uploaded as .xlsx, so an
    xlsx/xlsm-only filter silently missed it even though the name matched.
    drive_download handles exporting these to xlsx bytes on read.
    """
    q = ("'%s' in parents and trashed = false "
         "and (mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' "
         "or mimeType = 'application/vnd.ms-excel.sheet.macroenabled.12' "
         "or mimeType = 'application/vnd.google-apps.spreadsheet')") % parent_id
    result = service.files().list(
        q=q, fields="files(id, name)",
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    files = result.get("files", [])
    # prefer non-copy files; fall back to copy if nothing else found
    files.sort(key=lambda f: (1 if "copy" in f["name"].lower() else 0))
    for f in files:
        name_lower = f["name"].lower()
        if keyword.lower() in name_lower and "master" not in name_lower:
            return f["id"], f["name"]
    return None, None


def drive_download(service, file_id) -> bytes:
    """Download a file's bytes. Native Google Sheets (created directly in
    Drive rather than uploaded as .xlsx — confirmed real case: Hotel 1620's
    Forecast workbook) can't be read via get_media like a normal blob file;
    they must be exported to xlsx format instead."""
    meta = service.files().get(fileId=file_id, fields="mimeType", supportsAllDrives=True).execute()
    buf = io.BytesIO()
    if meta.get("mimeType") == "application/vnd.google-apps.spreadsheet":
        req = service.files().export_media(
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        req = service.files().get_media(fileId=file_id, supportsAllDrives=True)
    dl  = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        # MediaIoBaseDownload wraps the request rather than using it, so the
        # retry budget built into the request builder doesn't reach here and
        # has to be handed over per chunk. Downloads are the longest-lived
        # Drive calls the app makes and the likeliest to meet a dropped
        # connection.
        _, done = dl.next_chunk(num_retries=DRIVE_RETRIES)
    return buf.getvalue()


def drive_upload(service, file_id, file_bytes: bytes, file_name: str):
    """Overwrite an existing Drive file with new bytes."""
    buf   = io.BytesIO(file_bytes)
    media = MediaIoBaseUpload(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True,
    )
    # supportsAllDrives is required here even though drive_download's
    # get_media() works without it — confirmed real case: writing back to a
    # Shared Drive file (Hyannis Anchor In) 404'd on update() alone, for
    # every workbook type, despite the same file_id having just been read
    # successfully moments earlier.
    service.files().update(fileId=file_id, media_body=media, supportsAllDrives=True).execute()


def _ancillary_normalize_drive_name(value):
    return re.sub(r"[^A-Z0-9]+", " ", str(value or "").upper()).strip()


def ancillary_find_drive_report(service, hotel_id, hotel_name, report_month):
    """Find the hotel's existing Canary/SNT ancillary workbook in Drive.

    Important folder rule:
      The Ancillary Revenue folder is NOT inside the monthly Revenue Reports
      folder. It is a separate sibling/descendant under the hotel's Drive tree,
      e.g.:
          G: JUL2026 REVENUE REPORTS WOLFEBORO
          M: ANCILLARY REVENUE REPORTS WOLFEBORO

    Therefore, search from the HOTEL folder tree itself, not from the selected
    month's Revenue Reports folder.

    Folder match:
      - ANCILLARY REVENUE REPORTS
      - ANCILLARY REVENUE FILES

    Workbook match inside that folder:
      - filename contains REPORT
      - and contains CANARY or SNT (or both)

    Matching is case/punctuation insensitive. If multiple valid workbooks exist,
    use the most recently modified one.
    """

    # Keep the Revenue Reports folder only as a helpful diagnostic so the UI can
    # still show which year/month tree the hotel resolver sees.
    year_kw = str(report_month.year)
    month_kw = report_month.strftime("%b").upper()
    rev_id, rev_name = _find_rev_reports_folder_for_year(
        service,
        hotel_id,
        year_kw,
        month_kw,
    )

    # Search from every known hotel-scope folder, not from rev_id.
    scope_ids = _hotel_search_scope_ids(service, hotel_id)
    if not scope_ids:
        scope_ids = [hotel_id]

    ancillary_folders = []
    seen = set()
    frontier = [(sid, 0) for sid in scope_ids]
    max_depth = 4

    while frontier:
        parent_id, depth = frontier.pop(0)
        if not parent_id or parent_id in seen:
            continue
        seen.add(parent_id)

        if depth >= max_depth:
            continue

        q = (
            "mimeType='application/vnd.google-apps.folder' and trashed=false "
            f"and '{parent_id}' in parents"
        )

        page_token = None
        while True:
            resp = service.files().list(
                q=q,
                fields="nextPageToken,files(id,name,modifiedTime,parents)",
                pageSize=200,
                pageToken=page_token,
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
            ).execute()

            for f in resp.get("files", []):
                fid = f.get("id")
                if not fid:
                    continue

                n = _ancillary_normalize_drive_name(f.get("name"))
                if (
                    "ANCILLARY REVENUE REPORTS" in n
                    or "ANCILLARY REVENUE FILES" in n
                ):
                    ancillary_folders.append({
                        **f,
                        "_depth": depth + 1,
                    })
                else:
                    frontier.append((fid, depth + 1))

            page_token = resp.get("nextPageToken")
            if not page_token:
                break

    if not ancillary_folders:
        return None, (
            f"No folder containing 'Ancillary Revenue Reports' or "
            f"'Ancillary Revenue Files' was found in {hotel_name}'s Drive tree."
        )

    # Prefer the shallowest match; if tied, prefer most recently modified.
    ancillary_folders.sort(
        key=lambda f: (
            f.get("_depth", 99),
            str(f.get("modifiedTime", "")),
        )
    )
    min_depth = ancillary_folders[0].get("_depth", 99)
    same_depth = [
        f for f in ancillary_folders
        if f.get("_depth", 99) == min_depth
    ]
    same_depth.sort(
        key=lambda f: str(f.get("modifiedTime", "")),
        reverse=True,
    )
    anc_folder = same_depth[0]

    # Find the report workbook directly inside the ancillary folder.
    q = f"trashed=false and '{anc_folder['id']}' in parents"
    files = service.files().list(
        q=q,
        fields="files(id,name,mimeType,modifiedTime)",
        pageSize=200,
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute().get("files", [])

    candidates = []
    for f in files:
        if f.get("mimeType") == "application/vnd.google-apps.folder":
            continue

        n = _ancillary_normalize_drive_name(f.get("name"))
        if "REPORT" not in n:
            continue
        if "CANARY" not in n and "SNT" not in n:
            continue
        candidates.append(f)

    if not candidates:
        return None, (
            f"{anc_folder['name']}: no report workbook containing "
            f"'Canary' or 'SNT' was found."
        )

    candidates.sort(
        key=lambda f: str(f.get("modifiedTime", "")),
        reverse=True,
    )
    target = candidates[0]

    return {
        "file_id": target["id"],
        "file_name": target["name"],
        "mime_type": target.get("mimeType", ""),
        "folder_id": anc_folder["id"],
        "folder_name": anc_folder["name"],
        "revenue_folder_id": rev_id,
        "revenue_folder_name": rev_name or f"{report_month.year} Revenue Reports",
    }, None


def _ancillary_month_sheet_name(wb, report_month):
    """Return the existing month sheet name or the standard new name."""
    abbr = report_month.strftime("%b").upper()
    year = str(report_month.year)

    # Exact month abbreviation is the portfolio standard.
    if abbr in wb.sheetnames:
        return abbr

    # Reuse an existing month-specific variant such as "AUG 2026" or
    # "AUGUST" instead of creating a duplicate month tab.
    for name in wb.sheetnames:
        norm = _ancillary_normalize_drive_name(name)
        if not norm.startswith(abbr):
            continue
        if (
            "PIVOT" in norm
            or "RAW" in norm
            or "TEMPLATE" in norm
            or "CONTROL" in norm
        ):
            continue
        if year in norm or norm in {abbr, report_month.strftime("%B").upper()}:
            return name

    return abbr


def _ancillary_copy_generated_sheet(src_ws, dst_ws):
    """Copy a generated Report sheet into an existing monthly workbook."""
    from copy import copy as _copy

    # Column dimensions
    for key, dim in src_ws.column_dimensions.items():
        d = dst_ws.column_dimensions[key]
        d.width = dim.width
        d.hidden = dim.hidden
        d.bestFit = dim.bestFit
        d.outlineLevel = dim.outlineLevel

    # Row dimensions
    for idx, dim in src_ws.row_dimensions.items():
        d = dst_ws.row_dimensions[idx]
        d.height = dim.height
        d.hidden = dim.hidden
        d.outlineLevel = dim.outlineLevel

    # Cells/styles
    for row in src_ws.iter_rows():
        for src_cell in row:
            dst_cell = dst_ws[src_cell.coordinate]
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell._style = _copy(src_cell._style)
            if src_cell.number_format:
                dst_cell.number_format = src_cell.number_format
            dst_cell.font = _copy(src_cell.font)
            dst_cell.fill = _copy(src_cell.fill)
            dst_cell.border = _copy(src_cell.border)
            dst_cell.alignment = _copy(src_cell.alignment)
            dst_cell.protection = _copy(src_cell.protection)
            if src_cell.hyperlink:
                dst_cell._hyperlink = _copy(src_cell.hyperlink)
            if src_cell.comment:
                dst_cell.comment = _copy(src_cell.comment)

    # Merges
    for merged in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged))

    # Sheet-level settings
    dst_ws.freeze_panes = src_ws.freeze_panes
    dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    dst_ws.sheet_format.defaultColWidth = src_ws.sheet_format.defaultColWidth
    dst_ws.sheet_format.defaultRowHeight = src_ws.sheet_format.defaultRowHeight
    dst_ws.page_margins = _copy(src_ws.page_margins)
    dst_ws.page_setup = _copy(src_ws.page_setup)
    dst_ws.print_options = _copy(src_ws.print_options)
    dst_ws.sheet_properties = _copy(src_ws.sheet_properties)

    # Conditional formatting
    for cf_obj, rules in src_ws.conditional_formatting._cf_rules.items():
        for rule in rules:
            dst_ws.conditional_formatting.add(
                _copy(cf_obj),
                _copy(rule),
            )


def ancillary_insert_report_sheet(
    destination_bytes,
    generated_bytes,
    report_month,
    destination_name="",
):
    """Replace/create the monthly sheet inside the existing Canary/SNT report."""
    keep_vba = str(destination_name or "").lower().endswith(".xlsm")

    dest_wb = openpyxl.load_workbook(
        io.BytesIO(destination_bytes),
        data_only=False,
        keep_vba=keep_vba,
    )
    generated_wb = openpyxl.load_workbook(
        io.BytesIO(generated_bytes),
        data_only=False,
    )

    if "Report" not in generated_wb.sheetnames:
        raise ValueError("Generated ancillary workbook has no Report sheet.")

    sheet_name = _ancillary_month_sheet_name(dest_wb, report_month)

    # Replace only this month's sheet; preserve every other month and pivot tab.
    if sheet_name in dest_wb.sheetnames:
        old_idx = dest_wb.sheetnames.index(sheet_name)
        del dest_wb[sheet_name]
    else:
        pivot_idx = next(
            (
                i
                for i, name in enumerate(dest_wb.sheetnames)
                if "PIVOT" in name.upper()
            ),
            len(dest_wb.sheetnames),
        )
        old_idx = pivot_idx

    dst_ws = dest_wb.create_sheet(sheet_name, old_idx)
    _ancillary_copy_generated_sheet(generated_wb["Report"], dst_ws)

    out = io.BytesIO()
    dest_wb.save(out)
    return out.getvalue(), sheet_name


def drive_find_month_folder(service, parent_id: str, month_kw: str):
    """Return (folder_id, folder_name) for the month folder under parent_id,
    or (None, None) if it isn't there. Never creates a folder — the month
    folder is expected to already exist as part of the standard process;
    the caller is responsible for erroring out clearly when it's missing
    instead of the app silently creating one (which is how folders ended
    up nested in the wrong place before).
    """
    q = ("mimeType='application/vnd.google-apps.folder' and trashed=false "
         "and '%s' in parents") % parent_id
    result = service.files().list(
        q=q, fields="files(id,name)", pageSize=100,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    for f in result.get("files", []):
        if month_kw in f["name"].upper():
            return f["id"], f["name"]
    return None, None


def drive_copy_file(service, source_file_id: str, new_name: str, parent_folder_id: str):
    """Copy a Drive file to a new name in the given folder. Returns (new_file_id, new_name).
    supportsAllDrives=True is required when the destination is a shared drive —
    service accounts have no personal storage quota.
    """
    body = {"name": new_name, "parents": [parent_folder_id]}
    copied = service.files().copy(
        fileId=source_file_id, body=body, fields="id,name",
        supportsAllDrives=True,
    ).execute()
    return copied["id"], copied["name"]


def _hotel_search_scope_ids(service, hotel_id):
    """Return every folder id that could plausibly hold a hotel's MASTER
    template files: each of the hotel's own root candidate folder(s)
    (unwrapping a MULTI:<id>,<id>,... group) plus their direct children.

    Confirmed real case: master-file lookups used to search ALL of Drive
    with no folder scoping at all — Wolfeboro's 'Set Up New ROB' found and
    copied Hotel 1620's ROB master (Drive's unordered global search just
    happened to return it first), silently mislabeling the result 'JUL2026
    ROB PLYMOUTH.xlsx' inside Wolfeboro's own folder.
    """
    if hotel_id.startswith(MULTI_ID_PREFIX):
        root_ids = hotel_id[len(MULTI_ID_PREFIX):].split(",")
    else:
        root_ids = [hotel_id]

    scope_ids = list(root_ids)
    for rid in root_ids:
        q = ("mimeType = 'application/vnd.google-apps.folder' and trashed = false "
             "and '%s' in parents") % rid
        try:
            children = service.files().list(
                q=q, fields="files(id)", pageSize=100,
                supportsAllDrives=True, includeItemsFromAllDrives=True,
            ).execute().get("files", [])
        except Exception:
            children = []
        scope_ids.extend(c["id"] for c in children)
    return scope_ids


def find_rob_master(service, hotel_id: str, target_year=None):
    """Search the hotel's own Drive tree for the correct ROB master.

    If target_year is supplied, prefer a master explicitly associated with
    that year by either its filename or its parent-folder name. A master that
    is explicitly associated with a different year is never selected.
    """
    scope_ids = _hotel_search_scope_ids(service, hotel_id)
    if not scope_ids:
        return None, "Could not resolve hotel folder to search."

    parent_clause = " or ".join("'%s' in parents" % sid for sid in scope_ids)
    q = (
        "trashed=false and (%s) "
        "and (mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' "
        "or mimeType='application/vnd.ms-excel.sheet.macroenabled.12') "
        "and name contains 'MASTER' and name contains 'ROB'"
    ) % parent_clause

    result = service.files().list(
        q=q,
        fields="files(id,name,parents,modifiedTime)",
        pageSize=100,
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()

    candidates = result.get("files", [])
    if not candidates:
        return None, "No ROB master file found in Drive."

    def explicit_year(value):
        m = re.search(r"(?<!\\d)(20\\d{2})(?!\\d)", str(value or ""))
        return int(m.group(1)) if m else None

    parent_name_cache = {}

    def candidate_years(file_obj):
        years = set()
        fy = explicit_year(file_obj.get("name"))
        if fy:
            years.add(fy)
        for pid in file_obj.get("parents", []) or []:
            if pid not in parent_name_cache:
                try:
                    info = service.files().get(
                        fileId=pid,
                        fields="name",
                        supportsAllDrives=True,
                    ).execute()
                    parent_name_cache[pid] = info.get("name", "")
                except Exception:
                    parent_name_cache[pid] = ""
            py = explicit_year(parent_name_cache[pid])
            if py:
                years.add(py)
        return years

    eligible = []
    for f in candidates:
        years = candidate_years(f)

        # Explicitly wrong-year masters are not eligible.
        if target_year is not None and years and target_year not in years:
            continue

        score = 2 if (target_year is not None and target_year in years) else 1
        eligible.append((score, str(f.get("modifiedTime", "")), f))

    if not eligible:
        if target_year is not None:
            return None, f"No ROB master for {target_year} was found in this hotel's Drive folders."
        return None, "No eligible ROB master file found in Drive."

    eligible.sort(key=lambda x: (x[0], x[1]), reverse=True)
    best = eligible[0][2]
    return best["id"], best["name"]


def _is_rob_month_blank(ws, block_start):
    """Return True if cols 2,3,4 of the Revenue row are all empty (None, '', or formula)."""
    rev_row = block_start + 1
    return all(
        ws.cell(rev_row, c).value is None
        or ws.cell(rev_row, c).value == ""
        or is_formula(str(ws.cell(rev_row, c).value))
        for c in [2, 3, 4]
    )


def _rob_as_of_date(year, month):
    """Return the 1st of the month for a given year, advanced past Sat/Sun to Monday."""
    d = datetime.date(year, month, 1)
    if d.weekday() == 5:    # Saturday → Monday
        d += datetime.timedelta(days=2)
    elif d.weekday() == 6:  # Sunday → Monday
        d += datetime.timedelta(days=1)
    return d


def _resolve_cell(prev_wb_data, prev_wb_formulas, sheet_name, row, col):
    """Read a cell value, following simple cross-sheet formula references if needed.
    When data_only=True returns None (no cached value), reads the formula string
    from prev_wb_formulas and follows the reference (e.g. ='wk one'!B45)."""
    import re
    from openpyxl.utils import column_index_from_string

    def _read_data(wb, sname, r, c):
        if wb and sname in wb.sheetnames:
            return wb[sname].cell(r, c).value
        return None

    v = _read_data(prev_wb_data, sheet_name, row, col)
    if v is not None:
        return v

    # No cached value — try to follow the formula reference
    if prev_wb_formulas and sheet_name in prev_wb_formulas.sheetnames:
        formula = prev_wb_formulas[sheet_name].cell(row, col).value
        if formula and str(formula).startswith("="):
            m = re.match(r"^='?([^'!]+)'?!([A-Za-z]+)(\d+)$", str(formula).strip())
            if m:
                ref_sheet = m.group(1)
                ref_col   = column_index_from_string(m.group(2))
                ref_row   = int(m.group(3))
                v = _read_data(prev_wb_data, ref_sheet, ref_row, ref_col)
                if v is not None:
                    return v
                # Referenced sheet might itself have a formula — follow one more level
                v2 = _resolve_cell(prev_wb_data, prev_wb_formulas, ref_sheet, ref_row, ref_col)
                return v2
    return None


def _fill_rob_prev_table(wk1_ws, prev_wb, prev_wb_formulas, target_month):
    """Fill Week 1 Previous Sheet from one prior-month completed week.

    Source rule:
      1. Look at the previous month's ROB workbook.
      2. Starting with the latest week tab, find the last week where the
         PREVIOUS MONTH's current-year Revenue / Room Nights cells are
         actually populated in the data-only workbook.
      3. Once that week is selected, copy ONLY each month's Revenue row from
         that one sheet into the Week 1 Previous table.

    This table must never use Pickup WoW rows or mix source weeks.
    """
    month_abbrs = [
        "jan", "feb", "mar", "apr", "may", "jun",
        "jul", "aug", "sep", "oct", "nov", "dec"
    ]

    def _as_year(v):
        if isinstance(v, (datetime.datetime, datetime.date)):
            return v.year
        if isinstance(v, (int, float)):
            iv = int(v)
            if 2000 <= iv <= 2100:
                return iv
            if 40000 <= iv <= 60000:
                try:
                    return (
                        datetime.date(1899, 12, 30)
                        + datetime.timedelta(days=iv)
                    ).year
                except Exception:
                    pass
        if isinstance(v, str):
            m = re.search(r"\b(20\d{2})\b", v)
            if m:
                return int(m.group(1))
        return None

    # ----- Locate Week 1 Previous table -----
    hdr_row = hdr_col = None
    for r in range(1, min(wk1_ws.max_row + 1, 80)):
        for c in range(1, wk1_ws.max_column + 1):
            val = str(wk1_ws.cell(r, c).value or "").strip().lower()
            if (
                "week 1 previous" in val
                or ("calculation only" in val and "week" in val)
            ):
                hdr_row, hdr_col = r, c
                break
        if hdr_row:
            break

    if not hdr_row:
        return "Week 1 Previous Sheet table not found in wk one"

    # ----- Destination month rows -----
    month_label_col = None
    for r in range(hdr_row + 1, min(wk1_ws.max_row + 1, hdr_row + 40)):
        for c in range(
            max(1, hdr_col - 5),
            min(wk1_ws.max_column + 1, hdr_col + 8)
        ):
            if str(wk1_ws.cell(r, c).value or "").strip().lower() in month_abbrs:
                month_label_col = c
                break
        if month_label_col:
            break

    if not month_label_col:
        return "Could not find month label column in Week 1 Previous Sheet table"

    dest_month_row = {}
    for r in range(hdr_row + 1, min(wk1_ws.max_row + 1, hdr_row + 40)):
        v = str(wk1_ws.cell(r, month_label_col).value or "").strip().lower()
        if v in month_abbrs:
            dest_month_row[month_abbrs.index(v)] = r

    # ----- Destination year columns -----
    expected_years = [
        target_month.year - 3,
        target_month.year - 2,
        target_month.year - 1,
        target_month.year,
    ]

    dest_year_col = {}
    for r in range(hdr_row, min(wk1_ws.max_row + 1, hdr_row + 10)):
        parsed = {}
        for c in range(
            month_label_col + 1,
            min(wk1_ws.max_column + 1, month_label_col + 20)
        ):
            yr = _as_year(wk1_ws.cell(r, c).value)
            if yr:
                parsed[yr] = c
        if parsed:
            dest_year_col = parsed
            break

    # The standard Week 1 Previous table is four consecutive year columns.
    # Its headers may themselves be formulas (=B3, =C3, etc.), so if the
    # literal years cannot be read, use their known left-to-right order.
    if not dest_year_col:
        for c, year in zip(
            range(month_label_col + 1, month_label_col + 5),
            expected_years
        ):
            dest_year_col[year] = c

    # ----- Select ONE last completed week from previous month's ROB -----
    prev_month = (
        target_month - datetime.timedelta(days=1)
    ).replace(day=1)
    prev_month_idx = prev_month.month - 1

    source_sheet = None

    for candidate in reversed(ROB_SHEETS):
        if candidate not in prev_wb.sheetnames:
            continue

        src_ws = prev_wb[candidate]  # data_only=True workbook
        labels = rob_month_blocks(src_ws).get(prev_month_idx, {})
        rev_row = labels.get("revenue")
        rms_row = (
            labels.get("room nights")
            or labels.get("rms sold")
            or labels.get("rooms sold")
        )

        if not rev_row:
            continue

        # Current-year is column E in the ROB structure.
        rev_val = src_ws.cell(rev_row, 5).value
        rms_val = src_ws.cell(rms_row, 5).value if rms_row else None

        # "Completed/used" means the cells are populated, even if the real
        # numeric value happens to be zero. Blank cells indicate that week
        # was not used for that month.
        if rev_val is not None or rms_val is not None:
            source_sheet = candidate
            break

    if source_sheet is None:
        return (
            f"Could not identify the last completed week in the "
            f"{prev_month:%b %Y} ROB."
        )

    # ----- Copy literal Revenue values from that one sheet -----
    src_ws = prev_wb[source_sheet]  # data_only=True
    source_blocks = rob_month_blocks(src_ws)

    # ROB year columns are B:E. Row 4 may contain date headers instead of
    # literal year labels, so preserve the standard chronological mapping.
    src_year_col = {
        target_month.year - 3: 2,
        target_month.year - 2: 3,
        target_month.year - 1: 4,
        target_month.year: 5,
    }

    for month_idx, dest_row in dest_month_row.items():
        rev_row = source_blocks.get(month_idx, {}).get("revenue")
        if not rev_row:
            continue

        for year, dest_col in dest_year_col.items():
            src_col = src_year_col.get(year)
            if not src_col:
                continue

            # IMPORTANT: read only the cached/literal Revenue value from the
            # data-only workbook. Do not resolve formulas through Pickup WoW
            # or another row/sheet.
            value = src_ws.cell(rev_row, src_col).value

            if value is not None:
                wk1_ws.cell(dest_row, dest_col).value = value

    return None


def _fill_rob_sheet(new_ws, prev_ws, ly_ws, target_month, is_wk_one, wk_one_sheet_name):
    """Fill one ROB sheet tab with historical data.
    Preserves formulas from master template — only overwrites cells with values.
    """
    from openpyxl.utils import get_column_letter

    target_idx  = target_month.month - 1   # 0-based (Jul = 6)
    prev_idx    = target_idx - 1           # most recently completed month (Jun = 5)
    # LY col → new col shift: LY has [2022,2023,2024,2025], new needs [2023,2024,2025,2026]
    ly_to_new    = {3: 2, 4: 3, 5: 4}
    # Rows per month block, read off each sheet rather than assumed (8 for most
    # hotels, 11 where there's a Permanent-rooms section). Last year's workbook
    # is measured separately so a template change between years can't shift the
    # rows we read from.
    step         = rob_block_step(new_ws)
    ly_step      = rob_block_step(ly_ws) if ly_ws is not None else step
    data_offsets = list(range(1, step))    # offset 0 (date header) handled separately

    # ── As-of dates for this week tab ──────────────────────────────────────
    # Taken from row 4 of THIS sheet's counterpart in last year's ROB, so
    # AUG2026 'wk two' gets its 2023/2024/2025 reporting dates from AUG2025
    # 'wk two'. Each week was reported roughly seven days apart, so these must
    # never be shared between tabs.
    #
    # Row 4 specifically: it holds the literal dates, while every month block
    # below it chains upward (=B4, =B52, ...). Reading a lower block instead
    # picks up the chain formula itself on most sheets — and on some it picks
    # up whatever else is parked in that cell (one real file has a revenue
    # figure sitting in the Aug block's date position on 'wk two'). Requiring
    # a real date type keeps that kind of stray number out of the header.
    #
    # Col 5 (current year) is a placeholder; the weekly update stamps the real
    # as-of date when it runs.
    as_of_dates = {}
    if ly_ws:
        for ly_col, new_col in ly_to_new.items():
            v = ly_ws.cell(4, ly_col).value
            # A never-used week tab carries Excel's zero date (1899-12-30),
            # which is a real datetime and would otherwise be copied across as
            # though it were a reporting date.
            if isinstance(v, (datetime.datetime, datetime.date)) and v.year >= 2000:
                as_of_dates[new_col] = v
    prev_week_current_year_date = None
    if prev_ws is not None:
        v = prev_ws.cell(4, 5).value
        if isinstance(v, (datetime.datetime, datetime.date)) and v.year >= 2000:
            prev_week_current_year_date = v
    as_of_dates[5] = datetime.datetime(target_month.year, target_month.month, target_month.day)

    for month_idx in range(12):
        block_start = 4 + step * month_idx

        # ── Date header row (offset 0) ─────────────────────────────────────
        # Weeks 2+ used to get ='wk one'!<col><row> here, which forced every
        # week tab to display wk one's reporting dates.
        #
        # Row 4 always gets the real date. Month blocks below it keep a
        # same-sheet chain (=B4, =SUM(B12)) because that correctly carries
        # THIS tab's date down the sheet — but a cross-sheet reference is
        # overwritten, since pointing at another week's tab reintroduces the
        # very problem being fixed (one real template chains its lower blocks
        # to ='wk one'!B11).
        for col, date_val in as_of_dates.items():
            cell = new_ws.cell(block_start, col)
            cur = str(cell.value) if cell.value is not None else ""
            same_sheet_chain = is_formula(cur) and "!" not in cur
            if block_start != 4 and same_sheet_chain:
                continue
            write_date = date_val
            if col == 5 and month_idx <= prev_idx and prev_week_current_year_date is not None:
                write_date = prev_week_current_year_date
            cell.value = write_date
            cell.number_format = "mm/dd/yyyy"

        # ── Data rows (offsets 1–7) ───────────────────────────────────────────
        if month_idx < prev_idx:
            # Past months (Jan–May when target=Jul): copy all 4 cols from prev ROB
            # For new hotels with no historical data (e.g., first year), prev_ws will be None —
            # just skip and leave those months with master template values (no error).
            if not _is_rob_month_blank(new_ws, block_start):
                continue
            if is_wk_one:
                if prev_ws is None:
                    continue
                for dr in data_offsets:
                    r = block_start + dr
                    for c in [2, 3, 4, 5]:
                        # Skip if master template has a formula in this cell
                        if is_formula(str(new_ws.cell(r, c).value)):
                            continue
                        v = prev_ws.cell(r, c).value
                        if v is not None and not is_formula(str(v)) and not is_datelike(v):
                            new_ws.cell(r, c).value = v
            else:
                for dr in data_offsets:
                    r = block_start + dr
                    for c in [2, 3, 4, 5]:
                        # Skip if master template has a formula in this cell
                        if is_formula(str(new_ws.cell(r, c).value)):
                            continue
                        col_ltr = get_column_letter(c)
                        new_ws.cell(r, c).value = f"='{wk_one_sheet_name}'!{col_ltr}{r}"

        elif month_idx == prev_idx:
            # Prev month (Jun when target=Jul):
            # Cols 2,3,4 = historical years from LY ROB (same source as Jul+)
            # Col 5     = actual current-year data from prev ROB (built up weekly)
            if not _is_rob_month_blank(new_ws, block_start):
                continue
            if is_wk_one:
                if ly_ws:
                    for dr in data_offsets:
                        r = block_start + dr
                        for ly_col, new_col in ly_to_new.items():
                            # Skip if master template has a formula in this cell
                            if is_formula(str(new_ws.cell(r, new_col).value)):
                                continue
                            v = ly_ws.cell(r, ly_col).value
                            if v is not None and not is_formula(str(v)) and not is_datelike(v):
                                new_ws.cell(r, new_col).value = v
                    ly_sec_col = find_secondary_col(ly_ws, block_start) or 7
                    for dr in [4, 5, 6]:
                        r = block_start + dr
                        # Skip if master template has a formula in this cell
                        if is_formula(str(new_ws.cell(r, 8).value)):
                            continue
                        v = ly_ws.cell(r, ly_sec_col).value
                        if v is not None and not is_formula(str(v)) and not is_datelike(v):
                            new_ws.cell(r, 8).value = v
            else:
                for dr in data_offsets:
                    r = block_start + dr
                    for c in [2, 3, 4, 5]:
                        # Skip if master template has a formula in this cell
                        if is_formula(str(new_ws.cell(r, c).value)):
                            continue
                        col_ltr = get_column_letter(c)
                        new_ws.cell(r, c).value = f"='{wk_one_sheet_name}'!{col_ltr}{r}"

        else:
            # Current month and future months (Jul+): cols 2,3,4 from LY ROB
            # For new hotels with no historical data (e.g., Foxberry), ly_ws will be None —
            # just skip and leave those months with master template values (no error).
            if ly_ws is None:
                continue
            for dr in data_offsets:
                r = block_start + dr
                for ly_col, new_col in ly_to_new.items():
                    # Skip if master template has a formula in this cell
                    if is_formula(str(new_ws.cell(r, new_col).value)):
                        continue
                    v = ly_ws.cell(r, ly_col).value
                    if v is not None and not is_formula(str(v)) and not is_datelike(v):
                        new_ws.cell(r, new_col).value = v
            ly_sec_col = find_secondary_col(ly_ws, block_start) or 7
            for dr in [4, 5, 6]:
                r = block_start + dr
                # Skip if master template has a formula in this cell
                if is_formula(str(new_ws.cell(r, 8).value)):
                    continue
                v = ly_ws.cell(r, ly_sec_col).value
                if v is not None and not is_formula(str(v)) and not is_datelike(v):
                    new_ws.cell(r, 8).value = v


def _wk1_previous_table_refs(ws, target_year):
    months=['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']
    hdr=None
    for r in range(1,min(ws.max_row+1,80)):
        for c in range(1,ws.max_column+1):
            s=str(ws.cell(r,c).value or '').lower()
            if 'week 1 previous' in s or ('calculation only' in s and 'week' in s):
                hdr=(r,c); break
        if hdr: break
    if not hdr: return {}
    hr,hc=hdr; mcol=None
    for r in range(hr+1,min(ws.max_row+1,hr+40)):
        for c in range(max(1,hc-5),min(ws.max_column+1,hc+8)):
            if str(ws.cell(r,c).value or '').strip().lower() in months:
                mcol=c; break
        if mcol: break
    if not mcol: return {}
    mrows={}
    for r in range(hr+1,min(ws.max_row+1,hr+40)):
        s=str(ws.cell(r,mcol).value or '').strip().lower()
        if s in months: mrows[months.index(s)]=r
    ycols={}
    for r in range(hr,min(ws.max_row+1,hr+10)):
        for c in range(mcol+1,min(ws.max_column+1,mcol+20)):
            v=ws.cell(r,c).value; y=None
            if isinstance(v,(int,float)) and 2000<=int(v)<=2100: y=int(v)
            elif isinstance(v,str):
                m=re.search(r'\b(20\d{2})\b',v); y=int(m.group(1)) if m else None
            elif isinstance(v,(datetime.datetime,datetime.date)): y=v.year
            if y: ycols[y]=c
        if len(ycols)>=4: break
    if not ycols:
        for c,y in zip(range(mcol+1,mcol+5),range(target_year-3,target_year+1)): ycols[y]=c
    return {(mi,y):ws.cell(r,c).coordinate for mi,r in mrows.items() for y,c in ycols.items()}


def apply_rob_pickup_wow_formulas(wb, target_year):
    """All months, all year columns: current-week Revenue minus prior-week Revenue."""
    from openpyxl.utils import get_column_letter
    weeks=[s for s in ROB_SHEETS if s in wb.sheetnames]
    if not weeks: return ['Pickup WoW formulas: no ROB week tabs found.']
    wk1=wb[weeks[0]]; refs=_wk1_previous_table_refs(wk1,target_year)
    year_by_col={}
    for c in range(2,6):
        v=wk1.cell(4,c).value; y=None
        if isinstance(v,(datetime.datetime,datetime.date)): y=v.year
        elif isinstance(v,(int,float)) and 2000<=int(v)<=2100: y=int(v)
        elif isinstance(v,str):
            m=re.search(r'\b(20\d{2})\b',v); y=int(m.group(1)) if m else None
        year_by_col[c]=y or (target_year-5+c)
    for wi,sname in enumerate(weeks):
        ws=wb[sname]
        for mi,labels in rob_month_blocks(ws).items():
            rr=labels.get('revenue')
            pr=next((r for lab,r in labels.items() if 'pickup' in lab and 'wow' in lab.replace(' ','')),None)
            if not rr or not pr: continue
            for c in range(2,6):
                L=get_column_letter(c)
                if wi==0:
                    ref=refs.get((mi,year_by_col[c]))
                    if ref: ws.cell(pr,c).value=f'={L}{rr}-{ref}'
                else:
                    prev=weeks[wi-1]
                    ws.cell(pr,c).value=f"={L}{rr}-'{prev}'!{L}{rr}"
    return [] if refs else ['Pickup WoW formulas: WK1 previous table could not be mapped; WK1 formulas left unchanged.']


def setup_new_rob_month(service, hotel_id: str, hotel_name: str, target_month: datetime.date):
    """Full ROB new-month setup.
    Returns (new_file_name, error_or_warn_str, new_file_id, original_bytes).
    `original_bytes` is the file's content exactly as it was before this
    run's changes (None on any early-exit failure) — the caller can use it
    to offer an "undo"/reset-to-original for this setup, the same way the
    Strategy Report setup already does.
    """
    year_kw  = str(target_month.year)
    month_kw = target_month.strftime("%b%Y").upper()

    # ── Locate month folder (never created — must already exist) ────────────
    rev_id, rev_name = _find_rev_reports_folder_for_year(service, hotel_id, year_kw, month_kw)
    if not rev_id:
        return None, "No REVENUE REPORTS folder.", None, None
    month_id, month_name = _find_month_folder_under_rev(service, rev_id, year_kw, month_kw, target_month, hotel_name)
    if not month_id:
        return None, f"Could not find the {month_kw} folder for {hotel_name} — it should already exist.", None, None

    # Diagnostic: which folders are being used
    folder_diagnostic = f"Revenue Reports folder: {rev_name}; Month folder: {month_name}"

    # ── Find or copy the file ─────────────────────────────────────────────────
    # Diagnostic: list every file in month_id whose name contains "ROB" —
    # confirmed real risk this session (Wolfeboro's duplicate-named folders)
    # that two files with the same display name could sit in the same
    # folder, with drive_find_file silently picking one while a human
    # browsing to the file by name lands on the other.
    dup_check_warnings = []
    try:
        dup_q = ("'%s' in parents and trashed = false and name contains 'ROB'") % month_id
        dup_files = service.files().list(
            q=dup_q, fields="files(id,name)", pageSize=20,
            supportsAllDrives=True, includeItemsFromAllDrives=True,
        ).execute().get("files", [])
        if len(dup_files) > 1:
            dup_check_warnings.append(
                f"Multiple files matching 'ROB' found in the target month folder: "
                + ", ".join(f"'{f['name']}' (id: {f['id']})" for f in dup_files)
            )
    except Exception:
        pass

    existing_id, existing_name = drive_find_file(service, "ROB", month_id)
    is_fresh_copy = False
    if existing_id and "master" not in existing_name.lower():
        new_file_id, new_file_name = existing_id, existing_name
    else:
        is_fresh_copy = True
        master_id, master_name = find_rob_master(service, hotel_id, target_month.year)
        if not master_id:
            return None, master_name, None, None
        hotel_suffix = hotel_name.upper()
        name_upper   = master_name.upper()
        if "ROB" in name_upper:
            after = master_name[name_upper.find("ROB") + 3:].strip()
            after = after.replace(".xlsx","").replace(".xlsm","").replace(".XLSX","").replace(".XLSM","").strip()
            if after:
                hotel_suffix = after
        ext = ".xlsm" if master_name.lower().endswith(".xlsm") else ".xlsx"
        new_file_name = f"{month_kw} ROB {hotel_suffix}{ext}"
        try:
            new_file_id, new_file_name = drive_copy_file(service, master_id, new_file_name, month_id)
        except Exception as e:
            return None, str(e), None, None

    # ── Load all three workbooks ──────────────────────────────────────────────
    new_wb_bytes = drive_download(service, new_file_id)
    new_wb = openpyxl.load_workbook(io.BytesIO(new_wb_bytes), data_only=False)
    if is_fresh_copy:
        clear_tab_colors(new_wb, ROB_SHEETS)

    # Resolution/load failures here used to be swallowed silently — the
    # ROB workbook would just come back with July onward blank and no
    # indication why. Both lookup failures (resolve_drive_workbook returning
    # an error) and load failures (download/openpyxl exceptions) are now
    # captured into `warnings` and surfaced to the caller alongside the
    # success message, matching the diagnostic the SR flow already shows.
    warnings = [
        f"Target file id: {new_file_id} (name: {new_file_name})",
        f"Folder placement: {folder_diagnostic}",
    ] + dup_check_warnings

    prev_month_dt = (target_month - datetime.timedelta(days=1)).replace(day=1)
    prev_result, prev_err = resolve_drive_workbook(service, hotel_id, hotel_name, "ROB",
                                                     month_date=prev_month_dt)
    prev_wb = None
    prev_wb_formulas = None
    if prev_result:
        warnings.append(f"Prev month ({prev_month_dt.strftime('%b %Y')}) resolved to: {prev_result[1]}")
        try:
            prev_bytes = drive_download(service, prev_result[0])
            prev_wb = openpyxl.load_workbook(io.BytesIO(prev_bytes), data_only=True)
            prev_wb_formulas = openpyxl.load_workbook(io.BytesIO(prev_bytes), data_only=False)
        except Exception as e:
            warnings.append(f"Prev month ({prev_month_dt.strftime('%b %Y')}) workbook found but failed to load: {e}")
    else:
        warnings.append(f"Prev month ({prev_month_dt.strftime('%b %Y')}) not found: {prev_err}")

    ly_month_dt = target_month.replace(year=target_month.year - 1)
    ly_result, ly_err = resolve_drive_workbook(service, hotel_id, hotel_name, "ROB",
                                                month_date=ly_month_dt)
    ly_wb = None
    if ly_result:
        warnings.append(f"Last year ({ly_month_dt.strftime('%b %Y')}) resolved to: {ly_result[1]}")
        try:
            ly_wb = openpyxl.load_workbook(
                io.BytesIO(drive_download(service, ly_result[0])), data_only=True)
        except Exception as e:
            warnings.append(f"Last year ({ly_month_dt.strftime('%b %Y')}) workbook found but failed to load: {e}")
    else:
        warnings.append(f"Last year ({ly_month_dt.strftime('%b %Y')}) not found — future months' historical "
                         f"columns will be blank: {ly_err}")

    # Sheet lookups below are exact, case-sensitive matches against
    # ROB_SHEETS ("wk one", "wk two", ...) — a workbook can load successfully
    # above and still contribute nothing if its own tab names don't match
    # that exact casing/spacing (e.g. "Wk One", "WK1"). Checking only "wk
    # one" previously missed this: a source workbook can have a correctly
    # named "wk one" (so wk one fills fine) while missing/mismatching "wk
    # two" through "wk six" — confirmed real case where wk one populated but
    # wk two silently didn't. Check every ROB_SHEETS name, not just the first.
    for label, wb_obj in [("Prev month", prev_wb), ("Last year", ly_wb)]:
        if wb_obj is None:
            continue
        missing = [s for s in ROB_SHEETS if s not in wb_obj.sheetnames]
        if missing:
            warnings.append(
                f"{label} workbook is missing tabs: {missing} — "
                f"actual tabs: {wb_obj.sheetnames}"
            )

    # ── Fill each sheet ───────────────────────────────────────────────────────
    wk_one_name = ROB_SHEETS[0]
    for sheet_name in ROB_SHEETS:
        if sheet_name not in new_wb.sheetnames:
            continue
        new_ws  = new_wb[sheet_name]
        prev_ws = prev_wb[sheet_name] if prev_wb and sheet_name in prev_wb.sheetnames else None
        ly_ws   = ly_wb[sheet_name]   if ly_wb   and sheet_name in ly_wb.sheetnames   else None
        is_wk_one = (sheet_name == wk_one_name)
        _fill_rob_sheet(new_ws, prev_ws, ly_ws, target_month, is_wk_one, wk_one_name)

    # Normalize all ROB date headers to MM/DD/YYYY display.
    for _s in ROB_SHEETS:
        if _s not in new_wb.sheetnames:
            continue
        _ws = new_wb[_s]
        _step = rob_block_step(_ws)
        for _mi in range(12):
            _row = 4 + _step * _mi
            for _col in range(2, 6):
                _ws.cell(_row, _col).number_format = "mm/dd/yyyy"

    # Direct readback of the target month's own block, right after the fill
    # loop and before save/upload — confirms whether the in-memory write
    # actually happened per sheet, isolating "fill logic didn't write it
    # for this tab" from "something after this point (save/upload) lost
    # it". Checked across all six week tabs, not just WKONE — confirmed
    # real case where wk one's fill succeeded but wk two's silently didn't.
    # Also check Jan/Feb/Mar/Dec to verify backward/future months are filled.
    target_idx = target_month.month - 1
    target_block_start = 4 + 8 * target_idx
    readback = {
        s: new_wb[s].cell(target_block_start + 1, 2).value
        for s in ROB_SHEETS if s in new_wb.sheetnames
    }
    wk_one_sheet = ROB_SHEETS[0]
    if wk_one_sheet in new_wb.sheetnames:
        ws = new_wb[wk_one_sheet]
        month_check = {}
        for m, mlabel in [(1, "Jan"), (2, "Feb"), (3, "Mar"), (12, "Dec")]:
            block = 4 + 8 * (m - 1)
            month_check[mlabel] = ws.cell(block + 1, 2).value
        warnings.append(
            f"Months check (wk one only) — Jan/Feb/Mar/Dec Revenue (col B): {month_check!r}"
        )
    warnings.append(
        f"Readback check — {target_month.strftime('%b %Y')} Revenue (row "
        f"{target_block_start + 1}, col B) per tab after fill: {readback!r}"
    )

    # ── Fill Week 1 Previous Sheet table in wk one ───────────────────────────
    if prev_wb and wk_one_name in new_wb.sheetnames:
        err = _fill_rob_prev_table(new_wb[wk_one_name], prev_wb, prev_wb_formulas, target_month)
        if err:
            warnings.append(f"Prev table: {err}")

    warnings.extend(apply_rob_pickup_wow_formulas(new_wb, target_month.year))

    strip_tables(new_wb)
    out = io.BytesIO()
    new_wb.save(out)
    drive_upload(service, new_file_id, out.getvalue(), new_file_name)
    warn_str = "; ".join(warnings) if warnings else None
    return new_file_name, warn_str, new_file_id, new_wb_bytes


def find_forecast_master(service, hotel_id: str):
    """Search the hotel's own Drive tree for the Forecast master file (.xlsx or .xlsm)."""
    scope_ids = _hotel_search_scope_ids(service, hotel_id)
    if not scope_ids:
        return None, "Could not resolve hotel folder to search."
    parent_clause = " or ".join("'%s' in parents" % sid for sid in scope_ids)
    q = ("trashed=false and (%s) "
         "and (mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' "
         "or mimeType='application/vnd.ms-excel.sheet.macroenabled.12') "
         "and name contains 'MASTER' and name contains 'FORECAST'") % parent_clause
    result = service.files().list(
        q=q, fields="files(id,name,parents)", pageSize=50,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    for f in result.get("files", []):
        return f["id"], f["name"]
    return None, "No FORECAST master file found in Drive."


def setup_new_forecast_month(service, hotel_id: str, hotel_name: str, target_month: datetime.date):
    """
    Copy Forecast master → rename for target_month → place in month folder → set B4 date.
    Returns (new_file_name, error_str).
    """
    year_kw  = str(target_month.year)
    month_kw = target_month.strftime("%b%Y").upper()

    rev_id, _ = _find_rev_reports_folder_for_year(service, hotel_id, year_kw, month_kw)
    if not rev_id:
        return None, "No REVENUE REPORTS folder."

    month_id, _ = _find_month_folder_under_rev(service, rev_id, year_kw, month_kw, target_month, hotel_name)
    if not month_id:
        return None, f"Could not find the {month_kw} folder for {hotel_name} — it should already exist."

    # Check if Forecast already exists
    existing_id, existing_name = drive_find_file(service, "FORECAST", month_id)
    if existing_id and "master" not in existing_name.lower():
        return existing_name, None

    master_id, master_name = find_forecast_master(service, hotel_id)
    if not master_id:
        return None, master_name

    # Infer hotel suffix from master name
    hotel_suffix = hotel_name.upper()
    name_upper = master_name.upper()
    ext = ".xlsm" if master_name.lower().endswith(".xlsm") else ".xlsx"
    for kw in ("FORECAST",):
        if kw in name_upper:
            after = master_name[name_upper.find(kw) + len(kw):].strip()
            after = after.replace(".xlsx","").replace(".xlsm","").replace(".XLSX","").replace(".XLSM","").strip()
            if after:
                hotel_suffix = after
            break

    new_file_name = f"{month_kw} FORECAST {hotel_suffix}{ext}"
    try:
        new_file_id, created_name = drive_copy_file(service, master_id, new_file_name, month_id)
    except Exception as e:
        return None, str(e)

    # Set B4 = first day of target_month in FCST-WK1
    try:
        wb_bytes  = drive_download(service, new_file_id)
        wb        = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)
        clear_tab_colors(wb, FORECAST_SHEETS)
        sheet     = FORECAST_SHEETS[0] if FORECAST_SHEETS[0] in wb.sheetnames else wb.sheetnames[1] if len(wb.sheetnames) > 1 else wb.sheetnames[0]
        ws        = wb[sheet]
        # Find "Day of Week" cell → one right + one down = start date cell
        date_cell_row = date_cell_col = None
        for r in range(1, 15):
            for c in range(1, 10):
                if "day of week" in str(ws.cell(r, c).value or "").lower():
                    date_cell_row = r + 1
                    date_cell_col = c + 1
                    break
            if date_cell_row:
                break
        if date_cell_row and date_cell_col:
            ws.cell(date_cell_row, date_cell_col).value = datetime.datetime(
                target_month.year, target_month.month, 1)
        strip_tables(wb)
        out = io.BytesIO()
        wb.save(out)
        drive_upload(service, new_file_id, out.getvalue(), created_name)
    except Exception as e:
        return created_name, f"Copied OK but could not set start date: {e}"

    return created_name, None


def find_sr_master(service, hotel_id: str, target_year=None):
    """Search the hotel's Drive tree for the correct Strategy master.

    If target_year is supplied:
      - prefer a master explicitly tied to that year
      - allow a generic/no-year master
      - reject masters explicitly tied to another year

    This prevents a 2026 Strategy setup from accidentally copying a 2027
    master just because Drive returned that file first.
    """
    scope_ids = _hotel_search_scope_ids(service, hotel_id)
    if not scope_ids:
        return None, "Could not resolve hotel folder to search."

    parent_clause = " or ".join("'%s' in parents" % sid for sid in scope_ids)
    q = (
        "trashed=false and (%s) "
        "and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' "
        "and name contains 'MASTER' and name contains 'STRATEGY'"
    ) % parent_clause

    result = service.files().list(
        q=q,
        fields="files(id,name,parents,modifiedTime)",
        pageSize=100,
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()

    candidates = result.get("files", [])
    if not candidates:
        return None, "No STRATEGY master file found in Drive."

    def explicit_year(value):
        m = re.search(r"(?<!\\d)(20\\d{2})(?!\\d)", str(value or ""))
        return int(m.group(1)) if m else None

    parent_name_cache = {}

    def parent_years(file_obj):
        years = set()
        for pid in file_obj.get("parents", []) or []:
            if pid not in parent_name_cache:
                try:
                    info = service.files().get(
                        fileId=pid,
                        fields="name",
                        supportsAllDrives=True,
                    ).execute()
                    parent_name_cache[pid] = info.get("name", "")
                except Exception:
                    parent_name_cache[pid] = ""
            y = explicit_year(parent_name_cache[pid])
            if y:
                years.add(y)
        return years

    scored = []
    for f in candidates:
        years = set()
        fy = explicit_year(f.get("name"))
        if fy:
            years.add(fy)
        years.update(parent_years(f))

        if target_year is not None and years and target_year not in years:
            continue

        score = 2 if (target_year is not None and target_year in years) else 1
        scored.append((score, str(f.get("modifiedTime", "")), f))

    if not scored:
        return None, (
            f"No Strategy master for {target_year} was found in this hotel's Drive folders."
            if target_year
            else "No eligible Strategy master file found in Drive."
        )

    scored.sort(key=lambda x: (x[0], x[1]), reverse=True)
    best = scored[0][2]
    return best["id"], best["name"]


def setup_new_sr_month(service, hotel_id: str, hotel_name: str, target_month: datetime.date):
    """
    New month SR setup:
      1. Find SR master, copy it, rename to [MON][YEAR] STRATEGY [HOTEL].xlsx
      2. Find or create month folder under the year folder
      3. Move the copy into that folder
    Returns (new_file_name, error_str).
    """
    year_kw = str(target_month.year)
    month_kw = target_month.strftime("%b%Y").upper()

    # Resolve REVENUE REPORTS and month folder
    rev_id, rev_name = _find_rev_reports_folder_for_year(service, hotel_id, year_kw, month_kw)
    if not rev_id:
        return None, "No REVENUE REPORTS folder."

    # Find month folder — never created, must already exist
    month_id, month_name = _find_month_folder_under_rev(service, rev_id, year_kw, month_kw, target_month, hotel_name)
    if not month_id:
        return None, f"Could not find the {month_kw} folder for {hotel_name} — it should already exist. (Revenue Reports folder: {rev_name})"

    # Check if SR already exists in that folder
    existing_id, existing_name = drive_find_file(service, "STRATEGY", month_id)
    if existing_id and "master" not in existing_name.lower():
        return existing_name, None  # already set up

    # Find master
    master_id, master_name = find_sr_master(service, hotel_id, target_month.year)
    if not master_id:
        return None, master_name  # error string

    # Infer hotel suffix from master file name for the new file name
    # e.g. "MASTER 2026 STRATEGY PLYMOUTH.xlsx" → "PLYMOUTH"
    hotel_suffix = hotel_name.upper()
    name_upper = master_name.upper().replace(".XLSX", "")
    if "STRATEGY" in name_upper:
        after = name_upper[name_upper.find("STRATEGY") + len("STRATEGY"):].strip()
        if after:
            hotel_suffix = master_name[master_name.upper().find("STRATEGY") + len("STRATEGY"):].strip().replace(".xlsx", "").replace(".XLSX", "").strip()

    new_file_name = f"{month_kw} STRATEGY {hotel_suffix}.xlsx"
    try:
        _, created_name = drive_copy_file(service, master_id, new_file_name, month_id)
    except Exception as e:
        return None, str(e)
    return created_name, None


def get_prev_month_otb_trans(service, hotel_id: str, hotel_name: str, current_month: datetime.date):
    """Pull OTB TY Trans values for current_month dates from previous month's SR last filled tab.
    Returns {date: value} or {} on any failure.
    """
    prev_month = (current_month.replace(day=1) - datetime.timedelta(days=1)).replace(day=1)
    result, err = resolve_drive_workbook(service, hotel_id, hotel_name, "Strategy Report", month_date=prev_month)
    if err or not result:
        return {}

    file_id, _ = result
    try:
        file_bytes = drive_download(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    except Exception:
        return {}

    # Find last filled (colored) tab
    last_filled_sheet = None
    for sheet_name in STRATEGY_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        tab_color = ws.sheet_properties.tabColor
        if tab_color is not None:
            last_filled_sheet = sheet_name

    if not last_filled_sheet:
        return {}

    ws = wb[last_filled_sheet]
    col_map = detect_strategy_columns(ws)
    otb_col = col_map.get("otb_trans")
    date_col = detect_date_column(ws, wb=wb)
    if not otb_col:
        return {}

    result_map = {}
    for r in range(5, ws.max_row + 1):
        date_val = ws.cell(r, date_col).value
        if isinstance(date_val, datetime.datetime):
            d = date_val.date()
        elif isinstance(date_val, datetime.date):
            d = date_val
        else:
            continue
        if d < current_month:
            continue
        cell_val = ws.cell(r, otb_col).value
        if cell_val is not None and not is_formula(cell_val):
            result_map[d] = safe_float(cell_val)

    return result_map


def resolve_drive_workbook(service, hotel_id: str, hotel_name: str, workbook_type: str, month_date: datetime.date = None):
    """
    Walk Drive to find the target workbook. Handles two folder structures:
      A) Hotel > MMMYYYY REVENUE REPORTS HOTEL > files  (month in folder name, files direct)
      B) Hotel > REVENUE REPORTS > Year > Month > files (nested year/month subfolders)
    Returns ((file_id, file_name), None) or (None, error_message).
    Never touches files whose name contains 'master'.

    A "hotel" can also be a MULTI:<id>,<id>,... group — several candidate
    root folders sharing one dropdown entry, either because they're each a
    flat per-year/month folder shared directly (Hyannis Anchor In) or full
    duplicate top-level hotel folders from historical typos/copies
    (confirmed real case: "Provinceetown Surfside", "Provincertown
    Surfside", "Surfside (1)" all being the same hotel). Each candidate is
    resolved fully (structures A/B/C, recursing into its own REVENUE REPORTS
    child if it has one) rather than guessed from its name alone — whichever
    candidate actually contains the target file wins.
    """
    if month_date is None:
        month_date = datetime.date.today()

    month_kw        = month_date.strftime("%b%Y").upper()
    month_kw_2digit = month_date.strftime("%b%y").upper()
    year_kw         = str(month_date.year)
    wb_keyword      = WORKBOOK_KEYWORDS[workbook_type]

    def _list_subfolders(parent_id):
        q = (f"'{parent_id}' in parents and trashed = false and "
             f"mimeType = 'application/vnd.google-apps.folder'")
        return service.files().list(
            q=q, fields="files(id, name)", pageSize=100,
            supportsAllDrives=True, includeItemsFromAllDrives=True,
        ).execute().get("files", [])

    def _find_file_in(folder_id, folder_name):
        fid, fname = drive_find_file(service, wb_keyword, folder_id)
        if not fid:
            return None, f"No '{wb_keyword}' workbook found in '{folder_name}'."
        if "master" in fname.lower():
            return None, f"Resolved file '{fname}' looks like a master doc — aborting."
        return (fid, fname), None

    def _resolve_single(single_id, single_name):
        """Resolve within ONE candidate root — either a full hotel-parent
        folder (with its own REVENUE REPORTS > year > month nesting) or a
        folder that IS already the REVENUE REPORTS level directly (detected
        from its own name). Returns ((file_id, file_name), None) or (None, err).
        """
        self_is_rev = _is_rev_reports_name(single_name)

        # A0: candidate IS already the month-level folder — its own name
        # contains both "REVENUE REPORTS" and the target month, so the file
        # is directly inside it, not further down. Confirmed real case:
        # Tybee's month folders are individually named "G: JUL2026 REVENUE
        # REPORTS TYBEE"; because that name also contains "REVENUE REPORTS"
        # it gets swept into the same MULTI: candidate group as Tybee's root
        # and year-level folders (get_hotels_from_drive merges by extracted
        # hotel name, not by structural level). Without this check,
        # self_is_rev below always assumes further year/month descent is
        # needed, finds no subfolders (children are files here), and fails
        # even though this candidate is the correct one.
        if self_is_rev and month_kw in single_name.upper():
            result = _find_file_in(single_id, single_name)
            if result[0]:
                return result

        children = _list_subfolders(single_id)

        # A: Hotel > MMMYYYY REVENUE REPORTS HOTEL > file
        if not self_is_rev:
            a = next((f for f in children
                       if _is_rev_reports_name(f["name"]) and month_kw in f["name"].upper()), None)
            if a:
                result = _find_file_in(a["id"], a["name"])
                if result[0]:
                    return result

        # B: Hotel > REVENUE REPORTS > MMMYYYY ... > file
        if self_is_rev:
            rev = {"id": single_id, "name": single_name}
        else:
            rev_candidates = [
                f for f in children
                if _is_rev_reports_name(f["name"])
            ]
            rev = _pick_rev_reports_candidate(
                rev_candidates,
                year_kw,
                month_kw,
            )

        if rev:
            rev_children = children if self_is_rev else _list_subfolders(rev["id"])
            b1 = next((f for f in rev_children if month_kw in f["name"].upper()), None)
            if b1:
                result = _find_file_in(b1["id"], b1["name"])
                if result[0]:
                    return result
            b2_year = next((f for f in rev_children if year_kw in f["name"].upper()), None)
            if b2_year:
                b2_month_id, b2_month_name = drive_find_folder_by_keyword(
                    service, month_kw, parent_id=b2_year["id"])
                if b2_month_id:
                    result = _find_file_in(b2_month_id, b2_month_name)
                    if result[0]:
                        return result

        # C: Hotel > Year > Month > file  (no REVENUE REPORTS wrapper)
        c_year = next((f for f in children
                       if year_kw in f["name"].upper()
                       and not _is_rev_reports_name(f["name"])), None)
        if c_year:
            c_month_id, c_month_name = drive_find_folder_by_keyword(
                service, month_kw, parent_id=c_year["id"])
            if c_month_id:
                result = _find_file_in(c_month_id, c_month_name)
                if result[0]:
                    return result

        return None, f"Could not find '{month_kw}' workbook under '{single_name}'."

    if hotel_id.startswith(MULTI_ID_PREFIX):
        candidate_ids = hotel_id[len(MULTI_ID_PREFIX):].split(",")
        candidates = []

        for cid in candidate_ids:
            try:
                info = service.files().get(
                    fileId=cid,
                    fields="name",
                    supportsAllDrives=True,
                ).execute()
                candidates.append({"id": cid, "name": info["name"]})
            except Exception:
                continue

        if not candidates:
            return None, f"Could not read any of the shared folders for '{hotel_name}'."

        target_year = month_date.year
        candidates = [
            f for f in candidates
            if _explicit_folder_year(f["name"]) in (None, target_year)
        ]

        if not candidates:
            return None, (
                f"Could not find a {target_year} Revenue Reports folder "
                f"for '{hotel_name}'."
            )

        def _sort_key(f):
            name_up = f["name"].upper()
            if month_kw in name_up or month_kw_2digit in name_up:
                return 0
            if year_kw in name_up:
                return 1
            return 2

        ordered = sorted(candidates, key=_sort_key)

        last_err = None
        for cand in ordered:
            result, err = _resolve_single(cand["id"], cand["name"])
            if result:
                return result, None
            last_err = err

        return None, last_err or f"Could not find '{month_kw}' workbook for '{hotel_name}'."

    return _resolve_single(hotel_id, hotel_name)


DOW_ABBREVS = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]


def _count_sheet_data_rows(ws):
    """Count how many data rows exist from row 5 downward by finding the last
    row that has any non-empty content across the first 10 columns."""
    last_row = 4
    for r in range(5, ws.max_row + 1):
        if any(ws.cell(r, c).value is not None for c in range(1, 11)):
            last_row = r
    return max(0, last_row - 4)  # number of rows starting from row 5


def restructure_sr_dates(wb, target_month):
    """Restructure the three date columns in every strategy sheet for a new month.
    Row 4 (header):    col 1 = LY year, col 3 = TY year — else stale years carry
                        over from whatever the master template last had.
    Col 1 (LY date):   starts at day 2 of target_month, last year
    Col 2 (day of wk): abbreviation matching the TY date in col 3
    Col 3 (TY date):   starts at day 1 of target_month, this year
    Row count matches the master — hotels open only part of the year have fewer rows.
    """
    ty_start = target_month
    ly_start = datetime.date(ty_start.year - 1, ty_start.month, 2)

    for sheet_name in STRATEGY_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # Skip merged cells (can't set value on merged cells)
        try:
            ws.cell(4, 1).value = ly_start.year
        except (AttributeError, ValueError):
            pass
        try:
            ws.cell(4, 3).value = ty_start.year
        except (AttributeError, ValueError):
            pass
        num_rows = min(_count_sheet_data_rows(ws), 365)
        if num_rows == 0:
            continue
        for i in range(num_rows):
            row     = 5 + i
            ty_date = ty_start + datetime.timedelta(days=i)
            ly_date = ly_start + datetime.timedelta(days=i)
            dow     = DOW_ABBREVS[ty_date.weekday()]
            try:
                ws.cell(row, 1).value = datetime.datetime(ly_date.year, ly_date.month, ly_date.day)
            except (AttributeError, ValueError):
                pass
            try:
                ws.cell(row, 2).value = dow
            except (AttributeError, ValueError):
                pass
            try:
                ws.cell(row, 3).value = datetime.datetime(ty_date.year, ty_date.month, ty_date.day)
            except (AttributeError, ValueError):
                pass


def _load_wb_from_drive(svc, hotel_id, hotel_name, wb_type, month_date, data_only=True):
    """Download and parse a workbook from Drive. Returns openpyxl.Workbook or None.
    data_only=True (default) returns cached cell values — use for reference workbooks.
    data_only=False preserves formulas — use for workbooks we intend to write back.
    """
    result, err = resolve_drive_workbook(svc, hotel_id, hotel_name, wb_type, month_date=month_date)
    if err or not result:
        return None
    try:
        return openpyxl.load_workbook(io.BytesIO(drive_download(svc, result[0])), data_only=data_only)
    except Exception:
        return None



# ── Monthly Ancillary Revenue Report Builder ─────────────────────────────────
# Ported from the Google Apps Script prototype v11.12. The app version is
# upload/preview/download first so hotel-specific rules can be validated before
# any automatic Drive overwrite is enabled.
ANCILLARY_TEMPLATE_FILENAME = "Ancillary Revenue Report Builder.xlsx"

ANCILLARY_PROPERTY_PROFILES = {
    'ashworth by the sea': {'display':'Ashworth by the Sea','stlySource':'CANARY','journal':[
        {'label':'1005 Early Check-In','report':'Early Check In'},
        {'label':'1006 Late Checkout','report':'Late Checkout'},
        {'label':'1007 Very Early Check-In','report':'Very Early Check In'},
        {'label':'1008 Very Late Checkout','report':'Very Late Check Out'}]},
    'inn at middletown': {'display':'Inn at Middletown','stlySource':'CANARY','journal':[
        {'label':'6006 Early Check In / Late Checkout','report':'Early Check In / Late Checkout'}]},
    'crowne pointe inn and spa': {'display':'Crowne Pointe Inn & Spa','stlySource':'SNT','stlyJournal':True,'journal':[
        {'label':'1009 Early Arrival','report':'Early Check In'},
        {'label':'1008 Late Checkout Fee','report':'Late Checkout'}]},
    'anchor in': {'display':'Anchor In','stlySource':'CANARY','journal':[
        {'label':'4222 Early Check In Fee','report':'Early Check In'},
        {'label':'Late Checkout Fee','report':'Late Checkout'}]},
    'allegria hotel': {'display':'Allegria Hotel','stlySource':'SNT','stlyJournal':True,'journal':[
        {'label':'1030 Early Check In Fee','report':'Early Check In'},
        {'label':'1031 Late Check Out Fee','report':'Late Checkout'}]},
    'the brass key guesthouse': {'display':'The Brass Key Guesthouse','stlySource':'SNT','stlyJournal':True,'journal':[
        {'label':'112 Early Check In Fee','report':'Early Check In'},
        {'label':'121 Late Checkout','report':'Late Checkout'}]},
    'harbor hotel provincetown': {'display':'Harbor Hotel Provincetown','stlySource':'CANARY','journal':[
        {'label':'4006 Early Checkin/Late Departure Fee','report':'Early Check In / Late Checkout'}]},
    'provincetown inn': {'display':'Provincetown Inn','stlySource':'SNT','stlyJournal':True,'journal':[
        {'label':'1009 Early Arrival','report':'Early Check In'},
        {'label':'1008 Late Checkout Fee','report':'Late Checkout'}]},
    'surfside hotel and suites': {'display':'Surfside Hotel & Suites','stlySource':'CANARY','journal':[
        {'label':'4004 Early Arrival Fee','report':'Early Check In'},
        {'label':'1004 Late Checkout Fee 12PM $25.00','report':'Late Checkout'},
        {'label':'1005 Late Checkout 1PM $40.00','report':'Late Checkout'},
        {'label':'1006 Late Checkout 2PM $60.00','report':'Late Checkout'}]},
    'hotel tybee': {'display':'Hotel Tybee','stlySource':'CANARY','journal':[
        {'label':'4011 Early Check-In','report':'Early Check In'},
        {'label':'4012 Late Checkout','report':'Late Checkout'},
        {'label':'4013 Very Early Check-In','report':'Very Early Check In'},
        {'label':'4014 Very Late Checkout','report':'Very Late Check Out'}]},
    'pleasant view inn': {'display':'Pleasant View Inn','stlySource':'CANARY','journal':[
        {'label':'4004 Early Check In Fee','report':'Early Check In'},
        {'label':'4024 Late Check Out Fee','report':'Late Checkout'}]},
    'the wolfeboro inn': {'display':'The Wolfeboro Inn','stlySource':'CANARY','journal':[
        {'label':'1005 Early Check-In Fee Before 12:00 PM','report':'Early Check In'},
        {'label':'1006 Late Check-Out Fee 12:00 PM','report':'Late Checkout'},
        {'label':'1026 Early Check In Fee Before 1:00 PM','report':'Early Check In'},
        {'label':'1027 Late Check Out Fee 1:00 PM','report':'Late Checkout'}]},
}

# Ancillary display/profile name -> canonical hotel label used by the shared
# Revenue Reports Drive discovery.
ANCILLARY_DRIVE_HOTEL_MAP = {
    "ashworth by the sea": "Ashworth",
    "inn at middletown": "Middletown",
    "crowne pointe inn and spa": "Crowne Pointe",
    "anchor in": "Anchor Inn",
    "allegria hotel": "Long Beach",
    "the brass key guesthouse": "Brass Key",
    "harbor hotel provincetown": "Harbor Hotel",
    "provincetown inn": "Provincetown Inn",
    "surfside hotel and suites": "Surfside",
    "hotel tybee": "Tybee",
    "pleasant view inn": "Westerly",
    "the wolfeboro inn": "Wolfeboro",
}


ANCILLARY_PROPERTY_ALIASES = {
    'brass key guesthouse': 'the brass key guesthouse',
    'harbor hotel': 'harbor hotel provincetown',
    'provincetown surfside': 'surfside hotel and suites',
    'westerly': 'pleasant view inn',
    'wolfeboro inn': 'the wolfeboro inn',
}

ANCILLARY_RULES = {
    'ashworth by the sea': {
        'operational':['Parking Fee','Resort Fee- $25','Booking.com $25 Resort Fee Non-Taxable',
                       'Booking.com $20 Resort Fee Non-Taxable','Booking.com Resort Fee ADJUSTMENT -$5',
                       'Waive Parking Fee','Waive Resort Fee'],
        'itemized':['1pm Early Check-in','2pm Early Check-in','1pm Late Check Out','2pm Late Check-out'],
        'exclude':[]},
    'inn at middletown': {
        'operational':[],
        'itemized':['11am Very Early Check-in','1pm Early Check-in','1pm Late Checkout'],
        'exclude':['Amenity Fee']},
    'crowne pointe inn and spa': {
        'operational':[],
        'itemized':['11am Early Check In','1pm Early Check In','1 PM Late Check Out','2 PM Late Check Out'],
        'exclude':['Resort Fee Waived','Resort Fee Waived On Season']},
    'anchor in': {
        'operational':['Booking.com Resort Fee','Resort Fee - $25','Parking - $10'],
        'itemized':['Early Check in 12PM','Early check in 1PM'],
        'exclude':[]},
    'hotel tybee': {
        'operational':['Amenity Fee - $20','Group Amenity Fee $10.00','Parking Fee - $10',
                       '-$10 Parking Fee off season','Waive Amenity Fee'],
        'itemized':[], 'exclude':[]},
    'allegria hotel': {
        'operational':['Waive $55 Resort Fee (Incl Tax)'],
        'itemized':['Early Check In - 12pm - 3pm','Early Check In - 7am - 12pm',
                    'Relax & Revel until 1pm','Relax & Revel until 2pm'],
        'exclude':[]},
    'pleasant view inn': {
        'operational': ['$42 Resort Fee', 'Waive Resort Fee'],
        'itemized': [
            '1PM Late Check Out - On Season',
            '2PM Early Check In - On Season',
            '2PM Early Check In - On Season - Fr, Sat',
        ],
        'exclude': [],
    },

}

ANCILLARY_YOY_ALIASES = {
    'ashworth by the sea': [
        ('Pet Fee','Bring your furry friend'),('Cheese & Cracker Tray','Cheese and Cracker Tray'),
        ('Cookies & Milk','Cookies and Milk'),('Bottle of Champagne','Bottle of Champagne'),
        ('Early Check In','Early Check-in'),('Early Check In','Very Early Check-in'),
        ('Late Checkout','Late Checkout'),('Late Checkout','Very Late Checkout'),
        ('Very Early Check In','Very Early Check-in'),
        ('Oceanfront Room Two Queen Beds Balcony','Ocean Front, 2 Queen Beds, Balcony'),
        ('Oceanfront Room One King Bed Balcony','Ocean Front, 1 King Bed, Balcony'),
        ('Ashworth Oceanfront Room One King Bed Balcony','Ashworth Ocean Front, 1 King Bed, Balcony')],
    'inn at middletown': [
        ('Pet Fee','Pet Accommodation'),
        ('Early Check In / Late Checkout','Early Check-in'),
        ('Early Check In / Late Checkout','Very Early Check-in'),
        ('Early Check In / Late Checkout','Late Checkout'),
        ('Early Check In / Late Checkout','Very Late Checkout'),
        ('House Red Wine','Bottle of Red Wine')],
    'hotel tybee': [
        ('Pet Fee','Pet Accommodation'),('Rollaway Bed','Extra twin bed $25.00 per day'),
        ('Very Late Check Out','Very Late Checkout'),('Bottle of Prosecco','Bottle of Sparkling Wine'),
        ('Bottle or Red Wine','Bottle of Red Wine'),('All American Bucket of Beer','All American Beach Beer Bucket')],
}


def _ar_norm(v):
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9]+', ' ', str(v or '').lower().replace('&','and'))).strip()


def ancillary_profile(property_name):
    n = _ar_norm(property_name)
    n = ANCILLARY_PROPERTY_ALIASES.get(n, n)
    if n in ANCILLARY_PROPERTY_PROFILES:
        return ANCILLARY_PROPERTY_PROFILES[n], n
    for key, prof in ANCILLARY_PROPERTY_PROFILES.items():
        if n in key or key in n:
            return prof, key
    return ANCILLARY_PROPERTY_PROFILES['ashworth by the sea'], 'ashworth by the sea'


def _ar_num(v):
    if v is None or v == '':
        return None
    if isinstance(v, (int,float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip()
    neg = '(' in s and ')' in s
    s = re.sub(r'[$,%()\s,]', '', s)
    if not s:
        return None
    try:
        n = float(s)
        return -abs(n) if neg else n
    except Exception:
        return None


def _ar_date(v):
    if isinstance(v, datetime.datetime): return v
    if isinstance(v, datetime.date): return datetime.datetime.combine(v, datetime.time())
    if isinstance(v, (int,float)):
        try: return datetime.datetime(1899,12,30) + datetime.timedelta(days=float(v))
        except Exception: return None
    if not v: return None
    for fmt in ('%m/%d/%Y','%Y-%m-%d','%m/%d/%y','%b %d, %Y','%B %d, %Y'):
        try: return datetime.datetime.strptime(str(v).strip(), fmt)
        except Exception: pass
    try: return pd.to_datetime(v).to_pydatetime()
    except Exception: return None


def _ar_file_rows(uploaded_file):
    data = uploaded_file.getvalue() if hasattr(uploaded_file, 'getvalue') else uploaded_file
    name = getattr(uploaded_file, 'name', '').lower()
    if name.endswith('.csv'):
        txt = data.decode('utf-8-sig', errors='replace')
        return list(csv.reader(io.StringIO(txt)))
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _ar_header_map(row):
    return {_ar_norm(v): i for i,v in enumerate(row)}


def _ar_col(h, names):
    for name in names:
        k = _ar_norm(name)
        if k in h: return h[k]
    return -1


def ancillary_parse_addon(raw, property_key):
    rules = ANCILLARY_RULES.get(
        property_key,
        {'operational': [], 'itemized': [], 'exclude': []},
    )
    operational_set = {_ar_norm(x) for x in rules.get('operational', [])}
    itemized_set = {_ar_norm(x) for x in rules.get('itemized', [])}
    exclude_set = {_ar_norm(x) for x in rules.get('exclude', [])}

    header_row = -1
    for i, row in enumerate(raw[:12]):
        if (
            len(row) > 1
            and not str(row[0] or '').strip()
            and 'total count' in str(row[1] or '').lower()
        ):
            header_row = i
            break
    if header_row < 0:
        header_row = 2

    buckets = {'main': [], 'operational': [], 'itemized': []}

    for r in raw[header_row + 1:]:
        name = str(r[0] or '').strip() if r else ''
        if not name:
            continue

        count = _ar_num(r[1] if len(r) > 1 else None)
        revenue = _ar_num(r[2] if len(r) > 2 else None)
        average = _ar_num(r[3] if len(r) > 3 else None)
        if count is None and revenue is None:
            continue

        obj = {
            'name': name,
            'count': count or 0,
            'revenue': revenue or 0,
            'average': average or 0,
        }

        n = _ar_norm(name)
        if n in exclude_set:
            continue

        # Portfolio-wide operational/non-ancillary fee rules.
        is_operational = bool(re.search(
            r'\bwaiv(?:e|ed)\b'
            r'|\bresort\s+fee\b'
            r'|\bamenity\s+fee\b'
            r'|\bparking\s+fee\b'
            r'|\bdestination\s+fee\b'
            r'|\bgroup\s+destination\s+fee\b'
            r'|^parking\s*-?\s*\$'
            r'|booking\.com.*(?:resort|amenity|parking|destination)',
            name,
            re.I,
        ))

        # Detailed ECI/LCO timing products belong only in the itemized section.
        is_timing = bool(re.search(
            r'\bearly\s+(?:check\s*-?\s*in|arrival)\b'
            r'|\blate\s+(?:check\s*-?\s*out|checkout|departure)\b'
            r'|\bvery\s+early\s+check\s*-?\s*in\b'
            r'|\bvery\s+late\s+check\s*-?\s*out\b',
            name,
            re.I,
        ))

        # Main-table ECI/LCO revenue is Journal-only.
        # Raw generic Early Check In / Late Checkout rows are not retained.
        is_journal_equivalent = _ar_is_journal_equivalent(name)

        if n in operational_set or is_operational:
            buckets['operational'].append(obj)
        elif n in itemized_set or (is_timing and not is_journal_equivalent):
            buckets['itemized'].append(obj)
        elif is_journal_equivalent:
            continue
        else:
            buckets['main'].append(obj)

    def combine(rows):
        grouped = {}
        for r in rows:
            key = _ar_norm(r['name'])
            if key not in grouped:
                grouped[key] = {
                    'name': r['name'],
                    'count': 0,
                    'revenue': 0,
                    'average': 0,
                }
            grouped[key]['count'] += _ar_num(r['count']) or 0
            grouped[key]['revenue'] += _ar_num(r['revenue']) or 0

        for r in grouped.values():
            r['average'] = r['revenue'] / r['count'] if r['count'] else 0

        return sorted(
            grouped.values(),
            key=lambda x: x['revenue'],
            reverse=True,
        )

    return {k: combine(v) for k, v in buckets.items()}


def ancillary_parse_upsell(raw):
    if not raw: return {'byRoomType':[],'byStaff':[],'byLevel':[]}
    h=_ar_header_map(raw[0])
    user_c=_ar_col(h,['user']); channel_c=_ar_col(h,['channel']); from_c=_ar_col(h,['from level'])
    room_c=_ar_col(h,['to room type']); to_c=_ar_col(h,['to level']); nights_c=_ar_col(h,['nights']); total_c=_ar_col(h,['total amount'])
    if room_c<0 or total_c<0:
        raise ValueError('SNT Upsell file is missing To Room Type or Total Amount.')
    room={}; staff={}; level={}
    for r in raw[1:]:
        def rv(c): return r[c] if c>=0 and c<len(r) else None
        revenue=_ar_num(rv(total_c)) or 0
        producing=1 if revenue != 0 else 0
        nights=_ar_num(rv(nights_c)) or 0
        from_level=_ar_num(rv(from_c)) or 0; to_level=_ar_num(rv(to_c)) or 0
        lvl=int(to_level-from_level) if (to_level-from_level)>0 else 0
        room_name=str(rv(room_c) or '').strip(); channel=str(rv(channel_c) or '').strip(); user=str(rv(user_c) or '').strip()
        if not user and channel.upper() in ('URL','WEB'): user='WEB'
        if not user: user='Unknown'
        if room_name:
            z=room.setdefault(room_name,{'name':room_name,'count':0,'revenue':0,'actualNights':0})
            z['count']+=producing; z['revenue']+=revenue; z['actualNights']+=producing*nights
        z=staff.setdefault(user,{'name':user,'count':0,'revenue':0}); z['count']+=1; z['revenue']+=revenue
        if lvl>0: level['+'+str(lvl)] = level.get('+'+str(lvl),0)+1
    by_room=[]
    for x in room.values():
        if x['revenue'] or x['actualNights']:
            by_room.append({'name':x['name'],'count':x['actualNights'],'revenue':x['revenue'],'average':x['revenue']/x['actualNights'] if x['actualNights'] else 0})
    by_room.sort(key=lambda x:x['revenue'], reverse=True)
    by_staff=sorted(staff.values(), key=lambda x:(0 if x['name']=='WEB' else 1, x['name'].lower()))
    by_level=[{'level':k,'count':level[k]} for k in sorted(level,key=lambda z:int(z[1:]))]
    return {'byRoomType':by_room,'byStaff':by_staff,'byLevel':by_level}


def ancillary_parse_canary_history(raw, stly_month):
    if not raw: raise ValueError('Historical Canary file is empty.')
    h=_ar_header_map(raw[0]); ac=_ar_col(h,['arrival date']); sc=_ar_col(h,['status']); ic=_ar_col(h,['item']); rc=_ar_col(h,['revenue'])
    if min(ac,sc,ic,rc)<0: raise ValueError('Historical Canary file is missing Arrival Date, Status, Item, or Revenue.')
    agg={}
    for r in raw[1:]:
        dt=_ar_date(r[ac] if ac<len(r) else None)
        if not dt or dt.month!=stly_month.month or dt.year!=stly_month.year: continue
        name=str(r[ic] or '').strip(); status=str(r[sc] or '').strip().lower(); revenue=_ar_num(r[rc]) or 0
        if not name: continue
        z=agg.setdefault(name,{'name':name,'approved':0,'denied':0,'expired':0})
        if status=='approved': z['approved']+=revenue
        elif status=='denied': z['denied']+=revenue
        elif status=='expired': z['expired']+=revenue
    rows=[]
    for x in agg.values():
        x=dict(x); x['requested']=x['approved']+x['denied']+x['expired']; rows.append(x)
    rows.sort(key=lambda x:(-x['approved'],-x['requested']))
    totals={k:sum(x[k] for x in rows) for k in ('requested','approved','denied','expired')}
    return {'sourceType':'CANARY','rows':rows,'totals':totals,'byItem':{_ar_norm(x['name']):x for x in rows},'itemizedRows':[]}


def ancillary_parse_snt_history(addon_raw, upsell_raw, property_key):
    addon=ancillary_parse_addon(addon_raw, property_key); ups=ancillary_parse_upsell(upsell_raw)
    rows=sorted(addon['main']+ups['byRoomType'], key=lambda x:x['revenue'], reverse=True)
    out=[]; by={}
    for r in rows:
        x=dict(r,approved=r['revenue'],requested=r['revenue'],denied=0,expired=0); out.append(x); by[_ar_norm(x['name'])]=x
    totals={'count':sum(_ar_num(x.get('count')) or 0 for x in out),'revenue':sum(x['revenue'] for x in out)}
    totals.update({'requested':totals['revenue'],'approved':totals['revenue'],'denied':0,'expired':0})
    return {'sourceType':'SNT','rows':out,'totals':totals,'byItem':by,'itemizedRows':[dict(x) for x in addon['itemized']]}


def ancillary_apply_stly_journal(stly, journal_rows):
    remove={_ar_norm(x) for x in ['11am Early Check In','1pm Early Check In','11am Early Check-in','1pm Early Check-in','1 PM Late Check Out','2 PM Late Check Out','1pm Late Checkout','2pm Late Checkout']}
    rows=[dict(r) for r in stly.get('rows',[]) if _ar_norm(r['name']) not in remove]
    for j in journal_rows:
        if j.get('revenue') is None: continue
        rev=_ar_num(j['revenue']) or 0
        rows.append({'name':j['name'],'count':None,'revenue':rev,'average':None,'approved':rev,'requested':rev,'denied':0,'expired':0})
    rows.sort(key=lambda x:_ar_num(x.get('revenue')) or 0, reverse=True)
    by={_ar_norm(x['name']):x for x in rows}
    total_rev=sum(_ar_num(x.get('revenue')) or 0 for x in rows); total_count=sum(_ar_num(x.get('count')) or 0 for x in rows if x.get('count') is not None)
    return {'sourceType':'SNT','rows':rows,'totals':{'count':total_count,'revenue':total_rev,'requested':total_rev,'approved':total_rev,'denied':0,'expired':0},'byItem':by,'itemizedRows':stly.get('itemizedRows',[])}


def ancillary_property_corrections(stly, property_key):
    if property_key=='allegria hotel':
        for r in stly.get('rows',[]):
            if _ar_norm(r.get('name'))==_ar_norm('Rollaway Bed'):
                r['revenue']=r['approved']=r['requested']=800
                r['average']=800/(_ar_num(r.get('count')) or 1) if (_ar_num(r.get('count')) or 0) else None
        stly['rows'].sort(key=lambda x:_ar_num(x.get('revenue')) or 0, reverse=True)
        stly['byItem']={_ar_norm(x['name']):x for x in stly['rows']}
        rev=sum(_ar_num(x.get('revenue')) or 0 for x in stly['rows']); cnt=sum(_ar_num(x.get('count')) or 0 for x in stly['rows'] if x.get('count') is not None)
        stly['totals'].update({'count':cnt,'revenue':rev,'requested':rev,'approved':rev,'denied':0,'expired':0})
    return stly


def ancillary_parse_staff(raw):
    if not raw: return []
    h=_ar_header_map(raw[0]); nc=_ar_col(h,['staff name']); mc=_ar_col(h,['messages'])
    if nc<0 or mc<0: return []
    rows=[{'name':str(r[nc] or '').strip(),'messages':_ar_num(r[mc]) or 0} for r in raw[1:] if nc<len(r) and str(r[nc] or '').strip()]
    return sorted(rows,key=lambda x:x['messages'], reverse=True)


def ancillary_alias_map(property_key):
    out={}
    for cur,prev in ANCILLARY_YOY_ALIASES.get(property_key,[]): out.setdefault(_ar_norm(cur),[]).append(prev)
    return out


def ancillary_variance(current_rows, stly_by_item, property_key):
    aliases=ancillary_alias_map(property_key); is_ash='ashworth' in property_key; out=[]; matched={}; cur_by={_ar_norm(r['name']):r for r in current_rows}
    suppressed=set(); rollups={}
    if is_ash:
        early=_ar_norm('Early Check In'); very=_ar_norm('Very Early Check In'); rollups[early]=[early,very]; suppressed.add(very)
    for r in current_rows:
        n=_ar_norm(r['name'])
        if n in suppressed: continue
        current=sum(_ar_num(cur_by[x].get('revenue')) or 0 for x in rollups.get(n,[n]) if x in cur_by)
        candidates=aliases.get(n,[]) or [r['name']]; prev=0
        for name in candidates:
            sn=_ar_norm(name)
            if sn in matched: continue
            x=stly_by_item.get(sn)
            if x: prev+=_ar_num(x.get('approved')) or 0; matched[sn]=True
        if n not in aliases:
            x=stly_by_item.get(n)
            if x and n not in matched: prev+=_ar_num(x.get('approved')) or 0; matched[n]=True
        out.append({'name':r['name'],'current':current,'stly':prev,'variance':current-prev})
    for n,x in stly_by_item.items():
        if n in matched: continue
        prev=_ar_num(x.get('approved')) or 0
        if prev: out.append({'name':x['name'],'current':0,'stly':prev,'variance':-prev})
    return out


def _ar_is_journal_equivalent(name):
    n=_ar_norm(name)
    vals=['early check in','early check-in','late checkout','late check out','very early check in','very early check-in','very late checkout','very late check out','early check in / late checkout','early check in late checkout']
    return any(n==_ar_norm(x) for x in vals)


def _ar_copy_style_row(src_ws, src_row, dst_ws, dst_row, start_col=1, num_cols=5):
    for c in range(start_col,start_col+num_cols):
        s=src_ws.cell(src_row,c); d=dst_ws.cell(dst_row,c)
        if s.has_style:
            d._style=copy(s._style)
        d.number_format=s.number_format
        d.font=copy(s.font); d.fill=copy(s.fill); d.border=copy(s.border); d.alignment=copy(s.alignment); d.protection=copy(s.protection)
    dst_ws.row_dimensions[dst_row].height=src_ws.row_dimensions[src_row].height


def _ar_copy_style_shifted(src_ws, src_row, src_col, dst_ws, dst_row, dst_col, num_cols):
    """Copy a styled template row to a different column position."""
    for offset in range(num_cols):
        s = src_ws.cell(src_row, src_col + offset)
        d = dst_ws.cell(dst_row, dst_col + offset)
        if s.has_style:
            d._style = copy(s._style)
        d.number_format = s.number_format
        d.font = copy(s.font)
        d.fill = copy(s.fill)
        d.border = copy(s.border)
        d.alignment = copy(s.alignment)
        d.protection = copy(s.protection)
    dst_ws.row_dimensions[dst_row].height = src_ws.row_dimensions[src_row].height


def _ar_merge_name(ws,row,start_col=1,num_cols=5):
    try: ws.merge_cells(start_row=row,start_column=start_col,end_row=row,end_column=start_col+num_cols-1)
    except Exception: pass


def _ar_clear_output_sheet(ws, max_row=260, max_col=17):
    # Clear values and merges while retaining column widths. Styles will be
    # copied from Report Template section prototypes as sections are rendered.
    for rng in list(ws.merged_cells.ranges):
        try: ws.unmerge_cells(str(rng))
        except Exception: pass
    for row in ws.iter_rows(min_row=1,max_row=max_row,min_col=1,max_col=max_col):
        for cell in row:
            cell.value=None
            cell._style=copy(openpyxl.styles.Style()) if False else cell._style
    ws._charts=[]


def ancillary_render_report(
    template_bytes,
    property_name,
    property_key,
    report_month,
    main_rows,
    operational_rows,
    itemized_rows,
    upgrades,
    stly,
    variance_rows,
    staff_rows,
    messaging,
    engagement,
):
    """Render all hotels from one universal Report Template.

    Direct Python translation of the Apps Script renderReport_() structure:
      - A:J is cleared and rebuilt dynamically from style prototype rows.
      - M:P keeps the fixed Messaging Overview template layout.
      - Property differences remain in data/config rules, not formatting.
    """
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule, FormulaRule

    wb = openpyxl.load_workbook(io.BytesIO(template_bytes), data_only=False)
    if "Report Template" not in wb.sheetnames:
        raise ValueError('Template workbook is missing the "Report Template" sheet.')

    template = wb["Report Template"]
    if "Report" in wb.sheetnames:
        del wb["Report"]

    sh = wb.copy_worksheet(template)
    sh.title = "Report"

    # Apps Script equivalent:
    # A1:J250.breakApart().clearContent().clearFormat()
    blank_style = copy(Workbook().active["A1"]._style)

    # Only unmerge dynamic A:J. Do not destroy fixed Messaging Overview merges.
    for merged in list(sh.merged_cells.ranges):
        if merged.min_row <= 250 and merged.min_col <= 10 and merged.max_col >= 1:
            try:
                sh.unmerge_cells(str(merged))
            except Exception:
                pass

    # Clear content + formatting in the dynamic area, preserving dimensions.
    for row in sh.iter_rows(min_row=1, max_row=250, min_col=1, max_col=10):
        for cell in row:
            cell.value = None
            cell._style = copy(blank_style)

    # Clear month-specific right-side values but keep right-side formatting.
    for cell_range in ("M6:O18", "M22:P28", "M31:N38", "M40:O60"):
        for row in sh[cell_range]:
            for cell in row:
                cell.value = None

    month_name = report_month.strftime("%B")
    month_abbr = report_month.strftime("%b").upper()
    year = report_month.year

    short_prop = re.sub(
        r" by the Sea( Hotel)?$",
        "",
        property_name,
        flags=re.I,
    )
    # Keep the established display name, but not a separate template.
    if property_key == "pleasant view inn":
        short_prop = "Westerly"

    def unavailable_fill(cell):
        cell.fill = PatternFill(fill_type="solid", fgColor="CCCCCC")

    # ============================================================
    # LEFT SIDE
    # ============================================================
    left = 1

    _ar_copy_style_row(template, 1, sh, left, 1, 5)
    _ar_merge_name(sh, left, 1, 5)
    sh.cell(left, 1).value = f"{short_prop} Upsell Overview - {month_name}"
    left += 1

    _ar_copy_style_row(template, 2, sh, left, 1, 5)
    _ar_merge_name(sh, left, 1, 5)
    sh.cell(left, 1).value = year
    left += 1

    _ar_copy_style_row(template, 3, sh, left, 1, 5)
    sh.merge_cells(start_row=left, start_column=1, end_row=left, end_column=2)
    sh.cell(left, 1).value = "Name"
    sh.cell(left, 3).value = "Total Count"
    sh.cell(left, 4).value = "Total Revenue"
    sh.cell(left, 5).value = "Average revenue"
    left += 1

    for r in main_rows:
        _ar_copy_style_row(template, 4, sh, left, 1, 5)
        sh.merge_cells(start_row=left, start_column=1, end_row=left, end_column=2)
        sh.cell(left, 1).value = r["name"]

        count_blank = r.get("count") is None or r.get("count") == ""
        avg_blank = r.get("average") is None or r.get("average") == ""

        sh.cell(left, 3).value = "" if count_blank else r.get("count")
        sh.cell(left, 4).value = _ar_num(r.get("revenue")) or 0
        sh.cell(left, 5).value = "" if avg_blank else r.get("average")

        # Approved template rule: populated current-year metric cells are
        # light gray. Dark gray is reserved only for unavailable/blank metrics.
        light_gray = PatternFill(fill_type="solid", fgColor="EFEFEF")
        for metric_col in (3, 4, 5):
            sh.cell(left, metric_col).fill = copy(light_gray)
        if count_blank:
            unavailable_fill(sh.cell(left, 3))
        if avg_blank:
            unavailable_fill(sh.cell(left, 5))
        left += 1

    main_count = sum(_ar_num(r.get("count")) or 0 for r in main_rows)
    main_revenue = sum(_ar_num(r.get("revenue")) or 0 for r in main_rows)
    avg_rows = [
        r for r in main_rows
        if r.get("count") is not None and (_ar_num(r.get("count")) or 0) > 0
    ]
    avg_den = sum(_ar_num(r.get("count")) or 0 for r in avg_rows)
    avg_num = sum(_ar_num(r.get("revenue")) or 0 for r in avg_rows)

    _ar_copy_style_row(template, 24, sh, left, 1, 5)
    sh.merge_cells(start_row=left, start_column=1, end_row=left, end_column=2)
    sh.cell(left, 1).value = "TOTALS"
    sh.cell(left, 3).value = main_count
    sh.cell(left, 4).value = main_revenue
    sh.cell(left, 5).value = avg_num / avg_den if avg_den else 0
    left += 1

    _ar_copy_style_row(template, 25, sh, left, 1, 5)
    _ar_merge_name(sh, left, 1, 5)
    sh.cell(left, 1).value = "STLY"
    left += 1

    if stly["sourceType"] == "SNT":
        _ar_copy_style_row(template, 3, sh, left, 1, 5)
        sh.merge_cells(start_row=left, start_column=1, end_row=left, end_column=2)
        sh.cell(left, 1).value = "Upsell Name"
        sh.cell(left, 3).value = "Total Count"
        sh.cell(left, 4).value = "Total Revenue"
        sh.cell(left, 5).value = "Average Revenue"
        left += 1

        for r in stly["rows"]:
            _ar_copy_style_row(template, 4, sh, left, 1, 5)
            sh.merge_cells(start_row=left, start_column=1, end_row=left, end_column=2)
            sh.cell(left, 1).value = r["name"]

            count_blank = r.get("count") is None or r.get("count") == ""
            avg_blank = r.get("average") is None or r.get("average") == ""

            sh.cell(left, 3).value = "" if count_blank else r.get("count")
            sh.cell(left, 4).value = _ar_num(r.get("revenue")) or 0
            sh.cell(left, 5).value = "" if avg_blank else r.get("average")

            if count_blank:
                unavailable_fill(sh.cell(left, 3))
            if avg_blank:
                unavailable_fill(sh.cell(left, 5))
            left += 1

        _ar_copy_style_row(template, 24, sh, left, 1, 5)
        sh.merge_cells(start_row=left, start_column=1, end_row=left, end_column=2)
        sh.cell(left, 1).value = "TOTALS"
        sh.cell(left, 3).value = _ar_num(stly["totals"].get("count")) or 0
        sh.cell(left, 4).value = _ar_num(stly["totals"].get("revenue")) or 0

        avg_values = [
            _ar_num(r.get("average"))
            for r in stly["rows"]
            if r.get("average") not in (None, "")
        ]
        avg_values = [v for v in avg_values if v is not None]
        sh.cell(left, 5).value = sum(avg_values) / len(avg_values) if avg_values else 0
        left += 1

    else:
        _ar_copy_style_row(template, 26, sh, left, 1, 5)
        for col, value in enumerate(
            [
                "Name",
                "Revenue Requested $",
                "Revenue Approved $",
                "Revenue Denied $",
                "Revenue Expired $",
            ],
            1,
        ):
            sh.cell(left, col).value = value
        left += 1

        for r in stly["rows"]:
            _ar_copy_style_row(template, 27, sh, left, 1, 5)
            for col, value in enumerate(
                [r["name"], r["requested"], r["approved"], r["denied"], r["expired"]],
                1,
            ):
                sh.cell(left, col).value = value
            left += 1

        _ar_copy_style_row(template, 43, sh, left, 1, 5)
        for col, value in enumerate(
            [
                "TOTALS",
                stly["totals"]["requested"],
                stly["totals"]["approved"],
                stly["totals"]["denied"],
                stly["totals"]["expired"],
            ],
            1,
        ):
            sh.cell(left, col).value = value
        left += 1

    # Variance
    _ar_copy_style_row(template, 44, sh, left, 1, 5)
    _ar_merge_name(sh, left, 1, 5)
    sh.cell(left, 1).value = "Variance"
    left += 1

    _ar_copy_style_row(template, 45, sh, left, 1, 5)
    sh.cell(left, 1).value = "Name"
    sh.merge_cells(start_row=left, start_column=2, end_row=left, end_column=5)
    sh.cell(left, 2).value = "Total Revenue"
    left += 1

    variance_start = left
    for r in variance_rows:
        _ar_copy_style_row(template, 46, sh, left, 1, 5)
        sh.cell(left, 1).value = r["name"]
        sh.merge_cells(start_row=left, start_column=2, end_row=left, end_column=5)
        sh.cell(left, 2).value = _ar_num(r.get("variance")) or 0
        left += 1
    variance_end = left - 1

    variance_total = sum(_ar_num(r.get("variance")) or 0 for r in variance_rows)
    _ar_copy_style_row(template, 69, sh, left, 1, 5)
    sh.cell(left, 1).value = "TOTALS"
    sh.merge_cells(start_row=left, start_column=2, end_row=left, end_column=5)
    sh.cell(left, 2).value = variance_total
    sh.cell(left, 2).number_format = '$#,##0.00;$(#,##0.00);$-'

    sh.cell(left, 1).fill = PatternFill(fill_type="solid", fgColor="1C4587")
    sh.cell(left, 1).font = copy(template.cell(69, 1).font)
    sh.cell(left, 1).font = Font(
        name=sh.cell(left, 1).font.name or "Arial",
        size=sh.cell(left, 1).font.sz,
        bold=True,
        color="FFFFFF",
    )
    sh.cell(left, 1).alignment = Alignment(horizontal="center")

    variance_total_fill = "EA9999" if variance_total < 0 else "B6D7A8"
    for col in range(2, 6):
        sh.cell(left, col).fill = PatternFill(fill_type="solid", fgColor=variance_total_fill)
        base_font = sh.cell(left, col).font
        sh.cell(left, col).font = Font(
            name=base_font.name or "Arial",
            size=base_font.sz,
            bold=True,
            color="000000",
        )
    left += 1

    left += 3

    # Current-year itemized section
    _ar_copy_style_row(template, 72, sh, left, 1, 5)
    _ar_merge_name(sh, left, 1, 5)
    sh.cell(left, 1).value = year
    left += 1

    _ar_copy_style_row(template, 73, sh, left, 1, 5)
    sh.merge_cells(start_row=left, start_column=1, end_row=left, end_column=2)
    sh.cell(left, 1).value = "Early Check In & Late Checkout Itemized"
    sh.cell(left, 3).value = "Total Count"
    sh.cell(left, 4).value = "Total Revenue"
    sh.cell(left, 5).value = "Average revenue"
    left += 1

    for r in itemized_rows:
        _ar_copy_style_row(template, 74, sh, left, 1, 5)
        sh.merge_cells(start_row=left, start_column=1, end_row=left, end_column=2)
        sh.cell(left, 1).value = r["name"]
        sh.cell(left, 3).value = "" if r.get("count") is None else r.get("count")
        sh.cell(left, 4).value = _ar_num(r.get("revenue")) or 0
        sh.cell(left, 5).value = "" if r.get("average") is None else (_ar_num(r.get("average")) or 0)
        left += 1

    # STLY itemized section for SNT historical source
    if stly["sourceType"] == "SNT" and stly.get("itemizedRows"):
        left += 2

        _ar_copy_style_row(template, 72, sh, left, 1, 5)
        _ar_merge_name(sh, left, 1, 5)
        sh.cell(left, 1).value = year - 1
        left += 1

        _ar_copy_style_row(template, 73, sh, left, 1, 5)
        sh.merge_cells(start_row=left, start_column=1, end_row=left, end_column=2)
        sh.cell(left, 1).value = "Early Check In & Late Checkout Itemized"
        sh.cell(left, 3).value = "Total Count"
        sh.cell(left, 4).value = "Total Revenue"
        sh.cell(left, 5).value = "Average revenue"
        left += 1

        for r in stly["itemizedRows"]:
            _ar_copy_style_row(template, 74, sh, left, 1, 5)
            sh.merge_cells(start_row=left, start_column=1, end_row=left, end_column=2)
            sh.cell(left, 1).value = r["name"]
            sh.cell(left, 3).value = "" if r.get("count") is None else r.get("count")
            sh.cell(left, 4).value = _ar_num(r.get("revenue")) or 0
            sh.cell(left, 5).value = "" if r.get("average") is None else (_ar_num(r.get("average")) or 0)
            left += 1

    # ============================================================
    # MIDDLE SIDE — universal independent dynamic stack
    # ============================================================
    mid = 3

    _ar_copy_style_row(template, 3, sh, mid, 7, 4)
    for col, value in enumerate(
        ["Name", "Total Count", "Total Revenue", "Average revenue"],
        7,
    ):
        sh.cell(mid, col).value = value
    mid += 1

    for r in operational_rows:
        _ar_copy_style_row(template, 4, sh, mid, 7, 4)
        for col, value in enumerate(
            [
                r["name"],
                "" if r.get("count") is None else r.get("count"),
                _ar_num(r.get("revenue")) or 0,
                "" if r.get("average") is None else (_ar_num(r.get("average")) or 0),
            ],
            7,
        ):
            sh.cell(mid, col).value = value
        mid += 1

    mid += 3

    _ar_copy_style_row(template, 13, sh, mid, 7, 3)
    for col, value in enumerate(
        ["Staff Name", "Room Upgrades Produced", "Revenue Produced"],
        7,
    ):
        sh.cell(mid, col).value = value
    mid += 1

    for r in upgrades["byStaff"]:
        _ar_copy_style_row(template, 14, sh, mid, 7, 3)
        sh.cell(mid, 7).value = r["name"]
        sh.cell(mid, 8).value = r["count"]
        sh.cell(mid, 9).value = r["revenue"]
        mid += 1

    mid += 3

    _ar_copy_style_row(template, 25, sh, mid, 7, 2)
    sh.cell(mid, 7).value = "Room Level Increase"
    sh.cell(mid, 8).value = "Count"
    mid += 1

    for r in upgrades["byLevel"]:
        _ar_copy_style_row(template, 26, sh, mid, 7, 2)
        level = str(r.get("level") or "")
        match = re.match(r"^\+(\d+)$", level)
        sh.cell(mid, 7).value = f"{match.group(1)}+" if match else level
        sh.cell(mid, 8).value = r["count"]
        mid += 1

    mid += 3

    _ar_copy_style_row(template, 30, sh, mid, 7, 2)
    sh.cell(mid, 7).value = "Staff Name"
    sh.cell(mid, 8).value = "Messages"
    mid += 1

    for r in staff_rows:
        _ar_copy_style_row(template, 31, sh, mid, 7, 2)
        sh.cell(mid, 7).value = r["name"]
        sh.cell(mid, 8).value = r["messages"]
        mid += 1

    # ============================================================
    # RIGHT SIDE — fixed universal Messaging Overview M:P
    # ============================================================
    sh["M1"] = f"{short_prop} Messaging Overview - {month_name}"
    sh["M2"] = year
    sh["M21"] = ""
    sh["N21"] = f"{month_abbr} {year}"
    sh["O21"] = "STLY"
    sh["P21"] = "YoY"

    labels = [
        "Total Messages",
        "# of messages guest sent",
        "# of messages hotel sent",
        "% of your guests that sent a message",
        "Response Rate",
        "Average minutes to respond",
        "Median minutes to respond",
    ]
    current = [
        _ar_num(messaging.get("msgTotal")) or 0,
        _ar_num(messaging.get("msgGuest")) or 0,
        _ar_num(messaging.get("msgHotel")) or 0,
        messaging.get("msgGuestPct", 0) or 0,
        messaging.get("responseRate", 0) or 0,
        _ar_num(messaging.get("avgResponse")) or 0,
        _ar_num(messaging.get("medianResponse")) or 0,
    ]
    prior = [
        _ar_num(messaging.get("stlyMsgTotal")) or 0,
        _ar_num(messaging.get("stlyMsgGuest")) or 0,
        _ar_num(messaging.get("stlyMsgHotel")) or 0,
        messaging.get("stlyMsgGuestPct", 0) or 0,
        messaging.get("stlyResponseRate", 0) or 0,
        _ar_num(messaging.get("stlyAvgResponse")) or 0,
        _ar_num(messaging.get("stlyMedianResponse")) or 0,
    ]

    for i, label in enumerate(labels):
        row = 22 + i
        sh.cell(row, 13).value = label
        sh.cell(row, 14).value = current[i]
        sh.cell(row, 15).value = prior[i]
        sh.cell(row, 16).value = current[i] - prior[i]

    for row in range(22, 25):
        for col in range(14, 17):
            sh.cell(row, col).number_format = "0"
    for row in range(25, 27):
        for col in range(14, 17):
            sh.cell(row, col).number_format = "0.0%"
    for row in range(27, 29):
        for col in range(14, 17):
            sh.cell(row, col).number_format = "0.0"

    for i, (d, rate) in enumerate(engagement[:8], start=31):
        sh.cell(i, 13).value = d
        sh.cell(i, 14).value = rate
        sh.cell(i, 13).number_format = "m/d/yyyy"
        sh.cell(i, 14).number_format = "0.0%"

    # ============================================================
    # Conditional formatting
    # ============================================================
    sh.conditional_formatting = openpyxl.formatting.formatting.ConditionalFormattingList()

    green_fill = PatternFill(fill_type="solid", fgColor="B6D7A8")
    red_fill = PatternFill(fill_type="solid", fgColor="EA9999")

    if variance_end >= variance_start:
        # Every variance value is stored in the top-left cell of a B:E merged
        # box. Use formula rules that reference column B so the approved fill
        # applies across the ENTIRE merged box, not only the value cell.
        rng = f"B{variance_start}:E{variance_end}"
        sh.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f"$B{variance_start}>0"],
                fill=green_fill,
            ),
        )
        sh.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f"$B{variance_start}<0"],
                fill=red_fill,
            ),
        )

    sh.conditional_formatting.add(
        "P22:P26",
        CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill),
    )
    sh.conditional_formatting.add(
        "P22:P26",
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill),
    )
    sh.conditional_formatting.add(
        "P27:P28",
        CellIsRule(operator="greaterThan", formula=["0"], fill=red_fill),
    )
    sh.conditional_formatting.add(
        "P27:P28",
        CellIsRule(operator="lessThan", formula=["0"], fill=green_fill),
    )

    # Clone the approved charts from Report Template. copy_worksheet() does not
    # copy chart objects, so explicitly deep-copy and repoint each source range.
    sh._charts = []

    def _ar_chart_title_text(chart):
        try:
            return chart.title.tx.rich.p[0].r[0].t or ""
        except Exception:
            return ""

    def _ar_set_chart_series_range(chart, cat_formula, val_formula):
        if not getattr(chart, "ser", None):
            return
        ser = chart.ser[0]
        if getattr(ser, "cat", None) is not None:
            if getattr(ser.cat, "strRef", None) is not None:
                ser.cat.strRef.f = cat_formula
            elif getattr(ser.cat, "numRef", None) is not None:
                ser.cat.numRef.f = cat_formula
        if getattr(ser, "val", None) is not None and getattr(ser.val, "numRef", None) is not None:
            ser.val.numRef.f = val_formula

    for template_chart in template._charts:
        chart = deepcopy(template_chart)
        title = _ar_chart_title_text(chart).strip().lower()
        if "engagement rate" in title:
            _ar_set_chart_series_range(chart, "'Report'!$M$31:$M$38", "'Report'!$N$31:$N$38")
        elif "common guest" in title:
            _ar_set_chart_series_range(chart, "'Report'!$M$6:$M$13", "'Report'!$N$6:$N$13")
        elif "operational concerns" in title:
            _ar_set_chart_series_range(chart, "'Report'!$M$40:$M$44", "'Report'!$N$40:$N$44")
        sh._charts.append(chart)

    sh.freeze_panes = None

    # Put Report immediately after Report Template.
    try:
        wb._sheets.remove(sh)
        idx = wb._sheets.index(template)
        wb._sheets.insert(idx + 1, sh)
    except Exception:
        pass

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def ancillary_build_monthly_report(template_bytes, property_name, report_month, addon_file, upsell_file, stly_addon_file=None, stly_upsell_file=None, canary_history_file=None, staff_file=None, journal_values=None, stly_journal_values=None, messaging=None, engagement=None):
    profile,key=ancillary_profile(property_name); journal_values=journal_values or []; stly_journal_values=stly_journal_values or []; messaging=messaging or {}; engagement=engagement or []
    addon=ancillary_parse_addon(_ar_file_rows(addon_file),key); upgrades=ancillary_parse_upsell(_ar_file_rows(upsell_file))
    stly_month=datetime.datetime(report_month.year-1, report_month.month, 1)
    if profile.get('stlySource')=='SNT':
        if not stly_addon_file or not stly_upsell_file: raise ValueError('This property uses SNT for STLY and needs prior-year Add On Production + Upsell files.')
        stly=ancillary_parse_snt_history(_ar_file_rows(stly_addon_file),_ar_file_rows(stly_upsell_file),key)
        if profile.get('stlyJournal'):
            rows=[]
            for i,j in enumerate(profile.get('journal',[])[:2]): rows.append({'name':j['report'],'revenue':stly_journal_values[i] if i<len(stly_journal_values) else None})
            stly=ancillary_apply_stly_journal(stly,rows)
    else:
        if not canary_history_file: raise ValueError('This property uses Canary for STLY and needs the historical Canary upsell export.')
        stly=ancillary_parse_canary_history(_ar_file_rows(canary_history_file),stly_month)
    stly=ancillary_property_corrections(stly,key)
    journal_agg={}
    for i,j in enumerate(profile.get('journal',[])):
        v=journal_values[i] if i<len(journal_values) else None
        if v is None: continue
        journal_agg[j['report']]=journal_agg.get(j['report'],0)+(_ar_num(v) or 0)
    journal_rows=[{'name':n,'count':None,'revenue':v,'average':None} for n,v in journal_agg.items()]
    main=[r for r in addon['main'] if not _ar_is_journal_equivalent(r['name'])]
    main=journal_rows+main+upgrades['byRoomType']; main=sorted(main,key=lambda x:(-(_ar_num(x.get('revenue')) or 0), str(x['name']).lower()))
    variance=ancillary_variance(main,stly['byItem'],key); staff=ancillary_parse_staff(_ar_file_rows(staff_file)) if staff_file else []
    output=ancillary_render_report(template_bytes,property_name,key,report_month,main,addon['operational'],addon['itemized'],upgrades,stly,variance,staff,messaging,engagement)
    return output, {'mainRows':main,'stly':stly,'variance':variance,'operational':addon['operational'],'itemized':addon['itemized'],'upgrades':upgrades,'staff':staff}

# ── Ancillary Revenue (Plymouth/Hotel 1620 only, for now) ────────────────────
# Different shape from ROB/SR/Forecast: one workbook with a tab per month
# (not a file per hotel per month), and within a month's tab, up to 5 "weeks"
# sit side by side as column-pairs (WK1=cols B/C ... WK5=cols J/K) rather than
# as separate sheets. The folder/file are hardcoded rather than resolved
# through the general multi-hotel machinery since there's currently exactly
# one of these across all of Drive.
ANCILLARY_REVENUE_FOLDER_NAME = "M - Ancillary Revenue files PLYMOUTH - TEST"
ANCILLARY_UPSELL_HEADER_ROW   = 3   # "Name" / "Total number" / "Total Revenue" ...
ANCILLARY_UPSELL_FIRST_ROW    = 4
ANCILLARY_WEEK_COLS = {1: (2, 3), 2: (4, 5), 3: (6, 7), 4: (8, 9), 5: (10, 11)}  # week -> (count_col, revenue_col)
_YELLOW_FILL_RGB = "FFFFFF00"


def find_ancillary_revenue_file(service):
    """Locate Plymouth/Hotel 1620's Ancillary Revenue workbook.
    Returns (file_id, file_name) or (None, error_str)."""
    q = "mimeType = 'application/vnd.google-apps.folder' and trashed = false"
    result = service.files().list(
        q=q, fields="files(id, name)", pageSize=1000,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    target = ANCILLARY_REVENUE_FOLDER_NAME.strip().lower()
    folder_id = next((f["id"] for f in result.get("files", [])
                       if f["name"].strip().lower() == target), None)
    if not folder_id:
        return None, (f"Drive folder '{ANCILLARY_REVENUE_FOLDER_NAME}' not found. "
                       f"Create it and share it with the service account, then try again.")

    q2 = ("'%s' in parents and trashed = false and "
          "(mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' "
          "or mimeType='application/vnd.ms-excel.sheet.macroenabled.12')") % folder_id
    result2 = service.files().list(
        q=q2, fields="files(id, name)", pageSize=20,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    files = result2.get("files", [])
    if not files:
        return None, f"No workbook found in '{ANCILLARY_REVENUE_FOLDER_NAME}'."
    return files[0]["id"], files[0]["name"]


def parse_addon_production_csv(file_bytes: bytes) -> dict:
    """Parse the wide-format 'Add-on Production' export: one row per
    upsell item, columns are Name, Total, Revenue, Avg, then a
    (count, revenue) pair per day, in date order starting from the 1st of
    whatever month the export covers.

    Returns {normalized_name: [(count, revenue), ...]} — the per-day list
    is in calendar order starting day 1, so summing any 7-item slice gives
    that week's totals regardless of which week is being filled.

    Uses Python's built-in csv module rather than pandas — confirmed real
    case: this export's rows don't all have the same number of trailing
    empty columns (the two summary rows at the top are shorter than the
    per-item rows below), and pd.read_csv errors out ("Expected N fields
    ... saw M") the moment row lengths aren't uniform. The stdlib reader
    has no such expectation; each row is just whatever fields it has.
    """
    text = file_bytes.decode("utf-8-sig")
    reader = list(csv.reader(io.StringIO(text)))
    out = {}
    for row in reader[3:]:   # row 0 = date-range summary, row 1 = blank col header, row 2 = day-header row
        if not row:
            continue
        name = str(row[0] or "").strip()
        if not name:
            continue
        daily = []
        # Columns 4+ (0-indexed) are (count, revenue) pairs, one per day.
        c = 4
        while c + 1 < len(row):
            count_raw = row[c] if c < len(row) else ""
            rev_raw   = row[c + 1] if c + 1 < len(row) else ""
            count = safe_float(count_raw)
            rev   = safe_float(rev_raw)
            if count is None and rev is None:
                break
            daily.append((count or 0, rev or 0))
            c += 2
        out[_normalize_ancillary_name(name)] = daily
    return out


def _normalize_ancillary_name(name: str) -> str:
    """Loosen an item name for matching between the CSV export and the
    sheet's own row labels — confirmed real case: the CSV has trailing
    spaces on several names ('Bottle of House Wine ') that the sheet
    doesn't, and casing isn't always identical either."""
    return re.sub(r"\s+", " ", name).strip().upper()


def find_next_available_week(ws, header_row=ANCILLARY_UPSELL_HEADER_ROW,
                              first_data_row=ANCILLARY_UPSELL_FIRST_ROW):
    """Return the week number (1-5) whose column-pair is marked yellow
    anywhere in the Upsell table's data rows — i.e. whichever week the
    sheet itself is flagging as next to fill in. Scans from the header row
    down to the first row whose col-A label is 'TOTALS'. Returns None if
    no yellow cells are found (nothing left to fill, or already done)."""
    last_row = first_data_row
    for r in range(first_data_row, first_data_row + 30):
        label = str(ws.cell(r, 1).value or "").strip().upper()
        if label == "TOTALS":
            break
        last_row = r
    for week, (count_col, rev_col) in ANCILLARY_WEEK_COLS.items():
        for r in range(first_data_row, last_row + 1):
            for c in (count_col, rev_col):
                fill = ws.cell(r, c).fill
                rgb = fill.fgColor.rgb if fill and fill.fgColor and fill.fgColor.type == "rgb" else None
                if rgb == _YELLOW_FILL_RGB:
                    return week
    return None


def build_ancillary_addon_change_plan(daily_by_name: dict, ws, week: int,
                                       header_row=ANCILLARY_UPSELL_HEADER_ROW,
                                       first_data_row=ANCILLARY_UPSELL_FIRST_ROW):
    """Build changes for one week's (count, revenue) columns in the Upsell
    table, summing the Add-on Production CSV's daily figures for days
    (week-1)*7+1 .. week*7 of the month against each sheet row's own label
    (matched via _normalize_ancillary_name — row order isn't assumed to
    match the CSV's row order).
    """
    count_col, rev_col = ANCILLARY_WEEK_COLS[week]
    day_start = (week - 1) * 7
    day_end   = day_start + 7  # exclusive

    changes = []
    r = first_data_row
    while True:
        label = str(ws.cell(r, 1).value or "").strip()
        if not label or label.upper() == "TOTALS":
            break
        key = _normalize_ancillary_name(label)
        daily = daily_by_name.get(key)
        if daily is not None:
            slice_ = daily[day_start:day_end]
            total_count = sum(d[0] for d in slice_)
            total_rev   = sum(d[1] for d in slice_)
            changes.append({"row": r, "col": count_col, "label": f"{label} — count", "new_value": total_count})
            changes.append({"row": r, "col": rev_col,   "label": f"{label} — revenue", "new_value": total_rev})
        r += 1
    return changes


def apply_ancillary_changes(ws, changes):
    for ch in changes:
        ws.cell(ch["row"], ch["col"]).value = ch["new_value"]


# ── Monthly OOO Report (Sell-Out Efficiency Report .xlsm) ────────────────────
# This workbook has a macro-driven "Save Report" shape button on every one of
# its 170+ dated tabs. Confirmed by direct round-trip test: loading it with
# openpyxl and saving it back — even completely untouched — silently drops
# every one of those shape/button objects (openpyxl's drawing writer doesn't
# preserve non-image/non-chart shapes). So this feature never opens the file
# through openpyxl for writing. It only uses openpyxl (read-only) to read
# values off the existing dated tabs, then adds the new summary tab via
# direct zip/XML surgery — copying every original zip entry byte-for-byte
# and touching only the 3 small manifest parts that must change to register
# a new sheet (workbook.xml, workbook.xml.rels, [Content_Types].xml). This
# was verified to leave xl/vbaProject.bin and every xl/drawings/*.xml
# identical to the source file.
OOO_TAB_RGB = "FF00FF00"  # bright green, per explicit request — distinct from
                          # the app's other "done" green (FF00B050) used elsewhere
OOO_REPORT_FILE_KEYWORD = "SELL-OUT EFFICIENCY REPORT"


def find_ooo_report_file(service):
    """Locate the company-wide Sell-Out Efficiency Report workbook (currently
    shared as 'COPY Sell-Out Efficiency Report.xlsm - TEST'). Searched by
    filename across all of Drive rather than a fixed folder, since there's
    exactly one of these (unlike the per-hotel workbooks). Returns
    (file_id, file_name) or (None, error_str) — errors out on zero or
    multiple matches rather than guessing which one is right."""
    q = ("trashed = false and "
         "(mimeType='application/vnd.ms-excel.sheet.macroenabled.12' "
         "or mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')")
    result = service.files().list(
        q=q, fields="files(id, name)", pageSize=1000, corpora="allDrives",
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    candidates = [f for f in result.get("files", [])
                  if OOO_REPORT_FILE_KEYWORD in f["name"].strip().upper()]
    if not candidates:
        return None, (f"No workbook matching '{OOO_REPORT_FILE_KEYWORD.title()}' found anywhere "
                       f"in Drive — make sure it's been shared with the service account.")
    if len(candidates) > 1:
        names = ", ".join(f["name"] for f in candidates)
        return None, f"Found {len(candidates)} matching workbooks, expected exactly 1: {names}"
    return candidates[0]["id"], candidates[0]["name"]


def _parse_ooo_sheet_date(name):
    """Dated tabs are named like '07-28-2026' (M-D-YYYY, not always zero-
    padded). Anything that doesn't match — e.g. a stray old '10-22-22'
    2-digit-year tab found in a real file — is skipped rather than guessed."""
    m = re.match(r'^(\d{1,2})-(\d{1,2})-(\d{4})$', name.strip())
    if not m:
        return None
    month, day, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return datetime.date(year, month, day)
    except ValueError:
        return None


def list_ooo_available_months(wb):
    """Returns [(year, month, [sheet_names_in_date_order]), ...] sorted with
    the most recent month first. Derived from the tabs actually present in
    the uploaded file, never assumed from today's date."""
    months = {}
    for name in wb.sheetnames:
        d = _parse_ooo_sheet_date(name)
        if d:
            months.setdefault((d.year, d.month), []).append((d, name))
    out = []
    for (y, m), pairs in months.items():
        pairs.sort()
        out.append((y, m, [n for _, n in pairs]))
    out.sort(key=lambda t: (t[0], t[1]), reverse=True)
    return out


def build_ooo_monthly_totals(wb, year, month, sheet_names):
    """Sums column J ('End. OOO Rooms', per the row-7 header) per property
    (column C) across every daily tab in the given month. Row range is read
    per-day (row 8 down until column C is blank) rather than assumed fixed,
    since not every hotel appears on every day. Headers are checked on each
    sheet before trusting the columns — a sheet that doesn't match the
    expected 'Property' / 'End...OOO' layout is skipped, not guessed at."""
    totals = {}
    order = []
    days_included = 0
    skipped = []
    for name in sheet_names:
        ws = wb[name]
        prop_header = str(ws.cell(row=7, column=3).value or "").strip().lower()
        ooo_header = str(ws.cell(row=7, column=10).value or "").strip().lower().replace("\n", " ")
        if "propert" not in prop_header or "end" not in ooo_header or "ooo" not in ooo_header:
            skipped.append(name)
            continue
        days_included += 1
        row = 8
        while True:
            prop = ws.cell(row=row, column=3).value
            if prop is None or str(prop).strip() == "":
                break
            prop = str(prop).strip()
            val = ws.cell(row=row, column=10).value
            try:
                val = float(val) if val is not None else 0.0
            except (TypeError, ValueError):
                val = 0.0
            if prop not in totals:
                totals[prop] = 0.0
                order.append(prop)
            totals[prop] += val
            row += 1
    return order, totals, days_included, skipped


# ── OOO Report: per-property ADR pulled from each hotel's own ROB ────────────
# Each hotel's ROB is a full-year "trailing" sheet: every WK tab (wk one..wk
# six) contains one 8-row block per calendar month (Jan..Dec), laid out as
# Month-header / Revenue / Room Nights / ADR / Group Rms sold / Group Rm Rev
# / Group Rm ADR / PICKUP WoW, starting at row 4 and repeating every 8 rows —
# confirmed against a real file: block_start = 4 + 8*(month-1), Revenue =
# block_start+1, Room Nights = block_start+2, ADR = block_start+3 (verified
# July 2026 block landing at rows 52-59, ADR at row 55, matching a real
# screenshot of Hotel 1620's ROB with July 2026 ADR highlighted at that exact
# position). Year columns (2023/2024/2025/2026...) are consistent across
# every month block in a sheet, but how many trailing years — and therefore
# which column is "this year" — varies per hotel, so it's detected from the
# one header row that reliably holds literal (non-formula) dates: Jan's.
_OOO_HOTEL_MATCH_EXCLUDE = ("TEST",)


def _find_rob_year_column(ws, target_year):
    """Jan's header row (row 4) holds literal dates; every other month's
    header is a formula chain off the previous month and its cached value is
    often stale/blank on files this app has re-saved with openpyxl (which
    never recalculates formulas) — so only Jan's row is trusted here."""
    for c in range(2, 12):
        v = ws.cell(4, c).value
        if isinstance(v, (datetime.datetime, datetime.date)) and v.year == target_year:
            return c
        if isinstance(v, (int, float)) and int(v) == target_year:
            return c
    return None


def read_rob_adr(wb_bytes, sheet_name, target_year, target_month):
    """ADR is computed here as Revenue / Room Nights rather than read from
    the sheet's own ADR formula cell — Revenue and Room Nights are written
    as literal values by this app (always reliable), while the ADR formula
    cell's cached result is frequently blank on files this app has re-saved,
    since openpyxl never recalculates formulas on save. Returns (adr_or_None,
    error_str_or_None)."""
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=True)
    if sheet_name not in wb.sheetnames:
        return None, f"Sheet '{sheet_name}' not found."
    ws = wb[sheet_name]
    year_col = _find_rob_year_column(ws, target_year)
    if year_col is None:
        year_col = 5  # fall back to column E, the typical "current year" column
    block_start = 4 + 8 * (target_month - 1)
    rev = safe_float(ws.cell(block_start + 1, year_col).value)
    rms = safe_float(ws.cell(block_start + 2, year_col).value)
    if rev is None or rms is None or rms == 0:
        month_label = datetime.date(target_year, target_month, 1).strftime("%b %Y")
        return None, f"No {month_label} Revenue/Room Nights on '{sheet_name}' (col {get_column_letter(year_col)})."
    return rev / rms, None


def match_ooo_property_to_hotel(prop_name, hotels):
    """hotels: [(name, id), ...] from get_hotels_from_drive(). Matches by
    substring in either direction after normalizing (uppercase, strip
    punctuation) — the same 'read the actual name, don't assume a fixed
    mapping' approach used for every other keyword match in this app — and
    falls back to the KNOWN_MULTI_FOLDER_HOTELS keyword registry, which
    catches display-name mismatches like 'Plymouth 1620' (OOO report) vs
    'Hotel 1620' (Drive) that share no substring but share a registered
    keyword. Excludes '-TEST' sandbox copies. Returns (hotel_name, hotel_id,
    error_note) — error_note is None on a clean single match, otherwise
    explains why nothing was matched (no match / ambiguous / no hotels
    visible at all) rather than silently guessing."""
    if not hotels:
        return None, None, ("the app can see 0 hotel folders in Drive — nothing is shared "
                            "with this environment's service account (or the Drive listing failed)")

    def norm(s):
        return re.sub(r'[^A-Z0-9 ]', '', s.upper()).strip()
    prop_norm = norm(prop_name)
    candidates = []
    for name, hid in hotels:
        if any(x in name.upper() for x in _OOO_HOTEL_MATCH_EXCLUDE):
            continue
        hotel_norm = norm(name)
        if prop_norm and (prop_norm in hotel_norm or hotel_norm in prop_norm):
            candidates.append((name, hid))
            continue
        # Keyword-registry fallback: the property name and the hotel's Drive
        # display name may each contain a registered keyword without either
        # being a substring of the other.
        for reg_name, kws in KNOWN_MULTI_FOLDER_HOTELS.items():
            if norm(reg_name) == hotel_norm or reg_name.upper() in name.upper():
                if any(kw in prop_norm for kw in kws):
                    candidates.append((name, hid))
                break
    if not candidates:
        return None, None, "no matching hotel found in Drive"
    if len(candidates) > 1:
        return None, None, f"ambiguous — matched {len(candidates)} hotels: {', '.join(c[0] for c in candidates)}"
    return candidates[0][0], candidates[0][1], None


def _find_ooo_adr_source_workbook(service, hotel_id, hotel_name):
    """Finds the most recent week with data in the hotel's ROB file.
    Scans current month first (WK6 -> WK1), then previous month if current is empty.
    Returns (wb_bytes, sheet_name, file_name, error_str)."""
    today = datetime.date.today()
    cur_month_dt  = today.replace(day=1)
    prev_month_dt = (cur_month_dt - datetime.timedelta(days=1)).replace(day=1)
    tried = []
    for month_dt in (cur_month_dt, prev_month_dt):
        result, err = resolve_drive_workbook(service, hotel_id, hotel_name, "ROB", month_dt)
        if not result:
            tried.append(f"{month_dt.strftime('%b %Y')} ROB: {err}")
            continue
        fid, fname = result
        wb_bytes = drive_download(service, fid)
        wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)
        # First pass: look for tab with any color (marked/highlighted)
        for sheet_name in reversed(ROB_SHEETS):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            tc = ws.sheet_properties.tabColor
            # Any color means it's been marked/highlighted
            if tc is not None and getattr(tc, "rgb", None):
                return wb_bytes, sheet_name, fname, None
        # Second pass: if no tab marked done, use the first week with any data
        for sheet_name in reversed(ROB_SHEETS):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            # Check if sheet has any OTB data (column B, rows 5-14)
            if any(ws.cell(r, 2).value for r in range(5, 15)):
                return wb_bytes, sheet_name, fname, None
        tried.append(f"{fname}: no week with data found")
    return None, None, None, "; ".join(tried) or "No ROB file found for the current or previous month."


def build_ooo_adr_lookup(service, order, hotels, year, month):
    """For each property in `order`, resolves it to a Drive hotel and reads
    its most-recently-available ADR for (year, month) off that hotel's ROB.
    Returns {property: (adr_or_None, note_or_None)} — note explains why a
    property has no ADR (unmatched hotel, ambiguous match, no filled ROB
    week found, etc.) so gaps are visible rather than silently blank."""
    out = {}
    for prop in order:
        hotel_name, hotel_id, match_note = match_ooo_property_to_hotel(prop, hotels)
        if not hotel_name:
            out[prop] = (None, match_note)
            continue
        wb_bytes, sheet_name, file_name, src_err = _find_ooo_adr_source_workbook(service, hotel_id, hotel_name)
        if not wb_bytes:
            out[prop] = (None, src_err)
            continue
        adr, adr_err = read_rob_adr(wb_bytes, sheet_name, year, month)
        out[prop] = (adr, adr_err)
    return out


def _ooo_extract_style_ids(zin, wb_xml, rels_xml):
    """Sample the newest dated tab's cell styles so the summary tab visually
    matches the daily report — navy title bar, bordered bold headers, bordered
    property names, blue bordered numbers, blue $-formatted rates. Style
    indices are read from the actual sheet XML at build time (never
    hardcoded): cellXfs order isn't guaranteed stable across future edits to
    the workbook, same reason column positions are re-detected per sheet
    everywhere else in this app. Returns a dict of style ids (values may be
    None if a sample cell wasn't found — those cells just render unstyled)."""
    sample_path = None
    for m in re.finditer(r'<sheet [^>]*name="([^"]+)"[^>]*r:id="(rId\d+)"', wb_xml):
        if _parse_ooo_sheet_date(m.group(1)):
            rel = re.search(r'Id="%s"[^>]*Target="(worksheets/sheet\d+\.xml)"' % m.group(2), rels_xml)
            if rel:
                sample_path = 'xl/' + rel.group(1)
                break
    ids = {"title": None, "title_bar": None, "header_left": None,
           "header": None, "prop": None, "num": None, "money": None}
    if not sample_path:
        return ids
    try:
        sx = zin.read(sample_path).decode('utf-8')
    except KeyError:
        return ids

    def s_of(ref):
        m = re.search(r'<c [^>]*r="%s"[^>]*?s="(\d+)"' % ref, sx)
        if not m:  # s= can precede r= depending on the writer
            m = re.search(r'<c [^>]*s="(\d+)"[^>]*?r="%s"' % ref, sx)
        return m.group(1) if m else None

    ids["title"]       = s_of("C5")   # navy bar, bold white text
    ids["title_bar"]   = s_of("D5")   # navy bar continuation
    ids["header_left"] = s_of("C7")   # bold, bordered, left ("Property")
    ids["header"]      = s_of("D7")   # bold, bordered, centered
    ids["prop"]        = s_of("C8")   # property name, bordered
    ids["num"]         = s_of("F8")   # blue centered #,##0, bordered
    ids["money"]       = s_of("L8")   # blue centered $#,##0.00, bordered
    return ids


def _ooo_build_new_sheet_xml(month_label, order, totals, days_included,
                             adr_by_property=None, style_ids=None):
    adr_by_property = adr_by_property or {}
    sid = style_ids or {}

    def _s(kind):
        v = sid.get(kind)
        return f' s="{v}"' if v else ''

    def cell_str(ref, text, kind=None):
        return (f'<c r="{ref}"{_s(kind)} t="inlineStr">'
                f'<is><t xml:space="preserve">{escape(str(text))}</t></is></c>')

    def cell_num(ref, val, kind=None):
        return f'<c r="{ref}"{_s(kind)}><v>{val}</v></c>'

    def cell_empty(ref, kind):
        return f'<c r="{ref}"{_s(kind)}/>'

    # Row 1: navy title bar across A-D, matching the daily report's banner.
    rows_xml = [('<row r="1">'
                 + cell_str("A1", f"Monthly OOO Rooms Summary - {month_label}", "title")
                 + cell_empty("B1", "title_bar") + cell_empty("C1", "title_bar")
                 + cell_empty("D1", "title_bar") + '</row>')]
    r = 3
    rows_xml.append(f'<row r="{r}">'
                    + cell_str(f"A{r}", "Property", "header_left")
                    + cell_str(f"B{r}", "Total End. OOO Rooms", "header")
                    + cell_str(f"C{r}", "ADR", "header")
                    + cell_str(f"D{r}", "Revenue", "header") + '</row>')
    r += 1
    grand_total = 0.0
    grand_revenue = 0.0
    for prop in order:
        v = totals[prop]
        grand_total += v
        adr, _note = adr_by_property.get(prop, (None, None))
        adr_cell = cell_num(f"C{r}", round(adr, 2), "money") if adr is not None else cell_empty(f"C{r}", "prop")
        if adr is not None:
            revenue = adr * v
            grand_revenue += revenue
            rev_cell = cell_num(f"D{r}", round(revenue, 2), "money")
        else:
            rev_cell = cell_empty(f"D{r}", "prop")
        rows_xml.append(f'<row r="{r}">'
                        + cell_str(f"A{r}", prop, "prop")
                        + cell_num(f"B{r}", v, "num")
                        + adr_cell + rev_cell + '</row>')
        r += 1
    rows_xml.append(f'<row r="{r}">'
                    + cell_str(f"A{r}", "TOTAL", "header_left")
                    + cell_num(f"B{r}", grand_total, "num")
                    + cell_empty(f"C{r}", "prop")
                    + cell_num(f"D{r}", round(grand_revenue, 2), "money") + '</row>')
    r += 1
    rows_xml.append(f'<row r="{r}">{cell_str(f"A{r}", f"({days_included} daily reports included)")}</row>')

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<sheetPr><tabColor rgb="{OOO_TAB_RGB}"/></sheetPr>'
        f'<dimension ref="A1:D{r}"/>'
        '<sheetViews><sheetView workbookViewId="0"/></sheetViews>'
        '<sheetFormatPr defaultRowHeight="15"/>'
        '<cols><col min="1" max="1" width="30" customWidth="1"/><col min="2" max="2" width="22" customWidth="1"/>'
        '<col min="3" max="3" width="14" customWidth="1"/><col min="4" max="4" width="16" customWidth="1"/></cols>'
        '<sheetData>' + "".join(rows_xml) + '</sheetData>'
        '</worksheet>'
    )


def inject_ooo_monthly_sheet(original_bytes, year, month, order, totals, days_included, adr_by_property=None):
    """Adds the bright-green monthly summary tab to the OOO report, placed
    immediately to the right of the 'Report' tab (i.e. 2nd tab on the sheet),
    via raw zip/XML surgery. Never opens the file through openpyxl for
    writing — see the module note above for why. Returns (new_bytes,
    tab_name, error_str). error_str is set (and new_bytes is None) if a tab
    with that name already exists, rather than silently duplicating it."""
    month_label = datetime.date(year, month, 1).strftime("%B %Y")
    tab_name = datetime.date(year, month, 1).strftime("%b %Y").upper()  # e.g. "JUL 2026"
    tab_name = re.sub(r'[:\\/?*\[\]]', '', tab_name)[:31]

    zin = zipfile.ZipFile(io.BytesIO(original_bytes))

    wb_xml = zin.read('xl/workbook.xml').decode('utf-8')
    if f'name="{escape(tab_name)}"' in wb_xml:
        return None, tab_name, (f"A tab named '{tab_name}' already exists in this workbook — "
                                 f"delete it manually first if you want to regenerate it.")

    sheet_nums = [int(m.group(1)) for m in
                  (re.match(r'xl/worksheets/sheet(\d+)\.xml$', n) for n in zin.namelist()) if m]
    new_sheet_num = max(sheet_nums) + 1 if sheet_nums else 1
    new_sheet_path = f"xl/worksheets/sheet{new_sheet_num}.xml"

    existing_sheet_ids = [int(x) for x in re.findall(r'sheetId="(\d+)"', wb_xml)]
    new_sheet_id = max(existing_sheet_ids) + 1 if existing_sheet_ids else 1

    rels_xml = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8')
    existing_rids = [int(x) for x in re.findall(r'Id="rId(\d+)"', rels_xml)]
    new_rid = f"rId{max(existing_rids) + 1 if existing_rids else 1}"

    style_ids = _ooo_extract_style_ids(zin, wb_xml, rels_xml)
    new_sheet_xml = _ooo_build_new_sheet_xml(month_label, order, totals, days_included,
                                             adr_by_property, style_ids)

    new_sheet_el = f'<sheet state="visible" name="{escape(tab_name)}" sheetId="{new_sheet_id}" r:id="{new_rid}"/>'
    # Goes right after the "Report" tab (2nd position), not at the very front —
    # falls back to right after whichever tab is first if no "Report" tab is found,
    # rather than guessing at a name match.
    report_match = re.search(r'<sheet[^>]+name="Report"[^>]*/>', wb_xml)
    if not report_match:
        report_match = re.search(r'<sheet[^>]*/>', wb_xml)
    if report_match:
        insert_pos = report_match.end()
        wb_xml_new = wb_xml[:insert_pos] + new_sheet_el + wb_xml[insert_pos:]
    else:
        wb_xml_new = wb_xml.replace('<sheets>', '<sheets>' + new_sheet_el, 1)

    new_rel_el = (f'<Relationship Id="{new_rid}" '
                  f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
                  f'Target="worksheets/sheet{new_sheet_num}.xml"/>')
    rels_xml_new = rels_xml.replace('</Relationships>', new_rel_el + '</Relationships>', 1)

    ct_xml = zin.read('[Content_Types].xml').decode('utf-8')
    new_override = (f'<Override ContentType="application/vnd.openxmlformats-officedocument'
                    f'.spreadsheetml.worksheet+xml" PartName="/{new_sheet_path}"/>')
    ct_xml_new = ct_xml.replace('</Types>', new_override + '</Types>', 1)

    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == 'xl/workbook.xml':
                data = wb_xml_new.encode('utf-8')
            elif item.filename == 'xl/_rels/workbook.xml.rels':
                data = rels_xml_new.encode('utf-8')
            elif item.filename == '[Content_Types].xml':
                data = ct_xml_new.encode('utf-8')
            zout.writestr(item, data)
        zout.writestr(new_sheet_path, new_sheet_xml)
    return out_buf.getvalue(), tab_name, None


# ── User accounts (self-serve requests + admin approval) ─────────────────────
# Streamlit Cloud's filesystem is ephemeral and st.secrets is read-only at
# runtime, so per-person accounts created through the app can't live in either
# — store them in a small JSON file in Drive instead, using the exact same
# download/upload plumbing already used for every workbook in this app.

APP_DATA_FOLDER_NAME = "Workbook Updater App Data"
USERS_FILE_NAME      = "users.json"


def _find_app_data_folder(service):
    """Find the shared App Data folder — matched case-insensitively against
    APP_DATA_FOLDER_NAME (confirmed real case: exact-match search missed a
    folder created as 'workbook updater app data', different casing than the
    original name). Must be created once and shared with the service
    account, the same way every hotel's Drive folder is — can live anywhere
    inside a Shared Drive, including nested inside a hotel's folder."""
    q = "mimeType = 'application/vnd.google-apps.folder' and trashed = false"
    result = service.files().list(
        q=q, fields="files(id, name)", pageSize=1000,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    target = APP_DATA_FOLDER_NAME.strip().lower()
    for f in result.get("files", []):
        if f["name"].strip().lower() == target:
            return f["id"]
    return None


def _find_or_create_users_file(service):
    """Return (file_id, error). error is a user-facing string when the App
    Data folder itself hasn't been created/shared yet — creates users.json
    inside it (empty) on first use otherwise."""
    folder_id = _find_app_data_folder(service)
    if not folder_id:
        return None, (f"Drive folder '{APP_DATA_FOLDER_NAME}' not found. Create it in "
                       f"Google Drive and share it with the service account (same as a "
                       f"hotel folder), then try again.")
    q = "'%s' in parents and trashed = false and name = '%s'" % (folder_id, USERS_FILE_NAME)
    result = service.files().list(
        q=q, fields="files(id, name)", pageSize=5,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    files = result.get("files", [])
    if files:
        return files[0]["id"], None
    empty = json.dumps({"users": []}).encode("utf-8")
    media = MediaIoBaseUpload(io.BytesIO(empty), mimetype="application/json", resumable=True)
    created = service.files().create(
        body={"name": USERS_FILE_NAME, "parents": [folder_id]},
        media_body=media, fields="id", supportsAllDrives=True,
    ).execute()
    return created["id"], None


def _load_users(service, file_id):
    """Return the list of user dicts from users.json (empty list if unreadable)."""
    try:
        raw = drive_download(service, file_id)
        return json.loads(raw.decode("utf-8")).get("users", [])
    except Exception:
        return []


def _save_users(service, file_id, users):
    payload = json.dumps({"users": users}, indent=2).encode("utf-8")
    drive_upload(service, file_id, payload, USERS_FILE_NAME)


# ── UI ────────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Linchris Weekly Tools", layout="wide")


# ── Login gate ────────────────────────────────────────────────────────────────
def check_login(username: str, password: str):
    """Return (ok, is_admin). Checks the single admin account (Streamlit
    secrets) first, then falls back to approved entries in users.json."""
    admin_user = st.secrets["auth"]["username"]
    admin_hash = st.secrets["auth"]["password_hash"].encode()
    if username == admin_user and bcrypt.checkpw(password.encode(), admin_hash):
        return True, True

    try:
        svc = get_drive_service()
        file_id, err = _find_or_create_users_file(svc)
        if err:
            return False, False
        users = _load_users(svc, file_id)
    except Exception:
        return False, False

    for u in users:
        if (u.get("username") == username and u.get("status") == "approved"
                and bcrypt.checkpw(password.encode(), u.get("password_hash", "").encode())):
            return True, u.get("role") == "admin"
    return False, False


LOGIN_ENABLED = True
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False
    st.session_state["is_admin"] = False
    st.session_state["username"] = None

if LOGIN_ENABLED and not st.session_state["authenticated"]:
    st.title("Linchris Hotel Corporation")
    st.subheader("Please log in to continue")

    login_tab, request_tab = st.tabs(["Log In", "Request Access"])

    with login_tab:
        with st.form("login_form"):
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")
            submitted = st.form_submit_button("Log In")
            if submitted:
                ok, is_admin = check_login(username, password)
                if ok:
                    st.session_state["authenticated"] = True
                    st.session_state["is_admin"] = is_admin
                    st.session_state["username"] = username
                    st.rerun()
                else:
                    st.error("Incorrect username or password — or your account is still pending admin approval.")

    with request_tab:
        st.caption("Create your own login. An admin needs to approve it before you can log in.")
        with st.form("request_access_form"):
            req_username     = st.text_input("Choose a username", key="req_username")
            req_display_name = st.text_input("Your name", key="req_display_name")
            req_password     = st.text_input("Choose a password", type="password", key="req_password")
            req_confirm      = st.text_input("Confirm password", type="password", key="req_confirm")
            req_submitted    = st.form_submit_button("Request Access")
            if req_submitted:
                admin_user = st.secrets["auth"]["username"]
                uname = req_username.strip()
                if not uname or not req_password:
                    st.error("Username and password are required.")
                elif req_password != req_confirm:
                    st.error("Passwords don't match.")
                elif uname.lower() == admin_user.lower():
                    st.error("That username is taken.")
                else:
                    try:
                        svc = get_drive_service()
                        file_id, err = _find_or_create_users_file(svc)
                        if err:
                            st.error(err)
                        else:
                            existing_users = _load_users(svc, file_id)
                            if any(u.get("username", "").lower() == uname.lower() for u in existing_users):
                                st.error("That username is already taken or pending approval.")
                            else:
                                existing_users.append({
                                    "username":      uname,
                                    "display_name":  req_display_name.strip(),
                                    "password_hash": bcrypt.hashpw(req_password.encode(), bcrypt.gensalt()).decode(),
                                    "status":        "pending",
                                    "role":          "editor",
                                    "requested_at":  datetime.datetime.now().isoformat(),
                                    "decided_at":    None,
                                })
                                _save_users(svc, file_id, existing_users)
                                st.success("Request submitted — an admin needs to approve it before you can log in.")
                    except Exception as e:
                        st.error(f"Could not submit request: {e}")

    st.stop()

st.markdown("""
<style>
  .block-container { max-width: 100% !important; padding-left: 2rem !important; padding-right: 2rem !important; }

  .stTabs [data-baseweb="tab-list"] { gap: 8px; border-bottom: 1px solid #E5E7EB; }
  .stTabs [data-baseweb="tab"] {
    background: #F1F3F5; color: #1E293B;
    border-radius: 6px 6px 0 0; padding: 8px 20px; font-weight: 600;
  }
  .stTabs [aria-selected="true"] {
    background: #2563EB !important; color: #FFFFFF !important;
    box-shadow: inset 0 -3px 0 #C9A84C;
  }
  div[data-testid="metric-container"] {
    background: #F8F9FA; border: 1px solid #E5E7EB; border-left: 3px solid #C9A84C;
    border-radius: 8px; padding: 12px;
  }
  button[data-testid="stBaseButton-pills"],
  button[data-testid="stBaseButton-pillsActive"] {
    border-radius: 8px !important;
    border: 1.5px solid #94A3B8 !important;
    font-size: 1.5rem !important;
    padding: 1rem 1.8rem !important;
  }
  button[data-testid="stBaseButton-pillsActive"] {
    background-color: #2563EB !important;
    border-color: #2563EB !important;
    color: #FFFFFF !important;
  }

  /* Larger, consistent widget labels + selectbox text app-wide */
  label[data-testid="stWidgetLabel"] p { font-size: 1.2rem !important; }
  div[data-testid="stSelectbox"] div[data-baseweb="select"] div { font-size: 1.25rem !important; }
  [data-testid="stFileUploaderDropzone"] { font-size: 1.15rem !important; }
  [data-testid="stFileUploaderDropzoneInstructions"] span { font-size: 1.15rem !important; }
  ul[data-testid="stSelectboxVirtualDropdown"] li {
    font-size: 1.25rem !important;
    -webkit-font-smoothing: antialiased;
    text-rendering: optimizeLegibility;
  }
</style>
""", unsafe_allow_html=True)

if "view" not in st.session_state:
    st.session_state["view"] = "main"


def render_admin_settings(svc, users_file_id, users_err):
    st.title("Admin Settings")
    if st.button("← Back to app"):
        st.session_state["view"] = "main"
        st.rerun()

    if users_err:
        st.warning(f"Account requests unavailable: {users_err}")
        return

    all_users  = _load_users(svc, users_file_id)
    admin_user = st.secrets["auth"]["username"]

    st.subheader("Pending Requests")
    pending = [u for u in all_users if u.get("status") == "pending"]
    if not pending:
        st.caption("No pending requests.")
    for u in pending:
        c1, c2, c3, c4 = st.columns([3, 3, 1, 1])
        c1.write(f"**{u.get('username')}**")
        c2.write(u.get("display_name") or "—")
        if c3.button("Approve", key=f"approve_{u.get('username')}"):
            for uu in all_users:
                if uu.get("username") == u.get("username"):
                    uu["status"] = "approved"
                    uu["decided_at"] = datetime.datetime.now().isoformat()
            _save_users(svc, users_file_id, all_users)
            st.rerun()
        if c4.button("Reject", key=f"reject_{u.get('username')}"):
            all_users = [uu for uu in all_users if uu.get("username") != u.get("username")]
            _save_users(svc, users_file_id, all_users)
            st.rerun()

    st.divider()

    st.subheader("All Users")
    rows = [{"Username": admin_user, "Name": "—", "Role": "Admin", "Status": "approved"}]
    for u in all_users:
        rows.append({
            "Username": u.get("username"),
            "Name":     u.get("display_name") or "—",
            "Role":     "Admin" if u.get("role") == "admin" else "Editor",
            "Status":   u.get("status"),
        })
    st.dataframe(rows, use_container_width=True)

    approved = [u for u in all_users if u.get("status") == "approved"]
    if approved:
        st.caption("Change role / revoke access:")
        for u in approved:
            rc1, rc2, rc3 = st.columns([4, 2, 1])
            rc1.write(f"{u.get('username')} ({u.get('display_name') or '—'})")
            current_role = "Admin" if u.get("role") == "admin" else "Editor"
            new_role = rc2.selectbox("Role", ["Editor", "Admin"],
                                      index=["Editor", "Admin"].index(current_role),
                                      key=f"role_{u.get('username')}", label_visibility="collapsed")
            if new_role.lower() != u.get("role", "editor"):
                for uu in all_users:
                    if uu.get("username") == u.get("username"):
                        uu["role"] = new_role.lower()
                _save_users(svc, users_file_id, all_users)
                st.rerun()
            if rc3.button("Revoke", key=f"revoke_{u.get('username')}"):
                all_users = [uu for uu in all_users if uu.get("username") != u.get("username")]
                _save_users(svc, users_file_id, all_users)
                st.rerun()

    st.divider()

    st.subheader("Add New User")
    st.caption("Creates and approves an account directly — skips the request/approval step.")
    with st.form("admin_add_user_form"):
        new_username = st.text_input("Username", key="admin_new_username")
        new_display  = st.text_input("Name", key="admin_new_display")
        new_password = st.text_input("Password", type="password", key="admin_new_password")
        new_role     = st.selectbox("Role", ["Editor", "Admin"], key="admin_new_role",
                                     help="Editor: normal app access, no Admin Settings. Admin: full access including this page.")
        submitted = st.form_submit_button("Add User")
        if submitted:
            if not new_username or not new_password:
                st.error("Username and password are required.")
            elif new_username == admin_user or any(u.get("username") == new_username for u in all_users):
                st.error("That username is already taken.")
            else:
                all_users.append({
                    "username":      new_username,
                    "password_hash": bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode(),
                    "display_name":  new_display,
                    "role":          new_role.lower(),
                    "status":        "approved",
                    "requested_at":  datetime.datetime.now().isoformat(),
                    "decided_at":    datetime.datetime.now().isoformat(),
                })
                _save_users(svc, users_file_id, all_users)
                st.success(f"User '{new_username}' added and approved as {new_role}.")
                st.rerun()


title_col, toggle_col, profile_col = st.columns([5, 1, 1])
with title_col:
    st.title("Linchris Hotel Corporation — Weekly Update Tools")
with toggle_col:
    st.write("")
    test_mode = st.toggle("Test Mode", value=False, key="test_mode")
with profile_col:
    st.write("")
    with st.popover(f"👤 {st.session_state.get('username') or 'Account'}"):
        st.write(f"**{st.session_state.get('username')}**")
        st.caption("Admin" if st.session_state.get("is_admin") else "Editor")
        if st.session_state.get("is_admin"):
            if st.button("⚙️ Admin Settings", key="open_admin_settings"):
                st.session_state["view"] = "admin_settings"
                st.rerun()
        if st.button("Log Out", key="logout_btn"):
            st.session_state["authenticated"] = False
            st.session_state["is_admin"] = False
            st.session_state["username"] = None
            st.session_state["view"] = "main"
            st.rerun()

if st.session_state.get("view") == "admin_settings" and st.session_state.get("is_admin"):
    _admin_svc = None
    try:
        _admin_svc = get_drive_service()
        _users_file_id, _users_err = _find_or_create_users_file(_admin_svc)
    except Exception as e:
        _users_file_id, _users_err = None, str(e)
    render_admin_settings(_admin_svc, _users_file_id, _users_err)
    st.stop()

def render_portfolio_rob_month_setup(selected_hotels, key_prefix):
    """Shared ROB new-month setup for Hilton and IHG.

    Uses the exact same setup_new_rob_month() engine as Stay In Touch:
      - finds/copies the ROB master
      - carries prior-month / last-year values
      - preserves each week tab's own as-of date for completed months
      - rebuilds Pickup WoW formulas across every month/year
      - leaves an undo snapshot

    selected_hotels is [(display_name, drive_folder_id), ...].
    """
    if not selected_hotels:
        return

    setup_toggle = st.checkbox(
        "Set up new month",
        key=f"{key_prefix}_new_month",
        help="Uses the same ROB month-setup logic as the Stay In Touch portfolio."
    )
    if not setup_toggle:
        return

    with st.container(border=True):
        today = datetime.date.today()
        cur_month = today.replace(day=1)
        prev_month = (cur_month - datetime.timedelta(days=1)).replace(day=1)
        next_month = (cur_month + datetime.timedelta(days=32)).replace(day=1)

        options = {
            prev_month.strftime("%B %Y"): prev_month,
            cur_month.strftime("%B %Y"): cur_month,
            next_month.strftime("%B %Y"): next_month,
        }
        labels = list(options.keys())
        default_dt = next_month if today.day >= 22 else cur_month

        sel = st.selectbox(
            "Month to set up",
            labels,
            index=labels.index(default_dt.strftime("%B %Y")),
            key=f"{key_prefix}_setup_month",
        )
        target_month = options[sel]

        names = ", ".join(n for n, _ in selected_hotels)
        st.caption(f"ROB setup will run for: **{names}**")

        if st.button(
            "Set Up New ROB",
            key=f"{key_prefix}_setup_rob_btn",
            type="primary",
            use_container_width=True,
        ):
            svc = get_drive_service()
            undo_items = []
            successes = 0

            for hotel_name, hotel_id in selected_hotels:
                if not hotel_id:
                    st.error(f"{hotel_name}: no Drive folder found.")
                    continue
                try:
                    with st.spinner(f"Setting up {hotel_name} ROB..."):
                        name, err, file_id, original = setup_new_rob_month(
                            svc, hotel_id, hotel_name, target_month
                        )

                    if err and not name:
                        st.error(f"{hotel_name}: {err}")
                        continue

                    if err:
                        st.warning(f"{hotel_name}: {err}")

                    if file_id and original is not None:
                        undo_items.append({
                            "file_id": file_id,
                            "file_name": name,
                            "bytes": original,
                        })

                    st.success(
                        f"{hotel_name}: **{name}** ready for "
                        f"{target_month:%B %Y}."
                    )
                    successes += 1
                except Exception as e:
                    st.error(f"{hotel_name}: ROB setup error — {e}")

            if undo_items:
                st.session_state[f"{key_prefix}_setup_rob_undo"] = undo_items

            if successes:
                st.info(
                    "The same week-date carryover and Pickup WoW formula setup "
                    "used for Stay In Touch was applied."
                )

        undo_key = f"{key_prefix}_setup_rob_undo"
        if undo_key in st.session_state:
            if st.button(
                "↩ Reset ROB setup to original",
                key=f"{key_prefix}_setup_rob_reset",
                use_container_width=True,
            ):
                svc = get_drive_service()
                errors = []
                for item in st.session_state[undo_key]:
                    try:
                        drive_upload(
                            svc,
                            item["file_id"],
                            item["bytes"],
                            item["file_name"],
                        )
                    except Exception as e:
                        errors.append(f"{item['file_name']}: {e}")

                if errors:
                    for e in errors:
                        st.error(e)
                else:
                    del st.session_state[undo_key]
                    st.success("ROB workbook(s) restored to their original state.")


def render_hilton_update(hotels):
    """Hilton portfolio run.

    Differs from the other portfolios in two ways. Several properties are run
    at once, because a single SRP Activity export covers all of them and gets
    split by 'Property - InnCode'. And group figures come from a per-hotel
    Group Wash export rather than from the booking data, so each selected
    hotel needs its own wash file.

    Hilton properties do not run a Strategy Report.
    """
    if not hotels:
        st.info("No Hilton properties found in Drive.")
        return

    st.caption(
        "One SRP Activity export covers every property — it must be run with all "
        "the properties you tick below, or the ones it leaves out can't be "
        "updated. Each hotel also needs its own Group Wash export: Group and "
        "Permanent figures come from the Wash report, and the booking export "
        "supplies only the transient remainder. Both workbooks are built from "
        "the same stays — the ROB by month, the Forecast day by day."
    )

    with st.container(border=True):
        st.markdown("**Properties to run**")
        cols = st.columns(3)
        selected = []
        for i, (name, fid) in enumerate(hotels):
            with cols[i % 3]:
                if st.checkbox(name, key=f"hil_sel_{fid}"):
                    selected.append((name, fid))

        wb_sels = st.pills(
            "Workbooks to update",
            PORTFOLIO_WORKBOOKS["Hilton"],
            selection_mode="multi",
            default=PORTFOLIO_WORKBOOKS["Hilton"],
            key="hil_wb",
        ) or []

        hil_next_month = st.checkbox(
            "Include next month's Forecast",
            key="hil_fcst_next",
            disabled="Forecast" not in wb_sels,
            help="Fills next month's own Forecast workbook for every selected "
                 "property, from the same SRP export — it already runs a year "
                 "ahead. Everything that far out is still on the books, so it "
                 "all lands on the OTB rows and none on the actuals.",
        )

        if "ROB" in wb_sels and selected:
            render_portfolio_rob_month_setup(selected, "hil")

        srp_file = st.file_uploader(
            "SRP Activity — all Hilton properties (one file)",
            type=["xlsx"], key="hil_srp")

        wash_files = {}
        if selected:
            st.markdown("**Group Wash report — one per property**")
            wcols = st.columns(2)
            for i, (name, fid) in enumerate(selected):
                with wcols[i % 2]:
                    wash_files[name] = st.file_uploader(
                        name, type=["xlsx"], key=f"hil_wash_{fid}")

    if not selected:
        st.info("Select at least one property.")
        return
    if not srp_file:
        st.info("Upload the SRP Activity export to continue.")
        return

    try:
        srp = parse_srp_activity(srp_file)
    except Exception as e:
        st.error(f"Could not read the SRP Activity export: {e}")
        return

    st.success("SRP Activity covers {}: {}".format(
        f"{len(srp)} propert" + ("y" if len(srp) == 1 else "ies"),
        ", ".join(f"{c} ({srp[c]['name']})" if srp[c].get("name") else c
                  for c in sorted(srp))))

    # The export's own FILTERS block decides what the numbers can possibly mean.
    # Checked here so a stale filter is caught at upload rather than turning up
    # as odd figures in a workbook days later.
    try:
        srp_filters = parse_srp_filters(srp_file)
    except Exception:
        srp_filters = None
    for msg in srp_filter_warnings(srp_filters):
        st.warning(msg)
    hilton_as_of = (
        srp_filters.get("run_date")
        if srp_filters and srp_filters.get("run_date")
        else datetime.date.today()
    )

    if srp_filters and srp_filters.get("lines"):
        with st.expander("Filters this export was run with"):
            for line in srp_filters["lines"]:
                st.markdown(f"- {line}")
            st.caption("Booked Date should match the report run date. For the "
                       "current month, the ROB combines completed Forecast actuals "
                       "with the live SRP portion instead of expecting SRP to contain "
                       "the full month's completed stays.")

    # Resolved up front, not inside the run. A hotel the export doesn't cover
    # can't be updated, and finding that out only after pressing the button is
    # how a three-hotel run quietly came back having done one.
    codes = {name: _match_inncode(name, srp) for name, _ in selected}
    unmatched = [n for n, c in codes.items() if not c]
    if unmatched:
        st.warning(
            f"**Not in this SRP export, so {'it' if len(unmatched) == 1 else 'they'} "
            f"cannot be run: {', '.join(unmatched)}.** Re-run the SRP Activity "
            f"report with every selected property ticked, or deselect "
            f"{'it' if len(unmatched) == 1 else 'them'} here.")
        if len(unmatched) == len(selected):
            return

    missing_wash = [n for n, _ in selected if not wash_files.get(n)]
    if missing_wash:
        st.info(f"Waiting on Group Wash report for: {', '.join(missing_wash)}")
        return

    if st.button("Preview changes", key="hil_preview", type="primary"):
        svc = get_drive_service()
        next_month = (hilton_as_of.replace(day=1)
                      + datetime.timedelta(days=32)).replace(day=1)
        jobs, problems = [], []
        for name, fid in selected:
            if not fid:
                problems.append(
                    f"{name}: no Drive folder the app can see. Share its folder "
                    f"(or its REVENUE REPORTS subfolder) with "
                    f"{service_account_email() or 'the service account'} as Editor.")
                continue
            try:
                wash = parse_group_wash(wash_files[name])
            except Exception as e:
                problems.append(f"{name}: could not read the Group Wash report — {e}")
                continue

            inn = codes.get(name)
            if not inn:
                continue          # already reported above, before the button

            prop = srp[inn]
            for wb_type in wb_sels:
                if wb_type == "Forecast":
                    job = _hilton_forecast_job(
                        svc, fid, name, inn, prop, wash, problems,
                        as_of=hilton_as_of
                    )
                    if job:
                        jobs.append(job)
                    continue

                result, err = resolve_drive_workbook(svc, fid, name, wb_type)
                if err or not result:
                    problems.append(f"{name} — {wb_type}: {err}")
                    continue
                file_id, file_name = result
                raw = drive_download(svc, file_id)
                wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=False)
                avail = [s for s in ROB_SHEETS if s in wb.sheetnames]
                if not avail:
                    problems.append(f"{name} — ROB ({file_name}): none of the "
                                    f"week tabs were found.")
                    continue
                sheet = first_uncolored_sheet(wb, avail)
                current_month_total = None

                # Current-month Hilton ROB needs completed daily actuals from
                # the current Forecast workbook plus the live SRP tail.
                fcst_result, fcst_err = resolve_drive_workbook(
                    svc, fid, name, "Forecast", month_date=hilton_as_of.replace(day=1)
                )
                if fcst_result and not fcst_err:
                    fcst_id, fcst_name = fcst_result
                    try:
                        fcst_raw = drive_download(svc, fcst_id)
                        actuals = extract_hilton_mtd_actuals_from_forecast(
                            fcst_raw, hilton_as_of
                        )
                        current_month_total = hilton_current_month_total(
                            prop["days"], actuals, hilton_as_of
                        )
                        if current_month_total is None:
                            problems.append(
                                f"{name} — ROB: could not find completed Forecast "
                                f"actuals through {hilton_as_of - datetime.timedelta(days=2):%b %d}; "
                                f"current-month Revenue / Room Nights were left to "
                                f"the plain SRP total."
                            )
                    except Exception as e:
                        problems.append(
                            f"{name} — ROB: could not read current Forecast actuals "
                            f"for the current-month total — {e}"
                        )
                else:
                    problems.append(
                        f"{name} — ROB: current Forecast workbook was not found, so "
                        f"the current-month Revenue / Room Nights could not be "
                        f"reconciled with completed actuals."
                    )

                changes, rob_warns = build_hilton_rob_plan(
                    prop["months"],
                    wash["months"],
                    wb[sheet],
                    as_of=hilton_as_of,
                    current_month_total=current_month_total,
                )
                for w in rob_warns:
                    problems.append(f"{name} — ROB ({file_name}): {w}")
                note = f"  ·  InnCode {inn}"
                if current_month_total is not None:
                    note += (
                        f"  ·  current month = Forecast actuals through "
                        f"{current_month_total['actual_through']:%b %d} + SRP "
                        f"{current_month_total['srp_from']:%b %d}–month end"
                    )
                passed = [f"{n} ({w})" for n, w in rob_week_status(wb, avail)
                          if w and n != sheet]
                if passed:
                    note += "  ·  skipped " + "; ".join(passed)
                jobs.append({
                    "key": f"{name} — {wb_type}",
                    "file_id": file_id, "file_name": file_name,
                    "wb_bytes": raw, "sheet": sheet, "changes": changes,
                    "note": note,
                })

            if hil_next_month and "Forecast" in wb_sels:
                job = _hilton_forecast_job(
                    svc, fid, name, inn, prop, wash, problems,
                    month_date=next_month, as_of=hilton_as_of
                )
                if job:
                    jobs.append(job)
        st.session_state["hil_jobs"] = jobs
        st.session_state["hil_problems"] = problems

    jobs = st.session_state.get("hil_jobs") or []
    for msg in st.session_state.get("hil_problems") or []:
        st.warning(msg)

    _render_undo_bar("hil_undo")

    if not jobs:
        return

    for job in jobs:
        st.divider()
        st.subheader(job["key"] + f" — {job['sheet']}")
        st.caption(job["file_name"] + job.get("note", ""))
        _show_ihg_plan(job["changes"])

    st.divider()
    total = sum(len([c for c in j["changes"] if not c["skip_reason"]]) for j in jobs)
    st.write(f"**{total} cells across {len(jobs)} workbook(s).** "
             f"Applying writes to Drive and marks each tab done (green).")
    if st.button("Confirm and apply to Drive", key="hil_apply", type="primary"):
        saved, errors = apply_portfolio_plans(get_drive_service(), jobs, "hil_undo")
        for s in saved:
            st.success(f"Saved {s}")
        for e in errors:
            st.error(e)
        if saved:
            st.session_state.pop("hil_jobs", None)
            st.rerun()


def _render_undo_bar(undo_key: str):
    """Offer to roll back the last apply, for as long as there is one to roll
    back. Drawn outside the preview: applying clears the plan, so an undo
    button living inside it would disappear on exactly the rerun where someone
    wants it."""
    snapshot = st.session_state.get(undo_key)
    if not snapshot:
        return
    cells = sum(len(info.get("cells", {})) for info in snapshot.values())
    names = ", ".join(info["file_name"] for info in snapshot.values())
    with st.container(border=True):
        col_msg, col_btn = st.columns([4, 1])
        with col_msg:
            st.markdown(
                f"**Last apply can still be undone** — {cells} cell(s) across "
                f"{len(snapshot)} workbook(s).")
            st.caption(f"{names}. Restores the previous values and the tab's "
                       f"original colour.")
        with col_btn:
            st.write("")
            if st.button("Undo", key=f"{undo_key}_btn", use_container_width=True):
                saved, errors = undo_portfolio_plans(get_drive_service(), undo_key)
                for s in saved:
                    st.success(f"Reverted {s}")
                for e in errors:
                    st.error(e)
                if not errors:
                    st.rerun()


def render_ihg_strategy_month_setup(hotel_name, hotel_id):
    """Set up next month's IHG Strategy Report using the shared SR engine.

    Structural setup matches the Stay In Touch Strategy workflow:
      - locate/copy the correct-year Strategy master
      - create/use the next month's workbook
      - clear week-tab completion colors on a fresh copy
      - rebuild the Strategy date ranges for all week tabs
      - preload previous-month / LY reference fields where the shared
        Strategy headers exist
      - keep an undo snapshot

    Current IHG operational values are still filled later from the IHG PDFs.
    """
    st.markdown("**Set Up Next Month — Strategy Report**")

    today = datetime.date.today()
    cur_month = today.replace(day=1)
    next_month = (cur_month + datetime.timedelta(days=32)).replace(day=1)

    target_month = st.selectbox(
        "Strategy month to set up",
        [cur_month, next_month],
        index=1,
        format_func=lambda d: d.strftime("%B %Y"),
        key=f"ihg_sr_setup_month_{hotel_name}",
    )

    if st.button(
        "Set Up New Strategy Report",
        key=f"ihg_setup_sr_{hotel_name}",
        type="primary",
        use_container_width=True,
    ):
        if not hotel_id:
            st.error(f"{hotel_name}: no Drive folder found.")
            return

        try:
            svc = get_drive_service()
            month_kw = target_month.strftime("%b%Y").upper()

            # Step 1 — locate or create the target workbook.
            is_fresh_copy = False
            with st.spinner("Step 1 / 3 — locating or creating Strategy workbook..."):
                existing, find_err = resolve_drive_workbook(
                    svc,
                    hotel_id,
                    hotel_name,
                    "Strategy Report",
                    month_date=target_month,
                )

                if existing:
                    st.info(f"Found existing file: **{existing[1]}** — skipping copy.")
                else:
                    is_fresh_copy = True
                    created_name, create_err = setup_new_sr_month(
                        svc,
                        hotel_id,
                        hotel_name,
                        target_month,
                    )
                    if create_err:
                        master_id, master_name = find_sr_master(
                            svc,
                            hotel_id,
                            target_month.year,
                        )
                        hotel_suffix = ""
                        if master_name and "STRATEGY" in master_name.upper():
                            hotel_suffix = (
                                master_name[
                                    master_name.upper().find("STRATEGY")
                                    + len("STRATEGY"):
                                ]
                                .strip()
                                .replace(".xlsx", "")
                                .replace(".XLSX", "")
                                .strip()
                            )
                        suggested_name = (
                            f"{month_kw} STRATEGY {hotel_suffix}.xlsx".strip()
                        )

                        if "storageQuotaExceeded" in str(create_err):
                            st.warning(
                                f"Auto-copy requires a Shared Drive. In Google Drive:\n\n"
                                f"1. Right-click **{master_name or 'the Strategy master'}** → *Make a copy*\n"
                                f"2. Rename it to **`{suggested_name}`**\n"
                                f"3. Move it into the **{month_kw}** folder\n\n"
                                f"Then run **Set Up New Strategy Report** again."
                            )
                        else:
                            st.error(f"Could not create workbook: {create_err}")
                        return

            # Step 2 — load references.
            with st.spinner("Step 2 / 3 — loading Strategy references..."):
                prev_month = (
                    target_month - datetime.timedelta(days=1)
                ).replace(day=1)
                ly_month = target_month.replace(year=target_month.year - 1)

                prev_wb = _load_wb_from_drive(
                    svc,
                    hotel_id,
                    hotel_name,
                    "Strategy Report",
                    prev_month,
                    data_only=False,
                )
                ly_wb = _load_wb_from_drive(
                    svc,
                    hotel_id,
                    hotel_name,
                    "Strategy Report",
                    ly_month,
                )

            st.info(
                "Strategy references · "
                f"Previous month ({prev_month:%b %Y}): "
                f"{'found' if prev_wb else 'not found'} · "
                f"Last year ({ly_month:%b %Y}): "
                f"{'found' if ly_wb else 'not found'}"
            )

            # Step 3 — populate all week tabs.
            with st.spinner("Step 3 / 3 — preparing all Strategy week tabs..."):
                result, err = resolve_drive_workbook(
                    svc,
                    hotel_id,
                    hotel_name,
                    "Strategy Report",
                    month_date=target_month,
                )
                if err or not result:
                    st.error(f"Cannot open Strategy workbook: {err}")
                    return

                file_id, file_name = result
                wb_bytes = drive_download(svc, file_id)
                original_bytes = wb_bytes

                wb = openpyxl.load_workbook(
                    io.BytesIO(wb_bytes),
                    data_only=False,
                )

                if is_fresh_copy:
                    clear_tab_colors(wb, STRATEGY_SHEETS)

                restructure_sr_dates(wb, target_month)

                first_ws = (
                    wb[STRATEGY_SHEETS[0]]
                    if STRATEGY_SHEETS[0] in wb.sheetnames
                    else None
                )
                num_rows = _count_sheet_data_rows(first_ws) if first_ws else 365
                scope_start = target_month
                scope_end = target_month + datetime.timedelta(
                    days=max(0, num_rows - 1)
                )

                total_written = 0
                for sheet_name in STRATEGY_SHEETS:
                    if sheet_name not in wb.sheetnames:
                        continue

                    # No current IHG PDF data is written during setup.
                    # This call only carries forward prior-month / LY reference
                    # data that can be mapped by the shared Strategy headers.
                    changes = build_strategy_change_plan(
                        None,
                        wb,
                        sheet_name,
                        prev_month_wb=prev_wb,
                        ly_wb=ly_wb,
                        scope_start=scope_start,
                        scope_end=scope_end,
                    )
                    apply_strategy_changes(wb, sheet_name, changes)
                    total_written += len(
                        [c for c in changes if not c.get("skip_reason")]
                    )

                strip_tables(wb)
                out = io.BytesIO()
                wb.save(out)
                drive_upload(
                    svc,
                    file_id,
                    out.getvalue(),
                    file_name,
                )

                st.session_state["ihg_setup_undo_sr"] = {
                    "file_id": file_id,
                    "file_name": file_name,
                    "bytes": original_bytes,
                }

            st.success(
                f"**{file_name}** is set up for {target_month:%B %Y}. "
                f"Prepared all Strategy week tabs."
            )

        except Exception as e:
            st.error(f"Strategy setup error: {e}")

    if "ihg_setup_undo_sr" in st.session_state:
        if st.button(
            "↩ Reset Strategy Report to Original",
            key=f"ihg_setup_reset_sr_{hotel_name}",
            type="secondary",
            use_container_width=True,
        ):
            try:
                info = st.session_state["ihg_setup_undo_sr"]
                drive_upload(
                    get_drive_service(),
                    info["file_id"],
                    info["bytes"],
                    info["file_name"],
                )
                del st.session_state["ihg_setup_undo_sr"]
                st.success(
                    f"**{info['file_name']}** restored to its original state."
                )
            except Exception as e:
                st.error(f"Reset error: {e}")


def render_ihg_update(hotels):
    """IHG portfolio run.

    One hotel at a time, like Stay In Touch, but fed by the History and
    Forecast Business Block PDF rather than a CSV. That single report drives
    both workbooks: the ROB takes its Total row for the current month, and the
    Forecast takes the daily rows — completed days as actuals, on-the-books
    days as OTB rooms and rate.
    """
    if not hotels:
        st.info("No IHG properties found in Drive.")
        return

    hotel_names = [h[0] for h in hotels]
    id_map = {h[0]: h[1] for h in hotels}

    st.caption(
        "Two PDFs. History and Forecast covers the current month end to end — "
        "it fills that month of the ROB and the whole Forecast, split into "
        "actuals and on-the-books at its own Subtotal line. Business on the "
        "Books starts at the report date and runs a year out, filling every "
        "later month of the ROB."
    )

    with st.container(border=True):
        col_h, col_w = st.columns([3, 3])
        with col_h:
            hotel_sel = st.selectbox("Hotel", hotel_names, key="ihg_hotel")
        with col_w:
            wb_sels = st.pills(
                "Workbooks to update",
                WORKBOOK_TYPES,
                selection_mode="multi",
                default=WORKBOOK_TYPES,
                key="ihg_wb",
            ) or []
        c1, c2 = st.columns(2)
        with c1:
            pdf_file = st.file_uploader(
                "History and Forecast Business Block (PDF)",
                type=["pdf"], key=f"ihg_pdf_{hotel_sel}")
        with c2:
            bob_file = st.file_uploader(
                "Business on the Books (PDF)",
                type=["pdf"], key=f"ihg_bob_{hotel_sel}")
        ihg_next_month = st.checkbox(
            "Include next month's Forecast",
            key="ihg_fcst_next",
            help="Fills next month's Forecast workbook from the Business on the "
                 "Books daily rows. Needs that PDF.")

        if "ROB" in wb_sels:
            render_portfolio_rob_month_setup(
                [(hotel_sel, id_map.get(hotel_sel, ""))],
                "ihg",
            )

        if "Strategy Report" in wb_sels:
            ihg_sr_setup_toggle = st.checkbox(
                "Set up new month — Strategy Report",
                key="ihg_sr_new_month",
                help=(
                    "Show the next-month Strategy setup controls. "
                    "Leave unchecked for a normal Strategy update."
                ),
            )
            if ihg_sr_setup_toggle:
                st.divider()
                render_ihg_strategy_month_setup(
                    hotel_sel,
                    id_map.get(hotel_sel, ""),
                )

    if not pdf_file:
        st.info("Upload the History and Forecast PDF to continue.")
        return

    try:
        parsed = parse_ihg_history_forecast(pdf_file)
    except Exception as e:
        st.error(f"Could not read the History and Forecast PDF: {e}")
        return

    bob = None
    if bob_file:
        try:
            bob = parse_ihg_business_on_books(bob_file)
        except Exception as e:
            st.error(f"Could not read the Business on the Books PDF: {e}")
            return

    past = [d for d in parsed["days"] if d["is_past"]]
    future = [d for d in parsed["days"] if not d["is_past"]]
    t = parsed["total"]
    st.success(
        f"Report date {parsed['report_date']} — {len(parsed['days'])} days "
        f"({len(past)} completed, {len(future)} on the books). "
        f"Month total {t['total_occ']:,.0f} rooms / ${t['total_rev']:,.2f}."
    )

    if bob:
        # The two reports overlap from the report date to month end; if they
        # disagree there they were pulled at different moments and the ROB
        # would mix them.
        cur = bob["months"].get((parsed["report_date"].year, parsed["report_date"].month))
        sub = parsed["subtotals"][1] if len(parsed["subtotals"]) > 1 else None
        if cur and sub and abs(cur["total_rev"] - sub["total_rev"]) > 1:
            st.warning(
                f"The two reports disagree over their overlapping dates "
                f"(Business on the Books ${cur['total_rev']:,.2f} vs History and "
                f"Forecast ${sub['total_rev']:,.2f}). They were probably pulled at "
                f"different times — re-pull both together for a clean run."
            )
        st.success(
            f"Business on the Books — {len(bob['months'])} months, "
            f"{sum(1 for m in bob['months'].values() if m['rooms'])} with business on them."
        )
    else:
        st.info("No Business on the Books PDF — only the current month of the ROB will be filled.")

    if st.button("Preview changes", key="ihg_preview", type="primary"):
        svc = get_drive_service()
        hotel_id = id_map[hotel_sel]
        jobs, problems = [], []
        if not hotel_id:
            st.error(
                f"{hotel_sel} has no Drive folder the app can see. Share its "
                f"folder (or its REVENUE REPORTS subfolder) with "
                f"`{service_account_email() or 'the service account'}` as "
                f"Editor, then press ↺ to refresh."
            )
            st.stop()
        for wb_type in wb_sels:
            result, err = resolve_drive_workbook(svc, hotel_id, hotel_sel, wb_type)
            if err or not result:
                problems.append(f"{wb_type}: {err}")
                continue
            file_id, file_name = result
            raw = drive_download(svc, file_id)
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=False)
            if wb_type == "ROB":
                avail = [s for s in ROB_SHEETS if s in wb.sheetnames]
                sheet = first_uncolored_sheet(wb, avail)
                changes = build_ihg_rob_plan(parsed, wb[sheet], bob=bob)
                passed = [f"{n} ({w})" for n, w in rob_week_status(wb, avail)
                          if w and n != sheet]
                note = "  ·  skipped " + "; ".join(passed) if passed else ""
            elif wb_type == "Strategy Report":
                avail = [s for s in STRATEGY_SHEETS if s in wb.sheetnames]
                sheet = first_undone_strategy_sheet(wb, avail)
                # A second, values-only view: the date column is formulas on
                # some hotels' later week tabs, and formulas read back as text.
                wb_vals = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
                changes = build_ihg_sr_plan(parsed, bob, wb[sheet],
                                            ws_values=wb_vals[sheet])
                note = ""
                if not bob:
                    problems.append(
                        "Strategy Report: without the Business on the Books PDF "
                        "only the days before it starts can be filled.")
                if not changes:
                    # Say which of the two it was rather than guessing at one.
                    found = detect_ihg_sr_columns(wb[sheet])
                    if "trans_rms" not in found:
                        problems.append(
                            f"{file_name}: couldn't match the Strategy Report's "
                            f"column headings on '{sheet}'. Found: "
                            f"{sorted(found) or 'none'}.")
                    else:
                        problems.append(
                            f"{file_name}: headings matched on '{sheet}', but none "
                            f"of its dated rows line up with the report's dates "
                            f"({parsed['report_date']:%b %Y}). Is this the right "
                            f"month's workbook?")
                    continue
                # The as-of date sits above the grid rather than in it.
                changes.insert(0, {
                    "row": 2, "col": 5, "label": "As-of date", "month": None,
                    "new_value": parsed["report_date"],
                    "skip_reason": "formula" if is_formula(wb[sheet].cell(2, 5).value) else None,
                })
            else:
                avail = [s for s in FORECAST_SHEETS if s in wb.sheetnames]
                sheet = first_unhighlighted_forecast_sheet(wb, avail)
                changes = build_ihg_forecast_plan(parsed, wb[sheet], wb=wb)
                note = ""
                if not changes:
                    problems.append(
                        f"{file_name}: could not locate the OTB / actual rows on "
                        f"'{sheet}' — its column-A labels don't match the template.")
                    continue
            jobs.append({"key": f"{wb_type} ({parsed['report_date']:%b %Y})",
                         "file_id": file_id, "file_name": file_name,
                         "wb_bytes": raw, "sheet": sheet,
                         "changes": changes, "note": note})

        if ihg_next_month and "Forecast" in wb_sels:
            ref = parsed["report_date"] or datetime.date.today()
            nxt = (ref.replace(day=1) + datetime.timedelta(days=32)).replace(day=1)
            if not bob:
                problems.append("Next month's Forecast needs the Business on the "
                                "Books PDF — that's where its daily rows come from.")
            else:
                nm_result, nm_err = resolve_drive_workbook(
                    svc, hotel_id, hotel_sel, "Forecast", month_date=nxt)
                if nm_err or not nm_result:
                    problems.append(f"{nxt:%b %Y} Forecast: {nm_err}")
                else:
                    nm_id, nm_name = nm_result
                    nm_raw = drive_download(svc, nm_id)
                    nm_wb = openpyxl.load_workbook(io.BytesIO(nm_raw), data_only=False)
                    nm_avail = [s for s in FORECAST_SHEETS if s in nm_wb.sheetnames]
                    nm_sheet = first_unhighlighted_forecast_sheet(nm_wb, nm_avail)
                    nm_changes = build_ihg_next_month_forecast_plan(
                        bob, nm_wb[nm_sheet], nxt, wb=nm_wb)
                    if not nm_changes:
                        problems.append(
                            f"Nothing to write for {nxt:%B %Y} — no daily rows in "
                            f"Business on the Books, or '{nm_sheet}' doesn't match "
                            f"the template.")
                    else:
                        jobs.append({"key": f"Forecast ({nxt:%b %Y})",
                                     "file_id": nm_id, "file_name": nm_name,
                                     "wb_bytes": nm_raw, "sheet": nm_sheet,
                                     "changes": nm_changes, "note": ""})

        st.session_state["ihg_jobs"] = jobs
        st.session_state["ihg_problems"] = problems

    jobs = st.session_state.get("ihg_jobs") or []
    for msg in st.session_state.get("ihg_problems") or []:
        st.error(msg)

    # Drawn before the preview and outside it: applying clears the plan, so an
    # undo living inside the plan block would vanish on the one rerun where it
    # is wanted.
    _render_undo_bar("ihg_undo")

    if not jobs:
        return

    for job in jobs:
        st.divider()
        st.subheader(f"{job['key']} — {job['sheet']}")
        st.caption(job["file_name"] + job.get("note", ""))
        _show_ihg_plan(job["changes"])

    st.divider()
    total = sum(len([c for c in j["changes"] if not c["skip_reason"]]) for j in jobs)
    st.write(f"**{total} cells across {len(jobs)} workbook(s).** "
             f"Applying writes to Drive and marks each tab done (green).")
    if st.button("Confirm and apply to Drive", key="ihg_apply", type="primary"):
        saved, errors = apply_portfolio_plans(get_drive_service(), jobs, "ihg_undo")
        for s in saved:
            st.success(f"Saved {s}")
        for e in errors:
            st.error(e)
        if saved:
            st.session_state.pop("ihg_jobs", None)
            st.rerun()


def _show_ihg_plan(changes):
    writes = [c for c in changes if not c["skip_reason"]]
    skips = [c for c in changes if c["skip_reason"]]
    st.dataframe(pd.DataFrame([
        {"row": c["row"], "col": get_column_letter(c["col"]),
         "field": c["label"], "value": c["new_value"]}
        for c in writes]), use_container_width=True, hide_index=True)
    if skips:
        st.caption(f"{len(skips)} cell(s) left alone because they hold a formula.")


def _hilton_forecast_job(svc, hotel_id, hotel_name, inn, prop, wash, problems,
                         month_date=None, as_of=None):
    """Build the Forecast job for one Hilton hotel and one month.

    month_date picks which month's workbook to open; None means the current
    one. Nothing else differs between the two. The SRP export runs a year
    forward, and build_hilton_forecast_plan already sends every date that
    hasn't happened yet to the OTB rows, so next month needs no separate
    builder the way IHG's does — there, next month genuinely comes from a
    different report.

    Returns the job, or None having appended the reason to `problems`.
    """
    label = f"Forecast ({month_date:%b %Y})" if month_date else "Forecast"
    result, err = resolve_drive_workbook(
        svc, hotel_id, hotel_name, "Forecast", month_date=month_date)
    if err or not result:
        problems.append(f"{hotel_name} — {label}: {err}")
        return None

    file_id, file_name = result
    raw = drive_download(svc, file_id)
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=False)
    avail = [s for s in FORECAST_SHEETS if s in wb.sheetnames]
    if not avail:
        problems.append(f"{hotel_name} — {label} ({file_name}): none of "
                        f"FCST-WK1 – WK9 were found.")
        return None

    sheet = first_unhighlighted_forecast_sheet(wb, avail)
    changes, warns = build_hilton_forecast_plan(
        prop["days"], wb[sheet], as_of=as_of
    )
    for w in warns:
        problems.append(f"{hotel_name} — {label} ({file_name}, {sheet}): {w}")
    if not changes:
        return None

    return {
        "key": f"{hotel_name} — {label}",
        "file_id": file_id, "file_name": file_name,
        "wb_bytes": raw, "sheet": sheet, "changes": changes,
        "note": f"  ·  InnCode {inn}",
    }


def _match_inncode(hotel_name, srp):
    """Map a Drive hotel name to the InnCode used in the SRP export.

    The lookup table is tried first, then the export's own 'Property - Name'.
    That fallback is what keeps a hotel running the day Hilton reissues its
    code or a tenth property appears: an unrecognised code used to mean the
    hotel was silently dropped from the run.
    """
    up = (hotel_name or "").upper()
    labels = [label for label, kws in PORTFOLIO_HOTELS["Hilton"].items()
              if any(k in up for k in kws)]
    for label in [hotel_name, *labels]:
        code = HILTON_INNCODES.get(label)
        if code and code in srp:
            return code

    keywords = [k for label in labels
                for k in PORTFOLIO_HOTELS["Hilton"][label]] or [up]
    for code, prop in srp.items():
        name = (prop.get("name") or "").upper()
        if name and any(k in name for k in keywords):
            return code
    return None


# ── Manual upload (test mode only) ───────────────────────────────────────────
if test_mode:
 with st.expander("Manual Upload", expanded=False):
    with st.expander("ROB Update"):
        st.header("ROB Master Workbook Update")
        csv_file = st.file_uploader("Upload CSV (Business on the Books)", type=["csv", "xlsx"], key="rob_csv")
        xl_file  = st.file_uploader("Upload ROB Master Workbook (.xlsx)", type=["xlsx"], key="rob_xl")
        npu_compare_file = st.file_uploader(
            "Occupancy Statistics — with unpicked group revenue included (Margaritaville only, optional)",
            type=["xlsx"], key="rob_npu_compare")

        if csv_file and xl_file:
            xl_bytes  = xl_file.read()

            df = parse_bob_source(csv_file)
            npu_compare_df = parse_bob_source(npu_compare_file) if npu_compare_file else None
            wb = openpyxl.load_workbook(io.BytesIO(xl_bytes), data_only=False)

            auto_sheet = first_uncolored_sheet(wb, ROB_SHEETS)

            # A selectbox's `index=` is only honored the first time its `key` is
            # created — once rob_sheet exists in session_state, Streamlit ignores
            # index= on every rerun and keeps the old selection. Force a reset
            # whenever the uploaded file's bytes change (see the identical fix
            # applied to the Forecast manual-upload tab).
            rob_xl_hash = hashlib.md5(xl_bytes).hexdigest()
            if st.session_state.get("rob_xl_hash") != rob_xl_hash:
                st.session_state["rob_xl_hash"] = rob_xl_hash
                st.session_state["rob_sheet"] = auto_sheet

            sheet_choice = st.selectbox("Week tab", ROB_SHEETS, key="rob_sheet")
            st.caption(f"Auto-detected next tab: **{auto_sheet}**")
            _passed = [f"{n} — {why}" for n, why in rob_week_status(wb, ROB_SHEETS)
                       if why and n != auto_sheet]
            if _passed:
                st.caption("Skipped: " + "; ".join(_passed))
    
            if st.button("Preview Changes", key="rob_preview"):
                ws = wb[sheet_choice]
                grp_npu_rev_override = compute_grp_npu_rev_override(df, npu_compare_df)
                changes = build_rob_change_plan(df, ws, grp_npu_rev_override=grp_npu_rev_override)
                st.session_state["rob_changes"]   = changes
                st.session_state["rob_wb_bytes"]  = xl_bytes
                st.session_state["rob_sheet_sel"] = sheet_choice
    
            if "rob_changes" in st.session_state:
                changes    = st.session_state["rob_changes"]
                will_write = [c for c in changes if not c["skip_reason"]]
                skipped    = [c for c in changes if c["skip_reason"]]
    
                c1, c2 = st.columns(2)
                c1.metric("Cells to update", len(will_write))
                c2.metric("Skipped",         len(skipped))
    
                preview_rows = []
                for c in changes:
                    preview_rows.append({
                        "Month":  c["month"] or "—",
                        "Label":  c["label"],
                        "Row":    c["row"],
                        "Col":    c["col"],
                        "Value":  c["new_value"],
                        "Status": "✅ will write" if not c["skip_reason"] else f"⚠️ skip ({c['skip_reason']})",
                    })
                st.dataframe(preview_rows, use_container_width=True)
    
                if st.button("Confirm and Apply Changes", key="rob_apply"):
                    wb2 = openpyxl.load_workbook(io.BytesIO(st.session_state["rob_wb_bytes"]), data_only=False)
                    apply_rob_changes(wb2, st.session_state["rob_sheet_sel"], changes)
                    color_tab_done(wb2, st.session_state["rob_sheet_sel"])
                    strip_tables(wb2)
                    out = io.BytesIO()
                    wb2.save(out)
                    st.download_button(
                        "Download Updated ROB Workbook",
                        data=out.getvalue(),
                        file_name="ROB_Master_updated.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
    
    with st.expander("Strategy Report"):
        st.header("Strategy Report Update")

        col_a, col_b = st.columns(2)
        with col_a:
            csv_file2 = st.file_uploader("Upload CSV (Business on the Books)", type=["csv", "xlsx"], key="str_csv")
        with col_b:
            rate_file2 = st.file_uploader("Upload Rates & Restrictions CSV", type=["csv"], key="str_rate")
    
        col_c, col_d = st.columns(2)
        with col_c:
            xl_file2 = st.file_uploader("Upload Strategy Report Workbook (.xlsx)", type=["xlsx"], key="str_xl")
    
        if xl_file2:
            xl_bytes2 = xl_file2.read()
            wb2_peek  = openpyxl.load_workbook(io.BytesIO(xl_bytes2), data_only=False)

            auto_sheet2 = first_undone_strategy_sheet(wb2_peek, STRATEGY_SHEETS)

            # Same stale-selection issue as ROB/Forecast: force a reset whenever
            # the uploaded file's bytes change, since index= is ignored once
            # str_sheet already exists in session_state.
            str_xl_hash = hashlib.md5(xl_bytes2).hexdigest()
            if st.session_state.get("str_xl_hash") != str_xl_hash:
                st.session_state["str_xl_hash"] = str_xl_hash
                st.session_state["str_sheet"] = auto_sheet2

            sheet_choice2 = st.selectbox("Week tab", STRATEGY_SHEETS, key="str_sheet")
            st.caption(f"Auto-detected next tab: **{auto_sheet2}**")
    
            if (csv_file2 or rate_file2) and st.button("Preview Changes", key="str_preview"):
                wb2_full = openpyxl.load_workbook(io.BytesIO(xl_bytes2), data_only=False)
                all_changes = []
    
                if csv_file2:
                    df2 = parse_bob_source(csv_file2)
                    all_changes += build_strategy_change_plan(df2, wb2_full, sheet_choice2)
    
                rate_warnings = []
                if rate_file2:
                    rate_df = parse_rate_csv(rate_file2.read())
                    rate_changes, rate_warnings = build_rates_change_plan(
                        rate_df, wb2_full, sheet_choice2, hotel_name=None)
                    all_changes += rate_changes
    
                st.session_state["str_changes"]   = all_changes
                st.session_state["str_wb_bytes"]  = xl_bytes2
                st.session_state["str_sheet_sel"] = sheet_choice2
                st.session_state["str_warnings"]  = rate_warnings
    
            if "str_changes" in st.session_state:
                for w in st.session_state.get("str_warnings", []):
                    st.warning(w)
    
                changes2    = st.session_state["str_changes"]
                will_write2 = [c for c in changes2 if not c["skip_reason"]]
                skipped2    = [c for c in changes2 if c["skip_reason"]]
    
                c1, c2, c3 = st.columns(3)
                c1.metric("Cells to update", len(will_write2))
                c2.metric("Skipped",         len(skipped2))
                c3.metric("Days in scope",   len({c["date"] for c in changes2}))
    
                preview_rows2 = []
                for c in changes2:
                    preview_rows2.append({
                        "Date":   str(c["date"]),
                        "Label":  c["label"],
                        "Row":    c["row"],
                        "Col":    c["col"],
                        "Value":  c["new_value"],
                        "Status": "✅ will write" if not c["skip_reason"] else f"⚠️ skip ({c['skip_reason']})",
                    })
                st.dataframe(preview_rows2, use_container_width=True)
    
                if st.button("Confirm and Apply Changes", key="str_apply"):
                    wb3 = openpyxl.load_workbook(io.BytesIO(st.session_state["str_wb_bytes"]), data_only=False)
                    apply_strategy_changes(wb3, st.session_state["str_sheet_sel"], changes2)
                    color_tab_done(wb3, st.session_state["str_sheet_sel"])
                    strip_tables(wb3)
                    out2 = io.BytesIO()
                    wb3.save(out2)
                    st.download_button(
                        "Download Updated Strategy Workbook",
                        data=out2.getvalue(),
                        file_name="Strategy_Report_updated.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
    
    with st.expander("Forecast"):
        st.header("Forecast Update")
    
        fcst_csv     = st.file_uploader("Upload CSV (Business on the Books)", type=["csv", "xlsx"], key="fcst_csv")
        fcst_xl      = st.file_uploader("Upload Current Month Forecast Workbook (.xlsx/.xlsm)", type=["xlsx", "xlsm"], key="fcst_xl")
        st.caption("Weeks 3 & 4 only: also upload next month's forecast workbook.")
        fcst_xl_next = st.file_uploader("Upload Next Month Forecast Workbook (.xlsx/.xlsm)", type=["xlsx", "xlsm"], key="fcst_xl_next")
    
        if fcst_xl:
            fcst_bytes = fcst_xl.read()
            wb_fcst_peek = openpyxl.load_workbook(io.BytesIO(fcst_bytes), data_only=False)
    
            avail_fcst = [s for s in FORECAST_SHEETS if s in wb_fcst_peek.sheetnames]
            if not avail_fcst:
                st.error(f"None of the expected week tabs (FCST-WK1 – FCST-WK9) were found in "
                         f"this workbook. It has: {', '.join(wb_fcst_peek.sheetnames)}. Make sure "
                         f"you uploaded the destination Forecast workbook here, not the source "
                         f"data file.")
            auto_fcst  = first_unhighlighted_forecast_sheet(wb_fcst_peek, avail_fcst) if avail_fcst else None

            # A selectbox's `index=` is only honored the first time its `key` is
            # created — once fcst_sheet exists in session_state, Streamlit ignores
            # index= on every rerun and keeps the old selection. Re-uploading a new
            # file (this week's, with more tabs now done) would silently keep
            # writing to whatever tab was last picked instead of the newly
            # auto-detected one. Force a reset whenever the uploaded bytes change.
            fcst_hash = hashlib.md5(fcst_bytes).hexdigest()
            if st.session_state.get("fcst_xl_hash") != fcst_hash:
                st.session_state["fcst_xl_hash"] = fcst_hash
                if auto_fcst:
                    st.session_state["fcst_sheet"] = auto_fcst

            fcst_sheet = st.selectbox("Week tab", avail_fcst, key="fcst_sheet")
            if auto_fcst:
                st.caption(f"Auto-detected next tab: **{auto_fcst}**")
    
            if fcst_xl_next:
                fcst_next_bytes = fcst_xl_next.read()
                wb_next_peek    = openpyxl.load_workbook(io.BytesIO(fcst_next_bytes), data_only=False)
                avail_next      = [s for s in FORECAST_SHEETS if s in wb_next_peek.sheetnames]
                auto_next       = first_unhighlighted_forecast_sheet(wb_next_peek, avail_next) if avail_next else None

                fcst_next_hash = hashlib.md5(fcst_next_bytes).hexdigest()
                if st.session_state.get("fcst_xl_next_hash") != fcst_next_hash:
                    st.session_state["fcst_xl_next_hash"] = fcst_next_hash
                    if auto_next:
                        st.session_state["fcst_sheet_next"] = auto_next

                fcst_sheet_next = st.selectbox("Next month week tab", avail_next, key="fcst_sheet_next")
                if auto_next:
                    st.caption(f"Auto-detected next month tab: **{auto_next}**")
    
            if fcst_csv and fcst_sheet and st.button("Preview Changes", key="fcst_preview"):
                # Current month workbook
                wb_fcst_full = openpyxl.load_workbook(io.BytesIO(fcst_bytes), data_only=False)
                df_fcst = parse_bob_source(fcst_csv)
                ws_fcst = wb_fcst_full[fcst_sheet]
                fcst_changes, fcst_warnings = build_forecast_change_plan(df_fcst, ws_fcst)
    
                st.session_state["fcst_changes"]   = fcst_changes
                st.session_state["fcst_wb_bytes"]  = fcst_bytes
                st.session_state["fcst_sheet_sel"] = fcst_sheet
                st.session_state["fcst_warnings"]  = fcst_warnings
    
                # Next month workbook (weeks 3 & 4)
                if fcst_xl_next:
                    wb_next_full = openpyxl.load_workbook(io.BytesIO(fcst_next_bytes), data_only=False)
                    ws_next      = wb_next_full[fcst_sheet_next]
                    next_changes, next_warnings = build_next_month_forecast_plan(df_fcst, ws_next)
                    st.session_state["fcst_next_changes"]   = next_changes
                    st.session_state["fcst_next_wb_bytes"]  = fcst_next_bytes
                    st.session_state["fcst_next_sheet_sel"] = fcst_sheet_next
                    st.session_state["fcst_next_warnings"]  = next_warnings
                else:
                    st.session_state.pop("fcst_next_changes", None)
    
            if "fcst_changes" in st.session_state:
                for w in st.session_state.get("fcst_warnings", []):
                    st.warning(w)
    
                fcst_ch   = st.session_state["fcst_changes"]
                will_fcst = [c for c in fcst_ch if not c["skip_reason"]]
                skip_fcst = [c for c in fcst_ch if c["skip_reason"]]
    
                st.subheader("Current month changes")
                c1, c2 = st.columns(2)
                c1.metric("Cells to update", len(will_fcst))
                c2.metric("Skipped",         len(skip_fcst))
    
                preview_fcst = []
                for c in fcst_ch:
                    preview_fcst.append({
                        "Label":  c["label"],
                        "Row":    c["row"],
                        "Col":    c["col"],
                        "Value":  c["new_value"],
                        "Status": "✅ will write" if not c["skip_reason"] else f"⚠️ skip ({c['skip_reason']})",
                    })
                st.dataframe(preview_fcst, use_container_width=True)
    
                if "fcst_next_changes" in st.session_state:
                    next_ch = st.session_state["fcst_next_changes"]
                    for w in st.session_state.get("fcst_next_warnings", []):
                        st.warning(w)
                    will_next = [c for c in next_ch if not c["skip_reason"]]
                    skip_next = [c for c in next_ch if c["skip_reason"]]
                    st.subheader("Next month changes")
                    n1, n2 = st.columns(2)
                    n1.metric("Cells to update", len(will_next))
                    n2.metric("Skipped",         len(skip_next))
                    preview_next = []
                    for c in next_ch:
                        preview_next.append({
                            "Label":  c["label"],
                            "Row":    c["row"],
                            "Col":    c["col"],
                            "Value":  c["new_value"],
                            "Status": "✅ will write" if not c["skip_reason"] else f"⚠️ skip ({c['skip_reason']})",
                        })
                    st.dataframe(preview_next, use_container_width=True)
    
                if st.button("Confirm and Apply Changes", key="fcst_apply"):
                    # Apply current month
                    wb_out = openpyxl.load_workbook(io.BytesIO(st.session_state["fcst_wb_bytes"]), data_only=False)
                    apply_forecast_changes(wb_out, st.session_state["fcst_sheet_sel"], fcst_ch)
                    color_tab_done(wb_out, st.session_state["fcst_sheet_sel"])
                    strip_tables(wb_out)
                    out3 = io.BytesIO()
                    wb_out.save(out3)
                    st.download_button(
                        "Download Updated Current Month Forecast",
                        data=out3.getvalue(),
                        file_name="Forecast_current_updated.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
    
                    # Apply next month (if uploaded)
                    if "fcst_next_changes" in st.session_state:
                        wb_next_out = openpyxl.load_workbook(io.BytesIO(st.session_state["fcst_next_wb_bytes"]), data_only=False)
                        apply_forecast_changes(wb_next_out, st.session_state["fcst_next_sheet_sel"],
                                               st.session_state["fcst_next_changes"])
                        color_tab_done(wb_next_out, st.session_state["fcst_next_sheet_sel"])
                        strip_tables(wb_next_out)
                        out4 = io.BytesIO()
                        wb_next_out.save(out4)
                        st.download_button(
                            "Download Updated Next Month Forecast",
                            data=out4.getvalue(),
                            file_name="Forecast_next_month_updated.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
    
# ── Main menu ────────────────────────────────────────────────────────────────
# A landing page of two cards, then that section's tools as tabs.
#
# Streamlit runs the whole script top to bottom on every interaction, so the
# tool bodies further down execute whichever section is open. The ones that
# belong to the section you are NOT in are pointed at a placeholder, which is
# emptied at the end of the file — they run, as they always have, but nothing
# of theirs reaches the page.
SECTION_REVENUE = "Revenue Management"
SECTION_FINANCIAL = "Financial Tools"

_section = st.session_state.get("main_section")

if _section not in (SECTION_REVENUE, SECTION_FINANCIAL):
    st.markdown(
        """
        <style>
          div[data-testid="stButton"] > button {
              height: 200px;
              border: 2px solid #d6dce4;
              border-radius: 14px;
              background: #ffffff;
              color: #1F3864;
              transition: border-color .15s, box-shadow .15s, transform .15s,
                          background .15s;
          }
          /* Streamlit renders each blank-line-separated part of the label as
             its own <p>, so the icon, title and subtitle can be sized apart. */
          div[data-testid="stButton"] > button p {
              color: #1F3864;
              margin: 0;
          }
          div[data-testid="stButton"] > button p:nth-of-type(1) {
              font-size: 2.6rem;
              line-height: 1.5;
          }
          div[data-testid="stButton"] > button p:nth-of-type(2) {
              font-size: 1.3rem;
              font-weight: 700;
              line-height: 1.9;
          }
          div[data-testid="stButton"] > button p:nth-of-type(3) {
              font-size: .82rem;
              font-weight: 400;
              color: #5a6b8c;
              line-height: 1.6;
          }
          div[data-testid="stButton"] > button:hover {
              border-color: #1F3864;
              background: #f4f7fd;
              transform: translateY(-3px);
              box-shadow: 0 8px 22px rgba(31,56,100,.18);
              color: #1F3864;
          }
          div[data-testid="stButton"] > button:focus:not(:active) {
              border-color: #1F3864;
              color: #1F3864;
          }
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.write("")
    _pad_l, _card_l, _card_r, _pad_r = st.columns([1, 3, 3, 1])
    with _card_l:
        if st.button(
            "📈\n\nRevenue Management\n\nWeekly Update · Ancillary · OOO",
            key="menu_revenue", use_container_width=True,
        ):
            st.session_state["main_section"] = SECTION_REVENUE
            st.rerun()
    with _card_r:
        if st.button(
            "💵\n\nFinancial Tools\n\nP&L · 1-Year Projection",
            key="menu_financial", use_container_width=True,
        ):
            st.session_state["main_section"] = SECTION_FINANCIAL
            st.rerun()
    st.stop()

_back, _title = st.columns([1, 6])
with _back:
    if st.button("← Main menu", key="menu_back"):
        st.session_state.pop("main_section", None)
        st.rerun()
with _title:
    st.markdown(f"#### {_section}")

# Whichever section is not open is built into this placeholder and wiped at the
# end of the file.
_offstage = st.empty()
_offstage_box = _offstage.container()

if _section == SECTION_REVENUE:
    tab_weekly, tab_ancillary, tab_ooo = st.tabs(
        ["Weekly Workbook Update", "Ancillary Revenue", "Monthly OOO Report"])
    with _offstage_box:
        tab_pl, tab_projection = st.tabs(["P&L Spreadsheet", "1-Year Projection"])
else:
    tab_pl, tab_projection = st.tabs(["P&L Spreadsheet", "1-Year Projection"])
    with _offstage_box:
        tab_weekly, tab_ancillary, tab_ooo = st.tabs(
            ["Weekly Workbook Update", "Ancillary Revenue", "Monthly OOO Report"])

with tab_weekly:
    st.divider()
    # ── Google Drive Update ───────────────────────────────────────────────────────
    col_title, col_month = st.columns([5, 2])
    with col_title:
        st.header("Weekly Workbook Update")
    with col_month:
        st.write("")
        st.write("")
        st.caption(f"📅 Current month: **{datetime.date.today().strftime('%B %Y')}**")

    all_discovered = get_hotels_from_drive()

    # Each portfolio's data arrives in a different export, so they get their own
    # tab rather than sharing one hotel list.
    portfolio = st.radio("Portfolio", PORTFOLIOS, horizontal=True, key="drive_portfolio")

    hotels = hotels_in_portfolio(portfolio, all_discovered)
    hotel_names = [h[0] for h in hotels]
    hotel_id_map = {h[0]: h[1] for h in hotels}
    allowed_wbs = PORTFOLIO_WORKBOOKS[portfolio]

    missing = portfolio_hotels_missing(portfolio, all_discovered)
    if missing:
        who = service_account_email() or "the service account"
        st.warning(
            f"**No Drive folder found for:** {', '.join(missing)}\n\n"
            f"They are still listed below and can be selected, but a run will "
            f"stop with an error until this is sorted. This app signs in as "
            f"`{who}` — that is the address a folder has to be shared with, and "
            f"sharing with a different service account has no effect here."
        )
        with st.expander("Check what the app can actually see"):
            render_folder_diagnostic(portfolio, missing)

    if portfolio == "Hilton":
        render_hilton_update(hotels)
        st.stop()
    if portfolio == "IHG":
        render_ihg_update(hotels)
        st.stop()

    start_new_month = False
    with st.container(border=True):
        col_h, col_ref, col_w = st.columns([3, 1, 3])
        with col_h:
            hotel_sel = st.selectbox("Hotel", hotel_names if hotel_names else ["(no hotels found)"], key="drive_hotel")
        with col_ref:
            st.write("")
            if st.button("↺", key="refresh_hotels", help="Refresh hotel list"):
                get_hotels_from_drive.clear()
                _all_visible_folders.clear()
                st.rerun()
        with col_w:
            wb_sels = st.pills(
                "Workbooks to update",
                allowed_wbs,
                selection_mode="multi",
                default=allowed_wbs,
                key="drive_wb",
            ) or []

        # Keying these to the selected hotel clears any uploaded file the moment you
        # switch hotels — one hotel's BOB/R&R CSV should never carry over and get
        # applied to a different hotel.
        drive_csv = st.file_uploader("CSV — Business on the Books", type=["csv", "xlsx"], key=f"drive_csv_{hotel_sel}")
        drive_rate_csv = None
        drive_lighthouse_xlsx = None
        if "Strategy Report" in (wb_sels or []):
            drive_rate_csv = st.file_uploader(
                "CSV — SNT Rates & Restrictions",
                type=["csv"],
                key=f"drive_rate_csv_{hotel_sel}",
            )
            drive_lighthouse_xlsx = st.file_uploader(
                "Optional — Lighthouse Compset Rates (.xlsx)",
                type=["xlsx"],
                key=f"drive_lighthouse_{hotel_sel}",
                help=(
                    "Optional for SNT hotels that use Lighthouse. The selected "
                    "hotel's own rate still comes from SNT; Lighthouse only fills "
                    "competitors already listed in the Strategy Report."
                ),
            )
        drive_npu_compare_csv = None
        if "ROB" in (wb_sels or []) and "margaritaville" in hotel_sel.lower():
            drive_npu_compare_csv = st.file_uploader(
                "Occupancy Statistics — with unpicked group revenue included",
                type=["xlsx"], key=f"drive_npu_compare_{hotel_sel}", width=500)

        opt_col1, opt_col2 = st.columns(2)
        forecast_next_month = False
        if "Forecast" in (wb_sels or []):
            with opt_col1:
                forecast_next_month = st.checkbox("Include next month's Forecast", key="drive_fcst_next")
        with opt_col2:
            start_new_month = st.checkbox("Set up new month", key="drive_new_month")
    if start_new_month:
        with st.container(border=True):
            today         = datetime.date.today()
            cur_month_dt  = today.replace(day=1)
            prev_month_dt = (cur_month_dt - datetime.timedelta(days=1)).replace(day=1)
            next_month_dt = (cur_month_dt + datetime.timedelta(days=32)).replace(day=1)
            month_options = {
                prev_month_dt.strftime("%B %Y"): prev_month_dt,
                cur_month_dt.strftime("%B %Y"):  cur_month_dt,
                next_month_dt.strftime("%B %Y"): next_month_dt,
            }
            month_labels = list(month_options.keys())
            # Week 4+ of the current month → default to next month; week 1 → default
            # to the current month. Weeks 2-3 keep the plain current-month default.
            # All three options remain selectable regardless of the default.
            default_dt = next_month_dt if today.day >= 22 else cur_month_dt
            sel_month_label = st.selectbox("Month to set up", month_labels,
                                            index=month_labels.index(default_dt.strftime("%B %Y")),
                                            key="setup_month_sel")
            setup_month_dt  = month_options[sel_month_label]
            month_kw        = setup_month_dt.strftime("%b%Y").upper()

            rob_col, sr_col = st.columns(2)

            # ── ROB setup ──────────────────────────────────────────────────────────
            with rob_col:
                st.markdown("**ROB**")
                if st.button("Set Up New ROB", key="btn_setup_rob", type="primary", use_container_width=True):
                    try:
                        svc         = get_drive_service()
                        hotel_id_nm = hotel_id_map.get(hotel_sel, "")
                        with st.spinner("Setting up ROB — this may take a moment..."):
                            rob_name, rob_err, rob_file_id, rob_orig_bytes = setup_new_rob_month(
                                svc, hotel_id_nm, hotel_sel, setup_month_dt)
                        if rob_err and not rob_name:
                            if "storageQuotaExceeded" in str(rob_err):
                                _, master_name = find_rob_master(svc, hotel_id_nm, target_month.year)
                                rob_suffix = hotel_sel.upper()
                                if master_name and "ROB" in master_name.upper():
                                    after = master_name[master_name.upper().find("ROB") + 3:].strip()
                                    after = after.replace(".xlsx","").replace(".xlsm","").replace(".XLSX","").replace(".XLSM","").strip()
                                    if after:
                                        rob_suffix = after
                                ext = ".xlsm" if master_name and master_name.lower().endswith(".xlsm") else ".xlsx"
                                suggested_name = f"{month_kw} ROB {rob_suffix}{ext}"
                                st.warning(
                                    f"Auto-copy requires a Shared Drive. Do this in Google Drive first:\n\n"
                                    f"1. Right-click **{master_name or 'the ROB master'}** → *Make a copy*\n"
                                    f"2. Rename to: **`{suggested_name}`**\n"
                                    f"3. Move into the **{month_kw}** folder\n\n"
                                    f"Then click **Set Up New ROB** again."
                                )
                            else:
                                st.error(f"ROB setup error: {rob_err}")
                        else:
                            if rob_err:
                                st.warning(rob_err)
                            if rob_file_id and rob_orig_bytes is not None:
                                st.session_state["setup_undo_rob"] = {
                                    "file_id":   rob_file_id,
                                    "file_name": rob_name,
                                    "bytes":     rob_orig_bytes,
                                }
                            st.success(f"**{rob_name}** ready for {setup_month_dt.strftime('%B %Y')}.")
                    except Exception as e:
                        st.error(f"ROB setup error: {e}")

                # Reset button — shown after a successful ROB setup
                if "setup_undo_rob" in st.session_state:
                    if st.button("↩ Reset ROB to Original", key="setup_reset_rob", type="secondary", use_container_width=True):
                        try:
                            info = st.session_state["setup_undo_rob"]
                            with st.spinner("Restoring original ROB workbook..."):
                                drive_upload(get_drive_service(), info["file_id"], info["bytes"], info["file_name"])
                            del st.session_state["setup_undo_rob"]
                            st.success(f"**{info['file_name']}** restored to original state.")
                        except Exception as e:
                            st.error(f"Reset error: {e}")

            # ── Strategy Report setup ──────────────────────────────────────────────
            with sr_col:
                st.markdown("**Strategy Report**")
                if st.button("Set Up New SR", key="btn_setup_new_wb", type="primary", use_container_width=True):
                    try:
                        svc         = get_drive_service()
                        hotel_id_nm = hotel_id_map.get(hotel_sel, "")

                        # Step 1 — ensure the file exists; skip copy if it's already there
                        is_fresh_copy = False
                        with st.spinner("Step 1 / 3 — locating or creating workbook..."):
                            existing, find_err = resolve_drive_workbook(svc, hotel_id_nm, hotel_sel,
                                                                  "Strategy Report", month_date=setup_month_dt)
                            if existing:
                                st.info(f"Found existing file: **{existing[1]}** — skipping copy.")
                            else:
                                is_fresh_copy = True
                                created_name, create_err = setup_new_sr_month(svc, hotel_id_nm, hotel_sel, setup_month_dt)
                                if create_err:
                                    master_id, master_name = find_sr_master(svc, hotel_id_nm, setup_month_dt.year)
                                    hotel_suffix = ""
                                    if master_name and "STRATEGY" in master_name.upper():
                                        hotel_suffix = master_name[master_name.upper().find("STRATEGY") + len("STRATEGY"):].strip().replace(".xlsx","").replace(".XLSX","").strip()
                                    suggested_name = f"{month_kw} STRATEGY {hotel_suffix}.xlsx".strip()
                                    if "storageQuotaExceeded" in str(create_err):
                                        st.warning(
                                            f"Auto-copy requires a Shared Drive. Do this in Google Drive first:\n\n"
                                            f"1. Right-click **{master_name or 'the SR master'}** → *Make a copy*\n"
                                            f"2. Rename to: **`{suggested_name}`**\n"
                                            f"3. Move into the **{month_kw}** folder\n\n"
                                            f"Then click **Set Up New SR** again."
                                        )
                                    else:
                                        st.error(f"Could not create workbook: {create_err}")
                                    st.stop()

                        # Step 2 — load reference workbooks into memory
                        with st.spinner("Step 2 / 3 — loading reference workbooks..."):
                            prev_month_dt    = (setup_month_dt - datetime.timedelta(days=1)).replace(day=1)
                            ly_month_dt      = setup_month_dt.replace(year=setup_month_dt.year - 1)
                            prev_month_sr_wb = _load_wb_from_drive(svc, hotel_id_nm, hotel_sel, "Strategy Report", prev_month_dt, data_only=False)
                            ly_sr_wb         = _load_wb_from_drive(svc, hotel_id_nm, hotel_sel, "Strategy Report", ly_month_dt)
                        st.info(f"Prev month ({prev_month_dt.strftime('%b %Y')}): {'✓' if prev_month_sr_wb else '✗ not found'}")
                        st.info(f"Last year  ({ly_month_dt.strftime('%b %Y')}): {'✓' if ly_sr_wb else '✗ not found'}")

                        # Step 3 — populate all 5 weeks
                        with st.spinner("Step 3 / 3 — populating all weeks..."):
                            result, err = resolve_drive_workbook(svc, hotel_id_nm, hotel_sel,
                                                                 "Strategy Report", month_date=setup_month_dt)
                            if err:
                                st.error(f"Cannot open new workbook: {err}")
                                st.stop()
                            file_id, file_name = result
                            wb_bytes = drive_download(svc, file_id)
                            original_bytes = wb_bytes
                            wb       = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)
                            if is_fresh_copy:
                                clear_tab_colors(wb, STRATEGY_SHEETS)
                            restructure_sr_dates(wb, setup_month_dt)
                            first_ws = wb[STRATEGY_SHEETS[0]] if STRATEGY_SHEETS[0] in wb.sheetnames else None
                            num_rows = _count_sheet_data_rows(first_ws) if first_ws else 365
                            full_scope_start = setup_month_dt
                            full_scope_end   = setup_month_dt + datetime.timedelta(days=max(0, num_rows - 1))
                            total_written = 0
                            for sheet_name in STRATEGY_SHEETS:
                                if sheet_name not in wb.sheetnames:
                                    continue
                                changes = build_strategy_change_plan(None, wb, sheet_name,
                                                                      prev_month_wb=prev_month_sr_wb,
                                                                      ly_wb=ly_sr_wb,
                                                                      scope_start=full_scope_start,
                                                                      scope_end=full_scope_end)
                                apply_strategy_changes(wb, sheet_name, changes)
                                total_written += len([c for c in changes if not c.get("skip_reason")])
                            strip_tables(wb)
                            out = io.BytesIO()
                            wb.save(out)
                            drive_upload(svc, file_id, out.getvalue(), file_name)
                            st.session_state["setup_undo_sr"] = {
                                "file_id":   file_id,
                                "file_name": file_name,
                                "bytes":     original_bytes,
                            }

                        st.success(
                            f"**{file_name}** is set up for {setup_month_dt.strftime('%B %Y')}. "
                            f"Populated **{total_written}** cells across all weeks."
                        )
                    except Exception as e:
                        st.error(f"Setup error: {e}")

                # Reset button — shown after a successful SR setup
                if "setup_undo_sr" in st.session_state:
                    if st.button("↩ Reset SR to Original", key="setup_reset_sr", type="secondary", use_container_width=True):
                        try:
                            info = st.session_state["setup_undo_sr"]
                            with st.spinner("Restoring original SR workbook..."):
                                drive_upload(get_drive_service(), info["file_id"], info["bytes"], info["file_name"])
                            del st.session_state["setup_undo_sr"]
                            st.success(f"**{info['file_name']}** restored to original state.")
                        except Exception as e:
                            st.error(f"Reset error: {e}")




    def build_all_plans(
        svc,
        hotel_sel,
        hotel_id,
        wb_sels,
        df,
        rate_df,
        forecast_next_month=False,
        npu_compare_df=None,
        lighthouse_data=None,
    ):
        today = datetime.date.today()
        current_month = today.replace(day=1)
        all_plans = {}

        grp_npu_rev_override = compute_grp_npu_rev_override(df, npu_compare_df)

        # Pre-load reference workbooks into memory once — used for cross-sheet lookups
        prev_month_sr_wb = None
        ly_sr_wb         = None
        if "Strategy Report" in wb_sels:
            prev_month_dt = (current_month - datetime.timedelta(days=1)).replace(day=1)
            ly_month_dt   = current_month.replace(year=current_month.year - 1)
            prev_month_sr_wb = _load_wb_from_drive(svc, hotel_id, hotel_sel, "Strategy Report", prev_month_dt, data_only=False)
            ly_sr_wb         = _load_wb_from_drive(svc, hotel_id, hotel_sel, "Strategy Report", ly_month_dt)
            # Comp Set LY / OTB LY Trans / GRP LY etc. all come from ly_sr_wb — if it's
            # not found, those fields silently produce nothing (no warning previously),
            # which looked like "dates transferred but no text" with no explanation why.
            st.info(
                "Strategy references · "
                f"Previous month ({prev_month_dt:%b %Y}): "
                f"{'found' if prev_month_sr_wb else 'not found — Last Week OTB will stay blank'} · "
                f"Last year ({ly_month_dt:%b %Y}): "
                f"{'found' if ly_sr_wb else 'not found — LY fields will stay blank'}"
            )

        for wb_type in wb_sels:
            result, err = resolve_drive_workbook(svc, hotel_id, hotel_sel, wb_type)
            if err:
                st.error(f"{wb_type}: {err}")
                continue
            file_id, file_name = result
            wb_bytes = drive_download(svc, file_id)
            wb       = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)
            if wb_type == "ROB":
                avail    = [s for s in ROB_SHEETS if s in wb.sheetnames]
                auto     = first_uncolored_sheet(wb, avail)
                sheet    = auto or avail[0]
                changes  = build_rob_change_plan(df, wb[sheet], grp_npu_rev_override=grp_npu_rev_override)
                warnings = []
                # Say which weeks were passed over and why — a silently skipped
                # week tab is otherwise invisible until someone spots the gap.
                _passed = [f"{n} ({why})" for n, why in rob_week_status(wb, avail)
                           if why and n != sheet]
                if _passed:
                    warnings.append(f"Writing to '{sheet}'. Skipped: " + "; ".join(_passed))
            elif wb_type == "Strategy Report":
                avail    = [s for s in STRATEGY_SHEETS if s in wb.sheetnames]
                auto     = first_undone_strategy_sheet(wb, avail)
                sheet    = auto or avail[0]
                date_row_map_debug = build_date_row_map(wb, prefer_sheet=sheet)
                own_debug = build_date_row_map(
                    wb, prefer_sheet=sheet, fallback_to_wkone=False
                )

                if date_row_map_debug:
                    date_summary = (
                        f"{len(date_row_map_debug)} dates mapped "
                        f"({min(date_row_map_debug):%m/%d/%Y}–"
                        f"{max(date_row_map_debug):%m/%d/%Y})"
                    )
                else:
                    date_summary = "no dates mapped"

                if df is not None:
                    bob_daily = sum(
                        1 for _, r in df.iterrows()
                        if classify_row(str(r[0]).strip())[0] == "daily"
                    )
                    bob_summary = f" · BOB daily rows: {bob_daily}"
                else:
                    bob_summary = " · BOB file not loaded"

                st.info(
                    f"Strategy Report: **{file_name}** → **{sheet}** · "
                    f"{date_summary}{bob_summary}"
                )

                if not own_debug:
                    st.warning(
                        f"{sheet}: date rows could not be mapped on this tab, "
                        f"so daily Strategy values may not populate."
                    )

                # Only extract LY data during month setup, not on regular CSV uploads
                # (ly_sr_wb is already cleared of blanking logic if ly_data is empty)
                changes  = build_strategy_change_plan(df, wb, sheet,
                                                       prev_month_wb=prev_month_sr_wb,
                                                       ly_wb=None)
                warnings = []
                if rate_df is not None:
                    rate_changes, rate_warnings = build_rates_change_plan(
                        rate_df, wb, sheet, hotel_name=hotel_sel
                    )
                    changes += rate_changes
                    warnings += rate_warnings

                if lighthouse_data is not None:
                    lh_changes, lh_warnings = build_lighthouse_compset_change_plan(
                        lighthouse_data, wb, sheet, hotel_sel
                    )
                    changes += lh_changes
                    warnings += lh_warnings
            else:  # Forecast — current month (no Month Ending Forecast fill here)
                avail    = [s for s in FORECAST_SHEETS if s in wb.sheetnames]
                auto     = first_unhighlighted_forecast_sheet(wb, avail)
                sheet    = auto or avail[0]
                changes, warnings = build_forecast_change_plan(df, wb[sheet])
            all_plans[wb_type] = {
                "file_id":   file_id,
                "file_name": file_name,
                "wb_bytes":  wb_bytes,
                "sheet":     sheet,
                "changes":   changes,
                "warnings":  warnings,
            }

        # Next-month Forecast: only when checkbox is ticked
        if "Forecast" in wb_sels and forecast_next_month:
            next_month_dt = (datetime.date.today().replace(day=1) + datetime.timedelta(days=32)).replace(day=1)
            nm_result, nm_err = resolve_drive_workbook(svc, hotel_id, hotel_sel, "Forecast", month_date=next_month_dt)
            if nm_err:
                # Workbook not found — auto-create from master
                st.info(f"Next month Forecast not found — creating from master...")
                created_name, setup_err = setup_new_forecast_month(svc, hotel_id, hotel_sel, next_month_dt)
                if setup_err and not created_name:
                    st.warning(f"Next month Forecast: {setup_err}")
                    nm_result = None
                else:
                    if setup_err:
                        st.warning(setup_err)
                    else:
                        st.success(f"Created **{created_name}** for {next_month_dt.strftime('%B %Y')}.")
                    nm_result, nm_err = resolve_drive_workbook(svc, hotel_id, hotel_sel, "Forecast", month_date=next_month_dt)
                    if nm_err:
                        st.warning(f"Still could not find next month Forecast after creation: {nm_err}")
                        nm_result = None

            # Populate week 3/4's data into next month's WK1 regardless of
            # whether the file already existed or was just created above —
            # this used to live in the sibling `else` of `if nm_err`, so on
            # the very run that created the file, population never ran at
            # all (the new copy was created with dates only, no OTB data).
            if nm_result:
                nm_file_id, nm_file_name = nm_result
                nm_bytes = drive_download(svc, nm_file_id)
                nm_wb    = openpyxl.load_workbook(io.BytesIO(nm_bytes), data_only=False)
                nm_avail = [s for s in FORECAST_SHEETS if s in nm_wb.sheetnames]
                nm_auto  = first_unhighlighted_forecast_sheet(nm_wb, nm_avail)
                nm_sheet = nm_auto or nm_avail[0]
                nm_changes, nm_warnings = build_next_month_forecast_plan(df, nm_wb[nm_sheet])
                # Month Ending Forecast table — fill Budget + LY from next month's ROB
                nm_is_wk1 = (nm_sheet == FORECAST_SHEETS[0])
                if nm_is_wk1:
                    nm_rob_result, _ = resolve_drive_workbook(svc, hotel_id, hotel_sel, "ROB",
                                                               month_date=next_month_dt)
                    if nm_rob_result:
                        nm_rob_wb = openpyxl.load_workbook(
                            io.BytesIO(drive_download(svc, nm_rob_result[0])), data_only=True)
                        extra, extra_warn = build_forecast_change_plan(
                            df, nm_wb[nm_sheet], rob_wb=nm_rob_wb, is_wk1=True)
                        # Only keep the Month Ending Forecast entries from extra
                        nm_changes += [c for c in extra if "Month End Forecast" in c.get("label", "")]
                        nm_warnings += extra_warn
                all_plans["Forecast (next month)"] = {
                    "file_id":   nm_file_id,
                    "file_name": nm_file_name,
                    "wb_bytes":  nm_bytes,
                    "sheet":     nm_sheet,
                    "changes":   nm_changes,
                    "warnings":  nm_warnings,
                }

        return all_plans


    def _snapshot_changes(wb, sheet_name, changes):
        """Return {(sheet, row, col): original_value} for every cell in the change plan."""
        ws = wb[sheet_name]
        return {
            (sheet_name, ch["row"], ch["col"]): ws.cell(ch["row"], ch["col"]).value
            for ch in changes
            if not ch.get("skip_reason")
        }


    def apply_and_upload(svc, all_plans):
        saved, errors = [], []
        undo_snapshot = {}  # cumulative snapshot across all workbooks
        for wb_type, plan in all_plans.items():
            try:
                wb_apply = openpyxl.load_workbook(io.BytesIO(plan["wb_bytes"]), data_only=False)
                # Snapshot originals BEFORE writing
                snap = _snapshot_changes(wb_apply, plan["sheet"], plan["changes"])
                undo_snapshot[wb_type] = {
                    "file_id":   plan["file_id"],
                    "file_name": plan["file_name"],
                    "wb_bytes":  plan["wb_bytes"],   # clean pre-write bytes
                    "sheet":     plan["sheet"],
                    "cells":     snap,
                }
                if wb_type == "ROB":
                    apply_rob_changes(wb_apply, plan["sheet"], plan["changes"])
                elif wb_type == "Strategy Report":
                    apply_strategy_changes(wb_apply, plan["sheet"], plan["changes"])
                else:
                    apply_forecast_changes(wb_apply, plan["sheet"], plan["changes"])
                color_tab_done(wb_apply, plan["sheet"])
                strip_tables(wb_apply)
                out = io.BytesIO()
                wb_apply.save(out)
                drive_upload(svc, plan["file_id"], out.getvalue(), plan["file_name"])
                saved.append(plan["file_name"])
            except Exception as e:
                errors.append(f"{wb_type}: {e}")
        if saved:
            st.session_state["undo_snapshot"] = undo_snapshot
        return saved, errors


    def undo_all_changes(svc):
        """Restore every snapshotted cell to its original value and re-upload."""
        snapshot = st.session_state.get("undo_snapshot", {})
        if not snapshot:
            return [], ["No snapshot found — nothing to undo."]
        saved, errors = [], []
        for wb_type, info in snapshot.items():
            try:
                wb = openpyxl.load_workbook(io.BytesIO(info["wb_bytes"]), data_only=False)
                ws = wb[info["sheet"]]
                for (sheet, row, col), orig_val in info["cells"].items():
                    ws.cell(row, col).value = orig_val
                strip_tables(wb)
                out = io.BytesIO()
                wb.save(out)
                drive_upload(svc, info["file_id"], out.getvalue(), info["file_name"])
                saved.append(info["file_name"])
            except Exception as e:
                errors.append(f"{wb_type}: {e}")
        if saved:
            del st.session_state["undo_snapshot"]
        return saved, errors


    ready = drive_csv and wb_sels

    if test_mode:
        # ── Test mode: preview first, then confirm ────────────────────────────────
        if ready and st.button("Preview Changes", key="drive_preview"):
            try:
                svc = get_drive_service()
                df      = parse_bob_source(drive_csv) if drive_csv else None
                rate_df = parse_rate_csv(drive_rate_csv.read()) if drive_rate_csv else None
                lighthouse_data = (
                    parse_lighthouse_rates_xlsx(drive_lighthouse_xlsx.read())
                    if drive_lighthouse_xlsx else None
                )
                npu_compare_df = parse_bob_source(drive_npu_compare_csv) if drive_npu_compare_csv else None
                st.session_state["drive_plans"] = build_all_plans(
                    svc, hotel_sel, hotel_id_map.get(hotel_sel, ""), wb_sels,
                    df, rate_df, forecast_next_month, npu_compare_df, lighthouse_data
                )
                st.session_state["drive_hotel_sel"] = hotel_sel
            except Exception as e:
                st.error(f"Drive error: {e}")

        if "drive_plans" in st.session_state:
            all_plans = st.session_state["drive_plans"]
            for wb_type, plan in all_plans.items():
                st.subheader(wb_type)
                st.caption(f"File: **{plan['file_name']}** — Tab: **{plan['sheet']}**")
                for w in plan["warnings"]:
                    st.warning(w)
                ch = plan["changes"]
                will_write = [c for c in ch if not c["skip_reason"]]
                skipped    = [c for c in ch if c["skip_reason"]]
                c1, c2 = st.columns(2)
                c1.metric("Cells to update", len(will_write))
                c2.metric("Skipped",         len(skipped))
                st.dataframe([{
                    "Label":  c["label"],
                    "Row":    c["row"],
                    "Col":    c["col"],
                    "Value":  c["new_value"],
                    "Status": "✅ will write" if not c["skip_reason"] else f"⚠️ skip ({c['skip_reason']})",
                } for c in ch], use_container_width=True)

            if st.button("Confirm and Save All to Google Drive", key="drive_apply"):
                try:
                    saved, errors = apply_and_upload(get_drive_service(), all_plans)
                    for name in saved:
                        st.success(f"Saved **{name}** to Google Drive.")
                    for err in errors:
                        st.error(err)
                except Exception as e:
                    st.error(f"Drive error: {e}")
    else:
        # ── Normal mode: one click ────────────────────────────────────────────────
        if ready and st.button("Upload Data to Workbooks", key="drive_go", type="primary"):
            try:
                svc = get_drive_service()
                df      = parse_bob_source(drive_csv) if drive_csv else None
                rate_df = parse_rate_csv(drive_rate_csv.read()) if drive_rate_csv else None
                lighthouse_data = (
                    parse_lighthouse_rates_xlsx(drive_lighthouse_xlsx.read())
                    if drive_lighthouse_xlsx else None
                )
                npu_compare_df = parse_bob_source(drive_npu_compare_csv) if drive_npu_compare_csv else None
                with st.spinner("Updating workbooks in Google Drive..."):
                    all_plans = build_all_plans(
                        svc, hotel_sel, hotel_id_map.get(hotel_sel, ""), wb_sels,
                        df, rate_df, forecast_next_month, npu_compare_df, lighthouse_data
                    )
                    saved, errors = apply_and_upload(svc, all_plans)
                for name in saved:
                    st.success(f"Saved **{name}** to Google Drive.")
                for err in errors:
                    st.error(err)
            except Exception as e:
                st.error(f"Drive error: {e}")

    # ── Undo button (shown whenever a snapshot exists) ────────────────────────────
    if "undo_snapshot" in st.session_state:
        st.divider()
        undo_col, _ = st.columns([2, 5])
        with undo_col:
            if st.button("↩ Undo Last Upload", key="undo_all", type="secondary", use_container_width=True):
                try:
                    with st.spinner("Restoring original values..."):
                        saved, errors = undo_all_changes(get_drive_service())
                    for name in saved:
                        st.success(f"Restored **{name}** to original state.")
                    for err in errors:
                        st.error(err)
                except Exception as e:
                    st.error(f"Undo error: {e}")


with tab_ancillary:
    st.subheader("Monthly Ancillary Revenue Report Builder")
    st.caption(
        "Builds the monthly report from the universal template. After reviewing "
        "the result, you can save the month directly into the hotel's existing "
        "Canary and/or SNT Report workbook in Drive."
    )

    ar_properties = [p['display'] for p in ANCILLARY_PROPERTY_PROFILES.values()]
    # Remove duplicate display names while preserving order.
    ar_properties = list(dict.fromkeys(ar_properties))
    ar_property = st.selectbox("Property", ar_properties, key="ar_monthly_property")
    ar_profile, ar_key = ancillary_profile(ar_property)

    ar_month_date = st.date_input(
        "Report month",
        value=datetime.date.today().replace(day=1),
        key="ar_monthly_report_month",
    )
    ar_month_dt = datetime.datetime(ar_month_date.year, ar_month_date.month, 1)

    local_template = Path(__file__).resolve().with_name(ANCILLARY_TEMPLATE_FILENAME)
    ar_template_upload = None
    if local_template.exists():
        st.success(f"Using bundled template: {ANCILLARY_TEMPLATE_FILENAME}")
    else:
        st.info(
            f"Add **{ANCILLARY_TEMPLATE_FILENAME}** to the GitHub repo beside app.py "
            "for permanent use. For testing, upload it here."
        )
        ar_template_upload = st.file_uploader(
            "Ancillary Report Builder template workbook",
            type=["xlsx"],
            key="ar_template_upload",
        )

    st.markdown("**Current-year source files**")
    ar_addon = st.file_uploader(
        "SNT Add On Production",
        type=["csv", "xlsx"],
        key="ar_monthly_addon",
    )
    ar_upsell = st.file_uploader(
        "SNT Upsell By Day/User",
        type=["csv", "xlsx"],
        key="ar_monthly_upsell",
    )

    st.markdown(f"**STLY source — {ar_profile.get('stlySource')}**")
    ar_stly_addon = ar_stly_upsell = ar_canary_history = None
    if ar_profile.get('stlySource') == 'SNT':
        ar_stly_addon = st.file_uploader(
            f"{ar_month_dt.year - 1} SNT Add On Production",
            type=["csv", "xlsx"],
            key="ar_stly_addon",
        )
        ar_stly_upsell = st.file_uploader(
            f"{ar_month_dt.year - 1} SNT Upsell By Day/User",
            type=["csv", "xlsx"],
            key="ar_stly_upsell",
        )
    else:
        ar_canary_history = st.file_uploader(
            "Historical Canary Upsells",
            type=["csv", "xlsx"],
            key="ar_canary_history",
        )

    ar_staff = st.file_uploader(
        "Canary Message Count by Staff (optional)",
        type=["csv", "xlsx"],
        key="ar_staff_counts",
    )

    st.markdown("**SNT Journal revenue**")
    ar_journal_values = []
    journal_cols = st.columns(2)
    for i, journal in enumerate(ar_profile.get('journal', [])):
        with journal_cols[i % 2]:
            v = st.number_input(
                journal['label'],
                min_value=-1000000.0,
                max_value=1000000.0,
                value=0.0,
                step=1.0,
                key=f"ar_journal_{ar_key}_{i}",
                help=f"Main report line: {journal['report']}",
            )
            ar_journal_values.append(v)

    ar_stly_journal_values = []
    if ar_profile.get('stlySource') == 'SNT' and ar_profile.get('stlyJournal'):
        st.markdown(f"**{ar_month_dt.year - 1} STLY Journal revenue**")
        stly_cols = st.columns(2)
        for i, journal in enumerate(ar_profile.get('journal', [])[:2]):
            with stly_cols[i % 2]:
                v = st.number_input(
                    f"STLY — {journal['report']}",
                    min_value=-1000000.0,
                    max_value=1000000.0,
                    value=0.0,
                    step=1.0,
                    key=f"ar_stly_journal_{ar_key}_{i}",
                )
                ar_stly_journal_values.append(v)

    st.divider()
    st.markdown("### Canary Messaging Overview")
    st.caption(
        "Enter the monthly Canary Insights values here. Percentage fields use "
        "normal percentage points — enter 5.0 for 5%, not 0.05."
    )

    current_col, stly_col = st.columns(2)

    with current_col:
        st.markdown(f"**{ar_month_dt.strftime('%b').upper()} {ar_month_dt.year}**")
        msg_total = st.number_input(
            "Total Messages", value=0.0, key="ar_msg_total"
        )
        msg_guest = st.number_input(
            "Guest Messages", value=0.0, key="ar_msg_guest"
        )
        msg_hotel = st.number_input(
            "Hotel Messages", value=0.0, key="ar_msg_hotel"
        )
        msg_pct_ui = st.number_input(
            "% Guests Messaged",
            min_value=0.0,
            max_value=100.0,
            value=0.0,
            step=0.1,
            key="ar_msg_pct",
        )
        resp_ui = st.number_input(
            "Response Rate %",
            min_value=0.0,
            max_value=100.0,
            value=0.0,
            step=0.1,
            key="ar_resp",
        )
        avg = st.number_input(
            "Average Minutes to Respond", value=0.0, key="ar_avg"
        )
        med = st.number_input(
            "Median Minutes to Respond", value=0.0, key="ar_med"
        )

    with stly_col:
        st.markdown(f"**STLY — {ar_month_dt.year - 1}**")
        stly_total = st.number_input(
            "STLY Total Messages", value=0.0, key="ar_stly_total"
        )
        stly_guest = st.number_input(
            "STLY Guest Messages", value=0.0, key="ar_stly_guest"
        )
        stly_hotel = st.number_input(
            "STLY Hotel Messages", value=0.0, key="ar_stly_hotel"
        )
        stly_pct_ui = st.number_input(
            "STLY % Guests Messaged",
            min_value=0.0,
            max_value=100.0,
            value=0.0,
            step=0.1,
            key="ar_stly_pct",
        )
        stly_resp_ui = st.number_input(
            "STLY Response Rate %",
            min_value=0.0,
            max_value=100.0,
            value=0.0,
            step=0.1,
            key="ar_stly_resp",
        )
        stly_avg = st.number_input(
            "STLY Average Minutes to Respond",
            value=0.0,
            key="ar_stly_avg",
        )
        stly_med = st.number_input(
            "STLY Median Minutes to Respond",
            value=0.0,
            key="ar_stly_med",
        )

    msg_pct = msg_pct_ui / 100.0
    resp = resp_ui / 100.0
    stly_pct = stly_pct_ui / 100.0
    stly_resp = stly_resp_ui / 100.0

    engagement_rows = []
    with st.expander("Engagement Rate points (optional)"):
        st.caption(
            "Add up to eight Canary engagement-rate points for the monthly table."
        )
        for i in range(8):
            c_date, c_rate = st.columns([2, 1])
            with c_date:
                e_date = st.date_input(
                    f"Date {i + 1}",
                    value=None,
                    key=f"ar_eng_date_{i}",
                )
            with c_rate:
                e_rate = st.number_input(
                    f"Rate % {i + 1}",
                    min_value=0.0,
                    max_value=100.0,
                    value=0.0,
                    step=0.1,
                    key=f"ar_eng_rate_{i}",
                )
            if e_date is not None:
                engagement_rows.append(
                    (
                        datetime.datetime.combine(e_date, datetime.time()),
                        e_rate / 100.0,
                    )
                )

    ar_messaging={
        'msgTotal':msg_total,'msgGuest':msg_guest,'msgHotel':msg_hotel,'msgGuestPct':msg_pct,
        'responseRate':resp,'avgResponse':avg,'medianResponse':med,
        'stlyMsgTotal':stly_total,'stlyMsgGuest':stly_guest,'stlyMsgHotel':stly_hotel,
        'stlyMsgGuestPct':stly_pct,'stlyResponseRate':stly_resp,'stlyAvgResponse':stly_avg,'stlyMedianResponse':stly_med,
    }

    required_ok = ar_addon is not None and ar_upsell is not None
    if ar_profile.get('stlySource') == 'SNT':
        required_ok = required_ok and ar_stly_addon is not None and ar_stly_upsell is not None
    else:
        required_ok = required_ok and ar_canary_history is not None
    template_ok = local_template.exists() or ar_template_upload is not None

    if st.button(
        "Build Ancillary Revenue Report",
        type="primary",
        key="ar_build_monthly",
        disabled=not (required_ok and template_ok),
    ):
        try:
            with st.spinner("Building ancillary report..."):
                template_bytes = (
                    local_template.read_bytes()
                    if local_template.exists()
                    else ar_template_upload.getvalue()
                )
                ar_output, ar_summary = ancillary_build_monthly_report(
                    template_bytes=template_bytes,
                    property_name=ar_property,
                    report_month=ar_month_dt,
                    addon_file=ar_addon,
                    upsell_file=ar_upsell,
                    stly_addon_file=ar_stly_addon,
                    stly_upsell_file=ar_stly_upsell,
                    canary_history_file=ar_canary_history,
                    staff_file=ar_staff,
                    journal_values=ar_journal_values,
                    stly_journal_values=ar_stly_journal_values,
                    messaging=ar_messaging,
                    engagement=engagement_rows,
                )
                st.session_state['ar_monthly_output'] = ar_output
                st.session_state['ar_monthly_filename'] = (
                    f"{ar_month_dt.strftime('%b').upper()} {ar_month_dt.year} "
                    f"Ancillary Revenue - {ar_property}.xlsx"
                )
                st.session_state['ar_monthly_summary'] = ar_summary
            st.success("Report built. Review the summary below, then download the workbook for validation.")
        except Exception as e:
            st.error(f"Ancillary report build error: {e}")

    if 'ar_monthly_output' in st.session_state:
        summary=st.session_state.get('ar_monthly_summary',{})
        main=summary.get('mainRows',[]); stly=summary.get('stly',{}); variance=summary.get('variance',[])
        c1,c2,c3=st.columns(3)
        c1.metric("Current Revenue", f"${sum(_ar_num(x.get('revenue')) or 0 for x in main):,.2f}")
        c2.metric("STLY Revenue", f"${sum(_ar_num(x.get('approved')) or 0 for x in stly.get('rows',[])):,.2f}")
        c3.metric("YoY Variance", f"${sum(_ar_num(x.get('variance')) or 0 for x in variance):,.2f}")
        with st.expander("Preview current-year revenue rows"):
            st.dataframe(pd.DataFrame(main),use_container_width=True)
        download_col, drive_col = st.columns(2)

        with download_col:
            st.download_button(
                "Download Ancillary Revenue Report",
                data=st.session_state['ar_monthly_output'],
                file_name=st.session_state['ar_monthly_filename'],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="ar_download_monthly",
                use_container_width=True,
            )

        with drive_col:
            if st.button(
                f"Save {ar_month_dt.strftime('%b').upper()} Sheet to Drive",
                key="ar_save_drive",
                type="primary",
                use_container_width=True,
            ):
                try:
                    svc = get_drive_service()

                    discovered = dict(get_hotels_from_drive())
                    drive_label = ANCILLARY_DRIVE_HOTEL_MAP.get(ar_key)
                    hotel_id = discovered.get(drive_label, "")

                    if not drive_label:
                        raise ValueError(
                            f"No Drive hotel mapping is configured for {ar_property}."
                        )
                    if not hotel_id:
                        raise ValueError(
                            f"Could not find {drive_label}'s Revenue Reports folders "
                            "with the current Drive connection."
                        )

                    target, target_err = ancillary_find_drive_report(
                        svc,
                        hotel_id,
                        drive_label,
                        ar_month_dt,
                    )
                    if target_err or not target:
                        raise ValueError(target_err or "Ancillary Drive report not found.")

                    original_bytes = drive_download(svc, target["file_id"])
                    merged_bytes, month_sheet = ancillary_insert_report_sheet(
                        original_bytes,
                        st.session_state['ar_monthly_output'],
                        ar_month_dt,
                        destination_name=target["file_name"],
                    )

                    # Keep one-click undo until the next save.
                    st.session_state["ar_drive_undo"] = {
                        "file_id": target["file_id"],
                        "file_name": target["file_name"],
                        "bytes": original_bytes,
                    }

                    drive_upload(
                        svc,
                        target["file_id"],
                        merged_bytes,
                        target["file_name"],
                    )

                    st.session_state["ar_drive_last_target"] = {
                        "file_name": target["file_name"],
                        "folder_name": target["folder_name"],
                        "sheet_name": month_sheet,
                    }

                    st.success(
                        f"Saved **{month_sheet}** inside **{target['file_name']}** "
                        f"in **{target['folder_name']}**."
                    )

                except Exception as e:
                    st.error(f"Could not save ancillary report to Drive: {e}")

        if st.session_state.get("ar_drive_last_target"):
            t = st.session_state["ar_drive_last_target"]
            st.caption(
                f"Last Drive save: **{t['file_name']}** → "
                f"sheet **{t['sheet_name']}**"
            )

        if st.session_state.get("ar_drive_undo"):
            if st.button(
                "↩ Undo Last Ancillary Drive Save",
                key="ar_drive_undo_btn",
            ):
                try:
                    undo = st.session_state["ar_drive_undo"]
                    drive_upload(
                        get_drive_service(),
                        undo["file_id"],
                        undo["bytes"],
                        undo["file_name"],
                    )
                    st.session_state.pop("ar_drive_undo", None)
                    st.session_state.pop("ar_drive_last_target", None)
                    st.success("Restored the ancillary workbook to its previous version.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Could not undo ancillary Drive save: {e}")

with tab_ooo:
    st.caption(
        "Pulls the Sell-Out Efficiency Report straight from Drive, adds a bright-green "
        "summary tab — right after the Report tab (2nd tab on the sheet), named e.g. "
        "'JUL 2026' — totaling each property's End. OOO Rooms across every daily tab in "
        "the selected month, with an ADR column pulled from that property's own ROB "
        "(most recently-finalized week, current month falling back to previous month), "
        "and writes it back to Drive. The file is never re-saved through openpyxl, so the "
        "macro buttons on every existing tab are left completely untouched."
    )

    today = datetime.date.today()
    ooo_cur_month_dt  = today.replace(day=1)
    ooo_prev_month_dt = (ooo_cur_month_dt - datetime.timedelta(days=1)).replace(day=1)
    ooo_next_month_dt = (ooo_cur_month_dt + datetime.timedelta(days=32)).replace(day=1)
    ooo_month_options = {
        ooo_prev_month_dt.strftime("%B %Y"): ooo_prev_month_dt,
        ooo_cur_month_dt.strftime("%B %Y"):  ooo_cur_month_dt,
        ooo_next_month_dt.strftime("%B %Y"): ooo_next_month_dt,
    }
    ooo_month_labels = list(ooo_month_options.keys())
    # Default to the previous (just-completed) month — e.g. on Aug 1st this
    # defaults to July — since current/next are only there for override.
    ooo_sel_label = st.selectbox(
        "Month to summarize", ooo_month_labels,
        index=ooo_month_labels.index(ooo_prev_month_dt.strftime("%B %Y")),
        key="ooo_month_sel",
    )
    ooo_target_dt = ooo_month_options[ooo_sel_label]

    if st.button("Preview Monthly OOO Report", key="ooo_preview", type="primary"):
        try:
            svc = get_drive_service()
            ooo_file_id, ooo_file_name = find_ooo_report_file(svc)
            if not ooo_file_id:
                st.error(ooo_file_name)
            else:
                ooo_bytes = drive_download(svc, ooo_file_id)
                ooo_wb = openpyxl.load_workbook(io.BytesIO(ooo_bytes), data_only=True, read_only=True)
                ooo_months = list_ooo_available_months(ooo_wb)
                match = next((m for m in ooo_months
                              if m[0] == ooo_target_dt.year and m[1] == ooo_target_dt.month), None)
                if not match:
                    st.error(f"No dated tabs found for {ooo_target_dt.strftime('%B %Y')} in {ooo_file_name}.")
                else:
                    ooo_year, ooo_month, ooo_sheet_names = match
                    order, totals, days_included, skipped = build_ooo_monthly_totals(
                        ooo_wb, ooo_year, ooo_month, ooo_sheet_names)
                    if not order:
                        st.error(
                            "None of that month's tabs matched the expected layout "
                            "('Property' header in row 7 col C, 'End...OOO' in row 7 col J) — "
                            "nothing was totaled, so no report was built."
                        )
                    else:
                        with st.spinner("Looking up ADR for each property from its ROB..."):
                            hotels_for_adr = get_hotels_from_drive()
                            if not hotels_for_adr:
                                # A silent [] can also be a cached transient Drive
                                # failure (5-min TTL) — clear and retry once before
                                # concluding nothing is shared.
                                get_hotels_from_drive.clear()
                                hotels_for_adr = get_hotels_from_drive()
                            adr_lookup = build_ooo_adr_lookup(svc, order, hotels_for_adr, ooo_year, ooo_month)
                        st.session_state["ooo_hotels_seen"] = [h[0] for h in hotels_for_adr]
                        st.session_state["ooo_file_id"]       = ooo_file_id
                        st.session_state["ooo_file_name"]     = ooo_file_name
                        st.session_state["ooo_bytes"]         = ooo_bytes
                        st.session_state["ooo_year"]          = ooo_year
                        st.session_state["ooo_month"]         = ooo_month
                        st.session_state["ooo_order"]         = order
                        st.session_state["ooo_totals"]        = totals
                        st.session_state["ooo_days_included"] = days_included
                        st.session_state["ooo_skipped"]       = skipped
                        st.session_state["ooo_adr_lookup"]    = adr_lookup
                        matched = sum(1 for adr, _ in adr_lookup.values() if adr is not None)
                        st.success(f"Ready to summarize **{ooo_target_dt.strftime('%B %Y')}** "
                                   f"from **{ooo_file_name}** ({days_included} daily reports found, "
                                   f"ADR matched for {matched}/{len(order)} properties).")
        except Exception as e:
            st.error(f"Preview error: {e}")

    if "ooo_order" in st.session_state:
        order      = st.session_state["ooo_order"]
        totals     = st.session_state["ooo_totals"]
        adr_lookup = st.session_state["ooo_adr_lookup"]
        hotels_seen = st.session_state.get("ooo_hotels_seen", [])
        with st.expander(f"Hotel folders this app can see in Drive ({len(hotels_seen)})"):
            st.write(", ".join(hotels_seen) if hotels_seen else
                     "None — no hotel folders are shared with this environment's service account.")
        if st.session_state.get("ooo_skipped"):
            st.caption(f"Skipped tabs that didn't match the expected layout: "
                       f"{', '.join(st.session_state['ooo_skipped'])}")
        st.dataframe(
            [{"Property": p, "Total End. OOO Rooms": totals[p],
              "ADR": adr_lookup[p][0],
              "Revenue": round(adr_lookup[p][0] * totals[p], 2) if adr_lookup[p][0] is not None else None,
              "ADR note": adr_lookup[p][1] or ""} for p in order],
            use_container_width=True,
        )
        if st.button("Apply to Google Drive", key="ooo_apply", type="primary"):
            try:
                svc = get_drive_service()
                new_bytes, tab_name, err = inject_ooo_monthly_sheet(
                    st.session_state["ooo_bytes"], st.session_state["ooo_year"],
                    st.session_state["ooo_month"], order, totals,
                    st.session_state["ooo_days_included"], adr_lookup)
                if err:
                    st.error(err)
                else:
                    drive_upload(svc, st.session_state["ooo_file_id"], new_bytes,
                                 st.session_state["ooo_file_name"])
                    st.success(f"Added **{tab_name}** to **{st.session_state['ooo_file_name']}**.")
                    for key in ["ooo_file_id", "ooo_file_name", "ooo_bytes", "ooo_year", "ooo_month",
                                "ooo_order", "ooo_totals", "ooo_days_included", "ooo_skipped",
                                "ooo_adr_lookup", "ooo_hotels_seen"]:
                        st.session_state.pop(key, None)
            except Exception as e:
                st.error(f"Apply error: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# P&L Spreadsheet
# ══════════════════════════════════════════════════════════════════════════════
with tab_pl:
    st.divider()
    st.header("P&L Spreadsheet")
    st.caption(
        "Upload one Operating Statement (.xls) per hotel per year. The hotel name "
        "and year are read out of each report header, so filenames do not matter. "
        "Ten files for one hotel gives you ten years."
    )

    import hotel_pl_tool as PL
    st.caption("Builder version: **{}**".format(getattr(PL, "__version__", "unknown")))

    pl_files = st.file_uploader(
        "Operating statements (.xls)", type=["xls"],
        accept_multiple_files=True, key="pl_uploads",
    )

    if not pl_files:
        st.info("Drop your operating statements above to get started.")
    else:
        pl_parsed, pl_problems, pl_rows, pl_asof = {}, [], [], {}
        for f in pl_files:
            try:
                h, y, lines, asof = PL.parse_statement(data=f.getvalue(), name=f.name)
                used = "yes"
                prev = pl_asof.get((h, y))
                if prev is not None and prev >= asof:
                    used = "no - older as-of than another file for this year"
                else:
                    if prev is not None:
                        used = "yes - replaced an earlier as-of for this year"
                    pl_parsed.setdefault(h, {})[y] = lines
                    pl_asof[(h, y)] = asof
                pl_rows.append({"File": f.name, "Hotel": h, "Year": y,
                                "As of": "{}/{}".format(asof[0], asof[1]),
                                "Lines read": len(lines), "Used": used})
            except Exception as exc:
                pl_problems.append("{}: {}".format(f.name, exc))
                pl_rows.append({"File": f.name, "Hotel": "COULD NOT READ",
                                "Year": None, "As of": "", "Lines read": 0,
                                "Used": "no"})

        # always show what every file resolved to - this is where surprises show up
        with st.expander("Files read  ({} uploaded, {} hotel(s) found)".format(
                len(pl_files), len(pl_parsed)), expanded=len(pl_parsed) != 1):
            st.dataframe(pd.DataFrame(pl_rows), use_container_width=True, hide_index=True)
            if len(pl_parsed) > 1:
                st.warning(
                    "More than one hotel name was found, so the years are split "
                    "between them. If these are all the same property, the report "
                    "headers differ - check the Hotel column above."
                )

        for msg in pl_problems:
            st.warning("Skipped - " + msg)

        if not pl_parsed:
            st.error("No statements could be read.")
        else:
            hotel_pick = st.selectbox("Hotel", sorted(pl_parsed), key="pl_hotel")
            per_year = pl_parsed[hotel_pick]
            years = sorted(per_year)
            years = years[-PL.MAX_YEARS:]
            per_year = {y: per_year[y] for y in years}
            st.success("**{}** - {} year(s): {}".format(
                hotel_pick, len(years), ", ".join(str(y) for y in years)))
            if len(years) == 1:
                st.info(
                    "Each operating statement covers **one year**, so one file gives "
                    "one column. To get 10 years, upload 10 statements for this hotel "
                    "- one per year end. Check the 'Files read' table above to confirm "
                    "every file was picked up and resolved to the year you expect."
                )

            WANT = ["Total Revenue", "Room", "Food & Beverage", "Miscellaneous",
                    "Rental Income", "Operating Profit or Loss",
                    "Net Income or Loss", "A.D.R.", "Occupancy", "REV PAR"]
            recs = []
            for yr in years:
                row = {"Year": yr}
                for ln in per_year[yr]:
                    if ln.page == "Summary" and ln.label in WANT and ln.label not in row:
                        row[ln.label] = ln.act
                        row[ln.label + " (Budget)"] = ln.bud
                recs.append(row)
            pl_df = pd.DataFrame(recs).set_index("Year")

            latest = years[-1]
            mcols = st.columns(4)

            def _pl_metric(col, label, field, money=True):
                if field not in pl_df.columns:
                    return
                act = pl_df.loc[latest, field]
                bkey = field + " (Budget)"
                bud = pl_df.loc[latest, bkey] if bkey in pl_df.columns else None
                delta = None
                if bud not in (None, 0):
                    delta = "{:+.1%} vs budget".format((act - bud) / bud)
                shown = "${:,.0f}".format(act) if money else "{:,.2f}".format(act)
                col.metric(label, shown, delta)

            _pl_metric(mcols[0], "Total Revenue {}".format(latest), "Total Revenue")
            _pl_metric(mcols[1], "Operating Profit {}".format(latest),
                       "Operating Profit or Loss")
            _pl_metric(mcols[2], "Net Income {}".format(latest), "Net Income or Loss")
            _pl_metric(mcols[3], "ADR {}".format(latest), "A.D.R.", money=False)

            with st.expander("Show the underlying numbers"):
                st.dataframe(pl_df.style.format("{:,.2f}"), use_container_width=True)

            st.divider()
            if st.button("Build P&L Workbook", key="pl_build", type="primary"):
                try:
                    buf = io.BytesIO()
                    PL.build_workbook(hotel_pick, per_year, buf, {})
                    safe_name = re.sub(r"[^A-Za-z0-9 -]", "", hotel_pick).strip() or "Hotel"
                    st.download_button(
                        "Download P&L Workbook",
                        data=buf.getvalue(),
                        file_name=safe_name + " - P&L.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="pl_dl",
                    )
                    st.success("Built. Tabs: Summary, Rooms, Food, Beverage, "
                               "Miscellaneous, Fixed Expenses.")
                except Exception as exc:
                    st.error("Build error: {}".format(exc))


# ══════════════════════════════════════════════════════════════════════════════
# 1-Year Projection
# ══════════════════════════════════════════════════════════════════════════════
with tab_projection:
    st.divider()
    # Streamlit runs every tab body on every interaction, so importing the
    # projector here would pull altair and XlsxWriter — about 55 MB — into
    # memory on each page view, whether or not anyone opens this tab. This app
    # already runs close to the limit on Streamlit Cloud, so the import waits
    # behind a click and only happens for someone actually using the tool.
    if not st.session_state.get("projector_open"):
        st.header("1-Year Projection")
        st.caption(
            "Day-by-day rooms and ADR budget for the year ahead, built from a "
            "segmentation pivot export."
        )
        if st.button("Open the Budget Projector", key="projector_open_btn",
                     type="primary"):
            st.session_state["projector_open"] = True
            st.rerun()
    else:
        try:
            from projector import ui as projector_ui
            projector_ui.render()
        except Exception as exc:
            st.error(f"1-Year Projection failed to load: {exc}")
            st.caption(
                "It needs altair, numpy and XlsxWriter — check they installed "
                "with the rest of requirements.txt."
            )


# Everything the inactive section drew went into this placeholder; clear it so
# only the open section reaches the page. Must stay the last statement in the
# file — anything added after it would render into the void.
_offstage.empty()
